#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

from api.core.config import settings  # noqa: E402
from scripts.hybrid_seed_loader import (  # noqa: E402
    ensure_metadata,
    normalize_database_url,
    normalize_name,
    q,
    upsert_source,
    upsert_trial_result,
    upsert_variety,
)


SOURCE_CODE = "USER_XLSX_CORN_DEMO_2025"
DEFAULT_XLSX_PATH = "/Users/nikitasuvorov/Downloads/20260115_Общая_таблица_демо_кукурузы.xlsx"
SHEET_NAME = "БДА Капитал"


def clean_value(value: Any) -> Any:
    if value in (None, "", "н/д.", "н/д", "-", "—"):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def as_float(value: Any) -> float | None:
    value = clean_value(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def build_record_hash(payload: dict[str, Any]) -> str:
    ordered = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(ordered.encode("utf-8")).hexdigest()


def iter_rows(xlsx_path: str) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    headers = [str(item).strip() if item is not None else "" for item in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[idx]: clean_value(value) for idx, value in enumerate(values)}
        row["__row_number__"] = row_number
        rows.append(row)
    return rows


def build_trial_row(row: dict[str, Any], source_url: str) -> dict[str, Any] | None:
    hybrid_name = row.get("Гибрид")
    yield_14 = as_float(row.get("Урожайность на зерно при 14% влажности, ц/га"))
    if not hybrid_name or yield_14 is None:
        return None

    payload = {
        "provider": "User Upload",
        "source_code": SOURCE_CODE,
        "crop_code": "corn",
        "season_year": int(row.get("Год") or 2025),
        "macro_region": row.get("Регион допуска"),
        "variety_name": str(hybrid_name).strip(),
        "subject_rf": row.get("Субъект РФ"),
        "district": row.get("Район"),
        "sowing_date_raw": row.get("Дата посева"),
        "harvest_date_raw": row.get("Дата уборки"),
        "plant_density_ths_per_ha": None,
        "harvest_moisture_pct": as_float(row.get("Влажность при уборке на зерно, %")),
        "standard_moisture_pct": 14.0,
        "yield_standard_c_ha": yield_14,
        "source_url": source_url,
        "source_page": row["__row_number__"],
        "raw_text": f"{row.get('Субъект РФ') or ''} | {row.get('Район') or ''} | {hybrid_name}",
        "payload": row,
    }
    payload["record_hash"] = build_record_hash(payload)
    return payload


def main() -> None:
    database_url = normalize_database_url(os.getenv("DATABASE_URL", settings.DATABASE_URL))
    xlsx_path = os.getenv("HYBRID_XLSX_PATH", DEFAULT_XLSX_PATH)
    source_url = f"file://{xlsx_path}"

    engine = create_engine(database_url, pool_pre_ping=True)
    ensure_metadata(engine)

    rows = iter_rows(xlsx_path)
    loaded_trials = 0
    skipped_rows = 0

    with Session(engine) as db:
        source_id = upsert_source(
            db,
            code=SOURCE_CODE,
            provider="User Upload",
            source_type="excel_upload",
            name="Corn demo table 2025",
            source_url=source_url,
            crop_code="corn",
            season_year=2025,
        )
        for row in rows:
            trial_row = build_trial_row(row, source_url)
            if trial_row is None:
                skipped_rows += 1
                continue
            manufacturer = str(row.get("Производитель") or "Unknown").strip()
            variety_id = upsert_variety(
                db,
                crop_code="corn",
                product_type="hybrid",
                name_raw=str(row["Гибрид"]).strip(),
                manufacturer_norm=manufacturer,
            )
            upsert_trial_result(db, trial_row, source_id, variety_id)

            db.execute(
                text(
                    """
                    update hybrids.trial_results
                    set payload = coalesce(payload, '{}'::jsonb) || cast(:extra_payload as jsonb)
                    where record_hash = :record_hash
                    """
                ),
                {
                    "record_hash": trial_row["record_hash"],
                    "extra_payload": json.dumps(
                        {
                            "excel_row_number": row["__row_number__"],
                            "hybrid_name_norm": normalize_name(str(row["Гибрид"]).strip()),
                            "manufacturer_raw": row.get("Производитель"),
                            "settlement_name": row.get("Населенный пункт"),
                            "farm_name": row.get("Название хозяйства"),
                            "fao_raw": row.get("ФАО"),
                            "fao_group_raw": row.get("Группы ФАО"),
                            "soil_zone": row.get("Почвенная зона"),
                            "moisture_zone": row.get("Зона увлажнения"),
                            "executor": row.get("Исполнитель"),
                            "gross_yield_kg": row.get("Валовый сбор с участка, кг"),
                            "yield_harvest_c_ha": row.get("Урожайность на зерно при уборочной влажности, ц/га"),
                            "silage_yield_c_ha": row.get("Урожайность на силос при уборочной влажности, ц/га"),
                            "silage_moisture_pct": row.get("Влажность при уборке на силос, %"),
                            "cornage_yield_c_ha": row.get("Урожайность на корнаж при уборочной влажности, ц/га"),
                            "cornage_moisture_pct": row.get("Влажность при уборке на корнаж, %"),
                            "dry_matter_pct": row.get("Содержание сухого вещества, %"),
                            "starch_pct_dm": row.get("Содержание крахмала в сухом веществе, %"),
                            "starch_yield_c_ha": row.get("Выход крахмала, ц/га"),
                        },
                        ensure_ascii=False,
                    ),
                },
            )
            loaded_trials += 1
        db.commit()

        summary = db.execute(
            text(
                """
                select
                    count(*) as total_rows,
                    count(*) filter (where crop_code = 'corn' and season_year = 2025) as corn_2025_rows,
                    round((avg(yield_standard_c_ha) filter (where crop_code = 'corn' and season_year = 2025 and source_id = :source_id))::numeric, 2) as avg_yield_2025
                from hybrids.trial_results
                where source_id = :source_id
                """
            ),
            {"source_id": source_id},
        ).mappings().one()
        print(
            "[hybrids] excel append summary",
            {
                "xlsx_path": xlsx_path,
                "loaded_trials": loaded_trials,
                "skipped_rows": skipped_rows,
                **dict(summary),
            },
        )


if __name__ == "__main__":
    main()
