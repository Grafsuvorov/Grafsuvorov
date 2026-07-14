#!/usr/bin/env python3
from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

from api.core.config import settings  # noqa: E402

DEFAULT_SOURCE_XLSX = Path("/Users/nikitasuvorov/Downloads/20260115_Общая_таблица_демо_кукурузы.xlsx")
DEFAULT_OUTPUT_XLSX = ROOT / "reports" / "20260115_Общая_таблица_демо_кукурузы_merged.xlsx"
SOURCE_SOURCE_CODE = "USER_XLSX_CORN_DEMO_2025"
SOURCE_SOURCE_LABEL = "20260115_Общая_таблица_демо_кукурузы.xlsx"
SOURCE_SHEET_NAME = "БДА Капитал"

MAIN_COLUMNS = [
    "№",
    "Год",
    "Ответственный",
    "Регион допуска",
    "Субъект РФ",
    "Район",
    "Населенный пункт",
    "Почвенная зона",
    "Зона увлажнения",
    "Количество испытаний",
    "Название хозяйства",
    "Размер хозяйства, га",
    "Форма",
    "Исполнитель",
    "Телефон",
    "Почта",
    "Гибрид",
    "ФАО",
    "Группы ФАО",
    "Производитель",
    "Дата посева",
    "Предшественник",
    "Урожайность предшественника, ц/га",
    "Основная обработка почвы",
    "Предпосевная обработка",
    "Дата всходов",
    "Площадь участка, га ",
    "Дата уборки",
    "Валовый сбор с участка, кг",
    "Урожайность на зерно при уборочной влажности, ц/га",
    "Влажность при уборке на зерно, %",
    "Урожайность на зерно при 14% влажности, ц/га",
    "Урожайность на силос при уборочной влажности, ц/га",
    "Влажность при уборке на силос, %",
    "Урожайность на корнаж при уборочной влажности, ц/га",
    "Влажность при уборке на корнаж, %",
    "Содержание сухого вещества, %",
    "Содержание крахмала в сухом веществе, %",
    "Выход крахмала, ц/га",
    "Источник",
]

PRODUCT_COLUMNS = [
    "№",
    "Источник провайдер",
    "Источник тип",
    "Источник URL",
    "Источник код",
    "Культура",
    "Год",
    "Гибрид",
    "Производитель",
    "ФАО",
    "Группа спелости",
    "Стандартная влажность, %",
    "Белок, %",
    "Масло, %",
    "Крахмал, %",
    "Мин. урожайность, ц/га",
    "Макс. урожайность, ц/га",
    "Источник",
]


def normalize_database_url(url: str) -> str:
    return url.replace("host.docker.internal", "127.0.0.1")


def clean_value(value: Any) -> Any:
    if value in (None, "", "н/д.", "н/д", "-", "—"):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def load_source_workbook(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[SOURCE_SHEET_NAME]
    headers = [str(item).strip() if item is not None else "" for item in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[idx]: clean_value(value) for idx, value in enumerate(values)}
        if not any(v is not None and v != "" for v in row.values()):
            continue
        rows.append(row)

    df = pd.DataFrame(rows)
    for column in MAIN_COLUMNS:
        if column not in df.columns:
            df[column] = None

    df["Источник"] = SOURCE_SOURCE_LABEL
    return df[MAIN_COLUMNS]


def fetch_df(engine, sql: str) -> pd.DataFrame:
    return pd.read_sql_query(text(sql), engine)


def build_trial_db_frame(engine) -> pd.DataFrame:
    trial_df = fetch_df(
        engine,
        f"""
        select
          tr.id as row_id,
          tr.season_year as year,
          null::text as responsible,
          coalesce(s.macro_region, tr.macro_region, '') as region_admit,
          coalesce(tr.subject_rf, '') as subject_rf,
          coalesce(tr.district, '') as district,
          coalesce(tr.payload->>'settlement_name', tr.payload->>'site_name', '') as settlement,
          coalesce(tr.payload->>'soil_zone', '') as soil_zone,
          coalesce(tr.payload->>'moisture_zone', '') as moisture_zone,
          1 as trial_count,
          coalesce(tr.payload->>'farm_name', '') as farm_name,
          coalesce(tr.payload->>'farm_area_ha', '') as farm_area_ha,
          coalesce(tr.payload->>'legal_form', '') as legal_form,
          coalesce(tr.payload->>'executor', '') as executor,
          coalesce(tr.payload->>'phone', '') as phone,
          coalesce(tr.payload->>'email', '') as email,
          v.name_raw as hybrid_name,
          coalesce(ts.fao::text, tr.payload->>'fao_raw', '') as fao,
          coalesce(ts.maturity_label, tr.payload->>'fao_group_raw', '') as fao_group,
          v.manufacturer_norm as manufacturer,
          coalesce(tr.sowing_date_raw, tr.payload->>'sowing_date_raw', '') as sowing_date,
          coalesce(tr.payload->>'preceder', '') as preceder,
          coalesce(tr.payload->>'preceder_yield_c_ha', '') as preceder_yield_c_ha,
          coalesce(tr.payload->>'primary_tillage', '') as primary_tillage,
          coalesce(tr.payload->>'pre_sowing_tillage', '') as pre_sowing_tillage,
          coalesce(tr.payload->>'emergence_date', '') as emergence_date,
          coalesce(tr.payload->>'plot_area_ha', '') as plot_area_ha,
          coalesce(tr.harvest_date_raw, tr.payload->>'harvest_date_raw', '') as harvest_date,
          coalesce(tr.payload->>'gross_yield_kg', '') as gross_yield_kg,
          coalesce(tr.payload->>'yield_harvest_c_ha', '') as yield_harvest_c_ha,
          coalesce(tr.harvest_moisture_pct::text, tr.payload->>'harvest_moisture_pct', '') as harvest_moisture_pct,
          tr.yield_standard_c_ha as yield_standard_c_ha,
          coalesce(tr.payload->>'silage_yield_c_ha', '') as silage_yield_c_ha,
          coalesce(tr.payload->>'silage_moisture_pct', '') as silage_moisture_pct,
          coalesce(tr.payload->>'cornage_yield_c_ha', '') as cornage_yield_c_ha,
          coalesce(tr.payload->>'cornage_moisture_pct', '') as cornage_moisture_pct,
          coalesce(tr.payload->>'dry_matter_pct', '') as dry_matter_pct,
          coalesce(tr.payload->>'starch_pct_dm', tr.payload->>'starch_pct', '') as starch_pct_dm,
          coalesce(tr.payload->>'starch_yield_c_ha', '') as starch_yield_c_ha,
          case
            when s.code = '{SOURCE_SOURCE_CODE}' then '{SOURCE_SOURCE_LABEL}'
            else concat(s.provider, ' | ', s.source_type, ' | ', s.code)
          end as source_label
        from hybrids.trial_results tr
        join hybrids.sources s on s.id = tr.source_id
        join hybrids.varieties v on v.id = tr.variety_id
        left join hybrids.trait_snapshots ts
          on ts.source_id = tr.source_id and ts.variety_id = tr.variety_id
        where s.code <> '{SOURCE_SOURCE_CODE}'
        order by tr.season_year nulls last, s.provider, s.source_type, tr.crop_code, tr.subject_rf nulls last, tr.district nulls last, v.name_raw
        """
    )
    trial_df = trial_df.rename(
        columns={
            "row_id": "№",
            "year": "Год",
            "responsible": "Ответственный",
            "region_admit": "Регион допуска",
            "subject_rf": "Субъект РФ",
            "district": "Район",
            "settlement": "Населенный пункт",
            "soil_zone": "Почвенная зона",
            "moisture_zone": "Зона увлажнения",
            "trial_count": "Количество испытаний",
            "farm_name": "Название хозяйства",
            "farm_area_ha": "Размер хозяйства, га",
            "legal_form": "Форма",
            "executor": "Исполнитель",
            "phone": "Телефон",
            "email": "Почта",
            "hybrid_name": "Гибрид",
            "fao": "ФАО",
            "fao_group": "Группы ФАО",
            "manufacturer": "Производитель",
            "sowing_date": "Дата посева",
            "preceder": "Предшественник",
            "preceder_yield_c_ha": "Урожайность предшественника, ц/га",
            "primary_tillage": "Основная обработка почвы",
            "pre_sowing_tillage": "Предпосевная обработка",
            "emergence_date": "Дата всходов",
            "plot_area_ha": "Площадь участка, га ",
            "harvest_date": "Дата уборки",
            "gross_yield_kg": "Валовый сбор с участка, кг",
            "yield_harvest_c_ha": "Урожайность на зерно при уборочной влажности, ц/га",
            "harvest_moisture_pct": "Влажность при уборке на зерно, %",
            "yield_standard_c_ha": "Урожайность на зерно при 14% влажности, ц/га",
            "silage_yield_c_ha": "Урожайность на силос при уборочной влажности, ц/га",
            "silage_moisture_pct": "Влажность при уборке на силос, %",
            "cornage_yield_c_ha": "Урожайность на корнаж при уборочной влажности, ц/га",
            "cornage_moisture_pct": "Влажность при уборке на корнаж, %",
            "dry_matter_pct": "Содержание сухого вещества, %",
            "starch_pct_dm": "Содержание крахмала в сухом веществе, %",
            "starch_yield_c_ha": "Выход крахмала, ц/га",
            "source_label": "Источник",
        }
    )
    return trial_df[MAIN_COLUMNS]


def build_product_frame(engine) -> pd.DataFrame:
    product_df = fetch_df(
        engine,
        f"""
        select
          ts.id as "№",
          s.provider as "Источник провайдер",
          s.source_type as "Источник тип",
          s.source_url as "Источник URL",
          s.code as "Источник код",
          ts.crop_code as "Культура",
          s.season_year as "Год",
          v.name_raw as "Гибрид",
          v.manufacturer_norm as "Производитель",
          ts.fao as "ФАО",
          ts.maturity_label as "Группа спелости",
          ts.standard_moisture_pct as "Стандартная влажность, %",
          ts.protein_pct as "Белок, %",
          ts.oil_pct as "Масло, %",
          ts.starch_pct as "Крахмал, %",
          ts.yield_min_c_ha as "Мин. урожайность, ц/га",
          ts.yield_max_c_ha as "Макс. урожайность, ц/га",
          case
            when s.code = '{SOURCE_SOURCE_CODE}' then '{SOURCE_SOURCE_LABEL}'
            else concat(s.provider, ' | ', s.source_type, ' | ', s.code)
          end as "Источник"
        from hybrids.trait_snapshots ts
        join hybrids.sources s on s.id = ts.source_id
        join hybrids.varieties v on v.id = ts.variety_id
        order by s.provider, s.source_type, ts.crop_code, s.season_year nulls last, v.name_raw
        """
    )
    return product_df[PRODUCT_COLUMNS]


def set_sheet_format(ws, freeze_at: str = "A2") -> None:
    ws.freeze_panes = freeze_at
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    widths = {
        "№": 10,
        "Год": 10,
        "Ответственный": 16,
        "Регион допуска": 22,
        "Субъект РФ": 20,
        "Район": 18,
        "Населенный пункт": 20,
        "Почвенная зона": 20,
        "Зона увлажнения": 20,
        "Количество испытаний": 16,
        "Название хозяйства": 24,
        "Размер хозяйства, га": 18,
        "Форма": 14,
        "Исполнитель": 16,
        "Телефон": 16,
        "Почта": 22,
        "Гибрид": 22,
        "ФАО": 12,
        "Группы ФАО": 14,
        "Производитель": 18,
        "Дата посева": 14,
        "Предшественник": 18,
        "Урожайность предшественника, ц/га": 20,
        "Основная обработка почвы": 20,
        "Предпосевная обработка": 20,
        "Дата всходов": 14,
        "Площадь участка, га ": 18,
        "Дата уборки": 14,
        "Валовый сбор с участка, кг": 18,
        "Урожайность на зерно при уборочной влажности, ц/га": 22,
        "Влажность при уборке на зерно, %": 16,
        "Урожайность на зерно при 14% влажности, ц/га": 22,
        "Урожайность на силос при уборочной влажности, ц/га": 22,
        "Влажность при уборке на силос, %": 16,
        "Урожайность на корнаж при уборочной влажности, ц/га": 22,
        "Влажность при уборке на корнаж, %": 16,
        "Содержание сухого вещества, %": 16,
        "Содержание крахмала в сухом веществе, %": 18,
        "Выход крахмала, ц/га": 16,
        "Источник": 38,
        "Источник провайдер": 18,
        "Источник тип": 16,
        "Источник URL": 42,
        "Источник код": 22,
        "Культура": 14,
        "Группа спелости": 16,
        "Стандартная влажность, %": 16,
        "Белок, %": 12,
        "Масло, %": 12,
        "Крахмал, %": 12,
        "Мин. урожайность, ц/га": 18,
        "Макс. урожайность, ц/га": 18,
    }

    for idx, column in enumerate([cell.value for cell in ws[1]], start=1):
        width = widths.get(str(column), 18)
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    source_path = Path(os.getenv("AGRO_SOURCE_XLSX", str(DEFAULT_SOURCE_XLSX)))
    output_path = Path(os.getenv("AGRO_OUTPUT_XLSX", str(DEFAULT_OUTPUT_XLSX)))

    database_url = normalize_database_url(os.getenv("DATABASE_URL", settings.DATABASE_URL))
    engine = create_engine(database_url, pool_pre_ping=True)

    main_df = pd.concat([build_trial_db_frame(engine), load_source_workbook(source_path)], ignore_index=True)
    product_df = build_product_frame(engine)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
      main_df.to_excel(writer, sheet_name=SOURCE_SHEET_NAME, index=False)
      product_df.to_excel(writer, sheet_name="PRODUCT_LAYER", index=False)

    wb = load_workbook(output_path)
    set_sheet_format(wb[SOURCE_SHEET_NAME], freeze_at="A2")
    set_sheet_format(wb["PRODUCT_LAYER"], freeze_at="A2")
    wb.save(output_path)

    print(output_path)


if __name__ == "__main__":
    main()
