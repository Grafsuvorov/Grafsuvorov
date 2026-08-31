from __future__ import annotations

from fastapi import FastAPI, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from typing import List, Dict, Tuple, Set, Union, Any
from collections import Counter, deque, defaultdict
from pydantic import BaseModel
import os
import yaml

from datetime import datetime
from sqlalchemy import create_engine, text, bindparam
from typing import Optional
from pathlib import Path
from openpyxl import load_workbook
import traceback
from datetime import datetime, date, timedelta
from decimal import Decimal
import time
from typing import List, Dict, Tuple
from datetime import datetime
from sqlalchemy import text
from sqlalchemy.exc import OperationalError
import re
import json
import hashlib
import subprocess
import tempfile
from html import unescape
from itertools import combinations
from zoneinfo import ZoneInfo
from io import BytesIO
import threading
from uuid import uuid4
from posixpath import join as posix_join

from PIL import Image, ImageDraw, ImageFont

from fastapi import APIRouter, HTTPException



from .config import (
    TABLE_LOADING_HISTORY,
    TABLE_ENTITIES_META,
    TABLE_TABLES_META,
    TABLE_TABLE_COMPARE,
    TABLE_YT_SLA,
    TABLE_YTREK_INCIDENTS,
    TABLE_TABLES_META_CLICK,
    TABLE_DATA_QUALITY,
    TABLE_RELEASE_LOG,
    TABLE_RELEASE_OBJECTS,
    TABLE_YT_ISSUE_SNAPSHOT,
    TABLE_YT_ISSUE_CUSTOM,
    TABLE_YT_ISSUE_TIMELINE,
    TABLE_YT_ISSUE_WORKLOG,
    TABLE_YT_ISSUE_COMMENT,
    TABLE_CLICK_LOAD_RUN,
    TABLE_CLICK_LOAD_STAGE,
    CLICK_META_DIR,
    DEV_CLICK_META_DIR,
    ENTITY_META_DIR,
    DEV_ENTITY_META_DIR,
    DBT_MANIFEST_DIR,
    DBT_LOGS_DATABASE_URL,
    TABLE_DBT_MODEL_CATALOG,
    TABLE_DBT_MODEL_LOG,
    TABLE_DBT_RUN_LOG,
    DEV_COPY_SCHEMA_SYNC_DAG_ID,
    TABLE_SAY_COMPARE_GP_METADATA_LOG,
    TABLE_SAY_COMPARE_GP_METADATA_PROD_VS_DEV,
    ADMIN_CICD_SCRIPT,
    YTRACK_ISSUE_URL,
    DATABASE_URL,
    DEV_DATABASE_URL,
    DEV_META_DEPLOY_BASE_DIR,
    DEV_META_DEPLOY_HOST,
    DEV_META_DEPLOY_PASSWORD,
    DEV_META_DEPLOY_PORT,
    DEV_META_DEPLOY_SSH_KEY_PATH,
    DEV_META_DEPLOY_STRICT_HOST_KEY,
    DEV_META_DEPLOY_USER,
    ENTITY_META_GIT_META_ROOT,
    ENTITY_META_GIT_REPO,
    CLICK_META_GIT_ROOT,
    META_WORKSPACE_ROOT,
    TABLE_APP_FEEDBACK,
    GITLAB_API_URL,
    ANALYST_GITLAB_PROJECT,
    GITLAB_PROJECT,
    GITLAB_SSL_VERIFY,
    GITLAB_TOKEN,
    AIRFLOW_DEV_BASE_URL,
    AIRFLOW_DEV_DAG_ID,
    AIRFLOW_DEV_USERNAME,
    AIRFLOW_DEV_PASSWORD,
    DEV_META_LOCK_TTL_MIN,
    YOUTRACK_URL,
    YOUTRACK_PROJECT_ID,
    YOUTRACK_PROJECT,
    YOUTRACK_TOKEN,
    YOUTRACK_QUEUE,
    YOUTRACK_ISSUE_TYPE,
    YOUTRACK_SSL_VERIFY,
    YOUTRACK_DEFAULT_ESTIMATE_MINUTES,
    YOUTRACK_ESTIMATE_FIELD_NAME,
    YOUTRACK_CARD_TYPE_FIELD_NAME,
    YOUTRACK_CARD_TYPE_VALUE,
    YOUTRACK_ASSIGNEE_FIELD_NAME,
    YOUTRACK_ASSIGNEE_QUERY,
)


from .services.admin import refresh_application_caches, run_ci_cd_script
from .services.entities import fetch_entities
from .services.dev_meta import (
    acquire_dev_meta_lock,
    assert_dev_meta_lock_owner,
    deploy_dev_meta_file,
    generate_dev_meta_yaml,
    get_airflow_dev_dag_status,
    get_dev_meta_files,
    get_dev_meta_status,
    read_dev_meta_file,
    read_remote_dev_meta_file,
    release_dev_meta_lock,
    save_dev_meta_file,
    stop_airflow_dag_run,
    trigger_airflow_dev_dag,
    trigger_airflow_parametrized_dag,
    validate_dev_meta_content,
)
from .services.entity_dev_meta import (
    _dump_yaml,
    create_entity_meta_mr,
    delete_entity_dev_meta_bundle,
    execute_entity_dev_meta_sql,
    get_entity_dev_meta_status,
    init_entity_dev_meta_bundle,
    list_entity_reference_rows,
    list_entity_dev_catalog,
    lock_entity_dev_meta,
    move_entity_dev_meta_bundle,
    save_entity_dev_meta_bundle,
    unlock_entity_dev_meta,
    validate_entity_dev_meta_bundle,
)
from .services.meta_workspace import (
    BranchRevisionConflictError,
    create_meta_workspace_branch,
    _build_branch_catalog,
    build_meta_workspace_branch_tree,
    create_meta_workspace_mr,
    list_meta_workspace_branches,
    read_meta_workspace_branch_gp_bundle,
    read_meta_workspace_branch_file,
    save_meta_workspace_branch_gp_bundle,
    save_meta_workspace_branch_file,
    sync_meta_workspace_branch,
    validate_meta_workspace_branch,
)
from .services.feedback import list_feedback, save_feedback
from .services.corp_ai import enhance_assistant_response
from .services.prototype_review import (
    CREATE_OBJECT_PATTERNS,
    DROP_TARGET_PATTERNS,
    TARGET_PATTERNS,
    _normalize_fqn,
    add_ytrack_issue_comment,
    create_ytrack_issue,
    execute_sql_review_items_in_dev,
    extract_sql_dependencies,
    infer_final_target,
    infer_review_targets,
    load_merge_request_sql_bundle,
    parse_prototype_task_text,
    query_dev_table_checks,
    validate_prototype_sql,
)
from .services.dbt_manifest import (
    build_dbt_fallback_card,
    get_dbt_graph_snapshot,
    get_dbt_manifest_model,
    get_dbt_table_catalog,
)
from .services.dbt_logs import get_dbt_model_run_history




from .auth import auth_middleware, init_auth, router as auth_router, get_current_user_from_request


app = FastAPI()
# CORS для взаимодействия с фронтом
app.add_middleware(
  CORSMiddleware,
  allow_origins=[
      "http://rgm-s-dwhapp01.hq.root.ad:15312",
      "http://rgm-s-dwhapp01.hq.root.ad",
  ],
  allow_credentials=True,
  allow_methods=["*"],
  allow_headers=["*"],
)

# Подключение
engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=1800)
dbt_logs_engine = (
    create_engine(DBT_LOGS_DATABASE_URL, pool_pre_ping=True, pool_recycle=1800)
    if DBT_LOGS_DATABASE_URL
    else None
)
dev_engine = create_engine(DEV_DATABASE_URL, pool_pre_ping=True, pool_recycle=1800) if DEV_DATABASE_URL else engine
from fastapi import APIRouter, HTTPException

router = APIRouter()
print("BOOT FILE:", __file__)
DEV_COPY_DAG_ID = "load_from_prod_to_dev"
DEV_COPY_TZ = ZoneInfo("Europe/Moscow")
DEV_COPY_ALLOWED_HOUR_START = 8
DEV_COPY_ALLOWED_HOUR_END = 21

init_auth()
app.middleware("http")(auth_middleware)
app.include_router(auth_router)

# admin ci_cd status (in-memory)
_ci_cd_status = {
    "last_run_at": None,
    "status": None,
    "return_code": None,
    "stdout": None,
    "stderr": None,
}

_prototype_review_jobs: dict[str, dict[str, Any]] = {}
_prototype_review_jobs_lock = threading.Lock()


class DevMetaFilePayload(BaseModel):
    schema_name: str
    file_name: str
    source: Optional[str] = "dev"


class DevMetaLockPayload(BaseModel):
    schema_name: str
    file_name: str


class DevMetaSavePayload(BaseModel):
    schema_name: str
    file_name: str
    content: str
    task_id: Optional[str] = None


class DevMetaDagPayload(BaseModel):
    schema_name: str
    file_name: str


class DevMetaDagStatusPayload(BaseModel):
    schema_name: str
    file_name: str
    dag_id: Optional[str] = None
    dag_run_id: str
    auto_unpaused: bool = False


class DevCopyDagPayload(BaseModel):
    source_table_schema: str
    source_table_name: str
    target_table_schema: str
    target_table_name: str
    where: Optional[str] = ""


class DevCopyDagStatusPayload(BaseModel):
    dag_run_id: str
    auto_unpaused: bool = False


class DevCopySchemaSyncPayload(BaseModel):
    run_mode: Optional[str] = "self"
    check_table_schema: Optional[str] = ""
    check_table_name: Optional[str] = ""


class DevCopySchemaSyncDagStatusPayload(BaseModel):
    dag_run_id: str
    auto_unpaused: bool = False


class DevCopySchemaSyncReportPayload(BaseModel):
    check_table_schema: str
    check_table_name: str


class DevMetaDeployPayload(BaseModel):
    schema_name: str
    file_name: str
    content: str
    task_id: Optional[str] = None


class DevMetaGeneratePayload(BaseModel):
    schema_name_gp: str
    object_name: str
    schema_name_click: str = "dm"
    greenplum_table_name: Optional[str] = None
    order_by: List[str]
    dag_tags: List[str]


class EntityMetaInitPayload(BaseModel):
    entity_name: str
    schema_name: str
    table_name: str
    key_attributes: Optional[List[str]] = None


class EntityMetaLockPayload(BaseModel):
    entity_name: str
    schema_name: str
    table_name: str


class EntityMetaSavePayload(BaseModel):
    entity_name: str
    schema_name: str
    table_name: str
    task_id: str
    key_attributes: Optional[List[str]] = None
    source_object_key: Optional[str] = None
    replica_entity_names: Optional[List[str]] = None
    yaml_content: str
    recreate_sql: str
    insert_sql: str
    truncate_sql: str = ""


class EntityMetaDeletePayload(BaseModel):
    entity_name: str
    schema_name: str
    table_name: str
    task_id: str


class EntityMetaMovePayload(BaseModel):
    source_entity_name: str
    source_schema_name: str
    source_table_name: str
    target_entity_name: str
    target_schema_name: str
    target_table_name: str
    task_id: str


class EntityMetaMrPayload(BaseModel):
    task_id: str
    release_branch: str = "main"


class EntityMetaRunSqlPayload(BaseModel):
    entity_name: str
    schema_name: str
    table_name: str
    sql_kind: str
    sql_text: str


class MetaWorkspaceMrPayload(BaseModel):
    task_id: str
    release_branch: str = "main"
    branch_name: Optional[str] = None


class MetaWorkspaceValidatePayload(BaseModel):
    branch_name: str
    base_branch: str


class MetaWorkspaceCreateBranchPayload(BaseModel):
    branch_name: str
    base_branch: str = "main"


class MetaWorkspaceSyncPayload(BaseModel):
    task_id: str
    branch_name: str
    base_branch: str


class MetaWorkspaceBranchFilePayload(BaseModel):
    branch_name: str
    file_path: str


class MetaWorkspaceBranchFileSavePayload(BaseModel):
    branch_name: str
    base_branch: str
    file_path: str
    content: str
    task_id: Optional[str] = None
    expected_revision: Optional[dict] = None


class FeedbackPayload(BaseModel):
    topic: str
    message: str
    contact_email: Optional[str] = None
    page_path: Optional[str] = None


class MetaWorkspaceBranchGpBundlePayload(BaseModel):
    branch_name: str
    entity_name: str
    schema_name: str
    table_name: str


class MetaWorkspaceBranchGpBundleSavePayload(BaseModel):
    branch_name: str
    base_branch: str
    task_id: Optional[str] = None
    entity_name: str
    schema_name: str
    table_name: str
    yaml_content: str
    recreate_sql: str
    insert_sql: str
    truncate_sql: str
    expected_revision: Optional[dict] = None


class AssistantContextPayload(BaseModel):
    page: Optional[str] = None
    schema: Optional[str] = None
    table: Optional[str] = None
    source: Optional[str] = "current"
    fqn: Optional[str] = None


class AssistantQueryPayload(BaseModel):
    question: str
    context: Optional[AssistantContextPayload] = None


class PrototypeReviewRunPayload(BaseModel):
    mr_input: str
    key_attributes: Optional[List[str]] = None
    create_issue: bool = False
    task_text: Optional[str] = None
    target_table_fqn: Optional[str] = None
    entity_name: Optional[str] = None
    issue_summary: Optional[str] = None
    load_mode: Optional[str] = None
    stand_dev: bool = True
    stand_prod: bool = True
    copy_to_clickhouse: Optional[bool] = None
    dependent_views: Optional[List[str]] = None
    linked_issues: Optional[List[str]] = None


class PrototypeReviewTableCheckPayload(BaseModel):
    mr_input: str
    target_fqn: str
    entity_name: Optional[str] = None
    key_attributes: Optional[List[str]] = None


class PrototypeReviewItemPayload(BaseModel):
    path: Optional[str] = None
    target_fqn: str
    entity_name: Optional[str] = None
    key_attributes: Optional[List[str]] = None
    clickhouse_keys: Optional[List[str]] = None
    dependent_views: Optional[List[str]] = None
    is_new: Optional[bool] = None
    object_type: Optional[str] = None
    duration_sec: Optional[float] = None
    row_count: Optional[int] = None
    duplicate_groups: Optional[int] = None
    dependencies: Optional[List[str]] = None
    impact_tables: Optional[List[Dict[str, Any]]] = None
    yaml_content: Optional[str] = None
    stand_dev: Optional[bool] = True
    stand_prod: Optional[bool] = True
    copy_to_clickhouse: Optional[bool] = None


class PrototypeReviewCreateIssuePayload(BaseModel):
    mr_input: str
    task_text: Optional[str] = None
    issue_summary: Optional[str] = None
    load_mode: Optional[str] = None
    stand_dev: bool = True
    stand_prod: bool = True
    copy_to_clickhouse: Optional[bool] = None
    linked_issues: Optional[List[str]] = None
    review_items: List[PrototypeReviewItemPayload]


def _require_dev_meta_role(request: Request):
    user = get_current_user_from_request(request)
    if not user or not getattr(user, "email", None):
        raise HTTPException(status_code=403, detail="Authentication required")
    return user


def _require_admin(request: Request):
    user = get_current_user_from_request(request)
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    return user


def _require_meta_workspace_role(request: Request):
    user = get_current_user_from_request(request)
    if not user or user.role not in {"admin", "engineer"}:
        raise HTTPException(status_code=403, detail="Engineer or admin role required")
    return user


def _optional_user(request: Request):
    try:
        return get_current_user_from_request(request)
    except Exception:
        return None


def _require_authenticated(request: Request):
    user = get_current_user_from_request(request)
    if not user or not getattr(user, "email", None):
        raise HTTPException(status_code=403, detail="Authentication required")
    return user


def _get_dev_copy_window_status():
    now = datetime.now(DEV_COPY_TZ)
    current_hour = now.hour
    allowed = DEV_COPY_ALLOWED_HOUR_START <= current_hour < DEV_COPY_ALLOWED_HOUR_END
    return {
        "timezone": "Europe/Moscow",
        "now": now.isoformat(),
        "allowed": allowed,
        "allowed_from": f"{DEV_COPY_ALLOWED_HOUR_START:02d}:00",
        "allowed_to": f"{DEV_COPY_ALLOWED_HOUR_END:02d}:00",
    }


def _assert_dev_copy_window():
    window = _get_dev_copy_window_status()
    if not window["allowed"]:
        raise HTTPException(
            status_code=400,
            detail=(
                "Запуск DEV copy DAG разрешен только с "
                f"{window['allowed_from']} до {window['allowed_to']} по Москве."
            ),
        )


def _dev_copy_author(user) -> str:
    return str(getattr(user, "username", None) or getattr(user, "email", None) or "").strip()


def _get_schema_sync_latest_run(*, run_user: str, table_schema: str, table_name: str):
    with dev_engine.begin() as conn:
        row = conn.execute(
            text(
                f"""
                select
                    run_id,
                    run_timestamp,
                    run_user,
                    layer_filter,
                    table_filter,
                    prod_snapshot_last_dttm,
                    is_prod_snapshot_actual,
                    state_code,
                    state_name
                from {TABLE_SAY_COMPARE_GP_METADATA_LOG}
                where coalesce(deleted_flag, false) = false
                  and coalesce(run_user, '') = :run_user
                  and coalesce(layer_filter, '') = :table_schema
                  and coalesce(table_filter, '') = :table_name
                order by coalesce(run_timestamp, dttm_inserted) desc, run_id desc
                limit 1
                """
            ),
            {
                "run_user": str(run_user or "").strip(),
                "table_schema": str(table_schema or "").strip(),
                "table_name": str(table_name or "").strip(),
            },
        ).mappings().first()
    return dict(row) if row else None


def _get_schema_sync_report_rows(*, run_id: int):
    with dev_engine.begin() as conn:
        rows = conn.execute(
            text(
                f"""
                select
                    table_schema,
                    table_name,
                    column_name,
                    column_position,
                    diff_code,
                    diff_name,
                    data_type_prod,
                    data_type_dev
                from {TABLE_SAY_COMPARE_GP_METADATA_PROD_VS_DEV}
                where coalesce(deleted_flag, false) = false
                  and run_id = :run_id
                order by
                    coalesce(table_schema, ''),
                    coalesce(table_name, ''),
                    column_position nulls last,
                    coalesce(column_name, '')
                """
            ),
            {"run_id": run_id},
        ).mappings().all()
    return [dict(row) for row in rows]





@app.get("/api/health")
def healthcheck():
    return {"status": "ok"}


@router.post("/api/admin/refresh-cache")
def refresh_cache(request: Request):
    user = get_current_user_from_request(request)
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    globals()["_cached_meta_index"] = None
    globals()["_cache_timestamp"] = 0
    globals()["_order_breaches_cache"] = None
    globals()["_order_breaches_ts"] = 0
    globals()["_graph_snapshot"] = None
    globals()["_graph_snapshot_ts"] = 0
    globals()["_graph_snapshot_hash"] = None
    globals()["_graph_cache"].clear()
    globals()["_graph_cache_ts"] = 0
    globals()["_graph_cache_meta_ts"] = 0
    globals()["_logic_audit_cache_payload"] = None
    globals()["_logic_audit_cache_ts"] = 0
    globals()["_assistant_index_cache"] = None
    globals()["_assistant_index_ts"] = 0
    globals()["_table_sizes_cache_payload"] = None
    globals()["_table_sizes_cache_cycle"] = None

    try:
        get_cached_meta_and_index()
        get_cached_order_breaches()
        get_graph_snapshot()
        _build_logic_audit_cache()
        get_cached_table_sizes()
    except Exception as exc:
        print("❌ refresh cache error:", exc)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось обновить кеш")

    return {"ok": True}


@router.post("/api/feedback")
def submit_feedback(payload: FeedbackPayload, request: Request):
    user = _optional_user(request)
    try:
        result = save_feedback(
            engine=engine,
            table_name=TABLE_APP_FEEDBACK,
            topic=payload.topic,
            message=payload.message,
            user_email=getattr(user, "email", "") if user else "",
            user_name=getattr(user, "username", "") if user else "",
            contact_email=payload.contact_email or (getattr(user, "email", "") if user else ""),
            page_path=payload.page_path or request.url.path,
            meta_json=json.dumps(
                {
                    "user_role": getattr(user, "role", None) if user else None,
                    "referer": request.headers.get("referer"),
                    "user_agent": request.headers.get("user-agent"),
                },
                ensure_ascii=False,
            ),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return result


@router.get("/api/admin/feedback")
def get_feedback(request: Request, days: int = 30, topic: str = "", limit: int = 200):
    user = get_current_user_from_request(request)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        items = list_feedback(
            engine=engine,
            table_name=TABLE_APP_FEEDBACK,
            days=days,
            topic=topic,
            limit=limit,
        )
    except Exception as exc:
        print("❌ feedback list error:", exc)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось загрузить обратную связь")
    return {
        "items": items,
        "days": max(1, min(int(days or 30), 365)),
        "topic": str(topic or "").strip(),
        "limit": max(1, min(int(limit or 200), 1000)),
    }


@router.post("/api/admin/run-ci-cd")
def run_ci_cd(request: Request):
    user = get_current_user_from_request(request)
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    script_path = Path(ADMIN_CICD_SCRIPT)
    if not script_path.is_absolute():
        script_path = (BASE_DIR / script_path).resolve()

    if not script_path.exists():
        raise HTTPException(status_code=404, detail=f"Скрипт не найден: {script_path}")
    if not script_path.is_file():
        raise HTTPException(status_code=400, detail="Путь скрипта должен указывать на файл")

    _ci_cd_status.update(
        {
            "last_run_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status": "running",
            "return_code": None,
            "stdout": None,
            "stderr": None,
        }
    )

    try:
        result = subprocess.run(
            ["bash", str(script_path)],
            capture_output=True,
            text=True,
            timeout=900,
            cwd=str(script_path.parent),
            env=os.environ.copy(),
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=500, detail="Скрипт выполняется слишком долго (timeout)")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось запустить скрипт: {exc}")

    response = {
        "status": "ok" if result.returncode == 0 else "failed",
        "return_code": result.returncode,
        "stdout": (result.stdout or "").strip()[:2000],
        "stderr": (result.stderr or "").strip()[:2000],
        "last_run_at": _ci_cd_status.get("last_run_at"),
    }
    _ci_cd_status.update(response)
    if result.returncode != 0:
        raise HTTPException(status_code=500, detail=response)
    return response


@router.post("/api/admin/assistant/query")
def assistant_query(payload: AssistantQueryPayload, request: Request):
    _require_authenticated(request)
    try:
        local_response = _assistant_answer(payload.question, payload.context)
        context_payload = None
        if payload.context:
            context_payload = (
                payload.context.model_dump()
                if hasattr(payload.context, "model_dump")
                else payload.context.dict()
            )
        response = enhance_assistant_response(
            question=payload.question,
            context=context_payload,
            local_response=local_response,
        ) or local_response
        return JSONResponse(content=response, media_type="application/json; charset=utf-8")
    except Exception as exc:
        print("❌ /api/admin/assistant/query error:", exc)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить ответ ассистента")


def _prototype_find_meta_by_fqn(target_fqn: Optional[str]) -> Optional[dict[str, Any]]:
    normalized = norm(target_fqn)
    if not normalized or "." not in normalized:
        return None
    schema_name, table_name = normalized.split(".", 1)
    all_meta, _ = get_cached_meta_and_index()
    for item in all_meta:
        if item.get("table_schema") == schema_name and item.get("table_name") == table_name:
            return item
    return None


def _prototype_impact_summary(target_fqn: Optional[str]) -> dict[str, Any]:
    normalized = norm(target_fqn)
    if not normalized:
        return {"tables": [], "entities": [], "count": 0}
    snapshot = _get_table_graph_context("current")
    table_edges = snapshot["table_graph"]["edges"]
    table_nodes = snapshot["table_graph"]["nodes"]
    table_fqn_map = snapshot.get("table_fqn_map") or {}
    if "." not in normalized:
        return {"tables": [], "entities": [], "count": 0}
    schema_name, table_name = normalized.split(".", 1)
    key = _resolve_table_key(table_nodes, schema_name, table_name, fqn_map=table_fqn_map)
    if key not in table_nodes:
        return {"tables": [], "entities": [], "count": 0}

    visited, depth_map, truncated = _traverse_forward(key, table_edges, depth=3, max_nodes=300)
    visited.discard(key)
    downstream_tables: list[dict[str, Any]] = []
    entity_counter: dict[str, int] = {}

    for node_id in sorted(visited, key=lambda item: (depth_map.get(item, 0), item)):
        node = table_nodes.get(node_id) or {}
        entity_name = str(node.get("entity_name") or node.get("entity") or "").strip()
        if entity_name:
            entity_counter[entity_name] = entity_counter.get(entity_name, 0) + 1
        downstream_tables.append(
            {
                "fqn": str(node.get("fqn") or f"{node.get('schema')}.{node.get('table')}").lower(),
                "schema": node.get("schema"),
                "table": node.get("table"),
                "entity_name": entity_name or None,
                "table_id": node.get("table_id"),
                "depth": depth_map.get(node_id, 0),
            }
        )
    entities = [
        {"entity_name": name, "tables_count": count}
        for name, count in sorted(entity_counter.items(), key=lambda item: (-item[1], item[0]))[:12]
    ]
    return {
        "tables": downstream_tables[:40],
        "entities": entities,
        "count": len(downstream_tables),
        "truncated": truncated,
    }


def _prototype_issue_description(
    *,
    mr: dict[str, Any],
    final_target: Optional[str],
    entity_name: Optional[str],
    key_attributes: list[str],
    dependencies: list[str],
    preparation: dict[str, Any],
    execution: list[dict[str, Any]],
    checks: dict[str, Any],
    impact: dict[str, Any],
    task_context: dict[str, Any],
    initiator: dict[str, Any],
) -> str:
    def _format_duration(seconds: Any) -> str:
        try:
            numeric = float(seconds or 0)
        except Exception:
            return "—"
        if numeric >= 60:
            return f"{numeric / 60.0:.2f} мин"
        return f"{numeric:.3f} сек"

    def _format_count(value: Any) -> str:
        try:
            numeric = int(value)
        except Exception:
            return "—" if value in (None, "") else str(value)
        return f"{numeric:,}".replace(",", " ")

    total_execution_sec = sum(float(item.get("duration_sec") or 0) for item in (execution or []))
    duplicate_groups = (
        checks.get("duplicate_groups")
        if checks.get("duplicate_groups") is not None
        else "не проверялось"
    )
    lines = [
        "## MR",
        f"**Email инициатора:** {initiator.get('email') or '—'}",
        f"**MR:** {mr.get('web_url') or '—'}",
        f"**Ветка:** {mr.get('source_branch') or '—'} -> {mr.get('target_branch') or '—'}",
        f"**Автор MR:** {mr.get('author') or '—'}",
        "",
        "## Витрина",
        f"**Витрина:** {final_target or 'не определена'}",
        f"**Сущность:** {entity_name or 'не определена'}",
        f"**Ключевые поля:** {', '.join(key_attributes) if key_attributes else 'не заданы'}",
        f"**Количество строк:** {_format_count(checks.get('row_count') if checks else None)}",
        f"**Кол-во дублей:** {_format_count(duplicate_groups)}",
    ]
    if task_context:
        lines.extend(
            [
                "",
                "## Параметры загрузки",
                f"**Предметная область:** {task_context.get('subject_area') or '—'}",
                f"**Режим обновления:** {task_context.get('load_mode') or '—'}",
                f"**Git ref:** {task_context.get('git_reference') or '—'}",
                f"**Дата релиза:** {task_context.get('release_date') or '—'}",
                f"**Время выполнения SQL:** {_format_duration(total_execution_sec)}",
            ]
        )
    if dependencies:
        lines.extend(["", "## Зависимости SQL"])
        lines.extend(f"- {item}" for item in dependencies[:40])
    impact_tables = impact.get("tables") or []
    if impact_tables:
        lines.extend(["", "## Потенциальное downstream-влияние"])
        lines.extend(
            f"- **{item.get('fqn')}** · {item.get('entity_name') or '—'}"
            for item in impact_tables[:25]
        )
    return "\n".join(lines)


def _prototype_item_needs_attention(item: dict[str, Any]) -> tuple[bool, list[str]]:
    missing: list[str] = []
    if item.get("is_new") and not str(item.get("entity_name") or "").strip():
        missing.append("сущность")
    object_type = str(item.get("object_type") or "TABLE").upper()
    if object_type == "TABLE" and not [str(value).strip() for value in (item.get("key_attributes") or []) if str(value).strip()]:
        missing.append("ключевые поля")
    return bool(missing), missing


def _prototype_multi_issue_description(
    *,
    mr: dict[str, Any],
    task_context: dict[str, Any],
    initiator: dict[str, Any],
    review_items: list[dict[str, Any]],
) -> str:
    def _format_duration(seconds: Any) -> str:
        try:
            numeric = float(seconds or 0)
        except Exception:
            return "—"
        if numeric >= 60:
            return f"{numeric / 60.0:.2f} мин"
        return f"{numeric:.3f} сек"

    def _format_count(value: Any) -> str:
        try:
            numeric = int(value)
        except Exception:
            return "—" if value in (None, "") else str(value)
        return f"{numeric:,}".replace(",", " ")

    lines = [
        "## MR",
        f"**Email инициатора:** {initiator.get('email') or '—'}",
        f"**MR:** {mr.get('web_url') or '—'}",
        f"**Ветка:** {mr.get('source_branch') or '—'} -> {mr.get('target_branch') or '—'}",
        f"**Автор MR:** {mr.get('author') or '—'}",
    ]
    if task_context:
        lines.extend(
            [
                "",
                "## Общие параметры",
                f"**Git ref:** {task_context.get('git_reference') or '—'}",
                f"**Дата релиза:** {task_context.get('release_date') or '—'}",
            ]
        )
    for index, item in enumerate(review_items, start=1):
        needs_attention, missing = _prototype_item_needs_attention(item)
        item_row_count = item.get("row_count")
        if item_row_count is None:
            item_row_count = (item.get("checks") or {}).get("row_count")
        item_duplicate_groups = item.get("duplicate_groups")
        if item_duplicate_groups is None:
            item_duplicate_groups = (item.get("checks") or {}).get("duplicate_groups")
        item_stands = [
            label for enabled, label in ((item.get("stand_dev"), "DEV"), (item.get("stand_prod"), "PROD")) if enabled
        ]
        lines.extend(
            [
                "",
                f"## Таблица {index}",
                f"**Витрина:** {item.get('target_fqn') or 'не определена'}",
                f"**Сущность:** {item.get('entity_name') or '—'}",
                f"**Статус объекта:** {'новая таблица' if item.get('is_new') else 'существующий объект'}",
                f"**Ключевые поля:** {', '.join(item.get('key_attributes') or []) or '—'}",
                f"**Количество строк:** {_format_count(item_row_count)}",
                f"**Кол-во дублей:** {_format_count(item_duplicate_groups)}",
                f"**Время выполнения SQL:** {_format_duration(item.get('duration_sec'))}",
            ]
        )
        if item_stands:
            lines.append(f"**Стенды:** {', '.join(item_stands)}")
        if item.get("copy_to_clickhouse"):
            lines.append("**ClickHouse:** требуется")
            lines.append(
                f"**Ключевые поля для загрузки в ClickHouse:** {', '.join(item.get('clickhouse_keys') or []) or '—'}"
            )
        if needs_attention:
            lines.append(f"**Нужно заполнить вручную:** {', '.join(missing)}")
        if item.get("dependencies"):
            lines.append("")
            lines.append("**Зависимости SQL**")
            lines.extend(f"- {value}" for value in (item.get("dependencies") or [])[:40])
        impact_tables = (item.get("impact") or {}).get("tables") or item.get("impact_tables") or []
        if impact_tables:
            lines.append("")
            lines.append("**Потенциальное downstream-влияние**")
            lines.extend(
                f"- **{row.get('fqn')}** · {row.get('entity_name') or '—'}"
                for row in impact_tables[:25]
            )
    return "\n".join(lines)


def _prototype_review_job_update(job_id: str, **fields: Any) -> None:
    with _prototype_review_jobs_lock:
        current = dict(_prototype_review_jobs.get(job_id) or {})
        current.update(fields)
        _prototype_review_jobs[job_id] = current


def _prototype_review_attachment_name(review_item: dict[str, Any]) -> str:
    target_fqn = str((review_item or {}).get("target_fqn") or "").strip().lower()
    path_value = str((review_item or {}).get("path") or "").strip()
    if target_fqn:
        safe_name = re.sub(r"[^a-z0-9._-]+", "_", target_fqn).strip("._-")
        if safe_name:
            return f"{safe_name}__generated_yaml.sql"
    if path_value:
        safe_name = re.sub(r"[^a-zA-Z0-9._/-]+", "_", path_value).split("/")[-1].strip("._-")
        if safe_name:
            return f"{safe_name}__generated_yaml.sql"
    return "prototype_review_generated_yaml.sql"


def _prototype_review_collect_target_sql(target_fqn: str, files: list[dict[str, Any]]) -> dict[str, str]:
    target_norm = str(target_fqn or "").strip().lower()
    recreate_parts: list[str] = []
    insert_parts: list[str] = []
    truncate_parts: list[str] = []

    for file_item in files or []:
        path_value = str(file_item.get("path") or "").strip().lower()
        sql_text = str(file_item.get("sql") or "").strip()
        if not sql_text:
            continue
        statements = [
            str(statement).strip()
            for statement in (file_item.get("statements") or [])
            if str(statement).strip()
        ]
        if "recreate" in path_value:
            recreate_parts.append(sql_text)
            continue
        if "insert" in path_value:
            insert_parts.append(sql_text)
            continue
        if "truncate" in path_value:
            truncate_parts.append(sql_text)
            continue
        statement_recreate: list[str] = []
        statement_insert: list[str] = []
        statement_truncate: list[str] = []
        for statement in statements or [sql_text]:
            normalized_targets: set[str] = set()
            for pattern in TARGET_PATTERNS:
                for match in pattern.finditer(statement):
                    normalized = _normalize_fqn(match.group(1))
                    if normalized:
                        normalized_targets.add(normalized)
            for pattern in CREATE_OBJECT_PATTERNS:
                for match in pattern.finditer(statement):
                    normalized = _normalize_fqn(match.group(1))
                    if normalized:
                        normalized_targets.add(normalized)
            for pattern in DROP_TARGET_PATTERNS:
                for match in pattern.finditer(statement):
                    normalized = _normalize_fqn(match.group(1))
                    if normalized == target_norm:
                        statement_recreate.append(statement)
            normalized_statement = _strip_sql_comments(statement).lower()
            if target_norm in normalized_targets:
                if re.search(r"\bcreate\s+(?:or\s+replace\s+)?(?:table|view)\b", normalized_statement):
                    statement_recreate.append(statement)
                if re.search(r"\binsert\s+into\b", normalized_statement):
                    statement_insert.append(statement)
            if (
                target_norm in normalized_targets
                and re.search(r"\binsert\s+overwrite(?:\s+table)?\b", normalized_statement)
            ):
                statement_insert.append(statement)
            if re.search(r"\b(?:truncate\s+table|truncate|delete\s+from)\b", normalized_statement) and target_norm in normalized_targets:
                statement_truncate.append(statement)

        if statement_recreate:
            recreate_parts.append("\n\n".join(statement_recreate))
        if statement_insert:
            insert_parts.append("\n\n".join(statement_insert))
        if statement_truncate:
            truncate_parts.append("\n\n".join(statement_truncate))

    return {
        "recreate_sql": "\n\n".join(part for part in recreate_parts if part.strip()),
        "insert_sql": "\n\n".join(part for part in insert_parts if part.strip()),
        "truncate_sql": "\n\n".join(part for part in truncate_parts if part.strip()),
    }


def _prototype_review_group_dependencies(dependencies: list[str]) -> dict[str, list[str]]:
    grouped: dict[str, set[str]] = {}
    for item in dependencies or []:
        value = str(item or "").strip().lower()
        if "." not in value:
            continue
        schema_name, table_name = value.split(".", 1)
        if not schema_name or not table_name:
            continue
        grouped.setdefault(schema_name, set()).add(table_name)
    return {
        schema_name: sorted(table_names)
        for schema_name, table_names in sorted(grouped.items())
    }


def _prototype_review_apply_yaml_dependencies(yaml_content: str, dependencies: list[str]) -> str:
    try:
        payload = yaml.safe_load(yaml_content) or {}
    except Exception:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}
    payload["depends_on"] = _prototype_review_group_dependencies(dependencies)
    return _dump_yaml(payload)


def _prototype_review_resolve_item(
    *,
    target_fqn: str,
    path_value: str = "",
    execution_row: Optional[dict[str, Any]] = None,
    known_schemas: Optional[set[str]] = None,
    file_item: Optional[dict[str, Any]] = None,
    related_files: Optional[list[dict[str, Any]]] = None,
    fallback_entity_name: str = "",
    key_attributes_override: Optional[list[str]] = None,
) -> dict[str, Any]:
    meta = _prototype_find_meta_by_fqn(target_fqn)
    schema_name, table_name = target_fqn.split(".", 1)
    yaml_bundle = None
    yaml_key_attributes: list[str] = []
    yaml_entity_name = None
    entity_name_seed = (
        str((meta or {}).get("entity_name") or "").strip()
        or str(fallback_entity_name or "").strip()
    )
    try:
        yaml_bundle = init_entity_dev_meta_bundle(
            engine=engine,
            base_dir=BASE_DIR,
            prod_root_value=ENTITY_META_DIR,
            dev_root_value=DEV_ENTITY_META_DIR,
            entity_name=entity_name_seed,
            schema_name=schema_name,
            table_name=table_name,
            key_attributes=list(key_attributes_override or []) or None,
        )
    except Exception:
        yaml_bundle = None
    if yaml_bundle:
        yaml_key_attributes = list(yaml_bundle.get("key_attributes") or [])
        yaml_entity_name = str(yaml_bundle.get("entity_name") or "").strip() or None
    detected_keys = list(key_attributes_override or []) or yaml_key_attributes or list((meta or {}).get("key_attributes") or [])
    entity_name = (
        str(fallback_entity_name or "").strip()
        or yaml_entity_name
        or str((meta or {}).get("entity_name") or "").strip()
        or None
    )
    click_idx = get_click_meta_index()
    click_meta = (click_idx.get("meta") or {}).get((schema_name.lower(), table_name.lower())) or (click_idx.get("meta") or {}).get((schema_name.lower(), _clean_table_name(table_name.lower())))
    clickhouse_keys = list(((click_meta or {}).get("order_by") or []))
    yaml_payload = None
    if yaml_bundle and yaml_bundle.get("yaml_content"):
        try:
            yaml_payload = yaml.safe_load(yaml_bundle.get("yaml_content")) or {}
        except Exception:
            yaml_payload = {}
    current_files = [item for item in (related_files or []) if isinstance(item, dict)]
    if not current_files and file_item:
        current_files = [file_item]
    if yaml_bundle and entity_name_seed and current_files:
        sql_bundle = _prototype_review_collect_target_sql(target_fqn, current_files)
        normalized = validate_entity_dev_meta_bundle(
            engine=engine,
            base_dir=BASE_DIR,
            prod_root_value=ENTITY_META_DIR,
            dev_root_value=DEV_ENTITY_META_DIR,
            entity_name=entity_name_seed,
            schema_name=schema_name,
            table_name=table_name,
            key_attributes=detected_keys,
            source_object_key=None,
            yaml_content=str(yaml_bundle.get("yaml_content") or ""),
            recreate_sql=sql_bundle.get("recreate_sql", ""),
            insert_sql=sql_bundle.get("insert_sql", ""),
            truncate_sql=sql_bundle.get("truncate_sql", ""),
            dev_database_url=DEV_DATABASE_URL,
        )
        normalized_bundle = normalized.get("normalized") or {}
        if normalized_bundle.get("yaml_content"):
            yaml_bundle["yaml_content"] = normalized_bundle.get("yaml_content")
        if isinstance(normalized_bundle.get("key_attributes"), list):
            yaml_bundle["key_attributes"] = normalized_bundle.get("key_attributes")
            yaml_key_attributes = list(normalized_bundle.get("key_attributes") or [])
        try:
            yaml_payload = yaml.safe_load(yaml_bundle.get("yaml_content") or "") or {}
        except Exception:
            yaml_payload = {}
        detected_keys = list(key_attributes_override or []) or yaml_key_attributes or list((meta or {}).get("key_attributes") or [])
    table_load_mode = str((yaml_payload or {}).get("table_load_mode") or (meta or {}).get("table_load_mode") or "").strip()
    dependencies: list[str] = []
    if current_files and path_value:
        dependencies = extract_sql_dependencies(
            current_files,
            known_schemas=known_schemas,
            exclude_fqns={target_fqn},
        )
    if yaml_bundle and yaml_bundle.get("yaml_content") is not None:
        yaml_bundle["yaml_content"] = _prototype_review_apply_yaml_dependencies(
            str(yaml_bundle.get("yaml_content") or ""),
            dependencies,
        )
        try:
            yaml_payload = yaml.safe_load(yaml_bundle.get("yaml_content") or "") or {}
        except Exception:
            yaml_payload = {}
    impact = _prototype_impact_summary(target_fqn)
    is_new = bool(yaml_bundle and yaml_bundle.get("source") == "new") or not meta
    checks = {"row_count": None, "duplicate_groups": None}
    current_execution = execution_row or {"status": "skipped", "duration_sec": 0.0}
    if str(current_execution.get("status") or "") == "ok" and str((meta or {}).get("table_type") or "TABLE").upper() == "TABLE":
        try:
            checks = query_dev_table_checks(
                dev_database_url=DEV_DATABASE_URL,
                target_fqn=target_fqn,
                key_attributes=detected_keys,
            )
        except Exception:
            checks = {"row_count": None, "duplicate_groups": None}
    item_warnings: list[str] = []
    if not detected_keys:
        item_warnings.append("Ключевые поля не найдены автоматически")
    if current_execution.get("status") == "error":
        item_warnings.append(f"Ошибка в файле `{path_value}`: {current_execution.get('error_message') or 'SQL не выполнился'}")
    if checks.get("duplicate_groups") not in (None, 0):
        item_warnings.append(f"Обнаружены дубли по ключу: {checks.get('duplicate_groups')}")
    requires_item_input, missing_fields = _prototype_item_needs_attention({
        "is_new": is_new,
        "entity_name": entity_name,
        "key_attributes": detected_keys,
    })
    return {
        "path": path_value,
        "target_fqn": target_fqn,
        "entity_name": entity_name,
        "load_mode": table_load_mode,
        "key_attributes": detected_keys,
        "auto_detected_key_attributes": detected_keys,
        "clickhouse_keys": clickhouse_keys,
        "dependencies": dependencies,
        "execution": current_execution,
        "duration_sec": float(current_execution.get("duration_sec") or 0.0),
        "checks": checks,
        "impact": impact,
        "yaml_bundle": yaml_bundle,
        "is_new": is_new,
        "stand_dev": True,
        "stand_prod": True,
        "copy_to_clickhouse": bool(clickhouse_keys),
        "requires_user_input": requires_item_input,
        "missing_fields": missing_fields,
        "warnings": item_warnings,
    }


def _prototype_review_yaml_repo_path(entity_name: str, schema_name: str, table_name: str) -> str:
    return posix_join(
        Path(ENTITY_META_GIT_META_ROOT).as_posix().strip("/"),
        str(entity_name or "").strip(),
        str(schema_name or "").strip(),
        str(table_name or "").strip(),
        "meta_data_file.yaml",
    )


def _prototype_review_build_result(
    payload: PrototypeReviewRunPayload,
    user,
    progress_callback=None,
) -> dict[str, Any]:
    bundle = load_merge_request_sql_bundle(
        gitlab_api_url=GITLAB_API_URL,
        gitlab_project=GITLAB_PROJECT,
        gitlab_token=GITLAB_TOKEN,
        gitlab_ssl_verify=GITLAB_SSL_VERIFY,
        mr_input=payload.mr_input,
        default_project=ANALYST_GITLAB_PROJECT or GITLAB_PROJECT,
    )
    files = bundle.get("files") or []
    parsed_task = parse_prototype_task_text(payload.task_text or "")
    sql_validation = validate_prototype_sql(files)
    all_meta, _ = get_cached_meta_and_index()
    known_schemas = {
        str(meta.get("table_schema") or "").strip().lower()
        for meta in all_meta
        if str(meta.get("table_schema") or "").strip()
    }
    validation_errors = list(sql_validation.get("errors") or [])
    validation_warnings = list(sql_validation.get("warnings") or [])
    review_targets = infer_review_targets(files)
    if not any(item.get("target_fqn") for item in review_targets):
        validation_errors.append("Не удалось определить целевые таблицы по SQL-файлам MR")

    status_reason = "; ".join(validation_errors) if validation_errors else ""
    task_context = {
        **parsed_task,
        "summary": (payload.issue_summary or "").strip() or parsed_task.get("summary"),
        "dependent_views": payload.dependent_views or parsed_task.get("dependent_views") or [],
        "linked_issues": payload.linked_issues or parsed_task.get("linked_issues") or [],
    }
    preparation_rows: list[dict[str, Any]] = []
    execution_rows: list[dict[str, Any]] = []
    if not validation_errors:
        preparation_rows, execution_rows = execute_sql_review_items_in_dev(
            dev_database_url=DEV_DATABASE_URL,
            files=files,
            review_targets=review_targets,
            progress_callback=progress_callback,
        )

    exec_by_target = {str(item.get("target_fqn") or ""): item for item in execution_rows}
    prep_by_target = {str(item.get("target_fqn") or ""): item for item in preparation_rows}
    review_items: list[dict[str, Any]] = []
    all_dependencies: list[str] = []
    dependency_seen: set[str] = set()
    requires_user_input = False
    for target_item in review_targets:
        target_fqn = str(target_item.get("target_fqn") or "").strip()
        if not target_fqn:
            validation_warnings.append(f"Для файла `{target_item.get('path')}` не удалось определить целевой объект")
            continue
        execution_row = exec_by_target.get(target_fqn) or {"status": "skipped", "duration_sec": 0.0}
        path_value = str(target_item.get("path") or "")
        related_paths = [str(value).strip() for value in (target_item.get("paths") or []) if str(value).strip()]
        related_files = [row for row in files if str(row.get("path") or "") in set(related_paths)]
        file_item = related_files[0] if related_files else {}
        dependencies = extract_sql_dependencies(
            related_files or [file_item],
            known_schemas=known_schemas,
            exclude_fqns={target_fqn},
        )
        for dep in dependencies:
            if dep not in dependency_seen:
                dependency_seen.add(dep)
                all_dependencies.append(dep)
        item_result = _prototype_review_resolve_item(
            target_fqn=target_fqn,
            path_value=path_value,
            execution_row=execution_row,
            known_schemas=known_schemas,
            file_item=file_item,
            related_files=related_files,
            fallback_entity_name=str(payload.entity_name or parsed_task.get("entity_name") or "").strip(),
        )
        item_result["object_type"] = str(target_item.get("object_type") or "TABLE").upper()
        item_result["preparation"] = prep_by_target.get(target_fqn) or {"status": "skipped"}
        item_result["dependencies"] = dependencies
        requires_user_input = requires_user_input or bool(item_result.get("requires_user_input"))
        review_items.append(item_result)
    execution_errors = [item for item in execution_rows if item.get("status") == "error"]
    if execution_errors:
        for item in execution_errors:
            validation_errors.append(f"Файл `{item.get('path')}`: {item.get('error_message') or 'SQL не выполнился'}")
    if any(item.get("warnings") for item in review_items):
        for item in review_items:
            for warning in item.get("warnings") or []:
                if warning not in validation_warnings:
                    validation_warnings.append(warning)
    status = "error" if validation_errors else ("warning" if validation_warnings else "ok")
    if not status_reason:
        status_reason = "; ".join(validation_warnings) if validation_warnings else "Проверки завершены. Проверьте блоки по всем таблицам и затем создайте задачу."
    return {
        "status": status,
        "mr": bundle.get("mr") or {},
        "files": [{"path": item.get("path"), "statements_count": len(item.get("statements") or [])} for item in files],
        "final_target": review_items[0].get("target_fqn") if review_items else None,
        "review_items": review_items,
        "dependencies": all_dependencies,
        "preparation": preparation_rows,
        "execution": execution_rows,
        "validation_errors": validation_errors,
        "validation_warnings": validation_warnings,
        "status_reason": status_reason,
        "task_context": task_context,
        "issue": {"status": "skipped", "issue_id": None, "url": None, "link": None},
        "requires_user_input": requires_user_input,
    }


@router.post("/api/admin/prototype-review/run")
def run_admin_prototype_review(payload: PrototypeReviewRunPayload, request: Request):
    user = _require_authenticated(request)
    try:
        return _prototype_review_build_result(payload, user)
    except Exception as exc:
        print("❌ /api/admin/prototype-review/run error:", exc)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/api/admin/prototype-review/run-start")
def start_admin_prototype_review(payload: PrototypeReviewRunPayload, request: Request):
    user = _require_authenticated(request)
    job_id = uuid4().hex
    with _prototype_review_jobs_lock:
        _prototype_review_jobs[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "current": 0,
            "total": 0,
            "current_file": None,
            "current_target": None,
            "result": None,
            "error": None,
        }

    def _runner():
        try:
            _prototype_review_job_update(job_id, status="running")
            result = _prototype_review_build_result(
                payload,
                user,
                progress_callback=lambda event: _prototype_review_job_update(
                    job_id,
                    status="running",
                    current=int(event.get("current") or 0),
                    total=int(event.get("total") or 0),
                    current_file=event.get("path"),
                    current_target=event.get("target_fqn"),
                    last_event=event,
                ),
            )
            _prototype_review_job_update(
                job_id,
                status="completed",
                current=len(result.get("execution") or []),
                total=len(result.get("review_items") or []),
                result=result,
            )
        except Exception as exc:
            _prototype_review_job_update(job_id, status="error", error=str(exc))

    threading.Thread(target=_runner, daemon=True).start()
    return {"status": "queued", "job_id": job_id}


@router.get("/api/admin/prototype-review/run-status/{job_id}")
def get_admin_prototype_review_status(job_id: str, request: Request):
    _require_authenticated(request)
    with _prototype_review_jobs_lock:
        payload = dict(_prototype_review_jobs.get(job_id) or {})
    if not payload:
        raise HTTPException(status_code=404, detail="Job не найден")
    return payload


@router.post("/api/admin/prototype-review/check-table")
def check_admin_prototype_review_table(payload: PrototypeReviewTableCheckPayload, request: Request):
    _require_authenticated(request)
    try:
        bundle = load_merge_request_sql_bundle(
            gitlab_api_url=GITLAB_API_URL,
            gitlab_project=GITLAB_PROJECT,
            gitlab_token=GITLAB_TOKEN,
            gitlab_ssl_verify=GITLAB_SSL_VERIFY,
            mr_input=payload.mr_input,
            default_project=ANALYST_GITLAB_PROJECT or GITLAB_PROJECT,
        )
        files = bundle.get("files") or []
        all_meta, _ = get_cached_meta_and_index()
        known_schemas = {
            str(meta.get("table_schema") or "").strip().lower()
            for meta in all_meta
            if str(meta.get("table_schema") or "").strip()
        }
        related_targets = infer_review_targets(files)
        current_target = next(
            (
                item for item in related_targets
                if str(item.get("target_fqn") or "").strip().lower() == str(payload.target_fqn or "").strip().lower()
            ),
            None,
        )
        related_paths = [str(value).strip() for value in (current_target or {}).get("paths", []) if str(value).strip()]
        related_files = [row for row in files if str(row.get("path") or "") in set(related_paths)]
        item_result = _prototype_review_resolve_item(
            target_fqn=payload.target_fqn,
            path_value=str((current_target or {}).get("path") or ""),
            fallback_entity_name=str(payload.entity_name or "").strip(),
            key_attributes_override=[str(value).strip() for value in (payload.key_attributes or []) if str(value).strip()],
            execution_row={"status": "ok", "duration_sec": 0.0},
            known_schemas=known_schemas,
            file_item=related_files[0] if related_files else None,
            related_files=related_files,
        )
        return {"status": "ok", "item": item_result}
    except Exception as exc:
        print("❌ /api/admin/prototype-review/check-table error:", exc)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/api/admin/prototype-review/create-issue")
def create_admin_prototype_review_issue(payload: PrototypeReviewCreateIssuePayload, request: Request):
    user = _require_authenticated(request)
    try:
        bundle = load_merge_request_sql_bundle(
            gitlab_api_url=GITLAB_API_URL,
            gitlab_project=GITLAB_PROJECT,
            gitlab_token=GITLAB_TOKEN,
            gitlab_ssl_verify=GITLAB_SSL_VERIFY,
            mr_input=payload.mr_input,
            default_project=ANALYST_GITLAB_PROJECT or GITLAB_PROJECT,
        )
        parsed_task = parse_prototype_task_text(payload.task_text or "")
        review_items = [item.model_dump() for item in payload.review_items]
        incomplete = []
        for item in review_items:
            needs_attention, missing = _prototype_item_needs_attention(item)
            if needs_attention:
                incomplete.append(f"{item.get('target_fqn')}: {', '.join(missing)}")
        if incomplete:
            raise HTTPException(status_code=400, detail="Нужно заполнить вручную: " + "; ".join(incomplete))
        task_context = {
            **parsed_task,
            "summary": (payload.issue_summary or "").strip() or parsed_task.get("summary"),
            "linked_issues": payload.linked_issues or parsed_task.get("linked_issues") or [],
        }
        summary = (payload.issue_summary or "").strip() or parsed_task.get("summary") or f"[Prototype Review] {bundle.get('mr', {}).get('source_branch') or 'prototype'}"
        description = _prototype_multi_issue_description(
            mr=bundle.get("mr") or {},
            task_context=task_context,
            initiator={"email": getattr(user, "email", None), "username": getattr(user, "username", None)},
            review_items=review_items,
        )
        issue_result = create_ytrack_issue(
            base_url=YOUTRACK_URL,
            project_id=YOUTRACK_PROJECT_ID,
            project=YOUTRACK_PROJECT,
            token=YOUTRACK_TOKEN,
            queue=YOUTRACK_QUEUE,
            issue_type=YOUTRACK_ISSUE_TYPE,
            ssl_verify=YOUTRACK_SSL_VERIFY,
            summary=summary,
            description=description,
            default_estimate_minutes=YOUTRACK_DEFAULT_ESTIMATE_MINUTES,
            estimate_field_name=YOUTRACK_ESTIMATE_FIELD_NAME,
            card_type_field_name=YOUTRACK_CARD_TYPE_FIELD_NAME,
            card_type_value=YOUTRACK_CARD_TYPE_VALUE,
            assignee_field_name=YOUTRACK_ASSIGNEE_FIELD_NAME,
            assignee_query=YOUTRACK_ASSIGNEE_QUERY,
        )
        meta_branch = None
        meta_files = []
        meta_error = None
        meta_mr = None
        meta_mr_error = None
        raw_issue_id = str((issue_result.get("raw") or {}).get("id") or "").strip()
        if raw_issue_id:
            branch_name = f"feature/{str(issue_result.get('issue_id') or '').strip().upper()}"
            for item in review_items:
                yaml_content = str(item.get("yaml_content") or "").strip()
                if not yaml_content:
                    continue
                target_fqn = str(item.get("target_fqn") or "").strip().lower()
                entity_name = str(item.get("entity_name") or "").strip()
                if "." not in target_fqn or not entity_name:
                    continue
                schema_name, table_name = target_fqn.split(".", 1)
                try:
                    save_result = save_meta_workspace_branch_file(
                        git_repo_value=ENTITY_META_GIT_REPO,
                        workspace_root_value=META_WORKSPACE_ROOT,
                        workspace_owner=getattr(user, "email", None) or getattr(user, "username", None) or "prototype-review",
                        branch_name=branch_name,
                        base_branch="main",
                        file_path=_prototype_review_yaml_repo_path(entity_name, schema_name, table_name),
                        content=yaml_content,
                        task_id=str(issue_result.get("issue_id") or "").strip().upper(),
                        author=getattr(user, "email", None) or getattr(user, "username", None) or "prototype-review",
                        expected_revision=None,
                    )
                    meta_files.append(
                        {
                            "target_fqn": target_fqn,
                            "entity_name": entity_name,
                            "file_path": save_result.get("file_path"),
                            "branch_name": save_result.get("branch_name"),
                            "committed": bool(save_result.get("committed")),
                        }
                    )
                    meta_branch = save_result.get("branch_name") or meta_branch
                except Exception as exc:
                    meta_error = str(exc)
                    break
            if not meta_error and meta_branch:
                try:
                    meta_mr = create_meta_workspace_mr(
                        engine=engine,
                        base_dir=BASE_DIR,
                        entity_dev_root_value=DEV_ENTITY_META_DIR,
                        click_dev_root_value=DEV_CLICK_META_DIR,
                        git_repo_value=ENTITY_META_GIT_REPO,
                        entity_git_root_value=ENTITY_META_GIT_META_ROOT,
                        click_git_root_value=CLICK_META_GIT_ROOT,
                        gitlab_token=GITLAB_TOKEN,
                        gitlab_project=GITLAB_PROJECT,
                        gitlab_api_url=GITLAB_API_URL,
                        gitlab_ssl_verify=GITLAB_SSL_VERIFY,
                        task_id=str(issue_result.get("issue_id") or "").strip().upper(),
                        release_branch="main",
                        branch_name=meta_branch,
                        mr_title=f"{str(issue_result.get('issue_id') or '').strip().upper()}: Engineer MR to main",
                        author=getattr(user, "email", None) or getattr(user, "username", None) or "prototype-review",
                    )
                    if meta_mr.get("mr_url") and YOUTRACK_URL and YOUTRACK_TOKEN:
                        add_ytrack_issue_comment(
                            base_url=YOUTRACK_URL,
                            token=YOUTRACK_TOKEN,
                            issue_id=str(issue_result.get("issue_id") or "").strip().upper(),
                            ssl_verify=YOUTRACK_SSL_VERIFY,
                            text=(
                                "MR создан из Prototype Review для инженера.\n"
                                f"Ссылка: {meta_mr.get('mr_url')}\n"
                                f"Ветка: {meta_mr.get('feature_branch') or '—'} -> {meta_mr.get('release_branch') or 'main'}"
                            ),
                        )
                        meta_mr["task_link_attached"] = True
                except Exception as exc:
                    meta_mr_error = str(exc)
        if issue_result.get("issue_id"):
            issue_result["link"] = _build_ytrack_link(issue_result.get("issue_id"))
        return {
            "status": "ok",
            "issue": issue_result,
            "description": description,
            "meta_branch": meta_branch,
            "meta_files": meta_files,
            "meta_error": meta_error,
            "meta_mr": meta_mr,
            "meta_mr_error": meta_mr_error,
        }
    except HTTPException:
        raise
    except Exception as exc:
        print("❌ /api/admin/prototype-review/create-issue error:", exc)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/api/admin/ci-cd/status")
def get_ci_cd_status(request: Request):
    user = get_current_user_from_request(request)
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    return _ci_cd_status


@router.get("/api/admin/engineering-efficiency")
def get_admin_engineering_efficiency(
    request: Request,
    days: int = Query(90, ge=1, le=3650),
):
    user = get_current_user_from_request(request)
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    try:
        date_clause, params = _resolve_date_window(None, None, days)
        with engine.connect() as conn:
            base = f"""
                WITH rel AS (
                    SELECT r.release_id, r.started_at, r.finished_at, r.initiated_by, r.status
                    FROM {TABLE_RELEASE_LOG} r
                    WHERE {date_clause}
                ),
                ro AS (
                    SELECT ro.*
                    FROM {TABLE_RELEASE_OBJECTS} ro
                    JOIN rel r ON r.release_id = ro.release_id
                ),
                tasks AS (
                    SELECT DISTINCT task_id
                    FROM ro
                    WHERE task_id IS NOT NULL
                ),
                snap AS (
                    SELECT s.*
                    FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                    JOIN tasks t ON t.task_id = s.issue_id
                ),
                exec AS (
                    SELECT DISTINCT ON (issue_id)
                           issue_id, author AS executor, ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                dashboard AS (
                    SELECT issue_id, field_value AS dashboard_direction
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Дашборд КХД/Направление'
                ),
                engineer_issue AS (
                    SELECT
                        snap.issue_id,
                        COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан') AS engineer,
                        COALESCE(snap.created_by, 'Не указан') AS creator,
                        COALESCE(work.minutes, 0) AS minutes,
                        COALESCE(dashboard.dashboard_direction, 'Не указан') AS dashboard_direction
                    FROM snap
                    LEFT JOIN exec ON exec.issue_id = snap.issue_id
                    LEFT JOIN work ON work.issue_id = snap.issue_id
                    LEFT JOIN dashboard ON dashboard.issue_id = snap.issue_id
                ),
                engineer_objects AS (
                    SELECT
                        ei.engineer,
                        ei.issue_id,
                        ei.minutes,
                        ei.dashboard_direction,
                        ro.schema_name,
                        ro.table_name,
                        COUNT(*) OVER (PARTITION BY ro.schema_name, ro.table_name) AS changes_count
                    FROM engineer_issue ei
                    LEFT JOIN ro ON ro.task_id = ei.issue_id
                )
            """

            summary = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        COUNT(DISTINCT issue_id) AS tasks_count,
                        COUNT(schema_name || '.' || table_name) AS objects_count,
                        COALESCE(SUM(minutes), 0) / 60.0 AS hours,
                        COUNT(DISTINCT engineer) AS engineers_count
                    FROM engineer_objects
                    """
                ),
                params,
            ).mappings().first()

            engineers = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        engineer,
                        COUNT(DISTINCT issue_id) AS tasks_count,
                        COUNT(schema_name || '.' || table_name) AS objects_count,
                        COALESCE(SUM(minutes), 0) / 60.0 AS hours,
                        CASE
                            WHEN COUNT(DISTINCT issue_id) = 0 THEN 0
                            ELSE (COALESCE(SUM(minutes), 0) / 60.0) / COUNT(DISTINCT issue_id)
                        END AS avg_hours_per_task
                    FROM engineer_objects
                    GROUP BY engineer
                    ORDER BY hours DESC NULLS LAST, tasks_count DESC
                    """
                ),
                params,
            ).mappings().all()

            schema_breakdown = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        COALESCE(schema_name, 'UNKNOWN') AS schema_name,
                        COALESCE(SUM(minutes), 0) / 60.0 AS hours
                    FROM engineer_objects
                    WHERE schema_name IS NOT NULL
                    GROUP BY COALESCE(schema_name, 'UNKNOWN')
                    ORDER BY hours DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

            dashboards = conn.execute(
                text(
                    base
                    + """
                    ,
                    dashboard_agg AS (
                        SELECT
                            dashboard_direction,
                            engineer,
                            COUNT(DISTINCT issue_id) AS tasks_count,
                            COUNT(schema_name || '.' || table_name) AS objects_count,
                            COALESCE(SUM(minutes), 0) / 60.0 AS hours
                        FROM engineer_objects
                        GROUP BY dashboard_direction, engineer
                    )
                    SELECT
                        dashboard_direction,
                        SUM(hours) AS hours,
                        SUM(tasks_count) AS tasks_count,
                        SUM(objects_count) AS objects_count,
                        (ARRAY_AGG(engineer ORDER BY hours DESC NULLS LAST))[1] AS top_engineer
                    FROM dashboard_agg
                    GROUP BY dashboard_direction
                    ORDER BY hours DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

            top_objects = conn.execute(
                text(
                    base
                    + """
                    ,
                    object_agg AS (
                        SELECT
                            schema_name,
                            table_name,
                            engineer,
                            COUNT(DISTINCT issue_id) AS tasks_count,
                            COUNT(*) AS changes_count,
                            COALESCE(SUM(minutes), 0) / 60.0 AS hours
                        FROM engineer_objects
                        WHERE schema_name IS NOT NULL
                          AND table_name IS NOT NULL
                        GROUP BY schema_name, table_name, engineer
                    )
                    SELECT
                        schema_name,
                        table_name,
                        SUM(hours) AS hours,
                        SUM(tasks_count) AS tasks_count,
                        SUM(changes_count) AS changes_count,
                        (ARRAY_AGG(engineer ORDER BY hours DESC NULLS LAST))[1] AS top_engineer
                    FROM object_agg
                    GROUP BY schema_name, table_name
                    ORDER BY hours DESC NULLS LAST, changes_count DESC
                    LIMIT 30
                    """
                ),
                params,
            ).mappings().all()

            daily_engineers = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        DATE(COALESCE(snap.updated_at, snap.created_at))::text AS day,
                        COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан') AS engineer,
                        COUNT(DISTINCT snap.issue_id) AS tasks_count,
                        COALESCE(SUM(work.minutes), 0) / 60.0 AS hours
                    FROM snap
                    LEFT JOIN exec ON exec.issue_id = snap.issue_id
                    LEFT JOIN work ON work.issue_id = snap.issue_id
                    GROUP BY
                        DATE(COALESCE(snap.updated_at, snap.created_at))::text,
                        COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан')
                    ORDER BY day, hours DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

            engineer_schema_rows = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        engineer,
                        schema_name,
                        COALESCE(SUM(minutes), 0) / 60.0 AS hours
                    FROM engineer_objects
                    WHERE schema_name IS NOT NULL
                    GROUP BY engineer, schema_name
                    ORDER BY engineer, hours DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

        engineer_schema_map = {}
        for row in engineer_schema_rows:
            engineer_schema_map.setdefault(row["engineer"], []).append(
                {
                    "schema_name": row["schema_name"],
                    "hours": float(row["hours"] or 0),
                }
            )

        engineers_payload = []
        for row in engineers:
            hours = float(row["hours"] or 0)
            tasks_count = int(row["tasks_count"] or 0)
            avg = float(row["avg_hours_per_task"] or 0)
            if hours >= 120 or avg >= 12:
                load_status = "Перегружен"
            elif hours <= 16 and tasks_count <= 2:
                load_status = "Недогружен"
            elif tasks_count >= 5 and avg <= 6:
                load_status = "Эффективен"
            else:
                load_status = "Стабильно"
            engineers_payload.append(
                {
                    "engineer": row["engineer"],
                    "tasks_count": tasks_count,
                    "objects_count": int(row["objects_count"] or 0),
                    "hours": hours,
                    "avg_hours_per_task": avg,
                    "load_status": load_status,
                    "schemas": engineer_schema_map.get(row["engineer"], [])[:8],
                }
            )

        focus = {
            "overloaded": [row for row in engineers_payload if row["load_status"] == "Перегружен"][:5],
            "efficient": [row for row in engineers_payload if row["load_status"] == "Эффективен"][:5],
            "underloaded": [row for row in engineers_payload if row["load_status"] == "Недогружен"][:5],
        }

        return {
            "days": days,
            "summary": {
                "tasks_count": int((summary or {}).get("tasks_count") or 0),
                "objects_count": int((summary or {}).get("objects_count") or 0),
                "hours": float((summary or {}).get("hours") or 0),
                "engineers_count": int((summary or {}).get("engineers_count") or 0),
            },
            "engineers": engineers_payload,
            "focus": focus,
            "daily_engineers": [dict(row) for row in daily_engineers],
            "schema_breakdown": [dict(row) for row in schema_breakdown],
            "dashboard_report": [dict(row) for row in dashboards],
            "top_objects": [dict(row) for row in top_objects],
        }
    except Exception as e:
        print("❌ /api/admin/engineering-efficiency error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить страницу эффективности")


@router.get("/api/admin/reports/releases")
def get_admin_release_reports(
    request: Request,
    days: int = Query(180, ge=30, le=3650),
    debug: bool = Query(False),
):
    user = get_current_user_from_request(request)
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    try:
        date_clause, params = _resolve_date_window(None, None, days)
        with engine.connect() as conn:
            base = f"""
                WITH raw_rel AS (
                    SELECT
                        r.release_id,
                        r.release_type,
                        r.initiated_by,
                        r.started_at,
                        r.finished_at,
                        r.status,
                        r.total_objects,
                        r.ready_to_release,
                        CASE
                            WHEN lower(COALESCE(r.release_type, '') || ' ' || COALESCE(r.release_id, '')) LIKE '%hotfix%' THEN 'hotfix'
                            WHEN lower(COALESCE(r.release_type, '') || ' ' || COALESCE(r.release_id, '')) LIKE '%внерел%'
                              OR lower(COALESCE(r.release_type, '') || ' ' || COALESCE(r.release_id, '')) LIKE '%vnerelease%'
                              OR lower(COALESCE(r.release_type, '') || ' ' || COALESCE(r.release_id, '')) LIKE '%vne release%'
                              OR lower(COALESCE(r.release_type, '') || ' ' || COALESCE(r.release_id, '')) LIKE '%out_of_release%'
                              OR lower(COALESCE(r.release_type, '') || ' ' || COALESCE(r.release_id, '')) LIKE '%out of release%'
                              OR lower(COALESCE(r.release_type, '') || ' ' || COALESCE(r.release_id, '')) LIKE '%nonrelease%'
                              OR lower(COALESCE(r.release_type, '') || ' ' || COALESCE(r.release_id, '')) LIKE '%non-release%'
                            THEN 'outside_release'
                            ELSE 'release'
                        END AS fallback_release_bucket
                    FROM {TABLE_RELEASE_LOG} r
                    WHERE {date_clause}
                ),
                raw_ro AS (
                    SELECT
                        ro.*
                    FROM {TABLE_RELEASE_OBJECTS} ro
                    JOIN raw_rel ON raw_rel.release_id = ro.release_id
                ),
                release_tasks AS (
                    SELECT DISTINCT release_id, task_id
                    FROM raw_ro
                    WHERE task_id IS NOT NULL AND task_id <> ''
                ),
                task_ids AS (
                    SELECT DISTINCT task_id
                    FROM release_tasks
                ),
                task_custom_base AS (
                    SELECT issue_id, field_name, NULLIF(BTRIM(field_value), '') AS field_value
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name IN (
                          'Тип карточки',
                          'Тип внерелиза',
                          'Фактическая дата релиза',
                          'Дата выкатки',
                          'Номер релиза КХД'
                      )
                ),
                task_custom AS (
                    SELECT
                        issue_id,
                        MAX(CASE WHEN field_name = 'Тип карточки' THEN field_value END) AS card_type,
                        MAX(CASE WHEN field_name = 'Тип внерелиза' THEN field_value END) AS outside_type,
                        MAX(CASE WHEN field_name = 'Номер релиза КХД' THEN field_value END) AS release_slot_number,
                        MAX(CASE WHEN field_name = 'Фактическая дата релиза' THEN field_value END) AS actual_release_at_text,
                        MAX(CASE WHEN field_name = 'Дата выкатки' THEN field_value END) AS rollout_at_text
                    FROM task_custom_base
                    GROUP BY issue_id
                ),
                task_custom_norm AS (
                    SELECT
                        issue_id,
                        card_type,
                        outside_type,
                        release_slot_number,
                        LEFT(COALESCE(actual_release_at_text, rollout_at_text), 10) AS effective_release_at_key,
                        CASE
                            WHEN lower(COALESCE(card_type, '')) = 'release slot' THEN 'release'
                            WHEN lower(COALESCE(card_type, '')) = 'внерелиз'
                             AND (
                                 lower(COALESCE(outside_type, '')) LIKE '%хотфикс%'
                                 OR lower(COALESCE(outside_type, '')) LIKE '%hotfix%'
                             )
                            THEN 'hotfix'
                            WHEN lower(COALESCE(card_type, '')) = 'внерелиз' THEN 'outside_release'
                            ELSE NULL
                        END AS custom_bucket
                    FROM task_custom
                    WHERE lower(COALESCE(card_type, '')) IN ('release slot', 'внерелиз')
                ),
                release_match_by_date AS (
                    SELECT
                        rr.release_id,
                        MAX(tc.release_slot_number) AS release_slot_number,
                        COUNT(*) FILTER (WHERE tc.custom_bucket = 'release') AS release_slot_count,
                        COUNT(*) FILTER (WHERE tc.custom_bucket = 'hotfix') AS hotfix_count,
                        COUNT(*) FILTER (WHERE tc.custom_bucket = 'outside_release') AS outside_release_count
                    FROM raw_rel rr
                    JOIN task_custom_norm tc
                      ON tc.effective_release_at_key IS NOT NULL
                     AND tc.effective_release_at_key = to_char(rr.started_at, 'YYYY-MM-DD')
                    GROUP BY rr.release_id
                ),
                rel AS (
                    SELECT
                        rr.release_id,
                        COALESCE(
                            NULLIF(CASE WHEN COALESCE(rmd.release_slot_count, 0) > 0 THEN rmd.release_slot_number END, ''),
                            rr.release_type
                        ) AS release_type,
                        rr.initiated_by,
                        rr.started_at,
                        rr.finished_at,
                        rr.status,
                        rr.total_objects,
                        rr.ready_to_release,
                        CASE
                            WHEN COALESCE(rmd.release_slot_count, 0) > 0 THEN 'release'
                            ELSE rr.fallback_release_bucket
                        END AS release_bucket
                    FROM raw_rel rr
                    LEFT JOIN release_match_by_date rmd
                      ON rmd.release_id = rr.release_id
                ),
                ro AS (
                    SELECT
                        raw_ro.*,
                        rel.started_at,
                        rel.release_type,
                        rel.release_bucket,
                        rel.initiated_by,
                        rel.status
                    FROM raw_ro
                    JOIN rel ON rel.release_id = raw_ro.release_id
                ),
                snap AS (
                    SELECT s.*
                    FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                    JOIN task_ids t ON t.task_id = s.issue_id
                ),
                exec AS (
                    SELECT DISTINCT ON (issue_id)
                           issue_id, author AS executor, ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                task_fact AS (
                    SELECT
                        rt.release_id,
                        rt.task_id,
                        COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан') AS engineer,
                        COALESCE(snap.created_by, 'Не указан') AS creator,
                        COALESCE(work.minutes, 0) AS minutes
                    FROM release_tasks rt
                    LEFT JOIN snap ON snap.issue_id = rt.task_id
                    LEFT JOIN exec ON exec.issue_id = rt.task_id
                    LEFT JOIN work ON work.issue_id = rt.task_id
                ),
                task_rollup AS (
                    SELECT
                        release_id,
                        COUNT(DISTINCT task_id) AS tasks_count,
                        COUNT(DISTINCT engineer) AS engineers_count,
                        COUNT(DISTINCT creator) AS creators_count,
                        COALESCE(SUM(minutes), 0) AS minutes_total
                    FROM task_fact
                    GROUP BY release_id
                ),
                object_rollup AS (
                    SELECT
                        release_id,
                        COUNT(*) AS objects_count,
                        COUNT(DISTINCT entity_name) FILTER (WHERE entity_name IS NOT NULL AND entity_name <> '') AS entities_count,
                        COUNT(*) FILTER (WHERE lower(COALESCE(target_system, '')) LIKE '%click%') AS click_objects_count,
                        COUNT(*) FILTER (
                            WHERE lower(COALESCE(target_system, '')) LIKE '%greenplum%'
                               OR lower(COALESCE(target_system, '')) LIKE '%gp%'
                        ) AS gp_objects_count
                    FROM ro
                    GROUP BY release_id
                ),
                release_rollup AS (
                    SELECT
                        rel.release_id,
                        rel.release_type,
                        rel.release_bucket,
                        rel.initiated_by,
                        rel.started_at,
                        rel.finished_at,
                        rel.status,
                        rel.total_objects,
                        rel.ready_to_release,
                        COALESCE(obj.objects_count, 0) AS objects_count,
                        COALESCE(obj.entities_count, 0) AS entities_count,
                        COALESCE(obj.click_objects_count, 0) AS click_objects_count,
                        COALESCE(obj.gp_objects_count, 0) AS gp_objects_count,
                        COALESCE(task.tasks_count, 0) AS tasks_count,
                        COALESCE(task.engineers_count, 0) AS engineers_count,
                        COALESCE(task.creators_count, 0) AS creators_count,
                        COALESCE(task.minutes_total, 0) AS minutes_total,
                        EXTRACT(EPOCH FROM (COALESCE(rel.finished_at, now()) - rel.started_at)) / 60.0 AS duration_minutes
                    FROM rel
                    LEFT JOIN object_rollup obj ON obj.release_id = rel.release_id
                    LEFT JOIN task_rollup task ON task.release_id = rel.release_id
                )
            """

            summary = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        COUNT(*) AS releases_count,
                        COUNT(*) FILTER (WHERE release_bucket = 'hotfix') AS hotfix_count,
                        COUNT(*) FILTER (WHERE release_bucket = 'outside_release') AS outside_release_count,
                        COUNT(*) FILTER (WHERE release_bucket = 'release') AS release_count,
                        COALESCE(SUM(objects_count), 0) AS objects_count,
                        COALESCE(SUM(click_objects_count), 0) AS click_objects_count,
                        COALESCE(SUM(gp_objects_count), 0) AS gp_objects_count,
                        COALESCE(SUM(tasks_count), 0) AS tasks_count,
                        COALESCE(SUM(minutes_total), 0) / 60.0 AS hours_total,
                        COUNT(DISTINCT initiated_by) FILTER (WHERE initiated_by IS NOT NULL AND initiated_by <> '') AS initiators_count,
                        COUNT(DISTINCT started_at::date) AS release_days_count,
                        ROUND(COALESCE(AVG(objects_count)::numeric, 0), 1) AS avg_objects_per_release,
                        ROUND(COALESCE(AVG(duration_minutes)::numeric, 0), 1) AS avg_duration_minutes
                    FROM release_rollup
                    """
                ),
                params,
            ).mappings().first()

            cadence = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        to_char(date_trunc('week', started_at), 'DD.MM') AS week_label,
                        date_trunc('week', started_at)::date::text AS week_start,
                        COUNT(*) AS releases_count,
                        COUNT(*) FILTER (WHERE release_bucket = 'hotfix') AS hotfix_count,
                        COUNT(*) FILTER (WHERE release_bucket = 'outside_release') AS outside_release_count,
                        COUNT(*) FILTER (WHERE release_bucket = 'release') AS regular_release_count,
                        COALESCE(SUM(objects_count), 0) AS objects_count,
                        COALESCE(SUM(tasks_count), 0) AS tasks_count
                    FROM release_rollup
                    GROUP BY date_trunc('week', started_at)
                    ORDER BY date_trunc('week', started_at)
                    """
                ),
                params,
            ).mappings().all()

            weekday_heatmap = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        EXTRACT(ISODOW FROM started_at)::int AS weekday_no,
                        CASE EXTRACT(ISODOW FROM started_at)::int
                            WHEN 1 THEN 'Пн'
                            WHEN 2 THEN 'Вт'
                            WHEN 3 THEN 'Ср'
                            WHEN 4 THEN 'Чт'
                            WHEN 5 THEN 'Пт'
                            WHEN 6 THEN 'Сб'
                            ELSE 'Вс'
                        END AS weekday_label,
                        EXTRACT(HOUR FROM started_at)::int AS hour_of_day,
                        COUNT(*) AS releases_count,
                        COALESCE(SUM(objects_count), 0) AS objects_count,
                        COALESCE(SUM(tasks_count), 0) AS tasks_count
                    FROM release_rollup
                    GROUP BY EXTRACT(ISODOW FROM started_at), EXTRACT(HOUR FROM started_at)
                    ORDER BY weekday_no, hour_of_day
                    """
                ),
                params,
            ).mappings().all()

            type_breakdown = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        release_bucket,
                        COUNT(*) AS releases_count,
                        COALESCE(SUM(objects_count), 0) AS objects_count,
                        COALESCE(SUM(tasks_count), 0) AS tasks_count,
                        COALESCE(SUM(minutes_total), 0) / 60.0 AS hours_total
                    FROM release_rollup
                    GROUP BY release_bucket
                    ORDER BY releases_count DESC, objects_count DESC
                    """
                ),
                params,
            ).mappings().all()

            system_breakdown = conn.execute(
                text(
                    base
                    + """
                    SELECT 'ClickHouse' AS system_name, COALESCE(SUM(click_objects_count), 0) AS objects_count
                    FROM release_rollup
                    UNION ALL
                    SELECT 'Greenplum' AS system_name, COALESCE(SUM(gp_objects_count), 0) AS objects_count
                    FROM release_rollup
                    """
                ),
                params,
            ).mappings().all()

            ro_detail_rows = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        release_id,
                        release_type,
                        release_bucket,
                        initiated_by,
                        started_at,
                        status,
                        COALESCE(entity_name, 'Без сущности') AS entity_name,
                        schema_name,
                        table_name,
                        task_id
                    FROM ro
                    """
                ),
                params,
            ).mappings().all()

            task_detail_rows = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        release_id,
                        task_id,
                        engineer,
                        creator,
                        minutes
                    FROM task_fact
                    """
                ),
                params,
            ).mappings().all()

            release_detail_rows = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        release_id,
                        release_type,
                        release_bucket,
                        initiated_by,
                        started_at,
                        status,
                        objects_count,
                        tasks_count,
                        creators_count,
                        engineers_count,
                        minutes_total,
                        duration_minutes
                    FROM release_rollup
                    ORDER BY started_at DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

            exception_custom_rows = conn.execute(
                text(
                    f"""
                    SELECT
                        issue_id,
                        field_name,
                        NULLIF(BTRIM(field_value), '') AS field_value
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name IN (
                        'Тип карточки',
                        'Тип внерелиза',
                        'Фактическая дата релиза',
                        'Дата выкатки',
                        'Направление',
                        'Дашборд КХД/Направление'
                    )
                    """
                )
            ).mappings().all()

            exception_cards_by_issue = {}
            window_end = datetime.now().date()
            window_start = window_end - timedelta(days=days)
            for row in exception_custom_rows:
                card = exception_cards_by_issue.setdefault(
                    row.get("issue_id"),
                    {
                        "issue_id": row.get("issue_id"),
                        "card_type": None,
                        "outside_type": None,
                        "actual_release_at_text": None,
                        "rollout_at_text": None,
                        "direction": None,
                        "dashboard_direction": None,
                    },
                )
                field_name = row.get("field_name")
                if field_name == "Тип карточки":
                    card["card_type"] = row.get("field_value")
                elif field_name == "Тип внерелиза":
                    card["outside_type"] = row.get("field_value")
                elif field_name == "Фактическая дата релиза":
                    card["actual_release_at_text"] = row.get("field_value")
                elif field_name == "Дата выкатки":
                    card["rollout_at_text"] = row.get("field_value")
                elif field_name == "Направление":
                    card["direction"] = row.get("field_value")
                elif field_name == "Дашборд КХД/Направление":
                    card["dashboard_direction"] = row.get("field_value")

            exception_issue_ids = []
            exception_card_rows = []
            for card in exception_cards_by_issue.values():
                card_type = (card.get("card_type") or "").strip().lower()
                if card_type != "внерелиз":
                    continue
                effective_text = (card.get("rollout_at_text") or card.get("actual_release_at_text") or "")[:10]
                try:
                    effective_date = date.fromisoformat(effective_text)
                except Exception:
                    continue
                if effective_date < window_start or effective_date > window_end:
                    continue
                outside_type = (card.get("outside_type") or "").strip()
                outside_type_norm = outside_type.lower()
                release_bucket = (
                    "hotfix"
                    if ("хотфикс" in outside_type_norm or "hotfix" in outside_type_norm)
                    else "outside_release"
                )
                exception_issue_ids.append(card["issue_id"])
                exception_card_rows.append(
                    {
                        "issue_id": card["issue_id"],
                        "release_bucket": release_bucket,
                        "release_type": outside_type or "Внерелиз",
                        "started_at": datetime.combine(effective_date, datetime.min.time()),
                        "direction": card.get("direction") or card.get("dashboard_direction") or "Не указано",
                    }
                )

            exception_snap_map = {}
            exception_work_map = {}
            exception_object_map = {}
            if exception_issue_ids:
                exception_snap_rows = conn.execute(
                    text(
                        f"""
                        SELECT issue_id, created_by, assignee, current_state
                        FROM {TABLE_YT_ISSUE_SNAPSHOT}
                        WHERE issue_id = ANY(:ids)
                        """
                    ),
                    {"ids": exception_issue_ids},
                ).mappings().all()
                exception_snap_map = {row["issue_id"]: dict(row) for row in exception_snap_rows}

                exception_work_rows = conn.execute(
                    text(
                        f"""
                        SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                        FROM {TABLE_YT_ISSUE_WORKLOG}
                        WHERE issue_id = ANY(:ids)
                        GROUP BY issue_id
                        """
                    ),
                    {"ids": exception_issue_ids},
                ).mappings().all()
                exception_work_map = {row["issue_id"]: float(row.get("minutes") or 0) for row in exception_work_rows}

                exception_object_rows = conn.execute(
                    text(
                        f"""
                        SELECT
                            ro.task_id,
                            ro.release_id,
                            COALESCE(ro.entity_name, 'Без сущности') AS entity_name
                        FROM {TABLE_RELEASE_OBJECTS} ro
                        JOIN {TABLE_RELEASE_LOG} rl
                          ON rl.release_id = ro.release_id
                        WHERE ro.task_id = ANY(:ids)
                          AND rl.started_at >= (now() - (:days || ' days')::interval)
                        """
                    ),
                    {"ids": exception_issue_ids, "days": days},
                ).mappings().all()
                for row in exception_object_rows:
                    bucket = exception_object_map.setdefault(
                        row["task_id"],
                        {"objects_count": 0, "entity_names": set(), "release_ids": set()},
                    )
                    bucket["objects_count"] += 1
                    if row.get("entity_name"):
                        bucket["entity_names"].add(row["entity_name"])
                    if row.get("release_id"):
                        bucket["release_ids"].add(row["release_id"])

            exception_summary_rows = []
            for row in exception_card_rows:
                issue_id = row["issue_id"]
                snap_row = exception_snap_map.get(issue_id, {})
                object_row = exception_object_map.get(issue_id, {})
                creator = snap_row.get("created_by") or "Не указан"
                engineer = snap_row.get("assignee") or creator
                linked_release_ids = sorted(object_row.get("release_ids", set()))
                exception_summary_rows.append(
                    {
                        "release_id": linked_release_ids[0] if linked_release_ids else issue_id,
                        "task_id": issue_id,
                        "release_type": row["release_type"],
                        "release_bucket": row["release_bucket"],
                        "initiated_by": creator,
                        "direction": row.get("direction") or "Не указано",
                        "started_at": row["started_at"],
                        "status": snap_row.get("current_state") or "Не указан",
                        "objects_count": int(object_row.get("objects_count") or 0),
                        "tasks_count": 1,
                        "creators_count": 1,
                        "engineers_count": 1 if engineer else 0,
                        "minutes_total": float(exception_work_map.get(issue_id, 0)),
                        "hours_total": round(float(exception_work_map.get(issue_id, 0)) / 60.0, 1),
                        "duration_minutes": 0.0,
                        "entity_names": sorted(object_row.get("entity_names", set())),
                    }
                )

            entity_stats = defaultdict(
                lambda: {
                    "entity_name": "Без сущности",
                    "objects_count": 0,
                    "release_ids": set(),
                    "task_ids": set(),
                    "hotfix_release_ids": set(),
                    "outside_release_ids": set(),
                    "last_release_at": None,
                }
            )
            table_stats = defaultdict(
                lambda: {
                    "schema_name": None,
                    "table_name": None,
                    "objects_count": 0,
                    "release_ids": set(),
                    "task_ids": set(),
                    "last_change_at": None,
                    "entity_names": set(),
                }
            )
            entity_names_by_release = defaultdict(set)

            for row in ro_detail_rows:
                entity_name = row.get("entity_name") or "Без сущности"
                release_id = row.get("release_id")
                release_bucket = row.get("release_bucket")
                started_at = row.get("started_at")
                task_id = row.get("task_id")

                entity_names_by_release[release_id].add(entity_name)

                entity_bucket = entity_stats[entity_name]
                entity_bucket["entity_name"] = entity_name
                entity_bucket["objects_count"] += 1
                if release_id:
                    entity_bucket["release_ids"].add(release_id)
                if task_id:
                    entity_bucket["task_ids"].add(task_id)
                if release_bucket == "hotfix" and release_id:
                    entity_bucket["hotfix_release_ids"].add(release_id)
                if release_bucket == "outside_release" and release_id:
                    entity_bucket["outside_release_ids"].add(release_id)
                if started_at and (
                    entity_bucket["last_release_at"] is None or started_at > entity_bucket["last_release_at"]
                ):
                    entity_bucket["last_release_at"] = started_at

                schema_name = row.get("schema_name")
                table_name = row.get("table_name")
                if schema_name and table_name:
                    table_bucket = table_stats[(schema_name, table_name)]
                    table_bucket["schema_name"] = schema_name
                    table_bucket["table_name"] = table_name
                    table_bucket["objects_count"] += 1
                    if release_id:
                        table_bucket["release_ids"].add(release_id)
                    if task_id:
                        table_bucket["task_ids"].add(task_id)
                    if entity_name and entity_name != "Без сущности":
                        table_bucket["entity_names"].add(entity_name)
                    if started_at and (
                        table_bucket["last_change_at"] is None or started_at > table_bucket["last_change_at"]
                    ):
                        table_bucket["last_change_at"] = started_at

            top_entities = sorted(
                [
                    {
                        "entity_name": bucket["entity_name"],
                        "objects_count": bucket["objects_count"],
                        "releases_count": len(bucket["release_ids"]),
                        "tasks_count": len(bucket["task_ids"]),
                        "hotfix_count": len(bucket["hotfix_release_ids"]),
                        "outside_release_count": len(bucket["outside_release_ids"]),
                        "last_release_at": bucket["last_release_at"],
                    }
                    for bucket in entity_stats.values()
                ],
                key=lambda row: (
                    -(row["objects_count"] or 0),
                    -(row["releases_count"] or 0),
                    -(row["tasks_count"] or 0),
                    row["entity_name"] or "",
                ),
            )[:12]

            top_tables = sorted(
                [
                    {
                        "schema_name": bucket["schema_name"],
                        "table_name": bucket["table_name"],
                        "objects_count": bucket["objects_count"],
                        "releases_count": len(bucket["release_ids"]),
                        "tasks_count": len(bucket["task_ids"]),
                        "last_change_at": bucket["last_change_at"],
                        "entity_names": sorted(bucket["entity_names"]),
                        "primary_entity_name": min(bucket["entity_names"]) if bucket["entity_names"] else None,
                    }
                    for bucket in table_stats.values()
                ],
                key=lambda row: (
                    -(row["releases_count"] or 0),
                    -(row["objects_count"] or 0),
                    -(row["tasks_count"] or 0),
                    f"{row['schema_name']}.{row['table_name']}",
                ),
            )[:16]

            release_bucket_by_release = {
                row.get("release_id"): row.get("release_bucket")
                for row in release_detail_rows
            }

            user_stats = defaultdict(
                lambda: {
                    "engineer": "Не указан",
                    "task_ids": set(),
                    "release_ids": set(),
                    "creator_names": set(),
                    "minutes_total": 0,
                    "hotfix_release_ids": set(),
                    "outside_release_ids": set(),
                }
            )
            creator_stats = defaultdict(
                lambda: {
                    "creator": "Не указан",
                    "task_ids": set(),
                    "release_ids": set(),
                    "minutes_total": 0,
                    "hotfix_release_ids": set(),
                    "outside_release_ids": set(),
                }
            )

            for row in task_detail_rows:
                engineer = row.get("engineer") or "Не указан"
                creator = row.get("creator") or "Не указан"
                release_id = row.get("release_id")
                task_id = row.get("task_id")
                release_bucket = release_bucket_by_release.get(release_id)
                minutes = float(row.get("minutes") or 0)

                user_bucket = user_stats[engineer]
                user_bucket["engineer"] = engineer
                if task_id:
                    user_bucket["task_ids"].add(task_id)
                if release_id:
                    user_bucket["release_ids"].add(release_id)
                user_bucket["creator_names"].add(creator)
                user_bucket["minutes_total"] += minutes
                if release_bucket == "hotfix" and release_id:
                    user_bucket["hotfix_release_ids"].add(release_id)
                if release_bucket == "outside_release" and release_id:
                    user_bucket["outside_release_ids"].add(release_id)

                creator_bucket = creator_stats[creator]
                creator_bucket["creator"] = creator
                if task_id:
                    creator_bucket["task_ids"].add(task_id)
                if release_id:
                    creator_bucket["release_ids"].add(release_id)
                creator_bucket["minutes_total"] += minutes
                if release_bucket == "hotfix" and release_id:
                    creator_bucket["hotfix_release_ids"].add(release_id)
                if release_bucket == "outside_release" and release_id:
                    creator_bucket["outside_release_ids"].add(release_id)

            top_users = sorted(
                [
                    {
                        "engineer": bucket["engineer"],
                        "tasks_count": len(bucket["task_ids"]),
                        "releases_count": len(bucket["release_ids"]),
                        "creators_count": len(bucket["creator_names"]),
                        "hours_total": round(bucket["minutes_total"] / 60.0, 1),
                        "hotfix_count": len(bucket["hotfix_release_ids"]),
                        "outside_release_count": len(bucket["outside_release_ids"]),
                    }
                    for bucket in user_stats.values()
                ],
                key=lambda row: (-(row["hours_total"] or 0), -(row["tasks_count"] or 0), row["engineer"] or ""),
            )[:14]

            top_creators = sorted(
                [
                    {
                        "creator": bucket["creator"],
                        "tasks_count": len(bucket["task_ids"]),
                        "releases_count": len(bucket["release_ids"]),
                        "hours_total": round(bucket["minutes_total"] / 60.0, 1),
                        "hotfix_count": len(bucket["hotfix_release_ids"]),
                        "outside_release_count": len(bucket["outside_release_ids"]),
                    }
                    for bucket in creator_stats.values()
                ],
                key=lambda row: (-(row["tasks_count"] or 0), -(row["hours_total"] or 0), row["creator"] or ""),
            )[:14]

            top_initiators_map = defaultdict(
                lambda: {
                    "initiated_by": "Не указан",
                    "releases_count": 0,
                    "hotfix_count": 0,
                    "outside_release_count": 0,
                    "objects_count": 0,
                    "hours_total": 0.0,
                }
            )
            release_rows_enriched = []
            for row in release_detail_rows:
                release_row = dict(row)
                release_row["entity_names"] = sorted(entity_names_by_release.get(row.get("release_id"), set()))
                release_row["hours_total"] = round(float(row.get("minutes_total") or 0) / 60.0, 1)
                release_row["duration_minutes"] = round(float(row.get("duration_minutes") or 0), 1)
                release_rows_enriched.append(release_row)

                initiator = row.get("initiated_by") or "Не указан"
                initiator_bucket = top_initiators_map[initiator]
                initiator_bucket["initiated_by"] = initiator
                initiator_bucket["releases_count"] += 1
                initiator_bucket["objects_count"] += int(row.get("objects_count") or 0)
                initiator_bucket["hours_total"] += float(row.get("minutes_total") or 0) / 60.0
                if row.get("release_bucket") == "hotfix":
                    initiator_bucket["hotfix_count"] += 1
                if row.get("release_bucket") == "outside_release":
                    initiator_bucket["outside_release_count"] += 1

            top_initiators = sorted(
                [
                    {
                        **bucket,
                        "hours_total": round(bucket["hours_total"], 1),
                    }
                    for bucket in top_initiators_map.values()
                ],
                key=lambda row: (-(row["releases_count"] or 0), -(row["objects_count"] or 0), -(row["hours_total"] or 0)),
            )[:10]

            top_entity_names = [row["entity_name"] for row in top_entities[:6]]
            entity_timeline_map = defaultdict(
                lambda: {
                    "entity_name": "Без сущности",
                    "month_label": "",
                    "month_start": "",
                    "objects_count": 0,
                    "release_ids": set(),
                }
            )
            for row in ro_detail_rows:
                entity_name = row.get("entity_name") or "Без сущности"
                started_at = row.get("started_at")
                release_id = row.get("release_id")
                if entity_name not in top_entity_names or not started_at:
                    continue
                month_start = started_at.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                key = (entity_name, month_start)
                bucket = entity_timeline_map[key]
                bucket["entity_name"] = entity_name
                bucket["month_label"] = month_start.strftime("%m.%y")
                bucket["month_start"] = month_start.date().isoformat()
                bucket["objects_count"] += 1
                if release_id:
                    bucket["release_ids"].add(release_id)

            entity_timeline = sorted(
                [
                    {
                        "entity_name": bucket["entity_name"],
                        "month_label": bucket["month_label"],
                        "month_start": bucket["month_start"],
                        "objects_count": bucket["objects_count"],
                        "releases_count": len(bucket["release_ids"]),
                    }
                    for bucket in entity_timeline_map.values()
                ],
                key=lambda row: (row["month_start"], row["entity_name"]),
            )

            recent_releases = release_rows_enriched[:8]
            exception_releases = sorted(
                exception_summary_rows,
                key=lambda row: row.get("started_at") or datetime.min,
                reverse=True,
            )[:10]

            regular_release_rows = [row for row in release_rows_enriched if row.get("release_bucket") == "release"]
            regular_release_count = len(regular_release_rows)
            regular_objects_count = sum(int(row.get("objects_count") or 0) for row in regular_release_rows)
            regular_tasks_count = sum(int(row.get("tasks_count") or 0) for row in regular_release_rows)
            regular_hours_total = sum(float(row.get("minutes_total") or 0) for row in regular_release_rows) / 60.0

            hotfix_rows = [row for row in exception_summary_rows if row.get("release_bucket") == "hotfix"]
            outside_rows = [row for row in exception_summary_rows if row.get("release_bucket") == "outside_release"]
            exception_objects_total = sum(int(row.get("objects_count") or 0) for row in exception_summary_rows)
            exception_tasks_total = len(exception_summary_rows)
            exception_hours_total = sum(float(row.get("hours_total") or 0) for row in exception_summary_rows)

            type_rows = [
                {
                    "release_bucket": "release",
                    "releases_count": regular_release_count,
                    "objects_count": regular_objects_count,
                    "tasks_count": regular_tasks_count,
                    "hours_total": round(regular_hours_total, 1),
                },
                {
                    "release_bucket": "hotfix",
                    "releases_count": len(hotfix_rows),
                    "objects_count": sum(int(row.get("objects_count") or 0) for row in hotfix_rows),
                    "tasks_count": len(hotfix_rows),
                    "hours_total": round(sum(float(row.get("hours_total") or 0) for row in hotfix_rows), 1),
                },
                {
                    "release_bucket": "outside_release",
                    "releases_count": len(outside_rows),
                    "objects_count": sum(int(row.get("objects_count") or 0) for row in outside_rows),
                    "tasks_count": len(outside_rows),
                    "hours_total": round(sum(float(row.get("hours_total") or 0) for row in outside_rows), 1),
                },
            ]
            type_rows = [row for row in type_rows if row["releases_count"] > 0 or row["objects_count"] > 0 or row["tasks_count"] > 0]

            exception_creator_stats = defaultdict(
                lambda: {"creator": "Не указан", "count": 0, "objects_count": 0, "hours_total": 0.0}
            )
            exception_type_stats = defaultdict(
                lambda: {"release_type": "Внерелиз", "count": 0, "objects_count": 0, "hours_total": 0.0}
            )
            exception_entity_stats = defaultdict(
                lambda: {"entity_name": "Без сущности", "count": 0, "objects_count": 0}
            )
            exception_direction_stats = defaultdict(
                lambda: {"direction": "Не указано", "count": 0, "objects_count": 0, "hours_total": 0.0}
            )
            exception_day_stats = defaultdict(
                lambda: {"day": "", "count": 0, "hotfix_count": 0, "outside_release_count": 0, "objects_count": 0}
            )

            for row in exception_summary_rows:
                creator = row.get("initiated_by") or "Не указан"
                creator_bucket = exception_creator_stats[creator]
                creator_bucket["creator"] = creator
                creator_bucket["count"] += 1
                creator_bucket["objects_count"] += int(row.get("objects_count") or 0)
                creator_bucket["hours_total"] += float(row.get("hours_total") or 0)

                release_type_label = row.get("release_type") or "Внерелиз"
                type_bucket = exception_type_stats[release_type_label]
                type_bucket["release_type"] = release_type_label
                type_bucket["count"] += 1
                type_bucket["objects_count"] += int(row.get("objects_count") or 0)
                type_bucket["hours_total"] += float(row.get("hours_total") or 0)

                direction = row.get("direction") or "Не указано"
                direction_bucket = exception_direction_stats[direction]
                direction_bucket["direction"] = direction
                direction_bucket["count"] += 1
                direction_bucket["objects_count"] += int(row.get("objects_count") or 0)
                direction_bucket["hours_total"] += float(row.get("hours_total") or 0)

                for entity_name in row.get("entity_names") or ["Без сущности"]:
                    entity_bucket = exception_entity_stats[entity_name]
                    entity_bucket["entity_name"] = entity_name
                    entity_bucket["count"] += 1
                    entity_bucket["objects_count"] += int(row.get("objects_count") or 0)

                started_at = row.get("started_at")
                if started_at:
                    day_key = started_at.date().isoformat()
                    day_bucket = exception_day_stats[day_key]
                    day_bucket["day"] = day_key
                    day_bucket["count"] += 1
                    day_bucket["objects_count"] += int(row.get("objects_count") or 0)
                    if row.get("release_bucket") == "hotfix":
                        day_bucket["hotfix_count"] += 1
                    else:
                        day_bucket["outside_release_count"] += 1

            exception_insights = {
                "share_of_unplanned": round(
                    ((len(hotfix_rows) + len(outside_rows)) / max(regular_release_count + len(hotfix_rows) + len(outside_rows), 1)) * 100,
                    1,
                ),
                "top_type": (
                    sorted(
                        [
                            {
                                **row,
                                "hours_total": round(row["hours_total"], 1),
                            }
                            for row in exception_type_stats.values()
                        ],
                        key=lambda row: (-row["count"], -row["objects_count"], row["release_type"]),
                    )[0]
                    if exception_type_stats
                    else None
                ),
                "top_creator": (
                    sorted(
                        [
                            {
                                **row,
                                "hours_total": round(row["hours_total"], 1),
                            }
                            for row in exception_creator_stats.values()
                        ],
                        key=lambda row: (-row["count"], -row["objects_count"], row["creator"]),
                    )[0]
                    if exception_creator_stats
                    else None
                ),
                "top_entity": (
                    sorted(
                        list(exception_entity_stats.values()),
                        key=lambda row: (-row["count"], -row["objects_count"], row["entity_name"]),
                    )[0]
                    if exception_entity_stats
                    else None
                ),
                "by_type": sorted(
                    [
                        {
                            **row,
                            "hours_total": round(row["hours_total"], 1),
                        }
                        for row in exception_type_stats.values()
                    ],
                    key=lambda row: (-row["count"], -row["objects_count"], row["release_type"]),
                )[:6],
                "by_creator": sorted(
                    [
                        {
                            **row,
                            "hours_total": round(row["hours_total"], 1),
                        }
                        for row in exception_creator_stats.values()
                    ],
                    key=lambda row: (-row["count"], -row["objects_count"], row["creator"]),
                )[:8],
                "by_entity": sorted(
                    list(exception_entity_stats.values()),
                    key=lambda row: (-row["count"], -row["objects_count"], row["entity_name"]),
                )[:8],
                "by_direction": sorted(
                    [
                        {
                            **row,
                            "hours_total": round(row["hours_total"], 1),
                        }
                        for row in exception_direction_stats.values()
                    ],
                    key=lambda row: (-row["count"], -row["objects_count"], row["direction"]),
                )[:8],
                "by_day": sorted(
                    list(exception_day_stats.values()),
                    key=lambda row: row["day"],
                )[-12:],
            }

            cadence_map = {}
            for row in cadence:
                item = dict(row)
                cadence_map[item["week_start"]] = {
                    "week_label": item["week_label"],
                    "week_start": item["week_start"],
                    "releases_count": int(item.get("releases_count") or 0),
                    "hotfix_count": int(item.get("hotfix_count") or 0),
                    "outside_release_count": int(item.get("outside_release_count") or 0),
                    "regular_release_count": int(item.get("regular_release_count") or 0),
                    "objects_count": int(item.get("objects_count") or 0),
                    "tasks_count": int(item.get("tasks_count") or 0),
                }
            for row in exception_summary_rows:
                started_at = row.get("started_at")
                if not started_at:
                    continue
                week_start_dt = started_at - timedelta(days=started_at.weekday())
                week_start_dt = week_start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                week_start = week_start_dt.date().isoformat()
                bucket = cadence_map.setdefault(
                    week_start,
                    {
                        "week_label": week_start_dt.strftime("%d.%m"),
                        "week_start": week_start,
                        "releases_count": 0,
                        "hotfix_count": 0,
                        "outside_release_count": 0,
                        "regular_release_count": 0,
                        "objects_count": 0,
                        "tasks_count": 0,
                    },
                )
                bucket["releases_count"] += 1
                bucket["objects_count"] += int(row.get("objects_count") or 0)
                bucket["tasks_count"] += int(row.get("tasks_count") or 0)
                if row.get("release_bucket") == "hotfix":
                    bucket["hotfix_count"] += 1
                elif row.get("release_bucket") == "outside_release":
                    bucket["outside_release_count"] += 1
            cadence_rows = sorted(cadence_map.values(), key=lambda row: row["week_start"])

            release_day_values = {
                row.get("started_at").date().isoformat()
                for row in regular_release_rows
                if row.get("started_at")
            }
            release_day_values.update(
                row.get("started_at").date().isoformat()
                for row in exception_summary_rows
                if row.get("started_at")
            )

            key_blocks_empty = not top_entities or not top_tables or not top_users or not recent_releases
            if debug or key_blocks_empty:
                debug_counts = conn.execute(
                    text(
                        base
                        + """
                        SELECT 'raw_rel' AS step, COUNT(*) AS cnt FROM raw_rel
                        UNION ALL
                        SELECT 'raw_ro' AS step, COUNT(*) AS cnt FROM raw_ro
                        UNION ALL
                        SELECT 'release_tasks' AS step, COUNT(*) AS cnt FROM release_tasks
                        UNION ALL
                        SELECT 'task_ids' AS step, COUNT(*) AS cnt FROM task_ids
                        UNION ALL
                        SELECT 'task_custom_base' AS step, COUNT(*) AS cnt FROM task_custom_base
                        UNION ALL
                        SELECT 'task_custom' AS step, COUNT(*) AS cnt FROM task_custom
                        UNION ALL
                        SELECT 'task_custom_norm' AS step, COUNT(*) AS cnt FROM task_custom_norm
                        UNION ALL
                        SELECT 'task_custom_norm_with_bucket' AS step, COUNT(*) AS cnt
                        FROM task_custom_norm
                        WHERE custom_bucket IS NOT NULL
                        UNION ALL
                        SELECT 'release_match_by_date' AS step, COUNT(*) AS cnt FROM release_match_by_date
                        UNION ALL
                        SELECT 'rel' AS step, COUNT(*) AS cnt FROM rel
                        UNION ALL
                        SELECT 'ro' AS step, COUNT(*) AS cnt FROM ro
                        UNION ALL
                        SELECT 'task_fact' AS step, COUNT(*) AS cnt FROM task_fact
                        UNION ALL
                        SELECT 'release_rollup' AS step, COUNT(*) AS cnt FROM release_rollup
                        ORDER BY step
                        """
                    ),
                    params,
                ).mappings().all()
                print(
                    "ℹ️ /api/admin/reports/releases debug:",
                    {
                        "days": days,
                        "top_entities": len(top_entities),
                        "top_tables": len(top_tables),
                        "top_users": len(top_users),
                        "top_creators": len(top_creators),
                        "top_initiators": len(top_initiators),
                        "entity_timeline": len(entity_timeline),
                        "recent_releases": len(recent_releases),
                        "exception_releases": len(exception_releases),
                        "counts": [dict(row) for row in debug_counts],
                        "top_entity_sample": dict(top_entities[0]) if top_entities else None,
                        "top_table_sample": dict(top_tables[0]) if top_tables else None,
                        "top_user_sample": dict(top_users[0]) if top_users else None,
                        "recent_release_sample": dict(recent_releases[0]) if recent_releases else None,
                    },
                )

        summary_row = summary or {}
        if not locals().get("cadence_rows"):
            cadence_rows = [dict(row) for row in cadence]
        if not locals().get("type_rows"):
            type_rows = [dict(row) for row in type_breakdown]
        heatmap_rows = [dict(row) for row in weekday_heatmap]
        entities_rows = [dict(row) for row in top_entities]
        tables_rows = [dict(row) for row in top_tables]
        users_rows = [dict(row) for row in top_users]
        creators_rows = [dict(row) for row in top_creators]
        initiators_rows = [dict(row) for row in top_initiators]
        system_rows = [dict(row) for row in system_breakdown]
        timeline_rows = [dict(row) for row in entity_timeline]
        recent_rows = [dict(row) for row in recent_releases]
        exception_rows = [dict(row) for row in exception_releases]

        return {
            "days": days,
            "summary": {
                "releases_count": regular_release_count + len(hotfix_rows) + len(outside_rows),
                "release_count": regular_release_count,
                "hotfix_count": len(hotfix_rows),
                "outside_release_count": len(outside_rows),
                "objects_count": regular_objects_count + exception_objects_total,
                "click_objects_count": int(summary_row.get("click_objects_count") or 0),
                "gp_objects_count": int(summary_row.get("gp_objects_count") or 0),
                "tasks_count": regular_tasks_count + exception_tasks_total,
                "hours_total": round(regular_hours_total + exception_hours_total, 1),
                "initiators_count": int(summary_row.get("initiators_count") or 0),
                "release_days_count": len(release_day_values),
                "avg_objects_per_release": float(summary_row.get("avg_objects_per_release") or 0),
                "avg_duration_minutes": float(summary_row.get("avg_duration_minutes") or 0),
            },
            "cadence": cadence_rows,
            "weekday_heatmap": heatmap_rows,
            "type_breakdown": type_rows,
            "system_breakdown": system_rows,
            "top_entities": entities_rows,
            "top_tables": tables_rows,
            "top_users": users_rows,
            "top_creators": creators_rows,
            "top_initiators": initiators_rows,
            "entity_timeline": timeline_rows,
            "recent_releases": recent_rows,
            "exception_releases": exception_rows,
            "exception_insights": exception_insights,
            "focus": {
                "top_entity": entities_rows[0] if entities_rows else None,
                "top_table": tables_rows[0] if tables_rows else None,
                "top_user": users_rows[0] if users_rows else None,
            },
        }
    except Exception as e:
        print("❌ /api/admin/reports/releases error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить аналитику релизов")


@router.get("/api/admin/reports/incidents")
def get_admin_incident_reports(
    request: Request,
    days: int = Query(180, ge=1, le=3650),
):
    user = get_current_user_from_request(request)
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    try:
        detail_limit = 160
        with engine.connect() as conn:
            incident_rows = conn.execute(
                text(
                    """
                    SELECT
                        issue_id,
                        issue_type,
                        project_name,
                        summary,
                        state_name,
                        priority_name,
                        author_name,
                        assignee_name,
                        team_name,
                        direction_name,
                        dashboard_name,
                        component_name,
                        trigger_dttm,
                        incident_start_dttm,
                        detected_dttm,
                        incident_reason_name,
                        work_finished_dttm,
                        entity_name,
                        table_name_raw,
                        table_schema,
                        table_name,
                        alert_source,
                        spent_time_text,
                        actual_effort_text,
                        estimated_effort_text,
                        ai_saving_text,
                        created_at_yt,
                        updated_at_yt,
                        resolved_at_yt
                    FROM tech_etl.yt_incidents
                    WHERE COALESCE(incident_start_dttm, detected_dttm, trigger_dttm, created_at_yt, dttm_loaded)
                          >= (now() - (:days || ' days')::interval)
                    ORDER BY COALESCE(incident_start_dttm, detected_dttm, trigger_dttm, created_at_yt) DESC NULLS LAST,
                             issue_id DESC
                    """
                ),
                {"days": days},
            ).mappings().all()

            incident_ids = [row["issue_id"] for row in incident_rows if row.get("issue_id")]
            detail_issue_ids = incident_ids[:detail_limit]
            details_by_issue: dict[str, dict[str, Any]] = {}
            link_rows = []
            history_rows = []
            if detail_issue_ids:
                detail_query = text(
                    """
                    SELECT
                        issue_id,
                        description_text,
                        event_description,
                        root_cause,
                        fix_actions,
                        preventive_actions
                    FROM tech_etl.yt_incidents
                    WHERE issue_id IN :issue_ids
                    """
                ).bindparams(bindparam("issue_ids", expanding=True))
                detail_rows = conn.execute(detail_query, {"issue_ids": detail_issue_ids}).mappings().all()
                details_by_issue = {
                    str(row.get("issue_id")): dict(row)
                    for row in detail_rows
                    if row.get("issue_id")
                }
            if incident_ids:
                link_query = text(
                    """
                    SELECT
                        issue_id,
                        linked_issue_id,
                        linked_issue_type,
                        linked_issue_summary,
                        link_type,
                        dttm_loaded
                    FROM tech_etl.yt_incident_links
                    WHERE issue_id IN :issue_ids
                    ORDER BY issue_id, linked_issue_id
                    """
                ).bindparams(bindparam("issue_ids", expanding=True))
                link_rows = conn.execute(link_query, {"issue_ids": incident_ids}).mappings().all()

                history_query = text(
                    """
                    SELECT
                        issue_id,
                        event_dttm,
                        author_name,
                        event_type,
                        field_name,
                        old_value,
                        new_value,
                        comment_text,
                        dttm_loaded
                    FROM tech_etl.yt_incident_history
                    WHERE issue_id IN :issue_ids
                    ORDER BY issue_id, event_dttm
                    """
                ).bindparams(bindparam("issue_ids", expanding=True))
                history_rows = conn.execute(history_query, {"issue_ids": incident_ids}).mappings().all()

        links_by_issue: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in link_rows:
            links_by_issue[str(row.get("issue_id"))].append(dict(row))

        history_by_issue: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in history_rows:
            history_by_issue[str(row.get("issue_id"))].append(dict(row))

        incidents = []
        for row in incident_rows:
            issue_id = str(row.get("issue_id") or "")
            start_at = row.get("incident_start_dttm") or row.get("detected_dttm") or row.get("trigger_dttm") or row.get("created_at_yt")
            end_at = row.get("work_finished_dttm") or row.get("resolved_at_yt") or row.get("updated_at_yt")
            duration_minutes = 0
            if start_at and end_at and isinstance(start_at, datetime) and isinstance(end_at, datetime) and end_at >= start_at:
                duration_minutes = int(round((end_at - start_at).total_seconds() / 60.0))

            actual_effort_minutes = _parse_effort_minutes(row.get("actual_effort_text")) or _parse_effort_minutes(row.get("spent_time_text"))
            estimated_effort_minutes = _parse_effort_minutes(row.get("estimated_effort_text"))
            ai_saving_minutes = _parse_effort_minutes(row.get("ai_saving_text"))

            entity_names = _split_rich_multivalue(row.get("entity_name")) or ["Без сущности"]
            table_names = _split_rich_multivalue(row.get("table_name")) or _split_rich_multivalue(row.get("table_name_raw"))
            table_schema = str(row.get("table_schema") or "").strip() or None
            table_fqns = [f"{table_schema}.{name}" for name in table_names] if table_schema and table_names else table_names

            link_items = links_by_issue.get(issue_id, [])
            history_items = history_by_issue.get(issue_id, [])
            details = details_by_issue.get(issue_id, {})
            linked_issue_types = sorted({str(item.get("linked_issue_type") or "Не указано").strip() or "Не указано" for item in link_items})
            link_types = sorted({str(item.get("link_type") or "Не указано").strip() or "Не указано" for item in link_items})
            state_changes_count = sum(1 for item in history_items if str(item.get("event_type") or "").lower() == "state_change")
            assignee_changes_count = sum(1 for item in history_items if str(item.get("event_type") or "").lower() == "assignee_change")
            comments_count = sum(1 for item in history_items if str(item.get("comment_text") or "").strip())
            preventive_filled = bool(_normalize_rich_text(details.get("preventive_actions")))

            incidents.append(
                {
                    "issue_id": issue_id,
                    "link": _build_ytrack_link(issue_id),
                    "issue_type": row.get("issue_type"),
                    "project_name": row.get("project_name"),
                    "summary": str(row.get("summary") or "").strip(),
                    "description_text": _normalize_rich_text(details.get("description_text")),
                    "state_name": row.get("state_name"),
                    "status_bucket": _incident_status_bucket(row.get("state_name")),
                    "priority_name": row.get("priority_name") or "Normal",
                    "author_name": row.get("author_name") or "Не указан",
                    "assignee_name": row.get("assignee_name") or row.get("author_name") or "Не указан",
                    "team_name": row.get("team_name") or "Не указана",
                    "direction_name": row.get("direction_name") or "Не указано",
                    "dashboard_name": row.get("dashboard_name") or "Не указан",
                    "component_name": row.get("component_name") or "Не указан",
                    "trigger_dttm": serialize_datetime(row.get("trigger_dttm")),
                    "incident_start_dttm": serialize_datetime(start_at),
                    "detected_dttm": serialize_datetime(row.get("detected_dttm")),
                    "incident_reason_name": row.get("incident_reason_name") or "Не указана",
                    "work_finished_dttm": serialize_datetime(end_at),
                    "entity_names": entity_names,
                    "table_schema": table_schema,
                    "table_names": table_names,
                    "table_fqns": table_fqns,
                    "primary_table_fqn": table_fqns[0] if len(table_fqns) == 1 else None,
                    "affected_tables_count": len(table_fqns),
                    "event_description": _normalize_rich_text(details.get("event_description")),
                    "alert_source": row.get("alert_source") or "Не указан",
                    "root_cause": _normalize_rich_text(details.get("root_cause")) or _normalize_rich_text(row.get("incident_reason_name")),
                    "fix_actions": _normalize_rich_text(details.get("fix_actions")),
                    "preventive_actions": _normalize_rich_text(details.get("preventive_actions")),
                    "actual_effort_minutes": actual_effort_minutes,
                    "estimated_effort_minutes": estimated_effort_minutes,
                    "ai_saving_minutes": ai_saving_minutes,
                    "duration_minutes": duration_minutes,
                    "created_at_yt": serialize_datetime(row.get("created_at_yt")),
                    "updated_at_yt": serialize_datetime(row.get("updated_at_yt")),
                    "resolved_at_yt": serialize_datetime(row.get("resolved_at_yt")),
                    "linked_issues_count": len(link_items),
                    "linked_issue_types": linked_issue_types,
                    "link_types": link_types,
                    "history_events_count": len(history_items),
                    "state_changes_count": state_changes_count,
                    "assignee_changes_count": assignee_changes_count,
                    "comments_count": comments_count,
                    "preventive_filled": preventive_filled,
                }
            )

        incidents.sort(key=lambda row: row.get("incident_start_dttm") or "", reverse=True)

        unique_entities = set()
        unique_tables = set()
        resolved_count = 0
        open_count = 0
        linked_count = 0
        preventive_count = 0
        duration_values = []
        effort_values = []
        saving_values = []
        timeline_day = defaultdict(lambda: {"day": "", "count": 0, "resolved_count": 0, "open_count": 0, "duration_minutes": 0})
        timeline_week = defaultdict(lambda: {"week_start": "", "week_label": "", "count": 0, "resolved_count": 0, "open_count": 0, "duration_minutes": 0})

        def _agg(rows: list[dict[str, Any]], key_name: str, label_field: str) -> list[dict[str, Any]]:
            stats = defaultdict(lambda: {label_field: "Не указано", "count": 0, "duration_minutes": 0, "effort_minutes": 0, "open_count": 0, "resolved_count": 0, "objects_count": 0})
            for item in rows:
                label = str(item.get(key_name) or "Не указано").strip() or "Не указано"
                bucket = stats[label]
                bucket[label_field] = label
                bucket["count"] += 1
                bucket["duration_minutes"] += int(item.get("duration_minutes") or 0)
                bucket["effort_minutes"] += int(item.get("actual_effort_minutes") or 0)
                bucket["objects_count"] += int(item.get("affected_tables_count") or 0)
                if item.get("status_bucket") == "resolved":
                    bucket["resolved_count"] += 1
                else:
                    bucket["open_count"] += 1
            result = []
            for bucket in stats.values():
                count = max(bucket["count"], 1)
                result.append(
                    {
                        **bucket,
                        "avg_duration_hours": round((bucket["duration_minutes"] / count) / 60.0, 2),
                        "effort_hours": round(bucket["effort_minutes"] / 60.0, 2),
                    }
                )
            result.sort(key=lambda row: (-row["count"], -row["objects_count"], str(row[label_field])))
            return result

        for item in incidents:
            unique_entities.update(item.get("entity_names") or [])
            unique_tables.update(item.get("table_fqns") or [])
            if item.get("status_bucket") == "resolved":
                resolved_count += 1
            else:
                open_count += 1
            if item.get("linked_issues_count"):
                linked_count += 1
            if item.get("preventive_filled"):
                preventive_count += 1
            if item.get("duration_minutes"):
                duration_values.append(int(item["duration_minutes"]))
            if item.get("actual_effort_minutes"):
                effort_values.append(int(item["actual_effort_minutes"]))
            if item.get("ai_saving_minutes"):
                saving_values.append(int(item["ai_saving_minutes"]))

            start_iso = str(item.get("incident_start_dttm") or "")[:10]
            if start_iso:
                day_bucket = timeline_day[start_iso]
                day_bucket["day"] = start_iso
                day_bucket["count"] += 1
                day_bucket["duration_minutes"] += int(item.get("duration_minutes") or 0)
                if item.get("status_bucket") == "resolved":
                    day_bucket["resolved_count"] += 1
                else:
                    day_bucket["open_count"] += 1

            week_start = _incident_week_start(item.get("incident_start_dttm"))
            if week_start:
                week_bucket = timeline_week[week_start]
                week_bucket["week_start"] = week_start
                week_bucket["week_label"] = week_start[5:]
                week_bucket["count"] += 1
                week_bucket["duration_minutes"] += int(item.get("duration_minutes") or 0)
                if item.get("status_bucket") == "resolved":
                    week_bucket["resolved_count"] += 1
                else:
                    week_bucket["open_count"] += 1

        reason_rows = _agg(incidents, "incident_reason_name", "reason")
        source_rows = _agg(incidents, "alert_source", "source")
        direction_rows = _agg(incidents, "direction_name", "direction")
        assignee_rows = _agg(incidents, "assignee_name", "assignee")
        dashboard_rows = _agg(incidents, "dashboard_name", "dashboard")
        component_rows = _agg(incidents, "component_name", "component")

        entity_stats = defaultdict(lambda: {"entity_name": "Без сущности", "count": 0, "duration_minutes": 0, "effort_minutes": 0, "objects_count": 0})
        for item in incidents:
            for entity_name in item.get("entity_names") or ["Без сущности"]:
                bucket = entity_stats[entity_name]
                bucket["entity_name"] = entity_name
                bucket["count"] += 1
                bucket["duration_minutes"] += int(item.get("duration_minutes") or 0)
                bucket["effort_minutes"] += int(item.get("actual_effort_minutes") or 0)
                bucket["objects_count"] += int(item.get("affected_tables_count") or 0)
        entity_rows = sorted(
            [
                {
                    **bucket,
                    "avg_duration_hours": round((bucket["duration_minutes"] / max(bucket["count"], 1)) / 60.0, 2),
                    "effort_hours": round(bucket["effort_minutes"] / 60.0, 2),
                }
                for bucket in entity_stats.values()
            ],
            key=lambda row: (-row["count"], -row["objects_count"], row["entity_name"]),
        )

        link_type_stats = defaultdict(lambda: {"link_type": "Не указано", "count": 0})
        linked_issue_type_stats = defaultdict(lambda: {"linked_issue_type": "Не указано", "count": 0})
        for row in link_rows:
            link_type = str(row.get("link_type") or "Не указано").strip() or "Не указано"
            linked_issue_type = str(row.get("linked_issue_type") or "Не указано").strip() or "Не указано"
            link_type_stats[link_type]["link_type"] = link_type
            link_type_stats[link_type]["count"] += 1
            linked_issue_type_stats[linked_issue_type]["linked_issue_type"] = linked_issue_type
            linked_issue_type_stats[linked_issue_type]["count"] += 1

        link_type_rows = sorted(link_type_stats.values(), key=lambda row: (-row["count"], row["link_type"]))
        linked_issue_type_rows = sorted(linked_issue_type_stats.values(), key=lambda row: (-row["count"], row["linked_issue_type"]))
        week_rows = sorted(timeline_week.values(), key=lambda row: row["week_start"])[-12:]
        day_rows = sorted(timeline_day.values(), key=lambda row: row["day"])[-24:]

        longest_incident = max(
            incidents,
            key=lambda row: (int(row.get("duration_minutes") or 0), int(row.get("actual_effort_minutes") or 0)),
            default=None,
        )

        summary = {
            "total_incidents": len(incidents),
            "resolved_count": resolved_count,
            "open_count": open_count,
            "unique_entities": len(unique_entities),
            "unique_tables": len(unique_tables),
            "hours_spent_total": round(sum(effort_values) / 60.0, 2),
            "hours_saved_total": round(sum(saving_values) / 60.0, 2),
            "avg_resolution_hours": round((sum(duration_values) / max(len(duration_values), 1)) / 60.0, 2) if duration_values else 0,
            "linked_count": linked_count,
            "preventive_count": preventive_count,
            "preventive_share": round((preventive_count / max(len(incidents), 1)) * 100, 2) if incidents else 0,
        }

        return {
            "days": days,
            "source": "dev",
            "summary": summary,
            "weekly_timeline": week_rows,
            "daily_timeline": day_rows,
            "reason_breakdown": reason_rows[:10],
            "source_breakdown": source_rows[:8],
            "direction_breakdown": direction_rows[:8],
            "entity_breakdown": entity_rows[:10],
            "assignee_breakdown": assignee_rows[:8],
            "dashboard_breakdown": dashboard_rows[:8],
            "component_breakdown": component_rows[:8],
            "link_type_breakdown": link_type_rows[:8],
            "linked_issue_type_breakdown": linked_issue_type_rows[:8],
            "focus": {
                "top_reason": reason_rows[0] if reason_rows else None,
                "top_source": source_rows[0] if source_rows else None,
                "top_direction": direction_rows[0] if direction_rows else None,
                "top_entity": entity_rows[0] if entity_rows else None,
                "longest_incident": longest_incident,
            },
            "incidents": incidents,
        }
    except Exception as e:
        print("❌ /api/admin/reports/incidents error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить аналитику инцидентов")


@router.post("/api/admin/reports/export-pdf")
async def export_admin_report_pdf(request: Request):
    _require_admin(request)
    try:
        payload = await request.json()
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Некорректный payload для PDF")
        pdf_bytes = _render_report_pdf(payload)
        filename = str(payload.get("filename") or "reports-export.pdf").strip() or "reports-export.pdf"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)
    except HTTPException:
        raise
    except Exception as e:
        print("❌ /api/admin/reports/export-pdf error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось экспортировать PDF")


@router.get("/api/admin/dev-meta/status")
def get_admin_dev_meta_status(request: Request):
    _require_dev_meta_role(request)
    return get_dev_meta_status(
        engine=engine,
        base_dir=BASE_DIR,
        prod_root_value=CLICK_META_DIR,
        dev_root_value=DEV_CLICK_META_DIR,
        airflow_base_url=AIRFLOW_DEV_BASE_URL,
        airflow_dag_id=AIRFLOW_DEV_DAG_ID,
        lock_ttl_minutes=DEV_META_LOCK_TTL_MIN,
        dev_database_url=DEV_DATABASE_URL,
        deploy_host=DEV_META_DEPLOY_HOST,
        deploy_user=DEV_META_DEPLOY_USER,
        deploy_base_dir=DEV_META_DEPLOY_BASE_DIR,
    )

@router.get("/api/admin/dev-meta/files")
def get_admin_dev_meta_files(request: Request, schema_name: str = Query("dm")):
    _require_dev_meta_role(request)
    if schema_name not in {"dm", "dm_view"}:
        raise HTTPException(status_code=400, detail="schema_name must be dm or dm_view")
    return get_dev_meta_files(
        engine=engine,
        base_dir=BASE_DIR,
        prod_root_value=CLICK_META_DIR,
        dev_root_value=DEV_CLICK_META_DIR,
        schema_name=schema_name,
        deploy_host=DEV_META_DEPLOY_HOST,
        deploy_port=DEV_META_DEPLOY_PORT,
        deploy_user=DEV_META_DEPLOY_USER,
        deploy_password=DEV_META_DEPLOY_PASSWORD,
        deploy_base_dir=DEV_META_DEPLOY_BASE_DIR,
        deploy_ssh_key_path=DEV_META_DEPLOY_SSH_KEY_PATH,
        deploy_strict_host_key=DEV_META_DEPLOY_STRICT_HOST_KEY,
    )


@router.post("/api/admin/dev-meta/file")
def get_admin_dev_meta_file(payload: DevMetaFilePayload, request: Request):
    _require_dev_meta_role(request)
    root = DEV_CLICK_META_DIR if payload.source != "prod" else CLICK_META_DIR
    try:
        return read_dev_meta_file(
            base_dir=BASE_DIR,
            root_value=root,
            schema_name=payload.schema_name,
            file_name=payload.file_name,
        )
    except FileNotFoundError:
        if payload.source != "prod":
            try:
                return read_remote_dev_meta_file(
                    schema_name=payload.schema_name,
                    file_name=payload.file_name,
                    host=DEV_META_DEPLOY_HOST,
                    port=DEV_META_DEPLOY_PORT,
                    user=DEV_META_DEPLOY_USER,
                    password=DEV_META_DEPLOY_PASSWORD,
                    remote_base_dir=DEV_META_DEPLOY_BASE_DIR,
                    ssh_key_path=DEV_META_DEPLOY_SSH_KEY_PATH,
                    strict_host_key=DEV_META_DEPLOY_STRICT_HOST_KEY,
                )
            except Exception:
                pass
        raise HTTPException(status_code=404, detail="Файл не найден")


@router.post("/api/admin/dev-meta/generate")
def generate_admin_dev_meta(payload: DevMetaGeneratePayload, request: Request):
    user = _require_dev_meta_role(request)
    try:
        result = generate_dev_meta_yaml(
            database_url=DEV_DATABASE_URL or DATABASE_URL,
            schema_name_gp=payload.schema_name_gp,
            object_name=payload.object_name,
            schema_name_click=payload.schema_name_click,
            greenplum_table_name=payload.greenplum_table_name,
            order_by=payload.order_by,
            dag_tags=payload.dag_tags,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    payload_data = dict(result.get("payload") or {})
    attributes = payload_data.pop("attributes", [])
    payload_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    payload_data["created_by"] = user.email or user.username or "unknown"
    payload_data["created_role"] = user.role or "unknown"
    payload_data["attributes"] = attributes
    result["payload"] = payload_data
    result["content"] = yaml.dump(
        payload_data,
        default_flow_style=False,
        sort_keys=False,
        allow_unicode=True,
        width=float("inf"),
    )
    return {"status": "ok", **result}


@router.post("/api/admin/dev-meta/lock")
def lock_admin_dev_meta_file(payload: DevMetaLockPayload, request: Request):
    user = _require_dev_meta_role(request)
    try:
        return acquire_dev_meta_lock(
            engine=engine,
            schema_name=payload.schema_name,
            file_name=payload.file_name,
            author=user.email,
            ttl_minutes=DEV_META_LOCK_TTL_MIN,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))


@router.post("/api/admin/dev-meta/unlock")
def unlock_admin_dev_meta_file(payload: DevMetaLockPayload, request: Request):
    user = _require_dev_meta_role(request)
    release_dev_meta_lock(
        engine=engine,
        schema_name=payload.schema_name,
        file_name=payload.file_name,
        author=user.email,
    )
    return {"status": "ok"}


@router.post("/api/admin/dev-meta/validate")
def validate_admin_dev_meta(payload: DevMetaSavePayload, request: Request):
    _require_dev_meta_role(request)
    return validate_dev_meta_content(
        content=payload.content,
        schema_name=payload.schema_name,
        dev_database_url=DEV_DATABASE_URL,
    )


@router.post("/api/admin/dev-meta/save")
def save_admin_dev_meta(payload: DevMetaSavePayload, request: Request):
    user = _require_dev_meta_role(request)
    try:
        result = save_dev_meta_file(
            engine=engine,
            base_dir=BASE_DIR,
            dev_root_value=DEV_CLICK_META_DIR,
            schema_name=payload.schema_name,
            file_name=payload.file_name,
            content=payload.content,
            author=user.email,
            dev_database_url=DEV_DATABASE_URL,
            task_id=payload.task_id,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", **result}


@router.post("/api/admin/meta-workspace/mr")
def create_admin_meta_workspace_mr(payload: MetaWorkspaceMrPayload, request: Request):
    user = _require_meta_workspace_role(request)
    try:
        result = create_meta_workspace_mr(
            engine=engine,
            base_dir=BASE_DIR,
            entity_dev_root_value=DEV_ENTITY_META_DIR,
            click_dev_root_value=DEV_CLICK_META_DIR,
            git_repo_value=ENTITY_META_GIT_REPO,
            entity_git_root_value=ENTITY_META_GIT_META_ROOT,
            click_git_root_value=CLICK_META_GIT_ROOT,
            gitlab_token=GITLAB_TOKEN,
            gitlab_project=GITLAB_PROJECT,
            gitlab_api_url=GITLAB_API_URL,
            gitlab_ssl_verify=GITLAB_SSL_VERIFY,
            task_id=payload.task_id,
            release_branch=payload.release_branch,
            branch_name=payload.branch_name,
            author=user.email,
        )
        if result.get("mr_url") and YOUTRACK_URL and YOUTRACK_TOKEN:
            try:
                add_ytrack_issue_comment(
                    base_url=YOUTRACK_URL,
                    token=YOUTRACK_TOKEN,
                    issue_id=payload.task_id,
                    ssl_verify=YOUTRACK_SSL_VERIFY,
                    text=(
                        "MR создан для инженерных изменений.\n"
                        f"Ссылка: {result.get('mr_url')}\n"
                        f"Ветка: {result.get('feature_branch') or '—'} -> {result.get('release_branch') or 'main'}"
                    ),
                )
                result["task_link_attached"] = True
            except Exception as exc:
                result["task_link_attached"] = False
                result["task_link_error"] = str(exc)
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", **result}


@router.get("/api/admin/meta-workspace/branches")
def get_admin_meta_workspace_branches(request: Request):
    _require_meta_workspace_role(request)
    try:
        return list_meta_workspace_branches(git_repo_value=ENTITY_META_GIT_REPO)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/api/admin/meta-workspace/branches")
def create_admin_meta_workspace_branch(payload: MetaWorkspaceCreateBranchPayload, request: Request):
    _require_meta_workspace_role(request)
    try:
        return create_meta_workspace_branch(
            git_repo_value=ENTITY_META_GIT_REPO,
            branch_name=payload.branch_name,
            base_branch=payload.base_branch,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/api/admin/meta-workspace/branch-catalog")
def get_admin_meta_workspace_branch_catalog(
    request: Request,
    branch_name: str = Query(...),
    base_branch: str = Query("main"),
):
    user = _require_meta_workspace_role(request)
    try:
        return _build_branch_catalog(
            git_repo_root=Path(ENTITY_META_GIT_REPO).resolve(),
            entity_git_root_value=ENTITY_META_GIT_META_ROOT,
            click_git_root_value=CLICK_META_GIT_ROOT,
            workspace_root_value=META_WORKSPACE_ROOT,
            workspace_owner=user.email,
            branch_name=branch_name,
            base_branch=base_branch,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось построить каталог ветки: {exc}")


@router.get("/api/admin/meta-workspace/branch-tree")
def get_admin_meta_workspace_branch_tree(
    request: Request,
    branch_name: str = Query(...),
    base_branch: str = Query("main"),
):
    user = _require_meta_workspace_role(request)
    try:
        return build_meta_workspace_branch_tree(
            git_repo_value=ENTITY_META_GIT_REPO,
            entity_git_root_value=ENTITY_META_GIT_META_ROOT,
            click_git_root_value=CLICK_META_GIT_ROOT,
            workspace_root_value=META_WORKSPACE_ROOT,
            workspace_owner=user.email,
            branch_name=branch_name,
            base_branch=base_branch,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось прочитать структуру ветки: {exc}")


@router.post("/api/admin/meta-workspace/branch-file")
def get_admin_meta_workspace_branch_file(payload: MetaWorkspaceBranchFilePayload, request: Request):
    user = _require_meta_workspace_role(request)
    try:
        return read_meta_workspace_branch_file(
            git_repo_value=ENTITY_META_GIT_REPO,
            workspace_root_value=META_WORKSPACE_ROOT,
            workspace_owner=user.email,
            branch_name=payload.branch_name,
            base_branch="main",
            file_path=payload.file_path,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/api/admin/meta-workspace/branch-gp-bundle")
def get_admin_meta_workspace_branch_gp_bundle(payload: MetaWorkspaceBranchGpBundlePayload, request: Request):
    user = _require_meta_workspace_role(request)
    try:
        return read_meta_workspace_branch_gp_bundle(
            git_repo_value=ENTITY_META_GIT_REPO,
            entity_git_root_value=ENTITY_META_GIT_META_ROOT,
            workspace_root_value=META_WORKSPACE_ROOT,
            workspace_owner=user.email,
            branch_name=payload.branch_name,
            base_branch="main",
            entity_name=payload.entity_name,
            schema_name=payload.schema_name,
            table_name=payload.table_name,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/api/admin/meta-workspace/branch-gp-bundle/save")
def save_admin_meta_workspace_branch_gp_bundle(payload: MetaWorkspaceBranchGpBundleSavePayload, request: Request):
    user = _require_meta_workspace_role(request)
    try:
        return save_meta_workspace_branch_gp_bundle(
            git_repo_value=ENTITY_META_GIT_REPO,
            entity_git_root_value=ENTITY_META_GIT_META_ROOT,
            workspace_root_value=META_WORKSPACE_ROOT,
            workspace_owner=user.email,
            branch_name=payload.branch_name,
            base_branch=payload.base_branch,
            entity_name=payload.entity_name,
            schema_name=payload.schema_name,
            table_name=payload.table_name,
            yaml_content=payload.yaml_content,
            recreate_sql=payload.recreate_sql,
            insert_sql=payload.insert_sql,
            truncate_sql=payload.truncate_sql,
            task_id=payload.task_id or "",
            author=user.email,
            expected_revision=payload.expected_revision,
        )
    except BranchRevisionConflictError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось сохранить объект в ветку: {exc}")


@router.post("/api/admin/meta-workspace/branch-file/save")
def save_admin_meta_workspace_branch_file(payload: MetaWorkspaceBranchFileSavePayload, request: Request):
    user = _require_meta_workspace_role(request)
    try:
        return save_meta_workspace_branch_file(
            git_repo_value=ENTITY_META_GIT_REPO,
            workspace_root_value=META_WORKSPACE_ROOT,
            workspace_owner=user.email,
            branch_name=payload.branch_name,
            base_branch=payload.base_branch,
            file_path=payload.file_path,
            content=payload.content,
            task_id=payload.task_id or "",
            author=user.email,
            expected_revision=payload.expected_revision,
        )
    except BranchRevisionConflictError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось сохранить файл в ветку: {exc}")


@router.post("/api/admin/meta-workspace/validate-all")
def validate_admin_meta_workspace_branch(payload: MetaWorkspaceValidatePayload, request: Request):
    user = _require_meta_workspace_role(request)
    try:
        return validate_meta_workspace_branch(
            engine=engine,
            base_dir=BASE_DIR,
            git_repo_value=ENTITY_META_GIT_REPO,
            entity_git_root_value=ENTITY_META_GIT_META_ROOT,
            click_git_root_value=CLICK_META_GIT_ROOT,
            workspace_root_value=META_WORKSPACE_ROOT,
            workspace_owner=user.email,
            prod_root_value=ENTITY_META_DIR,
            dev_root_value=DEV_ENTITY_META_DIR,
            branch_name=payload.branch_name,
            base_branch=payload.base_branch,
            dev_database_url=DEV_DATABASE_URL,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/api/admin/meta-workspace/sync-branch")
def sync_admin_meta_workspace_branch(payload: MetaWorkspaceSyncPayload, request: Request):
    user = _require_meta_workspace_role(request)
    try:
        result = sync_meta_workspace_branch(
            engine=engine,
            base_dir=BASE_DIR,
            workspace_root_value=META_WORKSPACE_ROOT,
            prod_root_value=ENTITY_META_DIR,
            entity_dev_root_value=DEV_ENTITY_META_DIR,
            click_dev_root_value=DEV_CLICK_META_DIR,
            git_repo_value=ENTITY_META_GIT_REPO,
            entity_git_root_value=ENTITY_META_GIT_META_ROOT,
            click_git_root_value=CLICK_META_GIT_ROOT,
            task_id=payload.task_id,
            branch_name=payload.branch_name,
            base_branch=payload.base_branch,
            author=user.email,
            dev_database_url=DEV_DATABASE_URL,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", **result}


@router.post("/api/admin/dev-meta/run-dag")
def run_admin_dev_meta_dag(payload: DevMetaDagPayload, request: Request):
    user = _require_dev_meta_role(request)
    try:
        assert_dev_meta_lock_owner(
            engine=engine,
            schema_name=payload.schema_name,
            file_name=payload.file_name,
            author=user.email,
        )
        data = trigger_airflow_dev_dag(
            engine=engine,
            airflow_base_url=AIRFLOW_DEV_BASE_URL,
            dag_id=AIRFLOW_DEV_DAG_ID,
            username=AIRFLOW_DEV_USERNAME,
            password=AIRFLOW_DEV_PASSWORD,
            schema_name=payload.schema_name,
            file_name=payload.file_name,
            author=user.email,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", "response": data}


@router.post("/api/admin/dev-meta/dag-status")
def get_admin_dev_meta_dag_status(payload: DevMetaDagStatusPayload, request: Request):
    _require_dev_meta_role(request)
    try:
        dag_id = payload.dag_id or AIRFLOW_DEV_DAG_ID or Path(payload.file_name).stem
        data = get_airflow_dev_dag_status(
            airflow_base_url=AIRFLOW_DEV_BASE_URL,
            username=AIRFLOW_DEV_USERNAME,
            password=AIRFLOW_DEV_PASSWORD,
            dag_id=dag_id,
            dag_run_id=payload.dag_run_id,
            auto_unpaused=payload.auto_unpaused,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", "response": data}


@router.get("/api/admin/dev-copy/status")
def get_admin_dev_copy_status(request: Request):
    _require_authenticated(request)
    window = _get_dev_copy_window_status()
    return {
        "airflow": {
            "base_url": AIRFLOW_DEV_BASE_URL,
            "dag_id": DEV_COPY_DAG_ID,
            "configured": bool(AIRFLOW_DEV_BASE_URL),
        },
        "schema_sync": {
            "base_url": AIRFLOW_DEV_BASE_URL,
            "dag_id": DEV_COPY_SCHEMA_SYNC_DAG_ID,
            "configured": bool(AIRFLOW_DEV_BASE_URL),
        },
        "window": window,
    }


@router.post("/api/admin/dev-copy/run-dag")
def run_admin_dev_copy_dag(payload: DevCopyDagPayload, request: Request):
    user = _require_authenticated(request)
    _assert_dev_copy_window()
    try:
        values = {
            "source_table_schema": str(payload.source_table_schema or "").strip(),
            "source_table_name": str(payload.source_table_name or "").strip(),
            "target_table_schema": str(payload.target_table_schema or "").strip(),
            "target_table_name": str(payload.target_table_name or "").strip(),
            "where": str(payload.where or "").strip(),
        }
        missing = [key for key in ("source_table_schema", "source_table_name", "target_table_schema", "target_table_name") if not values[key]]
        if missing:
            raise ValueError("Нужно заполнить все параметры запуска DAG")
        data = trigger_airflow_parametrized_dag(
            airflow_base_url=AIRFLOW_DEV_BASE_URL,
            dag_id=DEV_COPY_DAG_ID,
            username=AIRFLOW_DEV_USERNAME,
            password=AIRFLOW_DEV_PASSWORD,
            conf={
                **values,
                "author": user.email,
            },
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", "response": data}


@router.post("/api/admin/dev-copy/dag-status")
def get_admin_dev_copy_dag_status(payload: DevCopyDagStatusPayload, request: Request):
    _require_authenticated(request)
    try:
        data = get_airflow_dev_dag_status(
            airflow_base_url=AIRFLOW_DEV_BASE_URL,
            username=AIRFLOW_DEV_USERNAME,
            password=AIRFLOW_DEV_PASSWORD,
            dag_id=DEV_COPY_DAG_ID,
            dag_run_id=payload.dag_run_id,
            auto_unpaused=payload.auto_unpaused,
        )
        window = _get_dev_copy_window_status()
        if (
            not window["allowed"]
            and str(data.get("dag_run_state") or "").lower() in {"queued", "running"}
        ):
            stop_airflow_dag_run(
                airflow_base_url=AIRFLOW_DEV_BASE_URL,
                username=AIRFLOW_DEV_USERNAME,
                password=AIRFLOW_DEV_PASSWORD,
                dag_id=DEV_COPY_DAG_ID,
                dag_run_id=payload.dag_run_id,
                state="failed",
            )
            data = get_airflow_dev_dag_status(
                airflow_base_url=AIRFLOW_DEV_BASE_URL,
                username=AIRFLOW_DEV_USERNAME,
                password=AIRFLOW_DEV_PASSWORD,
                dag_id=DEV_COPY_DAG_ID,
                dag_run_id=payload.dag_run_id,
                auto_unpaused=payload.auto_unpaused,
            )
            data["terminated_due_to_window"] = True
            data["window_message"] = (
                "DAG был остановлен, потому что вышел за разрешенное окно "
                f"{window['allowed_from']} - {window['allowed_to']} по Москве."
            )
        data["window"] = window
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", "response": data}


@router.post("/api/admin/dev-copy/schema-sync/run-dag")
def run_admin_dev_copy_schema_sync_dag(payload: DevCopySchemaSyncPayload, request: Request):
    user = _require_authenticated(request)
    try:
        run_mode = str(payload.run_mode or "self").strip().lower()
        if run_mode not in {"self", "all"}:
            raise ValueError("Допустимые значения run_mode: self или all")

        values = {}
        if run_mode == "self":
            values = {
                "author": _dev_copy_author(user),
                "check_table_schema": str(payload.check_table_schema or "").strip(),
                "check_table_name": str(payload.check_table_name or "").strip(),
            }
            missing = [key for key, value in values.items() if not value]
            if missing:
                raise ValueError("Нужно заполнить author, check_table_schema и check_table_name")

        data = trigger_airflow_parametrized_dag(
            airflow_base_url=AIRFLOW_DEV_BASE_URL,
            dag_id=DEV_COPY_SCHEMA_SYNC_DAG_ID,
            username=AIRFLOW_DEV_USERNAME,
            password=AIRFLOW_DEV_PASSWORD,
            conf=values,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", "response": data}


@router.post("/api/admin/dev-copy/schema-sync/dag-status")
def get_admin_dev_copy_schema_sync_dag_status(payload: DevCopySchemaSyncDagStatusPayload, request: Request):
    _require_authenticated(request)
    try:
        data = get_airflow_dev_dag_status(
            airflow_base_url=AIRFLOW_DEV_BASE_URL,
            username=AIRFLOW_DEV_USERNAME,
            password=AIRFLOW_DEV_PASSWORD,
            dag_id=DEV_COPY_SCHEMA_SYNC_DAG_ID,
            dag_run_id=payload.dag_run_id,
            auto_unpaused=payload.auto_unpaused,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", "response": data}


@router.post("/api/admin/dev-copy/schema-sync/report")
def get_admin_dev_copy_schema_sync_report(payload: DevCopySchemaSyncReportPayload, request: Request):
    user = _require_authenticated(request)
    table_schema = str(payload.check_table_schema or "").strip()
    table_name = str(payload.check_table_name or "").strip()
    if not table_schema or not table_name:
        raise HTTPException(status_code=400, detail="Нужно указать check_table_schema и check_table_name")
    try:
        run_info = _get_schema_sync_latest_run(
            run_user=_dev_copy_author(user),
            table_schema=table_schema,
            table_name=table_name,
        )
        if not run_info:
            raise HTTPException(
                status_code=404,
                detail=f"Не найден результат сверки для {table_schema}.{table_name} и пользователя {_dev_copy_author(user)}",
            )
        report_rows = _get_schema_sync_report_rows(run_id=run_info["run_id"])
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось прочитать результат schema sync: {exc}")

    return {
        "status": "ok",
        "run": run_info,
        "summary": {
            "diff_count": len(report_rows),
            "schema_name": table_schema,
            "table_name": table_name,
        },
        "items": report_rows,
    }


@router.post("/api/admin/dev-meta/deploy")
def deploy_admin_dev_meta(payload: DevMetaDeployPayload, request: Request):
    user = _require_dev_meta_role(request)
    try:
        result = deploy_dev_meta_file(
            engine=engine,
            base_dir=BASE_DIR,
            dev_root_value=DEV_CLICK_META_DIR,
            schema_name=payload.schema_name,
            file_name=payload.file_name,
            content=payload.content,
            author=user.email,
            dev_database_url=DEV_DATABASE_URL,
            host=DEV_META_DEPLOY_HOST,
            port=DEV_META_DEPLOY_PORT,
            user=DEV_META_DEPLOY_USER,
            password=DEV_META_DEPLOY_PASSWORD,
            remote_base_dir=DEV_META_DEPLOY_BASE_DIR,
            ssh_key_path=DEV_META_DEPLOY_SSH_KEY_PATH,
            strict_host_key=DEV_META_DEPLOY_STRICT_HOST_KEY,
            task_id=payload.task_id,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", **result}


@router.get("/api/admin/entity-meta/status")
def get_admin_entity_meta_status(request: Request):
    _require_admin(request)
    return get_entity_dev_meta_status(
        engine=engine,
        base_dir=BASE_DIR,
        prod_root_value=ENTITY_META_DIR,
        dev_root_value=DEV_ENTITY_META_DIR,
        lock_ttl_minutes=DEV_META_LOCK_TTL_MIN,
    )


@router.get("/api/admin/entity-meta/catalog")
def get_admin_entity_meta_catalog(request: Request):
    _require_admin(request)
    return list_entity_dev_catalog(
        base_dir=BASE_DIR,
        prod_root_value=ENTITY_META_DIR,
        dev_root_value=DEV_ENTITY_META_DIR,
    )


@router.get("/api/admin/entity-meta/reference/entities")
def get_admin_entity_meta_reference_entities(request: Request):
    _require_admin(request)
    return {"items": list_entity_reference_rows(engine=engine)}


@router.post("/api/admin/entity-meta/init")
def init_admin_entity_meta(payload: EntityMetaInitPayload, request: Request):
    _require_admin(request)
    return init_entity_dev_meta_bundle(
        engine=engine,
        base_dir=BASE_DIR,
        prod_root_value=ENTITY_META_DIR,
        dev_root_value=DEV_ENTITY_META_DIR,
        entity_name=payload.entity_name,
        schema_name=payload.schema_name,
        table_name=payload.table_name,
        key_attributes=payload.key_attributes,
    )


@router.post("/api/admin/entity-meta/lock")
def lock_admin_entity_meta(payload: EntityMetaLockPayload, request: Request):
    user = _require_admin(request)
    try:
        return lock_entity_dev_meta(
            engine=engine,
            entity_name=payload.entity_name,
            schema_name=payload.schema_name,
            table_name=payload.table_name,
            author=user.email,
            ttl_minutes=DEV_META_LOCK_TTL_MIN,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))


@router.post("/api/admin/entity-meta/unlock")
def unlock_admin_entity_meta(payload: EntityMetaLockPayload, request: Request):
    user = _require_admin(request)
    unlock_entity_dev_meta(
        engine=engine,
        entity_name=payload.entity_name,
        schema_name=payload.schema_name,
        table_name=payload.table_name,
        author=user.email,
    )
    return {"status": "ok"}


@router.post("/api/admin/entity-meta/validate")
def validate_admin_entity_meta(payload: EntityMetaSavePayload, request: Request):
    _require_admin(request)
    return validate_entity_dev_meta_bundle(
        engine=engine,
        base_dir=BASE_DIR,
        prod_root_value=ENTITY_META_DIR,
        dev_root_value=DEV_ENTITY_META_DIR,
        entity_name=payload.entity_name,
        schema_name=payload.schema_name,
        table_name=payload.table_name,
        key_attributes=payload.key_attributes,
        source_object_key=payload.source_object_key,
        yaml_content=payload.yaml_content,
        recreate_sql=payload.recreate_sql,
        insert_sql=payload.insert_sql,
        truncate_sql=payload.truncate_sql,
        dev_database_url=DEV_DATABASE_URL,
    )


@router.post("/api/admin/entity-meta/save")
def save_admin_entity_meta(payload: EntityMetaSavePayload, request: Request):
    user = _require_admin(request)
    try:
        result = save_entity_dev_meta_bundle(
            engine=engine,
            base_dir=BASE_DIR,
            prod_root_value=ENTITY_META_DIR,
            dev_root_value=DEV_ENTITY_META_DIR,
            entity_name=payload.entity_name,
            schema_name=payload.schema_name,
            table_name=payload.table_name,
            task_id=payload.task_id,
            key_attributes=payload.key_attributes,
            source_object_key=payload.source_object_key,
            replica_entity_names=payload.replica_entity_names,
            yaml_content=payload.yaml_content,
            recreate_sql=payload.recreate_sql,
            insert_sql=payload.insert_sql,
            truncate_sql=payload.truncate_sql,
            author=user.email,
            dev_database_url=DEV_DATABASE_URL,
            lock_ttl_minutes=DEV_META_LOCK_TTL_MIN,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", **result}


@router.post("/api/admin/entity-meta/run-sql")
def run_admin_entity_meta_sql(payload: EntityMetaRunSqlPayload, request: Request):
    user = _require_admin(request)
    try:
        result = execute_entity_dev_meta_sql(
            engine=engine,
            entity_name=payload.entity_name,
            schema_name=payload.schema_name,
            table_name=payload.table_name,
            sql_kind=payload.sql_kind,
            sql_text=payload.sql_text,
            dev_database_url=DEV_DATABASE_URL,
            author=user.email,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return result


@router.post("/api/admin/entity-meta/delete")
def delete_admin_entity_meta(payload: EntityMetaDeletePayload, request: Request):
    user = _require_admin(request)
    try:
        result = delete_entity_dev_meta_bundle(
            engine=engine,
            base_dir=BASE_DIR,
            dev_root_value=DEV_ENTITY_META_DIR,
            entity_name=payload.entity_name,
            schema_name=payload.schema_name,
            table_name=payload.table_name,
            task_id=payload.task_id,
            author=user.email,
            lock_ttl_minutes=DEV_META_LOCK_TTL_MIN,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", **result}


@router.post("/api/admin/entity-meta/move")
def move_admin_entity_meta(payload: EntityMetaMovePayload, request: Request):
    user = _require_admin(request)
    try:
        result = move_entity_dev_meta_bundle(
            engine=engine,
            base_dir=BASE_DIR,
            prod_root_value=ENTITY_META_DIR,
            dev_root_value=DEV_ENTITY_META_DIR,
            source_entity_name=payload.source_entity_name,
            source_schema_name=payload.source_schema_name,
            source_table_name=payload.source_table_name,
            target_entity_name=payload.target_entity_name,
            target_schema_name=payload.target_schema_name,
            target_table_name=payload.target_table_name,
            task_id=payload.task_id,
            author=user.email,
            lock_ttl_minutes=DEV_META_LOCK_TTL_MIN,
        )
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", **result}


@router.post("/api/admin/entity-meta/mr")
def create_admin_entity_meta_mr(payload: EntityMetaMrPayload, request: Request):
    user = _require_admin(request)
    try:
        result = create_entity_meta_mr(
            engine=engine,
            base_dir=BASE_DIR,
            dev_root_value=DEV_ENTITY_META_DIR,
            git_repo_value=ENTITY_META_GIT_REPO,
            git_meta_root_value=ENTITY_META_GIT_META_ROOT,
            gitlab_token=GITLAB_TOKEN,
            gitlab_project=GITLAB_PROJECT,
            gitlab_api_url=GITLAB_API_URL,
            gitlab_ssl_verify=GITLAB_SSL_VERIFY,
            task_id=payload.task_id,
            release_branch=payload.release_branch,
            author=user.email,
        )
        if result.get("mr_url") and YOUTRACK_URL and YOUTRACK_TOKEN:
            try:
                add_ytrack_issue_comment(
                    base_url=YOUTRACK_URL,
                    token=YOUTRACK_TOKEN,
                    issue_id=payload.task_id,
                    ssl_verify=YOUTRACK_SSL_VERIFY,
                    text=(
                        "MR создан для инженерных изменений.\n"
                        f"Ссылка: {result.get('mr_url')}\n"
                        f"Ветка: {result.get('feature_branch') or '—'} -> {result.get('release_branch') or 'main'}"
                    ),
                )
                result["task_link_attached"] = True
            except Exception as exc:
                result["task_link_attached"] = False
                result["task_link_error"] = str(exc)
    except PermissionError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"status": "ok", **result}

# Модель для ответа зависимостей
class DependencyItem(BaseModel):
    step: int
    schema: str
    table_name: str
    entity_id: int
    entity_name: str = None
    start_time: str = None
    avg_duration_minutes: Optional[float] = None
    depth: int = 0
    path: Optional[List[str]] = None


TOP_DIRS = [
    "BI_FI",
    "BI_INVESTMENT",
    "BI_TAXES",
    "CASE_4",
    "DICT_LOADER",
    "MISHKADEV_TABLES",
    "FI_COUNTERPARTY",
    "ISUIP_INVESTMENT",
    "LOGISTICS",
    "TRANSPORTATION",
    "BI_SB_WUC",
    "BI_FI_FACT_PAYMENTS",
    "STG_LOADER",
    "SD_STOCKS",
    "SALES_SHIPMENT_FROM_PLANT",
    "SALES_MM",
    "SALES_MARGIN",
    "MANAGEMENT_REPORTING_1",
    "TEST_SAP_ODATA_DELTA",
    "SALES_2HOUR",
]

ENTITY_GROUP_SUFFIX_RE = re.compile(r"^(.*?)(?:[_-]\d+)$", re.IGNORECASE)


def _normalize_entity_group(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    cleaned = str(name).strip()
    if not cleaned:
        return None
    lowered = cleaned.lower()
    match = ENTITY_GROUP_SUFFIX_RE.match(lowered)
    if match:
        base = match.group(1).strip()
        return base or lowered
    return lowered

_cached_meta_index = None
_cache_timestamp = 0

_graph_snapshot = None
_graph_snapshot_ts = 0
_graph_snapshot_hash = None
_GRAPH_SNAPSHOT_TTL = 86400  # 24 часа
_CACHE_TTL = 86400  # 24 часа

_order_breaches_cache = None
_order_breaches_ts = 0
_ORDER_BREACHES_TTL = 300  # 5 минут

_graph_cache = {}
_graph_cache_ts = 0
_GRAPH_CACHE_TTL = 86400  # 24 часа
_graph_cache_meta_ts = 0

_logic_audit_cache_payload = None
_logic_audit_cache_ts = 0
_LOGIC_AUDIT_CACHE_TTL = 86400
_table_sizes_cache_payload = None
_table_sizes_cache_cycle = None

SQL_STOPWORDS = {
    "select", "from", "where", "join", "left", "right", "inner", "outer", "full", "on",
    "and", "or", "not", "null", "is", "as", "case", "when", "then", "else", "end",
    "group", "by", "order", "limit", "with", "distinct", "union", "all", "into", "insert",
    "create", "table", "truncate", "having", "over", "partition", "rows", "range",
}

SQL_FUNCTION_BLACKLIST = {
    "select", "from", "where", "group", "order", "when", "then", "else", "end", "and", "or",
    "in", "on", "over", "partition", "by", "as",
}

SQL_FUNCTION_COMMON_PREFIXES = (
    "util_text_to_",
)


def _current_table_sizes_cache_cycle(now: datetime | None = None) -> str:
    current = now.astimezone(DEV_COPY_TZ) if now else datetime.now(DEV_COPY_TZ)
    cycle_date = current.date()
    if current.hour < 9:
        cycle_date -= timedelta(days=1)
    return cycle_date.isoformat()


def _build_table_sizes_cache() -> dict[str, Any]:
    query = """
        WITH created_ops AS (
            SELECT
                pso.schemaname AS table_schema,
                pso.objname AS table_name,
                MAX(pso.statime) AS dt_creation
            FROM pg_catalog.pg_stat_operations pso
            WHERE pso.actionname = 'CREATE'
              AND pso.subtype = 'TABLE'
            GROUP BY pso.schemaname, pso.objname
        )
        SELECT
            n.nspname AS table_schema,
            c.relname AS table_name,
            r.rolname AS owner_name,
            pg_total_relation_size(c.oid)::bigint AS size_bytes,
            created_ops.dt_creation AS dt_creation,
            CASE
                WHEN created_ops.dt_creation IS NULL THEN NULL
                ELSE EXTRACT(DAY FROM now() - created_ops.dt_creation)::int
            END AS days_old
        FROM pg_catalog.pg_class c
        JOIN pg_catalog.pg_namespace n
          ON n.oid = c.relnamespace
        JOIN pg_catalog.pg_roles r
          ON r.oid = c.relowner
        LEFT JOIN created_ops
          ON created_ops.table_schema = n.nspname
         AND created_ops.table_name = c.relname
        WHERE c.relkind IN ('r', 'p', 'm')
          AND n.nspname NOT IN ('pg_catalog', 'information_schema')
          AND n.nspname NOT LIKE 'pg_toast%%'
        ORDER BY size_bytes DESC NULLS LAST, n.nspname, c.relname
    """

    with engine.connect() as conn:
        rows = conn.execute(text(query)).mappings().all()

    normalized_rows = [
        {
            "table_schema": row.get("table_schema"),
            "table_name": row.get("table_name"),
            "owner_name": row.get("owner_name"),
            "size_bytes": int(row.get("size_bytes") or 0),
            "days_old": int(row.get("days_old")) if row.get("days_old") is not None else None,
            "dt_creation": serialize_datetime(row.get("dt_creation")),
        }
        for row in rows
    ]
    schemas = sorted({row["table_schema"] for row in normalized_rows if row.get("table_schema")})
    return {
        "generated_at": datetime.now(DEV_COPY_TZ).isoformat(),
        "schemas": schemas,
        "rows": normalized_rows,
    }


def get_cached_table_sizes() -> dict[str, Any]:
    global _table_sizes_cache_payload, _table_sizes_cache_cycle

    cycle = _current_table_sizes_cache_cycle()
    if _table_sizes_cache_payload is not None and _table_sizes_cache_cycle == cycle:
        return _table_sizes_cache_payload

    print("⚠️ rebuilding table sizes cache")
    _table_sizes_cache_payload = _build_table_sizes_cache()
    _table_sizes_cache_cycle = cycle
    return _table_sizes_cache_payload


def _strip_sql_comments(sql_text: str) -> str:
    if not sql_text:
        return ""
    text_wo_block = re.sub(r"/\*.*?\*/", " ", sql_text, flags=re.S)
    text_wo_inline = re.sub(r"--.*?$", " ", text_wo_block, flags=re.M)
    return text_wo_inline


def _normalize_sql(sql_text: str) -> str:
    text = _strip_sql_comments(sql_text).lower()
    text = text.replace("`", "").replace('"', "")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _tokenize_sql(sql_text: str) -> set[str]:
    tokens = set(re.findall(r"[a-z_][a-z0-9_]*", sql_text))
    return {t for t in tokens if t not in SQL_STOPWORDS and len(t) > 2}


def _canonical_source_name(name: str) -> str:
    raw = (name or "").strip().strip(",")
    if not raw:
        return ""
    raw = raw.split()[0]
    # remove typical temporary table patterns to group stable base logic
    raw = re.sub(r"^(tmp_|temp_|cte_)", "", raw)
    raw = re.sub(r"(_tmp|_temp)$", "", raw)
    raw = re.sub(r"_[0-9]{6,}$", "", raw)
    raw = raw.replace('"', "").replace("`", "")
    return raw


def _extract_source_tables(normalized_sql: str) -> set[str]:
    tables = set()
    for match in re.finditer(r"\b(from|join)\s+([a-z0-9_./]+)", normalized_sql):
        cleaned = _canonical_source_name(match.group(2) or "")
        if cleaned and cleaned not in {"select"}:
            tables.add(cleaned)
    return tables


def _extract_functions(normalized_sql: str) -> set[str]:
    funcs = set()
    for fn in re.findall(r"\b([a-z_][a-z0-9_]*)\s*\(", normalized_sql):
        if fn not in SQL_FUNCTION_BLACKLIST:
            funcs.add(fn)
    return funcs


def _is_common_function(fn: str, freq_map: dict[str, int], total_objects: int) -> bool:
    if not fn:
        return True
    if any(fn.startswith(prefix) for prefix in SQL_FUNCTION_COMMON_PREFIXES):
        return True
    if total_objects <= 0:
        return False
    frequency = freq_map.get(fn, 0) / total_objects
    return frequency >= 0.35


def _extract_where_clause(normalized_sql: str) -> str:
    match = re.search(r"\bwhere\b(.*?)(\bgroup\s+by\b|\border\s+by\b|\blimit\b|$)", normalized_sql, flags=re.S)
    return (match.group(1) or "").strip() if match else ""


def _split_top_level(text: str) -> list[str]:
    if not text:
        return []
    parts = []
    buf = []
    level = 0
    for ch in text:
        if ch == "(":
            level += 1
        elif ch == ")" and level > 0:
            level -= 1
        if ch == "," and level == 0:
            piece = "".join(buf).strip()
            if piece:
                parts.append(piece)
            buf = []
            continue
        buf.append(ch)
    tail = "".join(buf).strip()
    if tail:
        parts.append(tail)
    return parts


def _extract_select_targets(normalized_sql: str) -> list[dict]:
    match = re.search(r"\bselect\b(.*?)(\bfrom\b)", normalized_sql, flags=re.S)
    if not match:
        return []
    body = (match.group(1) or "").strip()
    targets = []
    for expr in _split_top_level(body):
        alias_match = re.search(r"\bas\s+([a-z_][a-z0-9_]*)\s*$", expr)
        alias = alias_match.group(1) if alias_match else None
        targets.append({
            "expression": expr,
            "alias": alias,
        })
    return targets


def _expression_signature(expr: str) -> tuple[str, set[str]]:
    normalized = re.sub(r"\s+", " ", (expr or "").strip().lower())
    normalized = normalized.replace('"', "").replace("`", "")
    tokens = _tokenize_sql(normalized)
    expr_hash = hashlib.sha1(normalized.encode("utf-8")).hexdigest()
    return expr_hash, tokens


def _split_sql_statements(sql: str) -> list[str]:
    if not sql:
        return []
    parts = []
    current = []
    depth = 0
    in_single = False
    in_double = False
    for idx, ch in enumerate(sql):
        prev = sql[idx - 1] if idx > 0 else ""
        if ch == "'" and not in_double and prev != "\\":
            in_single = not in_single
        elif ch == '"' and not in_single and prev != "\\":
            in_double = not in_double
        if not in_single and not in_double:
            if ch == "(":
                depth += 1
            elif ch == ")" and depth > 0:
                depth -= 1
            elif ch == ";" and depth == 0:
                stmt = "".join(current).strip()
                if stmt:
                    parts.append(stmt)
                current = []
                continue
        current.append(ch)
    tail = "".join(current).strip()
    if tail:
        parts.append(tail)
    return parts


def _normalize_statement_body(stmt: str) -> tuple[str, str, str]:
    raw = re.sub(r"\s+", " ", (stmt or "").strip())
    lowered = raw.lower()
    temp_match = re.match(
        r"^create\s+(?:temporary\s+|temp\s+)?table\s+([a-z0-9_\".`]+)\s+as\s+(select\b.+)$",
        raw,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if temp_match:
        return "temp_table", temp_match.group(1), temp_match.group(2).strip()

    create_as_match = re.match(
        r"^create\s+table\s+([a-z0-9_\".`]+)\s+as\s+(select\b.+)$",
        raw,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if create_as_match:
        return "create_as_select", create_as_match.group(1), create_as_match.group(2).strip()

    insert_match = re.match(
        r"^insert\s+into\s+([a-z0-9_\".`]+)\b",
        raw,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if insert_match:
        select_idx = lowered.find("select")
        body = raw[select_idx:].strip() if select_idx >= 0 else raw
        return "insert_select", insert_match.group(1), body

    if lowered.startswith("with "):
        return "query", "with_query", raw
    if lowered.startswith("select "):
        return "query", "select_query", raw
    return "statement", "statement", raw


def _extract_logic_blocks(normalized_sql: str) -> list[dict]:
    blocks = []
    for index, stmt in enumerate(_split_sql_statements(normalized_sql), start=1):
        block_type, block_label, body = _normalize_statement_body(stmt)
        if "select" not in body.lower():
            continue
        select_targets = _extract_select_targets(body)[:20]
        expr_hashes = set()
        expr_token_union = set()
        for target in select_targets:
            expr_hash, expr_tokens = _expression_signature(target.get("expression") or "")
            expr_hashes.add(expr_hash)
            expr_token_union |= expr_tokens
        source_tables = _extract_source_tables(body)
        functions = _extract_functions(body)
        where_clause = _extract_where_clause(body)
        tokens = _tokenize_sql(body)
        if len(tokens) < 18 and not source_tables:
            continue
        blocks.append({
            "block_id": f"{block_type}:{index}:{block_label}",
            "block_type": block_type,
            "block_label": block_label,
            "sql_hash": hashlib.sha1(body.encode("utf-8")).hexdigest(),
            "sql_preview": body[:500],
            "tokens": tokens,
            "source_tables": source_tables,
            "functions": functions,
            "signal_functions": set(functions),
            "where_clause": where_clause,
            "select_targets": select_targets,
            "expr_hashes": expr_hashes,
            "expr_token_union": expr_token_union,
        })
    return blocks


def _jaccard_similarity(a: set[str], b: set[str]) -> float:
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0


def _build_story(obj: dict) -> str:
    raw_checks = obj.get("verification") or []
    checks = []
    for item in raw_checks:
        if isinstance(item, str):
            checks.append(item)
        elif isinstance(item, dict):
            checks.append(
                item.get("name")
                or item.get("code")
                or item.get("type")
                or json.dumps(item, ensure_ascii=False, sort_keys=True)
            )
        else:
            checks.append(str(item))
    keys = obj.get("key_attributes") or []
    sources = sorted(obj.get("source_tables") or [])
    funcs = sorted(obj.get("signal_functions") or obj.get("functions") or [])
    depends_on = obj.get("depends_on") or {}
    layers = ", ".join(sorted(depends_on.keys())) if depends_on else "не заданы"
    parts = [
        f"{obj.get('fqn')} ({obj.get('entity_name') or 'UNKNOWN'})",
        f"Режим загрузки: {obj.get('table_load_mode') or 'N/A'}",
        f"Слои зависимостей: {layers}",
        f"Ключевые поля: {', '.join(keys[:8]) if keys else 'не указаны'}",
        f"Проверки: {', '.join(checks) if checks else 'не указаны'}",
        f"SQL-функции: {', '.join(funcs[:10]) if funcs else 'не найдены'}",
        f"Источники SQL: {', '.join(sources[:8]) if sources else 'не найдены'}",
    ]
    return " | ".join(parts)


def _extract_field_descriptions(meta: dict) -> list[dict]:
    result = []
    for key in ("columns", "fields", "attributes", "column_descriptions"):
        value = meta.get(key)
        if isinstance(value, list):
            for item in value:
                if not isinstance(item, dict):
                    continue
                name = item.get("name") or item.get("column") or item.get("field")
                descr = item.get("description") or item.get("comment") or item.get("caption")
                if name and descr:
                    result.append({"name": str(name), "description": str(descr)})
        elif isinstance(value, dict):
            for name, descr in value.items():
                if name and descr:
                    result.append({"name": str(name), "description": str(descr)})
    dedup = {}
    for row in result:
        dedup[row["name"]] = row["description"]
    return [{"name": k, "description": v} for k, v in dedup.items()]


def _build_diff_hints(left: dict, right: dict) -> list[str]:
    hints = []
    src_left = left.get("source_tables") or set()
    src_right = right.get("source_tables") or set()
    fn_left = left.get("signal_functions") or set()
    fn_right = right.get("signal_functions") or set()
    where_left = left.get("where_clause") or ""
    where_right = right.get("where_clause") or ""
    if src_left != src_right:
        hints.append("Разные источники в FROM/JOIN")
    if fn_left != fn_right:
        hints.append("Отличаются используемые SQL-функции")
    if where_left != where_right:
        hints.append("Отличаются условия WHERE")
    if not hints:
        hints.append("Логика почти идентична, различия минимальны")
    return hints


def _build_pair_comparison(left: dict, right: dict) -> dict:
    left_sources = left.get("source_tables") or set()
    right_sources = right.get("source_tables") or set()
    left_functions = left.get("signal_functions") or set()
    right_functions = right.get("signal_functions") or set()

    left_aliases = {x.get("alias") for x in (left.get("select_targets") or []) if x.get("alias")}
    right_aliases = {x.get("alias") for x in (right.get("select_targets") or []) if x.get("alias")}

    left_keys = {str(x).lower() for x in (left.get("key_attributes") or []) if str(x).strip()}
    right_keys = {str(x).lower() for x in (right.get("key_attributes") or []) if str(x).strip()}

    left_where = (left.get("where_clause") or "").strip()
    right_where = (right.get("where_clause") or "").strip()

    left_common_fields = {
        str(row.get("name")).strip().lower()
        for row in (left.get("field_descriptions") or [])
        if row.get("name")
    }
    right_common_fields = {
        str(row.get("name")).strip().lower()
        for row in (right.get("field_descriptions") or [])
        if row.get("name")
    }

    same = []
    if left_sources & right_sources:
        same.append({"label": "Общие источники", "items": sorted(left_sources & right_sources)[:14]})
    if left_functions & right_functions:
        same.append({"label": "Общие бизнес-функции", "items": sorted(left_functions & right_functions)[:14]})
    if left_aliases & right_aliases:
        same.append({"label": "Одинаковые алиасы в SELECT", "items": sorted(left_aliases & right_aliases)[:14]})
    if left_keys & right_keys:
        same.append({"label": "Общие ключевые поля", "items": sorted(left_keys & right_keys)[:14]})
    if left_common_fields & right_common_fields:
        same.append({"label": "Совпадающие описанные поля", "items": sorted(left_common_fields & right_common_fields)[:14]})
    if left_where and right_where and left_where == right_where:
        same.append({"label": "WHERE совпадает", "items": [left_where[:220]]})

    different = []
    if left_sources - right_sources:
        different.append({"label": "Источники только в левом объекте", "items": sorted(left_sources - right_sources)[:14]})
    if right_sources - left_sources:
        different.append({"label": "Источники только в правом объекте", "items": sorted(right_sources - left_sources)[:14]})
    if left_functions - right_functions:
        different.append({"label": "Функции только в левом объекте", "items": sorted(left_functions - right_functions)[:14]})
    if right_functions - left_functions:
        different.append({"label": "Функции только в правом объекте", "items": sorted(right_functions - left_functions)[:14]})
    if left_aliases - right_aliases:
        different.append({"label": "Алиасы только в левом объекте", "items": sorted(left_aliases - right_aliases)[:14]})
    if right_aliases - left_aliases:
        different.append({"label": "Алиасы только в правом объекте", "items": sorted(right_aliases - left_aliases)[:14]})
    if left_where != right_where:
        if left_where:
            different.append({"label": "WHERE (левый объект)", "items": [left_where[:220]]})
        if right_where:
            different.append({"label": "WHERE (правый объект)", "items": [right_where[:220]]})

    return {
        "same": same,
        "different": different,
    }


def _build_pair_explanation(record: dict, left: dict, right: dict, comparison: dict) -> dict:
    score = record.get("score") or 0
    expr_overlap = record.get("expression_overlap_count") or 0
    merge_potential = record.get("merge_potential") or "LOW"
    diff_hints = record.get("diff_hints") or []

    if merge_potential == "HIGH":
        decision = "Пара выглядит как хороший кандидат на объединение в один расчёт."
    elif merge_potential == "MEDIUM":
        decision = "Логику лучше унифицировать, но сначала сверить бизнес-правила."
    else:
        decision = "Пока лучше оставить отдельно и явно зафиксировать различия."

    left_fields = left.get("field_descriptions") or []
    right_fields = right.get("field_descriptions") or []
    common_field_names = sorted(
        {row.get("name") for row in left_fields if row.get("name")}
        & {row.get("name") for row in right_fields if row.get("name")}
    )

    same_labels = [x.get("label") for x in (comparison.get("same") or []) if x.get("label")]
    diff_labels = [x.get("label") for x in (comparison.get("different") or []) if x.get("label")]
    diff_text = ", ".join(diff_labels[:2]) if diff_labels else ", ".join(diff_hints[:2]) or "критичных отличий не видно"
    summary = (
        f"{left.get('fqn')} и {right.get('fqn')} похожи на {round(score * 100)}%. "
        f"Совпадающих выражений SELECT: {expr_overlap}. "
        f"Совпадает: {', '.join(same_labels[:2]) if same_labels else 'базовый SQL-паттерн'}. "
        f"Отличается: {diff_text}."
    )

    return {
        "summary": summary,
        "decision": decision,
        "common_fields": common_field_names[:12],
        "left_field_docs_count": len(left_fields),
        "right_field_docs_count": len(right_fields),
    }


def _calc_logic_similarity(left: dict, right: dict) -> float:
    token_sim = _jaccard_similarity(left["tokens"], right["tokens"])
    source_sim = _jaccard_similarity(left["source_tables"], right["source_tables"])
    function_sim = _jaccard_similarity(left.get("signal_functions") or set(), right.get("signal_functions") or set())
    expr_exact_count = len((left.get("expr_hashes") or set()) & (right.get("expr_hashes") or set()))
    expr_exact_den = max(min(len(left.get("expr_hashes") or []), len(right.get("expr_hashes") or [])), 1)
    expr_exact_sim = expr_exact_count / expr_exact_den
    expr_token_sim = _jaccard_similarity(left.get("expr_token_union") or set(), right.get("expr_token_union") or set())
    expr_sim = 0.6 * expr_exact_sim + 0.4 * expr_token_sim

    score = 0.35 * token_sim + 0.25 * source_sim + 0.15 * function_sim + 0.25 * expr_sim
    if left["sql_hash"] == right["sql_hash"]:
        score = 1.0
    return round(score, 4), {
        "expr_exact_count": expr_exact_count,
        "expr_exact_sim": round(expr_exact_sim, 4),
        "expr_token_sim": round(expr_token_sim, 4),
    }


def _resolve_sql_path_from_meta(meta_path: Path, meta: dict, field_name: str, fallback_file_name: str) -> Optional[Path]:
    raw_path = str(meta.get(field_name) or "").strip()
    candidates = []
    if raw_path:
        raw = Path(raw_path)
        candidates.append(raw)
        candidates.append(meta_path.parent / raw_path)
        candidates.append(BASE_DIR / raw_path)
        candidates.append(BASE_DIR.parent / raw_path)
    candidates.append(meta_path.parent / fallback_file_name)

    seen = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve(strict=False)
        except Exception:
            resolved = candidate
        key = str(resolved)
        if key in seen:
            continue
        seen.add(key)
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _read_sql_from_meta(meta_path: Path, meta: dict, field_name: str, fallback_file_name: str) -> str:
    sql_path = _resolve_sql_path_from_meta(meta_path, meta, field_name, fallback_file_name)
    if not sql_path:
        return f"-- {fallback_file_name} not found"
    try:
        return sql_path.read_text(encoding="utf-8")
    except Exception as exc:
        return f"-- failed to read {sql_path}: {exc}"


def _build_logic_object(meta_path: Path) -> Optional[dict]:
    try:
        meta = yaml.safe_load(meta_path.read_text("utf-8")) or {}
    except Exception:
        return None

    schema = (meta.get("table_schema") or "").strip().lower()
    table = (meta.get("table_name") or "").strip().lower()
    if not schema or not table:
        return None
    entity_name = str(meta.get("entity_name") or "").strip()
    if entity_name.upper() == "DQ":
        return None

    sql_path = (
        _resolve_sql_path_from_meta(meta_path, meta, "sql_query_insert_init", "sql_query_insert_init.sql")
        or _resolve_sql_path_from_meta(meta_path, meta, "sql_query_recreate_init", "sql_query_recreate_init.sql")
    )
    sql_text = sql_path.read_text("utf-8", errors="ignore") if sql_path else ""
    normalized_sql = _normalize_sql(sql_text)
    if not normalized_sql:
        return None
    statement_blocks = _extract_logic_blocks(normalized_sql)[:24]
    select_targets = _extract_select_targets(normalized_sql)[:25]
    expr_hashes = set()
    expr_token_union = set()
    for target in select_targets:
        expr_hash, expr_tokens = _expression_signature(target.get("expression") or "")
        expr_hashes.add(expr_hash)
        expr_token_union |= expr_tokens

    return {
        "fqn": f"{schema}.{table}",
        "schema": schema,
        "table": table,
        "entity_name": entity_name,
        "entity_id": meta.get("entity_id"),
        "table_id": meta.get("table_id"),
        "table_load_mode": meta.get("table_load_mode"),
        "depends_on": meta.get("depends_on") or {},
        "field_descriptions": _extract_field_descriptions(meta),
        "key_attributes": list(meta.get("key_attributes") or []),
        "verification": list(meta.get("verification") or []),
        "sql_path": str(sql_path) if sql_path else None,
        "sql_hash": hashlib.sha1(normalized_sql.encode("utf-8")).hexdigest(),
        "sql_preview": normalized_sql[:700],
        "tokens": _tokenize_sql(normalized_sql),
        "source_tables": _extract_source_tables(normalized_sql),
        "functions": _extract_functions(normalized_sql),
        "where_clause": _extract_where_clause(normalized_sql),
        "select_targets": select_targets,
        "expr_hashes": expr_hashes,
        "expr_token_union": expr_token_union,
        "statement_blocks": statement_blocks,
    }


def _build_logic_block_candidates(objects: list[dict]) -> tuple[list[dict], list[dict], dict, dict]:
    exact_groups = defaultdict(list)
    source_groups = defaultdict(list)
    all_blocks = []
    pair_index = {}

    for obj in objects:
        for block in obj.get("statement_blocks") or []:
            block_row = {
                **block,
                "fqn": obj.get("fqn"),
                "entity_name": obj.get("entity_name"),
            }
            if len(block_row.get("tokens") or []) < 18:
                continue
            all_blocks.append(block_row)
            exact_groups[block_row["sql_hash"]].append(block_row)
            source_tables = sorted(block_row.get("source_tables") or [])
            if source_tables:
                source_groups["|".join(source_tables[:4])].append(block_row)

    exact_clusters = []
    for rows in exact_groups.values():
        object_fqns = sorted({row["fqn"] for row in rows if row.get("fqn")})
        if len(object_fqns) < 2:
            continue
        entities = sorted({str(row.get("entity_name") or "Без сущности") for row in rows})
        source_tables = sorted({src for row in rows for src in (row.get("source_tables") or set())})
        functions = sorted({fn for row in rows for fn in (row.get("signal_functions") or set())})
        block_type_counts = Counter(row.get("block_type") or "statement" for row in rows)
        sample_pair_id = None
        sample_pair_label = None
        for left, right in combinations(rows, 2):
            if left.get("fqn") == right.get("fqn"):
                continue
            pair_key = "|".join(sorted((left["fqn"], right["fqn"], left["block_id"], right["block_id"])))
            sample_pair_id = hashlib.sha1(pair_key.encode("utf-8")).hexdigest()[:16]
            sample_pair_label = f"{left['fqn']} ↔ {right['fqn']}"
            comparison = _build_pair_comparison(left, right)
            explanation = _build_pair_explanation(
                {
                    "score": 1.0,
                    "expression_overlap_count": len((left.get("expr_hashes") or set()) & (right.get("expr_hashes") or set())),
                    "merge_potential": "HIGH",
                    "diff_hints": _build_diff_hints(left, right),
                },
                left,
                right,
                comparison,
            )
            pair_index[sample_pair_id] = {
                "pair_id": sample_pair_id,
                "pair_kind": "block_exact",
                "score": 1.0,
                "expression_overlap_count": len((left.get("expr_hashes") or set()) & (right.get("expr_hashes") or set())),
                "merge_potential": "HIGH",
                "left_fqn": left.get("fqn"),
                "right_fqn": right.get("fqn"),
                "left_entity": left.get("entity_name"),
                "right_entity": right.get("entity_name"),
                "left_block_id": left.get("block_id"),
                "right_block_id": right.get("block_id"),
                "left_block_type": left.get("block_type"),
                "right_block_type": right.get("block_type"),
                "comparison": comparison,
                "explanation": explanation,
                "left": left,
                "right": right,
                "left_features": {
                    "tokens_count": len(left.get("tokens") or []),
                    "source_tables": sorted(left.get("source_tables") or []),
                    "functions": sorted(left.get("signal_functions") or []),
                },
                "right_features": {
                    "tokens_count": len(right.get("tokens") or []),
                    "source_tables": sorted(right.get("source_tables") or []),
                    "functions": sorted(right.get("signal_functions") or []),
                },
            }
            break
        exact_clusters.append({
            "kind": "exact",
            "block_type": block_type_counts.most_common(1)[0][0],
            "occurrences_count": len(rows),
            "objects_count": len(object_fqns),
            "entities": entities[:8],
            "sample_objects": object_fqns[:6],
            "sample_sources": source_tables[:6],
            "sample_functions": functions[:6],
            "sql_preview": rows[0].get("sql_preview"),
            "sample_pair_id": sample_pair_id,
            "sample_pair_label": sample_pair_label,
        })

    similar_pairs = []
    seen_pairs = set()
    for rows in source_groups.values():
        if len(rows) < 2 or len(rows) > 18:
            continue
        for left, right in combinations(rows, 2):
            if left["fqn"] == right["fqn"]:
                continue
            pair_key = tuple(sorted((left["fqn"], right["fqn"], left["block_id"], right["block_id"])))
            if pair_key in seen_pairs:
                continue
            seen_pairs.add(pair_key)
            score, sim_meta = _calc_logic_similarity(left, right)
            if score < 0.84 or sim_meta.get("expr_exact_count", 0) < 1:
                continue
            pair_key = "|".join(sorted((left["fqn"], right["fqn"], left["block_id"], right["block_id"])))
            pair_id = hashlib.sha1(pair_key.encode("utf-8")).hexdigest()[:16]
            diff_hints = _build_diff_hints(left, right)
            comparison = _build_pair_comparison(left, right)
            explanation = _build_pair_explanation(
                {
                    "score": score,
                    "expression_overlap_count": sim_meta.get("expr_exact_count", 0),
                    "merge_potential": "HIGH" if score >= 0.9 else "MEDIUM",
                    "diff_hints": diff_hints,
                },
                left,
                right,
                comparison,
            )
            pair_index[pair_id] = {
                "pair_id": pair_id,
                "pair_kind": "block_similar",
                "score": score,
                "expression_overlap_count": sim_meta.get("expr_exact_count", 0),
                "merge_potential": "HIGH" if score >= 0.9 else "MEDIUM",
                "left_fqn": left.get("fqn"),
                "right_fqn": right.get("fqn"),
                "left_entity": left.get("entity_name"),
                "right_entity": right.get("entity_name"),
                "left_block_id": left.get("block_id"),
                "right_block_id": right.get("block_id"),
                "left_block_type": left.get("block_type"),
                "right_block_type": right.get("block_type"),
                "comparison": comparison,
                "explanation": explanation,
                "left": left,
                "right": right,
                "left_features": {
                    "tokens_count": len(left.get("tokens") or []),
                    "source_tables": sorted(left.get("source_tables") or []),
                    "functions": sorted(left.get("signal_functions") or []),
                },
                "right_features": {
                    "tokens_count": len(right.get("tokens") or []),
                    "source_tables": sorted(right.get("source_tables") or []),
                    "functions": sorted(right.get("signal_functions") or []),
                },
            }
            similar_pairs.append({
                "pair_id": pair_id,
                "left_fqn": left["fqn"],
                "right_fqn": right["fqn"],
                "left_entity": left.get("entity_name"),
                "right_entity": right.get("entity_name"),
                "left_block_id": left.get("block_id"),
                "right_block_id": right.get("block_id"),
                "left_block_type": left.get("block_type"),
                "right_block_type": right.get("block_type"),
                "score": score,
                "expression_overlap_count": sim_meta.get("expr_exact_count", 0),
                "diff_hints": diff_hints,
                "sample_sources": sorted((left.get("source_tables") or set()) & (right.get("source_tables") or set()))[:6],
                "sample_functions": sorted((left.get("signal_functions") or set()) & (right.get("signal_functions") or set()))[:6],
            })

    exact_clusters.sort(
        key=lambda item: (
            -item["objects_count"],
            -item["occurrences_count"],
            item["sample_objects"][0] if item["sample_objects"] else "",
        ),
    )
    similar_pairs.sort(
        key=lambda item: (
            -(item["score"] or 0),
            -(item["expression_overlap_count"] or 0),
            item["left_fqn"],
            item["right_fqn"],
        ),
    )
    summary = {
        "objects_with_blocks": len({row["fqn"] for row in all_blocks}),
        "blocks_count": len(all_blocks),
        "exact_clusters_count": len(exact_clusters),
        "similar_pairs_count": len(similar_pairs),
    }
    return exact_clusters[:16], similar_pairs[:20], summary, pair_index


def _build_logic_audit_cache():
    global _logic_audit_cache_payload, _logic_audit_cache_ts
    now = time.time()
    if _logic_audit_cache_payload and now - _logic_audit_cache_ts < _LOGIC_AUDIT_CACHE_TTL:
        return _logic_audit_cache_payload

    objects = []
    for root_dir in iter_meta_dirs():
        for root, _, files in os.walk(root_dir):
            if "meta_data_file.yaml" not in files:
                continue
            obj = _build_logic_object(Path(root) / "meta_data_file.yaml")
            if obj:
                objects.append(obj)

    function_freq = {}
    total_objects = len(objects)
    for obj in objects:
        for fn in obj.get("functions") or set():
            function_freq[fn] = function_freq.get(fn, 0) + 1

    for obj in objects:
        filtered = {
            fn for fn in (obj.get("functions") or set())
            if not _is_common_function(fn, function_freq, total_objects)
        }
        obj["signal_functions"] = filtered
        for block in obj.get("statement_blocks") or []:
            block_filtered = {fn for fn in (block.get("functions") or set()) if fn in filtered}
            block["signal_functions"] = block_filtered or set(block.get("functions") or set())
        obj["story"] = _build_story(obj)

    objects_index = {}
    for obj in objects:
        objects_index[obj["fqn"]] = {
            "fqn": obj.get("fqn"),
            "schema": obj.get("schema"),
            "table": obj.get("table"),
            "entity_name": obj.get("entity_name"),
            "table_load_mode": obj.get("table_load_mode"),
            "story": obj.get("story"),
            "source_tables": sorted(obj.get("source_tables") or []),
            "functions": sorted(obj.get("signal_functions") or []),
            "where_clause": obj.get("where_clause"),
            "select_targets": (obj.get("select_targets") or [])[:12],
            "field_descriptions": (obj.get("field_descriptions") or [])[:40],
            "sql_path": obj.get("sql_path"),
        }

    pairs = []
    pair_index = {}
    for left, right in combinations(objects, 2):
        if left["fqn"] == right["fqn"]:
            continue
        # dm_view рассматриваем только внутри dm_view (1-1 view слой)
        if (left["schema"] == "dm_view") != (right["schema"] == "dm_view"):
            continue
        # быстрый pre-filter
        if not (left["source_tables"] & right["source_tables"] or left["signal_functions"] & right["signal_functions"]):
            continue

        score, sim_meta = _calc_logic_similarity(left, right)
        if score < 0.72:
            continue

        if left["sql_hash"] == right["sql_hash"]:
            issue_type = "duplicate_exact"
            merge_potential = "HIGH"
        elif score >= 0.86:
            issue_type = "duplicate_candidate"
            merge_potential = "HIGH"
        elif score >= 0.78:
            issue_type = "similar_candidate"
            merge_potential = "MEDIUM"
        else:
            issue_type = "similar_candidate"
            merge_potential = "LOW"

        pair_key = "|".join(sorted([left["fqn"], right["fqn"]]))
        pair_id = hashlib.sha1(pair_key.encode("utf-8")).hexdigest()[:16]
        diff_hints = _build_diff_hints(left, right)
        record = {
            "pair_id": pair_id,
            "left_fqn": left["fqn"],
            "right_fqn": right["fqn"],
            "left_entity": left.get("entity_name"),
            "right_entity": right.get("entity_name"),
            "score": score,
            "expression_overlap_count": sim_meta.get("expr_exact_count", 0),
            "expression_overlap_score": sim_meta.get("expr_exact_sim", 0),
            "issue_type": issue_type,
            "merge_potential": merge_potential,
            "diff_hints": diff_hints,
        }
        pairs.append(record)
        comparison = _build_pair_comparison(left, right)
        explanation = _build_pair_explanation(record, left, right, comparison)
        pair_index[pair_id] = {
            **record,
            "explanation": explanation,
            "comparison": comparison,
            "left": {
                k: v for k, v in left.items()
                if k not in {"tokens", "source_tables", "functions", "expr_hashes", "expr_token_union"}
            },
            "right": {
                k: v for k, v in right.items()
                if k not in {"tokens", "source_tables", "functions", "expr_hashes", "expr_token_union"}
            },
            "left_features": {
                "tokens_count": len(left["tokens"]),
                "source_tables": sorted(left["source_tables"]),
                "functions": sorted(left.get("signal_functions") or []),
            },
            "right_features": {
                "tokens_count": len(right["tokens"]),
                "source_tables": sorted(right["source_tables"]),
                "functions": sorted(right.get("signal_functions") or []),
            },
        }

    block_exact_clusters, block_similar_pairs, block_summary, block_pair_index = _build_logic_block_candidates(objects)

    pairs.sort(key=lambda row: (row["score"], row["merge_potential"]), reverse=True)
    payload = {
        "generated_at": datetime.utcnow().isoformat(),
        "objects_count": len(objects),
        "pairs_count": len(pairs),
        "pairs": pairs,
        "pair_index": pair_index,
        "objects_index": objects_index,
        "block_exact_clusters": block_exact_clusters,
        "block_similar_pairs": block_similar_pairs,
        "block_summary": block_summary,
        "block_pair_index": block_pair_index,
    }
    _logic_audit_cache_payload = payload
    _logic_audit_cache_ts = now
    return payload

def compute_order_breaches():
    """
    ТЯЖЁЛАЯ логика расчёта order breaches.
    НИЧЕГО НЕ ЗНАЕТ ПРО HTTP.
    """
    resp = get_dependency_violations()
    rows = json.loads(resp.body)

    entity_map = _entity_map_from_meta()
    grouped = {}

    for r in rows:
        target = f"{r['dependent_schema']}.{r['dependent_table']}"
        source = f"{r['source_schema']}.{r['source_table']}"

        src_time = datetime.fromisoformat(r["source_last_load"])
        tgt_time = datetime.fromisoformat(r["dependent_last_load"])
        gap_sec = (src_time - tgt_time).total_seconds()

        g = grouped.setdefault(target, {
            "target_fqn": target,
            "target_last_load": r["dependent_last_load"],
            "worst_upstream": None,
            "worst_upstream_time": None,
            "worst_gap_sec": 0,
            "violations": [],
            "_seen": set(),
        })

        dedupe_key = (source, r["source_last_load"], r["dependent_last_load"])
        if dedupe_key in g["_seen"]:
            continue
        g["_seen"].add(dedupe_key)

        g["violations"].append({
            "source_fqn": source,
            "gap_sec": gap_sec,
            "source_last_load": r["source_last_load"],
            "dependent_last_load": r["dependent_last_load"],
        })

        if gap_sec > g["worst_gap_sec"]:
            g["worst_gap_sec"] = gap_sec
            g["worst_upstream"] = source
            g["worst_upstream_time"] = r["source_last_load"]

    result = []
    for g in grouped.values():
        gap_min = g["worst_gap_sec"] / 60
        if gap_min > 30:
            sev = "CRITICAL"
        elif gap_min > 5:
            sev = "MAJOR"
        else:
            sev = "WARNING"

        g["severity"] = sev
        g["gap_minutes"] = round(gap_min, 1)
        g["violations_count"] = len(g["violations"])
        g.pop("_seen", None)
        result.append(g)

    result.sort(key=lambda x: x["worst_gap_sec"], reverse=True)
    return result

def get_cached_order_breaches():
    global _order_breaches_cache, _order_breaches_ts

    now = time.time()
    if _order_breaches_cache and now - _order_breaches_ts < _ORDER_BREACHES_TTL:
        return _order_breaches_cache

    print("⚠️ rebuilding orderbreaches cache")
    result = compute_order_breaches()

    _order_breaches_cache = result
    _order_breaches_ts = now
    return result

def get_cached_meta_and_index():
    global _cached_meta_index, _cache_timestamp
    now = time.time()
    if _cached_meta_index and now - _cache_timestamp < _CACHE_TTL:
        return _cached_meta_index

    all_meta = []
    seen = set()
    for entity_root in iter_meta_dirs():
        for root, _, files in os.walk(entity_root):
            if "meta_data_file.yaml" not in files:
                continue
            path = Path(root) / "meta_data_file.yaml"
            try:
                meta = yaml.safe_load(path.read_text("utf-8")) or {}
                schema = norm(meta.get("table_schema"))
                table = norm(meta.get("table_name"))
                if not schema or not table:
                    continue
                key = f"{schema}.{table}"
                seen.add(key)
                depends_on = {}
                for src_schema, tables in (meta.get("depends_on") or {}).items():
                    src_schema_norm = norm(src_schema)
                    if not src_schema_norm:
                        continue
                    cleaned = [norm(t) for t in (tables or []) if t]
                    depends_on[src_schema_norm] = [t for t in cleaned if t]
                all_meta.append({
                    "table_schema": schema,
                    "table_name": table,
                    "entity_id": meta.get("entity_id"),
                    "entity_name": meta.get("entity_name"),
                    "depends_on": depends_on,
                    "table_id": meta.get("table_id"),
                })
            except Exception as e:
                print("META ERROR:", path, e)

    print("META COUNT:", len(seen))
    print("META SAMPLE:", sorted(list(seen))[:30])

    meta_lookup = {
        (m.get("table_schema"), m.get("table_name"))
        for m in all_meta
        if m.get("table_schema") and m.get("table_name")
    }
    reverse = {}
    for m in all_meta:
        consumer = (m["table_schema"], m["table_name"])
        for src_schema, tables in m["depends_on"].items():
            for src_table in tables:
                if (src_schema or "").lower() in ("raw_ext", "dict_raw_ext"):
                    continue
                if (src_schema, src_table) not in meta_lookup:
                    print(
                        "❌ BROKEN DEP:",
                        f"{consumer[0]}.{consumer[1]}",
                        "depends on",
                        f"{src_schema}.{src_table}",
                        "BUT META NOT FOUND",
                    )
                reverse.setdefault((src_schema, src_table), []).append({
                    "schema": consumer[0],
                    "table_name": consumer[1],
                    "entity_id": m["entity_id"],
                    "entity_name": m["entity_name"],
                    "table_id": m["table_id"],
                })

    _cached_meta_index = (all_meta, reverse)
    _cache_timestamp = now
    return _cached_meta_index

def norm(s: Optional[str]) -> Optional[str]:
    return s.lower() if isinstance(s, str) else s


def _hash_meta(all_meta: list[dict]) -> str:
    items = []
    for m in all_meta:
        items.append({
            "table_schema": m.get("table_schema"),
            "table_name": m.get("table_name"),
            "entity_name": m.get("entity_name"),
            "depends_on": m.get("depends_on") or {},
            "table_id": m.get("table_id"),
        })
    payload = json.dumps(items, ensure_ascii=True, sort_keys=True)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def _estimate_node_width(label: str, min_width: int = 160, max_width: int = 420) -> int:
    base = 120
    per_char = 7
    width = base + (len(label or "") * per_char)
    return max(min_width, min(max_width, width))


def _find_sccs(nodes: list[str], edges: list[dict]) -> list[list[str]]:
    adj = {n: [] for n in nodes}
    for e in edges:
        src = e.get("source")
        tgt = e.get("target")
        if src in adj and tgt:
            adj[src].append(tgt)

    index = 0
    stack = []
    on_stack = set()
    indices = {}
    lowlinks = {}
    sccs = []

    def strongconnect(v):
        nonlocal index
        indices[v] = index
        lowlinks[v] = index
        index += 1
        stack.append(v)
        on_stack.add(v)

        for w in adj.get(v, []):
            if w not in indices:
                strongconnect(w)
                lowlinks[v] = min(lowlinks[v], lowlinks[w])
            elif w in on_stack:
                lowlinks[v] = min(lowlinks[v], indices[w])

        if lowlinks[v] == indices[v]:
            scc = []
            while stack:
                w = stack.pop()
                on_stack.remove(w)
                scc.append(w)
                if w == v:
                    break
            sccs.append(scc)

    for n in nodes:
        if n not in indices:
            strongconnect(n)

    return sccs


def _table_fqn_from_node(node_or_fqn: Any) -> str:
    if isinstance(node_or_fqn, dict):
        schema = str(node_or_fqn.get("schema") or "").strip()
        table = str(node_or_fqn.get("table") or "").strip()
        if schema and table:
            return f"{schema}.{table}"
        return str(node_or_fqn.get("fqn") or node_or_fqn.get("id") or "")
    return str(node_or_fqn or "")


def _layer_of_table(node_or_fqn: Any) -> str:
    fqn = _table_fqn_from_node(node_or_fqn)
    if not fqn or "." not in fqn:
        return "other"
    schema = fqn.split(".", 1)[0]
    schema_norm = schema.lower()
    if schema_norm in ("dict", "dict_stg", "dict_ods", "dict_dds"):
        if "dict_ods" in fqn or schema_norm == "dict_ods":
            return "dict_ods"
        if "dict_dds" in fqn or schema_norm == "dict_dds":
            return "dict_dds"
        return "dict_stg"
    if schema_norm in ("raw_ext", "dict_raw_ext"):
        return "raw_ext"
    if schema_norm in ("stg", "ods", "dds", "dm", "dm_calc", "dm_view", "landing", "raw_ext"):
        return schema_norm
    return "other"


def _grid_layout_table(table_nodes: dict, edges: list[dict]) -> dict:
    order = ["raw_ext", "landing", "dict_stg", "dict_ods", "dict_dds", "stg", "ods", "dds", "dm_calc", "dm", "dm_view", "other"]
    columns = {key: [] for key in order}
    for node_id in table_nodes:
        layer = _layer_of_table(table_nodes.get(node_id))
        columns.setdefault(layer, []).append(node_id)

    col_gap = 180
    row_gap = 140
    layout = {}
    cursor_x = 0
    layer_index = {layer: idx for idx, layer in enumerate(order)}
    neighbors = {node_id: [] for node_id in table_nodes}
    for edge in edges or []:
        src = edge.get("source")
        tgt = edge.get("target")
        if src in table_nodes and tgt in table_nodes:
            neighbors[src].append(layer_index.get(_layer_of_table(table_nodes.get(tgt)), 0))
            neighbors[tgt].append(layer_index.get(_layer_of_table(table_nodes.get(src)), 0))
    for layer in order:
        items = columns.get(layer) or []
        if not items:
            continue
        items.sort(key=lambda n: (sum(neighbors.get(n) or [0]) / max(len(neighbors.get(n) or [1]), 1), n))
        max_width = max(table_nodes[n].get("width") or 0 for n in items)
        column_center = cursor_x + (max_width / 2)
        for row_idx, node_id in enumerate(items):
            layout[node_id] = {"x": column_center, "y": row_idx * row_gap}
        cursor_x += max_width + col_gap

    return layout


def _grid_layout_subset(table_nodes: dict, edges: list[dict], node_ids: set) -> dict:
    subset = {nid: table_nodes[nid] for nid in node_ids if nid in table_nodes}
    subset_edges = [e for e in edges if e.get("source") in subset and e.get("target") in subset]
    return _grid_layout_table(subset, subset_edges)


def _normalize_layer_widths(nodes: list[dict]) -> list[dict]:
    if not nodes:
        return []
    layers = {}
    for node in nodes:
        layer = _layer_of_table(node)
        layers.setdefault(layer, []).append(node)

    max_widths = {
        layer: max(n.get("width") or 0 for n in items)
        for layer, items in layers.items()
    }
    out = []
    for node in nodes:
        layer = _layer_of_table(node)
        width = max_widths.get(layer) or node.get("width") or 200
        updated = dict(node)
        updated["width"] = width
        updated["height"] = 64
        out.append(updated)
    return out


def _dagre_layout(nodes: list[dict], edges: list[dict], rankdir: str = "LR") -> dict:
    if not nodes:
        return {}

    base_dir = Path(__file__).resolve().parent.parent
    script_path = base_dir / "scripts" / "dagre_layout.cjs"
    if not script_path.exists():
        raise FileNotFoundError(script_path)

    payload = {
        "nodes": nodes,
        "edges": edges,
        "rankdir": rankdir,
        "nodesep": 32,
        "ranksep": 90,
        "marginx": 20,
        "marginy": 20,
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / "dagre_input.json"
        output_path = Path(tmpdir) / "dagre_output.json"
        input_path.write_text(json.dumps(payload), encoding="utf-8")

        result = subprocess.run(
            ["node", str(script_path), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(f"dagre layout failed: {result.stderr.strip()}")

        out = json.loads(output_path.read_text(encoding="utf-8"))
        layout = out.get("nodes", {})

        # Normalize ranks to a grid to avoid excessive vertical drift.
        buckets = {}
        for node_id, pos in layout.items():
            if not pos:
                continue
            key = int(round(pos.get("x", 0)))
            buckets.setdefault(key, []).append((node_id, pos.get("y", 0)))

        if not buckets:
            return layout

        col_gap = 240
        row_gap = 90
        new_layout = {}
        for col_idx, x_key in enumerate(sorted(buckets.keys())):
            items = sorted(buckets[x_key], key=lambda item: item[1])
            for row_idx, (node_id, _y) in enumerate(items):
                new_layout[node_id] = {"x": col_idx * col_gap, "y": row_idx * row_gap}

        return new_layout


def build_graph_snapshot():
    all_meta, _ = get_cached_meta_and_index()
    meta_hash = _hash_meta(all_meta)

    entries = []
    meta_tables = set()
    fqn_to_node_ids: dict[str, list[str]] = {}
    for m in all_meta:
        schema = norm(m.get("table_schema"))
        table = norm(m.get("table_name"))
        entity = m.get("entity_name")
        table_id = m.get("table_id")
        if not schema or not table:
            continue
        if schema in ("raw_ext", "dict_raw_ext"):
            continue
        if isinstance(entity, str) and entity.lower() == "raw_ext":
            continue

        fqn = f"{schema}.{table}"
        node_id = f"T::{table_id}" if table_id else f"F::{entity or 'UNKNOWN'}::{fqn}"
        meta_tables.add(node_id)
        depends = {}
        for src_schema, tables in (m.get("depends_on") or {}).items():
            src_schema_norm = norm(src_schema)
            if not src_schema_norm:
                continue
            cleaned = [norm(t) for t in (tables or []) if t]
            depends[src_schema_norm] = [t for t in cleaned if t]

        entries.append({
            "node_id": node_id,
            "fqn": fqn,
            "table_schema": schema,
            "table_name": table,
            "entity_name": entity or "UNKNOWN",
            "table_id": table_id,
            "depends_on": depends,
        })

    table_entities: dict[str, set[str]] = {}
    table_info: dict[str, dict] = {}

    def register_table(node_id: str, fqn: str, schema: str, table: str, entity: str, table_id):
        if node_id not in table_info:
            table_info[node_id] = {
                "id": node_id,
                "fqn": fqn,
                "schema": schema,
                "table": table,
                "table_id": table_id,
            }
        table_entities.setdefault(node_id, set()).add(entity)
        node_ids = fqn_to_node_ids.setdefault(fqn, [])
        if node_id not in node_ids:
            node_ids.append(node_id)

    for m in entries:
        register_table(m["node_id"], m["fqn"], m["table_schema"], m["table_name"], m["entity_name"], m.get("table_id"))

    edges_set: set[tuple[str, str]] = set()
    for m in entries:
        target = m["node_id"]
        target_entity_group = _normalize_entity_group(m.get("entity_name"))
        for src_schema, tables in (m.get("depends_on") or {}).items():
            for src_table in tables:
                source_fqn = f"{src_schema}.{src_table}"
                source_ids = fqn_to_node_ids.get(source_fqn)
                if not source_ids:
                    schema_val, table_val = source_fqn.split(".", 1)
                    placeholder_id = f"X::{source_fqn}"
                    register_table(placeholder_id, source_fqn, schema_val, table_val, "UNKNOWN", None)
                    source_ids = [placeholder_id]
                if len(source_ids) > 1:
                    same_group = [
                        source_id
                        for source_id in source_ids
                        if _normalize_entity_group(next(iter(table_entities.get(source_id) or []), None)) == target_entity_group
                    ]
                    source_ids = same_group or source_ids
                for source_id in source_ids:
                    edges_set.add((source_id, target))

    table_nodes = {}
    for node_id, info in table_info.items():
        entities = sorted(table_entities.get(node_id) or [])
        width = _estimate_node_width(info["fqn"], min_width=200, max_width=520)
        table_nodes[node_id] = {
            "id": node_id,
            "fqn": info["fqn"],
            "schema": info["schema"],
            "table": info["table"],
            "entity": entities[0] if entities else "UNKNOWN",
            "entities": entities,
            "table_id": info.get("table_id"),
            "shared": len(entities) > 1,
            "width": width,
            "height": 64,
        }

    table_edges = [{"source": s, "target": t} for s, t in sorted(edges_set)]

    entity_nodes = {}
    entity_tables: dict[str, set[str]] = {}
    for fqn, entities in table_entities.items():
        for ent in entities:
            if not ent or ent == "UNKNOWN":
                continue
            entity_tables.setdefault(ent, set()).add(fqn)

    for ent, tables in entity_tables.items():
        node_id = f"ENTITY::{ent}"
        width = _estimate_node_width(ent, min_width=140, max_width=320)
        entity_nodes[node_id] = {
            "id": node_id,
            "label": ent,
            "tables_count": len(tables),
            "width": width,
            "height": 56,
        }

    entity_edges_set: set[tuple[str, str]] = set()
    for edge in edges_set:
        src, tgt = edge
        src_entities = table_entities.get(src) or set()
        tgt_entities = table_entities.get(tgt) or set()
        for src_ent in src_entities:
            for tgt_ent in tgt_entities:
                if (
                    not src_ent
                    or not tgt_ent
                    or src_ent == tgt_ent
                    or src_ent == "UNKNOWN"
                    or tgt_ent == "UNKNOWN"
                ):
                    continue
                if _normalize_entity_group(src_ent) == _normalize_entity_group(tgt_ent):
                    continue
                entity_edges_set.add((f"ENTITY::{src_ent}", f"ENTITY::{tgt_ent}"))

    entity_edges = [{"source": s, "target": t} for s, t in sorted(entity_edges_set)]

    entity_ids = list(entity_nodes.keys())
    sccs = _find_sccs(entity_ids, entity_edges)
    entity_cycles = []
    for scc in sccs:
        if len(scc) <= 1:
            continue
        labels = [entity_nodes.get(node_id, {}).get("label", node_id) for node_id in scc]
        entity_cycles.append({"nodes": labels, "size": len(labels)})

    edge_set = {(e["source"], e["target"]) for e in entity_edges}
    entity_mutual = []
    entity_mutual_any = []
    mutual_pairs = set()
    mutual_pairs_any = set()
    for src, tgt in edge_set:
        if (tgt, src) in edge_set:
            pair_key = tuple(sorted([src, tgt]))
            mutual_pairs_any.add(pair_key)

    table_pair_edges = {}
    table_pair_edges_rev = {}
    table_pair_edges_any = {}
    table_pair_edges_any_rev = {}
    exclusive_pair_edges = set()
    def is_exclusive_edge(src_ents: set, tgt_ents: set, left_ent: str, right_ent: str) -> bool:
        return (
            left_ent in src_ents
            and right_ent in tgt_ents
            and right_ent not in src_ents
            and left_ent not in tgt_ents
            and len(src_ents) == 1
            and len(tgt_ents) == 1
        )

    def is_pair_edge(src_ents: set, tgt_ents: set, left_ent: str, right_ent: str) -> bool:
        if _normalize_entity_group(left_ent) == _normalize_entity_group(right_ent):
            return False
        src_norm = {e.lower() for e in src_ents if isinstance(e, str)}
        tgt_norm = {e.lower() for e in tgt_ents if isinstance(e, str)}
        left_norm = left_ent.lower()
        right_norm = right_ent.lower()
        allowed = {left_norm, right_norm}
        return (
            left_norm in src_norm
            and right_norm in tgt_norm
            and left_norm not in tgt_norm
            and right_norm not in src_norm
            and src_norm.issubset(allowed)
            and tgt_norm.issubset(allowed)
        )

    for edge in table_edges:
        src = edge["source"]
        tgt = edge["target"]
        src_entities = table_entities.get(src) or set()
        tgt_entities = table_entities.get(tgt) or set()
        for src_ent in src_entities:
            for tgt_ent in tgt_entities:
                if not src_ent or not tgt_ent or src_ent == tgt_ent:
                    continue
                pair_key = (f"ENTITY::{src_ent}", f"ENTITY::{tgt_ent}")
                if tuple(sorted(pair_key)) in mutual_pairs_any and is_pair_edge(src_entities, tgt_entities, src_ent, tgt_ent):
                    table_pair_edges_any.setdefault(pair_key, set()).add((src, tgt))
                    table_pair_edges_any_rev.setdefault(tuple(sorted(pair_key)), {}).setdefault(pair_key, set()).add((src, tgt))
                if is_exclusive_edge(src_entities, tgt_entities, src_ent, tgt_ent):
                    exclusive_pair_edges.add(pair_key)
                    table_pair_edges.setdefault(pair_key, set()).add((src, tgt))
                    table_pair_edges_rev.setdefault(tuple(sorted(pair_key)), {}).setdefault(pair_key, set()).add((src, tgt))

    for src, tgt in exclusive_pair_edges:
        if (tgt, src) in exclusive_pair_edges:
            mutual_pairs.add(tuple(sorted([src, tgt])))

    for pair_key in sorted(mutual_pairs_any):
        left, right = pair_key
        edges_forward = sorted(table_pair_edges_any_rev.get(pair_key, {}).get((left, right), set()))
        edges_backward = sorted(table_pair_edges_any_rev.get(pair_key, {}).get((right, left), set()))
        entity_mutual_any.append({
            "a": entity_nodes.get(left, {}).get("label", left),
            "b": entity_nodes.get(right, {}).get("label", right),
            "edges_ab_count": len(edges_forward),
            "edges_ba_count": len(edges_backward),
            "edges_ab_sample": [{"source": s, "target": t} for s, t in edges_forward[:10]],
            "edges_ba_sample": [{"source": s, "target": t} for s, t in edges_backward[:10]],
        })

    for pair_key in sorted(mutual_pairs):
        left, right = pair_key
        edges_forward = sorted(table_pair_edges_rev.get(pair_key, {}).get((left, right), set()))
        edges_backward = sorted(table_pair_edges_rev.get(pair_key, {}).get((right, left), set()))
        entity_mutual.append({
            "a": entity_nodes.get(left, {}).get("label", left),
            "b": entity_nodes.get(right, {}).get("label", right),
            "edges_ab_count": len(edges_forward),
            "edges_ba_count": len(edges_backward),
            "edges_ab_sample": [{"source": s, "target": t} for s, t in edges_forward[:10]],
            "edges_ba_sample": [{"source": s, "target": t} for s, t in edges_backward[:10]],
        })

    table_ids = list(table_nodes.keys())
    table_sccs = _find_sccs(table_ids, table_edges)
    table_cycles = []
    for scc in table_sccs:
        if len(scc) <= 1:
            continue
        sample = sorted(scc)[:8]
        table_cycles.append({"size": len(scc), "nodes": sample})
    table_cycles.sort(key=lambda c: c["size"], reverse=True)
    table_cycles = table_cycles[:10]

    entity_layout_nodes = []
    for node_id, node in entity_nodes.items():
        label = node["label"]
        entity_layout_nodes.append({
            "id": node_id,
            "width": node.get("width") or _estimate_node_width(label, min_width=140, max_width=320),
            "height": node.get("height") or 56,
        })

    table_layout_nodes = []
    for fqn, node in table_nodes.items():
        label = fqn
        table_layout_nodes.append({
            "id": fqn,
            "width": node.get("width") or _estimate_node_width(label, min_width=170, max_width=460),
            "height": node.get("height") or 56,
        })

    entity_layout = _dagre_layout(entity_layout_nodes, entity_edges, rankdir="LR")
    table_layout = _grid_layout_table(table_nodes, table_edges)

    return {
        "meta_hash": meta_hash,
        "meta_ts": _cache_timestamp,
        "generated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        "table_graph": {"nodes": table_nodes, "edges": table_edges},
        "table_fqn_map": {k: sorted(v) for k, v in fqn_to_node_ids.items()},
        "entity_graph": {"nodes": entity_nodes, "edges": entity_edges},
        "layouts": {"entity": entity_layout, "table": table_layout},
        "table_entity_map": {k: sorted(v) for k, v in table_entities.items()},
        "table_meta": sorted(meta_tables),
        "entity_cycles": entity_cycles,
        "entity_mutual": entity_mutual,
        "entity_mutual_any": entity_mutual_any,
        "table_cycles": table_cycles,
    }


def get_graph_snapshot():
    global _graph_snapshot, _graph_snapshot_ts, _graph_snapshot_hash
    now = time.time()

    if _graph_snapshot and now - _graph_snapshot_ts < _GRAPH_SNAPSHOT_TTL:
        if _graph_snapshot.get("meta_ts") == _cache_timestamp:
            return _graph_snapshot

    snapshot = build_graph_snapshot()
    _graph_snapshot = snapshot
    _graph_snapshot_ts = now
    _graph_snapshot_hash = snapshot.get("meta_hash")
    return snapshot


def _get_table_graph_context(source: str = "current") -> dict:
    source_name = (source or "current").strip().lower()
    if source_name and source_name != "current":
        return get_dbt_graph_snapshot(BASE_DIR, DBT_MANIFEST_DIR, source=source_name)
    return get_graph_snapshot()


def _cap_graph(nodes: dict, edges: list, layout: dict, max_nodes: int, max_edges: int, sort_key: str = None):
    node_list = list(nodes.values())
    if sort_key:
        node_list.sort(key=lambda n: n.get(sort_key) or 0, reverse=True)
    node_ids = {n["id"] for n in node_list[:max_nodes]}

    truncated = len(nodes) > max_nodes
    filtered_edges = [e for e in edges if e["source"] in node_ids and e["target"] in node_ids]

    if len(filtered_edges) > max_edges:
        filtered_edges = filtered_edges[:max_edges]
        truncated = True

    filtered_nodes = {nid: nodes[nid] for nid in node_ids if nid in nodes}
    filtered_layout = {nid: layout.get(nid) for nid in node_ids if nid in layout}
    return filtered_nodes, filtered_edges, filtered_layout, truncated


def _compute_orphan_tables(snapshot: dict, final_schemas: set[str], meta_only: bool = False) -> dict:
    nodes = snapshot.get("table_graph", {}).get("nodes", {}) or {}
    edges = snapshot.get("table_graph", {}).get("edges", []) or []
    meta_tables = set(snapshot.get("table_meta") or [])
    if meta_only:
        nodes = {k: v for k, v in nodes.items() if k in meta_tables}
        edges = [e for e in edges if e.get("source") in meta_tables and e.get("target") in meta_tables]
    if not nodes:
        return {
            "final_schemas": sorted(final_schemas),
            "meta_only": meta_only,
            "total_tables": 0,
            "final_count": 0,
            "reachable_count": 0,
            "orphan_count": 0,
            "coverage_pct": 0.0,
            "orphans": [],
            "count_by_schema": {},
        }

    incoming = {}
    outgoing = {}
    reverse = {}
    for edge in edges:
        src = edge.get("source")
        tgt = edge.get("target")
        if not src or not tgt:
            continue
        outgoing[src] = outgoing.get(src, 0) + 1
        incoming[tgt] = incoming.get(tgt, 0) + 1
        reverse.setdefault(tgt, []).append(src)

    finals = {
        node_id
        for node_id, node in nodes.items()
        if node.get("schema") in final_schemas
    }

    reachable = set()
    stack = list(finals)
    while stack:
        current = stack.pop()
        if current in reachable:
            continue
        reachable.add(current)
        for parent in reverse.get(current, []):
            if parent not in reachable:
                stack.append(parent)

    orphans = [node_id for node_id in nodes.keys() if node_id not in reachable]
    count_by_schema = {}
    for node_id in orphans:
        schema = nodes[node_id].get("schema") or "unknown"
        count_by_schema[schema] = count_by_schema.get(schema, 0) + 1

    coverage_pct = (len(reachable) / len(nodes)) * 100 if nodes else 0.0
    return {
        "final_schemas": sorted(final_schemas),
        "meta_only": meta_only,
        "total_tables": len(nodes),
        "final_count": len(finals),
        "reachable_count": len(reachable),
        "orphan_count": len(orphans),
        "coverage_pct": round(coverage_pct, 2),
        "orphans": orphans,
        "count_by_schema": count_by_schema,
        "incoming": incoming,
        "outgoing": outgoing,
    }


@router.get("/api/entities/shared")
def get_shared_tables_by_entity(limit: int = Query(5, ge=0, le=50)):
    shared_map = {}
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT entity_id, table_schema, table_name
                    FROM {TABLE_TABLES_META}
                    WHERE entity_id IS NOT NULL
                    """
                )
            ).mappings().all()

        table_entity_map = {}
        for row in rows:
            entity_id = row.get("entity_id")
            schema = row.get("table_schema")
            table = row.get("table_name")
            if entity_id is None or not schema or not table:
                continue
            fqn = f"{schema}.{table}"
            table_entity_map.setdefault(fqn, set()).add(str(entity_id))

        for table_fqn, entity_ids in table_entity_map.items():
            if len(entity_ids) < 2:
                continue
            for entity_id in entity_ids:
                entry = shared_map.setdefault(entity_id, {"count": 0, "tables": []})
                entry["count"] += 1
                if limit and len(entry["tables"]) < limit:
                    entry["tables"].append(table_fqn)
    except Exception:
        shared_map = {}

    return JSONResponse(content=shared_map, media_type="application/json; charset=utf-8")


@router.get("/api/entities/intersections")
def get_entity_intersections(
    limit: int = Query(120, ge=1, le=1000),
    min_score: int = Query(1, ge=1, le=1000),
):
    try:
        snapshot = get_graph_snapshot()
        all_meta_list, _ = get_cached_meta_and_index()

        entity_by_id: dict[str, dict[str, Any]] = {}
        for row in all_meta_list:
            entity_id = row.get("entity_id")
            entity_name = row.get("entity_name")
            if entity_id is None or not entity_name:
                continue
            entity_by_id[str(entity_id)] = {
                "entity_id": int(entity_id),
                "entity_name": str(entity_name),
            }

        table_entities: dict[str, set[str]] = {}
        for row in all_meta_list:
            entity_id = row.get("entity_id")
            schema_name = row.get("table_schema")
            table_name = row.get("table_name")
            if entity_id is None or not schema_name or not table_name:
                continue
            fqn = f"{schema_name}.{table_name}"
            table_entities.setdefault(fqn, set()).add(str(entity_id))

        pair_map: dict[tuple[str, str], dict[str, Any]] = {}
        entity_table_counts: dict[str, int] = {}

        for entity_ids in table_entities.values():
            for entity_id in entity_ids:
                entity_table_counts[entity_id] = entity_table_counts.get(entity_id, 0) + 1

        def ensure_pair(left_id: str, right_id: str) -> dict[str, Any]:
            pair_key = tuple(sorted((left_id, right_id)))
            entry = pair_map.get(pair_key)
            if entry is None:
                left = entity_by_id.get(pair_key[0], {"entity_id": int(pair_key[0]), "entity_name": pair_key[0]})
                right = entity_by_id.get(pair_key[1], {"entity_id": int(pair_key[1]), "entity_name": pair_key[1]})
                entry = {
                    "pair_key": f"{pair_key[0]}::{pair_key[1]}",
                    "a": left,
                    "b": right,
                    "shared_tables": set(),
                    "links_ab": 0,
                    "links_ba": 0,
                    "edge_samples_ab": [],
                    "edge_samples_ba": [],
                }
                pair_map[pair_key] = entry
            return entry

        for table_fqn, entity_ids in table_entities.items():
            if len(entity_ids) < 2:
                continue
            for left_id, right_id in combinations(sorted(entity_ids), 2):
                entry = ensure_pair(left_id, right_id)
                entry["shared_tables"].add(table_fqn)

        table_edges = snapshot.get("table_graph", {}).get("edges", []) or []
        table_nodes = snapshot.get("table_graph", {}).get("nodes", {}) or {}
        for edge in table_edges:
            source_node = table_nodes.get(edge.get("source")) or {}
            target_node = table_nodes.get(edge.get("target")) or {}
            source_fqn = f"{source_node.get('schema')}.{source_node.get('table')}".strip(".")
            target_fqn = f"{target_node.get('schema')}.{target_node.get('table')}".strip(".")
            source_entities = table_entities.get(source_fqn) or set()
            target_entities = table_entities.get(target_fqn) or set()
            if not source_entities or not target_entities:
                continue
            for left_id in source_entities:
                for right_id in target_entities:
                    if left_id == right_id:
                        continue
                    entry = ensure_pair(left_id, right_id)
                    left_key = str(entry["a"]["entity_id"])
                    sample = {"source": source_fqn, "target": target_fqn}
                    if left_id == left_key:
                        entry["links_ab"] += 1
                        if len(entry["edge_samples_ab"]) < 4 and sample not in entry["edge_samples_ab"]:
                            entry["edge_samples_ab"].append(sample)
                    else:
                        entry["links_ba"] += 1
                        if len(entry["edge_samples_ba"]) < 4 and sample not in entry["edge_samples_ba"]:
                            entry["edge_samples_ba"].append(sample)

        rows = []
        entity_rank: dict[str, dict[str, Any]] = {}
        involved_entities: set[str] = set()
        pairs_with_shared = 0
        pairs_with_links = 0
        total_links = 0

        for entry in pair_map.values():
            shared_tables = sorted(entry["shared_tables"])
            shared_count = len(shared_tables)
            links_ab = int(entry["links_ab"] or 0)
            links_ba = int(entry["links_ba"] or 0)
            links_total = links_ab + links_ba
            total_score = shared_count + links_total
            if total_score < min_score:
                continue

            if shared_count > 0:
                pairs_with_shared += 1
            if links_total > 0:
                pairs_with_links += 1
            total_links += links_total
            involved_entities.add(str(entry["a"]["entity_id"]))
            involved_entities.add(str(entry["b"]["entity_id"]))

            rows.append(
                {
                    "pair_key": entry["pair_key"],
                    "a": entry["a"],
                    "b": entry["b"],
                    "shared_tables_count": shared_count,
                    "shared_tables_sample": shared_tables[:6],
                    "links_ab_count": links_ab,
                    "links_ba_count": links_ba,
                    "links_total": links_total,
                    "score": total_score,
                    "edge_samples_ab": entry["edge_samples_ab"],
                    "edge_samples_ba": entry["edge_samples_ba"],
                }
            )

            for side, peer, outbound, inbound in (
                (entry["a"], entry["b"], links_ab, links_ba),
                (entry["b"], entry["a"], links_ba, links_ab),
            ):
                entity_key = str(side["entity_id"])
                rank = entity_rank.setdefault(
                    entity_key,
                    {
                        "entity_id": side["entity_id"],
                        "entity_name": side["entity_name"],
                        "peer_ids": set(),
                        "shared_tables_total": 0,
                        "outbound_links": 0,
                        "inbound_links": 0,
                    },
                )
                rank["peer_ids"].add(str(peer["entity_id"]))
                rank["shared_tables_total"] += shared_count
                rank["outbound_links"] += outbound
                rank["inbound_links"] += inbound

        rows.sort(
            key=lambda item: (
                -item["score"],
                -item["links_total"],
                -item["shared_tables_count"],
                item["a"]["entity_name"],
                item["b"]["entity_name"],
            )
        )

        top_entities = []
        for item in entity_rank.values():
            peer_count = len(item["peer_ids"])
            top_entities.append(
                {
                    "entity_id": item["entity_id"],
                    "entity_name": item["entity_name"],
                    "table_count": entity_table_counts.get(str(item["entity_id"]), 0),
                    "peer_count": peer_count,
                    "shared_tables_total": item["shared_tables_total"],
                    "outbound_links": item["outbound_links"],
                    "inbound_links": item["inbound_links"],
                    "total_links": item["outbound_links"] + item["inbound_links"],
                    "score": peer_count + item["shared_tables_total"] + item["outbound_links"] + item["inbound_links"],
                }
            )
        top_entities.sort(
            key=lambda item: (
                -item["score"],
                -item["peer_count"],
                -item["total_links"],
                item["entity_name"],
            )
        )

        entity_items = []
        for item in top_entities:
            entity_key = str(item["entity_id"])
            related = []
            for row in rows:
                if str(row["a"]["entity_id"]) == entity_key:
                    peer = row["b"]
                    outbound = row["links_ab_count"]
                    inbound = row["links_ba_count"]
                elif str(row["b"]["entity_id"]) == entity_key:
                    peer = row["a"]
                    outbound = row["links_ba_count"]
                    inbound = row["links_ab_count"]
                else:
                    continue
                related.append(
                    {
                        "entity_id": peer["entity_id"],
                        "entity_name": peer["entity_name"],
                        "shared_tables_count": row["shared_tables_count"],
                        "links_out": outbound,
                        "links_in": inbound,
                        "links_total": outbound + inbound,
                        "score": row["score"],
                        "shared_tables_sample": row["shared_tables_sample"][:4],
                    }
                )
            related.sort(
                key=lambda rel: (
                    -rel["score"],
                    -rel["links_total"],
                    -rel["shared_tables_count"],
                    rel["entity_name"],
                )
            )
            entity_items.append(
                {
                    **item,
                    "related_entities": related,
                }
            )

        payload = {
            "summary": {
                "pairs": len(rows),
                "entities_involved": len(involved_entities),
                "pairs_with_shared_tables": pairs_with_shared,
                "pairs_with_links": pairs_with_links,
                "direct_links": total_links,
            },
            "entities": entity_items,
            "top_entities": top_entities[:18],
            "pairs": rows[:limit],
        }
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        print("❌ /api/entities/intersections error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/graph/orphans")
def get_orphan_tables(
    final_schemas: str = Query("dm"),
    meta_only: bool = Query(True),
    offset: int = Query(0, ge=0),
    limit: int = Query(40, ge=0, le=500),
):
    snapshot = get_graph_snapshot()
    requested = {norm(s) for s in (final_schemas or "").split(",") if s}
    final_set = {s for s in requested if s}
    data = _compute_orphan_tables(snapshot, final_set, meta_only=meta_only)

    table_nodes = snapshot.get("table_graph", {}).get("nodes", {}) or {}
    table_entity_map = snapshot.get("table_entity_map") or {}
    incoming = data.get("incoming", {})
    outgoing = data.get("outgoing", {})

    orphans_sorted = sorted(data["orphans"])
    total_orphans = len(orphans_sorted)
    start = min(offset, total_orphans)
    end = total_orphans if limit == 0 else min(start + limit, total_orphans)

    orphans = []
    for node_id in orphans_sorted[start:end]:
        node = table_nodes.get(node_id) or {}
        orphans.append({
            "id": node_id,
            "schema": node.get("schema"),
            "table": node.get("table"),
            "table_id": node.get("table_id"),
            "entities": table_entity_map.get(node_id, []),
            "incoming": incoming.get(node_id, 0),
            "outgoing": outgoing.get(node_id, 0),
        })
    has_more = end < total_orphans

    payload = {
        "final_schemas": data["final_schemas"],
        "meta_only": data["meta_only"],
        "total_tables": data["total_tables"],
        "final_count": data["final_count"],
        "reachable_count": data["reachable_count"],
        "orphan_count": data["orphan_count"],
        "coverage_pct": data["coverage_pct"],
        "count_by_schema": data["count_by_schema"],
        "offset": start,
        "limit": limit,
        "has_more": has_more,
        "orphans": orphans,
    }
    return JSONResponse(content=payload, media_type="application/json; charset=utf-8")

def normalize_fqn(table_fqn: str) -> tuple[str, str]:
    """
    Всегда возвращает (schema, table) в lowercase
    """
    s = table_fqn.strip().lower()

    if "/" in s and "." in s:
        schema, rest = s.split(".", 1)
        rest = rest.replace("/", "").replace("-", "").replace(" ", "")
        s = f"{schema}.{rest}"

    if "." not in s:
        raise ValueError("Expected schema.table")

    return tuple(s.split(".", 1))


def _parse_numeric(text_val: Optional[str]) -> Optional[float]:
    if text_val is None:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", str(text_val))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _entity_map_from_meta() -> dict:
    all_meta, _ = get_cached_meta_and_index()
    entity_map = {}
    for meta in all_meta:
        schema = norm(meta.get("table_schema"))
        table = norm(meta.get("table_name"))
        entity = meta.get("entity_name")
        if schema and table and entity:
            entity_map[f"{schema}.{table}"] = entity
    return entity_map


def _normalize_table_param(schema: str, table: str) -> tuple[Optional[str], Optional[str]]:
    schema_norm = norm(schema)
    table_norm = norm(table)
    if not table_norm:
        return schema_norm, table_norm
    table_norm = table_norm.strip()
    return schema_norm, table_norm


def _clean_table_name(table_norm: Optional[str]) -> Optional[str]:
    if not table_norm:
        return table_norm
    return table_norm.replace("/", "").replace("-", "").replace(" ", "")


def _resolve_table_key(
    mapping: dict,
    schema: str,
    table: str,
    table_id: Optional[int] = None,
    fqn_map: Optional[dict[str, list[str]]] = None,
) -> Optional[str]:
    schema_norm = norm(schema)
    table_norm = norm(table)
    direct = f"{schema_norm}.{table_norm}"
    candidates: list[str] = []
    if fqn_map:
        candidates.extend(fqn_map.get(direct) or [])
    if not candidates:
        with_leading_slash = f"{schema_norm}./{table_norm}" if table_norm and not table_norm.startswith("/") else None
        if with_leading_slash and fqn_map:
            candidates.extend(fqn_map.get(with_leading_slash) or [])
    table_clean = _clean_table_name(table_norm)
    if not candidates:
        cleaned = f"{schema_norm}.{table_clean}" if table_clean else None
        if cleaned and fqn_map:
            candidates.extend(fqn_map.get(cleaned) or [])
    if table_id is not None:
        for candidate in candidates:
            node = mapping.get(candidate) or {}
            if node.get("table_id") == table_id:
                return candidate
    if candidates:
        meta_candidates = [candidate for candidate in candidates if str(candidate).startswith("T::")]
        return (meta_candidates or candidates)[0]
    return None

@app.on_event("startup")
def warm_up_cache():
    try:
        get_cached_meta_and_index()
        get_cached_order_breaches()  # 🔥 прогрев orderbreaches
        get_graph_snapshot()
        get_cached_table_sizes()
    except Exception as e:
        print("Ошибка при старте приложения:", e)


BASE_DIR = Path(__file__).resolve().parent.parent
# Можно переопределить путь к метаданным через переменную окружения:
# export META_PARENT_DIR=/path/to/meta_info/database/greenplum/schema_name/tech_etl/etl_load_entity
# Локальный Windows путь (раскомментируй на Win):
# META_PARENT_DIRS = [Path(r"C:\\GIT\\meta_info\\database\\greenplum\\schema_name\\tech_etl\\etl_load_entity")]
META_PARENT_DIRS = [Path(os.getenv("META_PARENT_DIR", BASE_DIR / "etl_loads_entity"))]

# ClickHouse meta configs (config_files/meta)
_click_meta_cache = None
_click_meta_ts = 0
_CLICK_META_TTL = 86400  # 24 часа
_assistant_index_cache = None
_assistant_index_ts = 0
_ASSISTANT_INDEX_TTL = 900
_ASSISTANT_LAYER_NAMES = {"stg", "ods", "dds", "dict_stg", "dict_dds", "dm", "dm_view", "dm_calc"}
_ASSISTANT_STOPWORDS = {
    "и", "или", "что", "где", "как", "какие", "какая", "какой", "покажи", "найди", "нужны",
    "нужно", "есть", "для", "про", "это", "таблица", "таблицы", "слой", "слое", "самая",
    "самые", "долгая", "долгие", "загрузка", "загрузки", "зависит", "влияет", "от", "на",
    "она", "него", "нее", "по", "описанию", "частично", "мне", "с", "со",
}


def _load_click_meta_index():
    root = Path(CLICK_META_DIR)
    if not root.is_absolute():
        root = (BASE_DIR / root).resolve()

    meta_index: Dict[Tuple[str, str], Dict[str, Any]] = {}
    view_sql_index: Dict[Tuple[str, str], str] = {}
    view_refs: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}

    dm_dir = root / "dm"
    if dm_dir.exists():
        for path in dm_dir.rglob("*.yaml"):
            try:
                data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
            except Exception:
                continue
            schema_gp = str(data.get("schema_name_gp") or "").strip().lower()
            obj_name = str(data.get("object_name") or "").strip().lower()
            if schema_gp and obj_name:
                meta_index[(schema_gp, obj_name)] = data

    dm_view_dir = root / "dm_view"
    if dm_view_dir.exists():
        for path in dm_view_dir.rglob("*.sql"):
            obj_name = path.stem.strip().lower()
            if obj_name:
                sql_text = path.read_text(encoding="utf-8")
                view_sql_index[("dm_view", obj_name)] = sql_text

                # parse FROM/JOIN references
                for match in re.finditer(
                    r"(from|join)\s+([A-Za-z0-9_\"/]+)\.([A-Za-z0-9_\"/]+)",
                    sql_text,
                    flags=re.IGNORECASE,
                ):
                    schema_ref = match.group(2).replace('"', "").strip().lower()
                    table_ref = match.group(3).replace('"', "").strip().lower()
                    schema_ref = schema_ref.replace("/", "")
                    table_ref = _clean_table_name(table_ref)
                    if not schema_ref or not table_ref:
                        continue
                    key = (schema_ref, table_ref)
                    view_refs.setdefault(key, []).append(
                        {
                            "view_schema": "dm_view",
                            "view_name": obj_name,
                            "reason": "from_match",
                        }
                    )

    return {"meta": meta_index, "view_sql": view_sql_index, "view_refs": view_refs, "root": str(root)}


def get_click_meta_index():
    global _click_meta_cache, _click_meta_ts
    now = time.time()
    if _click_meta_cache and now - _click_meta_ts < _CLICK_META_TTL:
        return _click_meta_cache
    _click_meta_cache = _load_click_meta_index()
    _click_meta_ts = now
    return _click_meta_cache


def _assistant_normalize_text(value: Any) -> str:
    text_value = str(value or "").lower()
    text_value = re.sub(r"[^a-zа-я0-9_.]+", " ", text_value, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text_value).strip()


def _assistant_extract_terms(value: str) -> list[str]:
    terms: list[str] = []
    for token in _assistant_normalize_text(value).split():
        if len(token) < 2 or token in _ASSISTANT_STOPWORDS:
            continue
        if token not in terms:
            terms.append(token)
    return terms


def _assistant_compact_description(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    parts = [part.strip() for part in raw.split("|") if str(part).strip()]
    if not parts:
        return raw
    filtered = [
        part for part in parts
        if not re.search(r"[A-Za-z0-9_/$]+\.[A-Za-z0-9_/$\"]+", part)
    ]
    normalized: list[str] = []
    for part in (filtered or parts):
        key = part.lower()
        if key not in {item.lower() for item in normalized}:
            normalized.append(part)
    return normalized[0] if normalized else raw


def _assistant_parse_recreate_comments(meta_path: Path, meta: dict) -> tuple[str, list[str]]:
    sql_text = _read_sql_from_meta(meta_path, meta, "sql_query_recreate_init", "sql_query_recreate_init.sql")
    if not sql_text or sql_text.lstrip().startswith("--"):
        return "", []

    table_match = re.search(
        r"COMMENT\s+ON\s+TABLE\s+.+?\s+IS\s+'((?:''|[^'])*)'\s*;",
        sql_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    table_comment = table_match.group(1).replace("''", "'").strip() if table_match else ""

    column_comments: list[str] = []
    column_pattern = re.compile(
        r"COMMENT\s+ON\s+COLUMN\s+.+?\.(?:\"[^\"]+\"|[A-Za-z_][A-Za-z0-9_/$]*)\s+IS\s+'((?:''|[^'])*)'\s*;",
        flags=re.IGNORECASE | re.DOTALL,
    )
    for match in column_pattern.finditer(sql_text):
        comment = match.group(1).replace("''", "'").strip()
        compact = _assistant_compact_description(comment)
        if compact:
            column_comments.append(compact)

    deduped_columns: list[str] = []
    seen: set[str] = set()
    for item in column_comments:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped_columns.append(item)
    return _assistant_compact_description(table_comment), deduped_columns


def _assistant_build_current_description(meta: dict) -> str:
    parts: list[str] = []
    for key in ("table_comment", "description", "entity_name"):
        raw = str(meta.get(key) or "").strip()
        if raw:
            parts.append(raw)
    for item in (meta.get("attributes") or [])[:12]:
        if not isinstance(item, dict):
            continue
        descr = str(item.get("description") or "").strip()
        if descr:
            parts.append(descr)
    for item in _extract_field_descriptions(meta)[:12]:
        descr = str(item.get("description") or "").strip()
        if descr:
            parts.append(descr)
    return " ".join(parts).strip()


def _build_assistant_table_index() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for entity_root in iter_meta_dirs():
        for root, _, files in os.walk(entity_root):
            if "meta_data_file.yaml" not in files:
                continue
            path = Path(root) / "meta_data_file.yaml"
            try:
                meta = yaml.safe_load(path.read_text("utf-8")) or {}
            except Exception:
                continue
            schema = norm(meta.get("table_schema"))
            table = norm(meta.get("table_name"))
            if not schema or not table:
                continue
            key = ("current", f"{schema}.{table}")
            if key in seen:
                continue
            seen.add(key)
            sql_table_comment, sql_column_comments = _assistant_parse_recreate_comments(path, meta)
            yaml_description = _assistant_build_current_description(meta)
            description = sql_table_comment or yaml_description
            depends_on = meta.get("depends_on") or {}
            search_parts = [
                f"{schema}.{table}",
                str(meta.get("entity_name") or ""),
                description,
                " ".join(sql_column_comments[:20]),
                " ".join(str(x) for x in (meta.get("key_attributes") or [])),
            ]
            rows.append(
                {
                    "source": "current",
                    "schema": schema,
                    "table": table,
                    "fqn": f"{schema}.{table}",
                    "entity_name": meta.get("entity_name"),
                    "description": description,
                    "depends_on": depends_on,
                    "key_attributes": list(meta.get("key_attributes") or []),
                    "search_blob": _assistant_normalize_text(
                        " ".join(search_parts)
                    ),
                }
            )

    for item in get_dbt_table_catalog(BASE_DIR, DBT_MANIFEST_DIR, source="ohd"):
        key = ("ohd", item["fqn"].lower())
        if key in seen:
            continue
        seen.add(key)
        description = str(item.get("description") or "").strip()
        rows.append(
            {
                "source": "ohd",
                "schema": str(item.get("schema") or "").lower(),
                "table": str(item.get("table") or "").lower(),
                "fqn": str(item.get("fqn") or "").lower(),
                "entity_name": item.get("entity_name"),
                "description": description,
                "depends_on": {},
                "key_attributes": [],
                "search_blob": _assistant_normalize_text(
                    " ".join(
                        [
                            str(item.get("fqn") or ""),
                            str(item.get("label") or ""),
                            description,
                            str(item.get("entity_name") or ""),
                        ]
                    )
                ),
            }
        )
    return rows


def get_assistant_table_index() -> list[dict[str, Any]]:
    global _assistant_index_cache, _assistant_index_ts
    now = time.time()
    if _assistant_index_cache and now - _assistant_index_ts < _ASSISTANT_INDEX_TTL:
        return _assistant_index_cache
    _assistant_index_cache = _build_assistant_table_index()
    _assistant_index_ts = now
    return _assistant_index_cache


def _assistant_find_table(schema: str | None, table: str | None, source: str = "current") -> Optional[dict[str, Any]]:
    schema_norm = norm(schema)
    table_norm = norm(table)
    if not schema_norm or not table_norm:
        return None
    for item in get_assistant_table_index():
        if item.get("source") == source and item.get("schema") == schema_norm and item.get("table") == table_norm:
            return item
    return None


def _assistant_extract_fqn(question: str) -> tuple[Optional[str], Optional[str]]:
    match = re.search(r"\b([a-z_][a-z0-9_]*)\.([a-z_][a-z0-9_]*)\b", _assistant_normalize_text(question))
    if not match:
        return None, None
    return match.group(1), match.group(2)


def _assistant_extract_layer(question: str, context: AssistantContextPayload | None = None) -> Optional[str]:
    normalized = _assistant_normalize_text(question)
    for layer in _ASSISTANT_LAYER_NAMES:
        if re.search(rf"(^|\s){re.escape(layer)}($|\s)", normalized):
            return layer
    if "слое" in normalized or "слой" in normalized:
        candidate = norm(getattr(context, "schema", None))
        if candidate in _ASSISTANT_LAYER_NAMES:
            return candidate
    return None


def _assistant_search_tables(question: str, limit: int = 8) -> list[dict[str, Any]]:
    terms = _assistant_extract_terms(question)
    if not terms:
        return []
    results: list[tuple[float, dict[str, Any]]] = []
    for item in get_assistant_table_index():
        blob = item.get("search_blob") or ""
        score = 0.0
        for term in terms:
            if term in (item.get("fqn") or ""):
                score += 3.0
            if term in str(item.get("entity_name") or "").lower():
                score += 2.0
            if term in blob:
                score += 1.0
        if score <= 0:
            continue
        results.append((score, item))
    results.sort(key=lambda row: (-row[0], row[1].get("fqn") or "", row[1].get("source") or ""))
    out = []
    for _, item in results[:limit]:
        out.append(
            {
                "schema": item.get("schema"),
                "table": item.get("table"),
                "source": item.get("source"),
                "fqn": item.get("fqn"),
                "entity_name": item.get("entity_name"),
                "description": str(item.get("description") or "")[:280],
            }
        )
    return out


def _assistant_format_table_refs(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for item in items:
        schema = item.get("schema")
        table = item.get("table")
        if not schema or not table:
            continue
        out.append(
            {
                "schema": schema,
                "table": table,
                "source": item.get("source") or "current",
                "fqn": item.get("fqn") or f"{schema}.{table}",
                "entity_name": item.get("entity_name"),
                "description": item.get("description"),
            }
        )
    return out


def _assistant_answer_summary(schema: str, table: str, source: str) -> dict[str, Any]:
    item = _assistant_find_table(schema, table, source)
    if not item:
        return {
            "mode": "summary",
            "title": "Контекст не найден",
            "answer": "Не удалось найти таблицу в текущем контексте.",
            "tables": [],
            "stats": [],
            "suggestions": ["Самая долгая загрузка", "От чего зависит таблица?", "На что влияет таблица?"],
        }

    stats = [
        {"label": "Источник", "value": "OHD / dbt" if source != "current" else "Current"},
        {"label": "Слой", "value": str(item.get("schema") or "—").upper()},
    ]
    if item.get("entity_name"):
        stats.append({"label": "Сущность", "value": item.get("entity_name")})
    if item.get("key_attributes"):
        stats.append({"label": "Ключи", "value": ", ".join(item.get("key_attributes")[:3])})

    answer = f"Текущий контекст: {item.get('fqn')}."
    if item.get("description"):
        answer += f" Описание: {str(item.get('description'))[:320]}"
    else:
        answer += " Описание в метаданных почти отсутствует."
    return {
        "mode": "summary",
        "title": f"Контекст {item.get('fqn')}",
        "answer": answer,
        "tables": _assistant_format_table_refs([item]),
        "stats": stats,
        "suggestions": ["От чего зависит таблица?", "На что влияет таблица?", "Что есть по похожему описанию?"],
    }


def _assistant_answer_upstream(schema: str, table: str, source: str) -> dict[str, Any]:
    schema_norm = norm(schema)
    table_norm = norm(table)
    if source != "current":
        model = get_dbt_manifest_model(
            base_dir=BASE_DIR,
            manifest_dir=DBT_MANIFEST_DIR,
            schema_name=schema_norm,
            table_name=table_norm,
            source=source,
        )
        upstream = [
            {
                "schema": item.get("schema"),
                "table": item.get("table_name"),
                "source": source,
                "fqn": f"{item.get('schema')}.{item.get('table_name')}" if item.get("schema") and item.get("table_name") else item.get("unique_id"),
                "entity_name": item.get("model_name"),
            }
            for item in (model or {}).get("upstream_models") or []
            if item.get("schema") and item.get("table_name")
        ]
        answer = (
            f"У модели {schema_norm}.{table_norm} найдено {len(upstream)} upstream-зависимостей."
            if upstream else f"Для модели {schema_norm}.{table_norm} upstream-зависимости не найдены."
        )
        return {
            "mode": "upstream",
            "title": "Upstream зависимости",
            "answer": answer,
            "tables": _assistant_format_table_refs(upstream[:12]),
            "stats": [{"label": "Upstream", "value": len(upstream)}, {"label": "Источник", "value": "OHD / dbt"}],
            "suggestions": ["На что влияет таблица?", "Найди похожие таблицы"],
        }

    item = _assistant_find_table(schema_norm, table_norm, source)
    upstream_rows = []
    for src_schema, tables in (item or {}).get("depends_on", {}).items():
        for src_table in tables or []:
            upstream_rows.append(
                {
                    "schema": src_schema,
                    "table": src_table,
                    "source": "current",
                    "fqn": f"{src_schema}.{src_table}",
                }
            )
    answer = (
        f"У таблицы {schema_norm}.{table_norm} найдено {len(upstream_rows)} upstream-зависимостей."
        if upstream_rows else f"Для таблицы {schema_norm}.{table_norm} upstream-зависимости не найдены."
    )
    return {
        "mode": "upstream",
        "title": "Зависимости вверх по потоку",
        "answer": answer,
        "tables": _assistant_format_table_refs(upstream_rows[:12]),
        "stats": [{"label": "Upstream", "value": len(upstream_rows)}, {"label": "Источник", "value": "Current"}],
        "suggestions": ["На что влияет таблица?", "Что это за таблица?"],
    }


def _assistant_answer_impact(schema: str, table: str, source: str) -> dict[str, Any]:
    summary = get_impact_summary(schema, table, depth=3, max_nodes=300, limit=12, source=source)
    if isinstance(summary, JSONResponse):
        return {
            "mode": "impact",
            "title": "Влияние не найдено",
            "answer": "Не удалось построить граф влияния для этой таблицы.",
            "tables": [],
            "stats": [],
            "suggestions": ["От чего зависит таблица?", "Найди похожие таблицы"],
        }
    tables = [
        {
            "schema": row.get("schema"),
            "table": row.get("table"),
            "source": source,
            "fqn": f"{row.get('schema')}.{row.get('table')}" if row.get("schema") and row.get("table") else row.get("id"),
            "entity_name": row.get("entity"),
            "description": f"Depth {row.get('depth')}" if row.get("depth") is not None else None,
        }
        for row in summary.get("tables") or []
    ]
    return {
        "mode": "impact",
        "title": "Влияние вниз по потоку",
        "answer": (
            f"Таблица {schema}.{table} влияет на {summary.get('total_tables') or 0} таблиц "
            f"и {summary.get('total_entities') or 0} сущностей."
        ),
        "tables": _assistant_format_table_refs(tables[:12]),
        "stats": [
            {"label": "Таблиц", "value": summary.get("total_tables") or 0},
            {"label": "Сущностей", "value": summary.get("total_entities") or 0},
            {"label": "Глубина", "value": summary.get("depth") or 0},
        ],
        "suggestions": ["От чего зависит таблица?", "Что это за таблица?"],
    }


def _assistant_answer_slowest(question: str, context: AssistantContextPayload | None = None, days: int = 30) -> dict[str, Any]:
    layer = _assistant_extract_layer(question, context)
    query = f"""
        WITH base AS (
            SELECT
                COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                e.entity_name,
                EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration
            FROM {TABLE_LOADING_HISTORY} l
            LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
            LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
            WHERE l.object_type = 'table'
              AND l.loading_state IN ('SUCCESS', 'LOADED')
              AND l.loading_start_dttm IS NOT NULL
              AND l.loading_finish_dttm IS NOT NULL
              AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
        )
        SELECT
            table_schema,
            table_name,
            entity_name,
            COUNT(*) AS runs_count,
            AVG(duration) AS avg_duration,
            percentile_cont(0.95) WITHIN GROUP (ORDER BY duration) AS p95_duration,
            MAX(duration) AS max_duration
        FROM base
        WHERE (:layer IS NULL OR lower(table_schema) = :layer)
        GROUP BY table_schema, table_name, entity_name
        ORDER BY p95_duration DESC NULLS LAST, max_duration DESC NULLS LAST
        LIMIT 8
    """
    with engine.connect() as conn:
        rows = conn.execute(text(query), {"days": days, "layer": layer}).mappings().all()
    tables = [
        {
            "schema": row.get("table_schema"),
            "table": row.get("table_name"),
            "source": "current",
            "fqn": f"{row.get('table_schema')}.{row.get('table_name')}",
            "entity_name": row.get("entity_name"),
            "description": f"P95 {round(float(row.get('p95_duration') or 0), 2)} мин · runs {int(row.get('runs_count') or 0)}",
        }
        for row in rows
        if row.get("table_schema") and row.get("table_name")
    ]
    layer_text = f" на слое {layer}" if layer else ""
    answer = (
        f"Показываю самые долгие загрузки{layer_text} за последние {days} дней по P95 длительности."
        if rows else f"Не нашел долгих загрузок{layer_text} за последние {days} дней."
    )
    return {
        "mode": "slowest",
        "title": "Самые долгие загрузки",
        "answer": answer,
        "tables": _assistant_format_table_refs(tables),
        "stats": [{"label": "Слой", "value": (layer or "Все").upper()}, {"label": "Период", "value": f"{days} дн"}],
        "suggestions": ["Самая долгая загрузка на слое dm", "От чего зависит таблица?", "На что влияет таблица?"],
    }


def _assistant_answer_click_slowest(question: str, context: AssistantContextPayload | None = None, days: int = 30) -> dict[str, Any]:
    layer = _assistant_extract_layer(question, context)
    query = (
        _clickhouse_run_agg_cte(
            run_filter_sql="AND r.start_dttm >= now() - (:days || ' days')::interval",
            stage_filter_sql="",
        )
        + """
        SELECT
            lower(schema_name) AS table_schema,
            lower(table_name) AS table_name,
            COUNT(*) AS runs_count,
            AVG(actual_duration_seconds) / 60.0 AS avg_duration,
            percentile_cont(0.95) WITHIN GROUP (ORDER BY actual_duration_seconds / 60.0) AS p95_duration,
            MAX(actual_duration_seconds) / 60.0 AS max_duration
        FROM run_agg
        WHERE (:layer IS NULL OR lower(schema_name) = :layer)
        GROUP BY lower(schema_name), lower(table_name)
        ORDER BY p95_duration DESC NULLS LAST, max_duration DESC NULLS LAST
        LIMIT 8
        """
    )
    with engine.connect() as conn:
        rows = conn.execute(text(query), {"days": days, "layer": layer}).mappings().all()
    tables = [
        {
            "schema": row.get("table_schema"),
            "table": row.get("table_name"),
            "source": "current",
            "fqn": f"{row.get('table_schema')}.{row.get('table_name')}",
            "description": f"ClickHouse P95 {round(float(row.get('p95_duration') or 0), 2)} мин · runs {int(row.get('runs_count') or 0)}",
        }
        for row in rows
        if row.get("table_schema") and row.get("table_name")
    ]
    layer_text = f" на слое {layer}" if layer else ""
    answer = (
        f"Показываю самые долгие загрузки в ClickHouse{layer_text} за последние {days} дней по P95 длительности."
        if rows else f"Не нашел длительных загрузок в ClickHouse{layer_text} за последние {days} дней."
    )
    return {
        "mode": "slowest_click",
        "title": "Самые долгие загрузки ClickHouse",
        "answer": answer,
        "tables": _assistant_format_table_refs(tables),
        "stats": [{"label": "Контур", "value": "ClickHouse"}, {"label": "Слой", "value": (layer or "Все").upper()}],
        "suggestions": ["Самая долгая загрузка", "Покажи последние ошибки", "На что влияет таблица?"],
    }


def _assistant_answer_primary_key(schema: str, table: str, source: str) -> dict[str, Any]:
    item = _assistant_find_table(schema, table, source)
    key_attributes = list((item or {}).get("key_attributes") or [])
    if not key_attributes:
        return {
            "mode": "primary_key",
            "title": "Первичный ключ",
            "answer": f"Для таблицы {schema}.{table} первичный ключ или ключевые атрибуты в метаданных не заданы.",
            "tables": _assistant_format_table_refs([item]) if item else [],
            "stats": [{"label": "Ключей", "value": 0}],
            "suggestions": ["Что это за таблица?", "От чего зависит таблица?", "Покажи последние ошибки"],
        }

    formatted = []
    for raw in key_attributes:
        raw_text = str(raw or "").strip()
        parts = [part.strip() for part in raw_text.split("|") if part.strip()]
        source_ref = next((part for part in parts if "." in part), "")
        field_name = source_ref.split(".")[-1] if source_ref else raw_text
        formatted.append({"schema": schema, "table": table, "source": source, "fqn": f"{schema}.{table}", "description": f"{field_name} · {_assistant_compact_description(raw_text)}"})

    return {
        "mode": "primary_key",
        "title": "Первичный ключ / key attributes",
        "answer": f"Для таблицы {schema}.{table} найдено {len(key_attributes)} ключевых атрибутов.",
        "tables": formatted[:8],
        "stats": [{"label": "Ключей", "value": len(key_attributes)}],
        "suggestions": ["Что это за таблица?", "От чего зависит таблица?", "На что влияет таблица?"],
    }


def _assistant_answer_recent_errors(schema: str | None, table: str | None, source: str, limit: int = 6) -> dict[str, Any]:
    schema_norm = norm(schema)
    table_norm = norm(table)
    if source != "current" and schema_norm and table_norm and dbt_logs_engine:
        try:
            payload = get_dbt_model_run_history(
                engine=dbt_logs_engine,
                base_dir=BASE_DIR,
                manifest_dir=DBT_MANIFEST_DIR,
                schema_name=schema_norm,
                table_name=table_norm,
                source=source,
                limit=limit,
                table_model_catalog=TABLE_DBT_MODEL_CATALOG,
                table_model_log=TABLE_DBT_MODEL_LOG,
                table_run_log=TABLE_DBT_RUN_LOG,
            )
            runs = [row for row in (payload.get("runs") or []) if str(row.get("model_status") or "").upper() == "FAILED"][:limit]
        except Exception:
            runs = []
        return {
            "mode": "errors",
            "title": "Последние ошибки dbt",
            "answer": (
                f"Для {schema_norm}.{table_norm} нашел {len(runs)} последних ошибок dbt."
                if runs else f"Для {schema_norm}.{table_norm} ошибок dbt не найдено."
            ),
            "tables": [
                {
                    "schema": schema_norm,
                    "table": table_norm,
                    "source": source,
                    "fqn": f"{schema_norm}.{table_norm}",
                    "description": f"dbt FAILED · {serialize_datetime(row.get('finish_dttm')) or row.get('finish_dttm') or '—'} · {str(row.get('error_message') or '')[:180]}",
                }
                for row in runs
            ],
            "stats": [{"label": "Источник", "value": "dbt"}, {"label": "Ошибок", "value": len(runs)}],
            "suggestions": ["Что это за таблица?", "От чего зависит таблица?", "На что влияет таблица?"],
        }

    gp_rows: list[dict[str, Any]] = []
    click_rows: list[dict[str, Any]] = []
    with engine.connect() as conn:
        if schema_norm and table_norm:
            table_clean = _clean_table_name(table_norm)
            gp_query = f"""
                SELECT
                    :schema AS table_schema,
                    :table AS table_name,
                    loading_finish_dttm AS error_time,
                    message
                FROM {TABLE_LOADING_HISTORY}
                WHERE object_type = 'table'
                  AND loading_state = 'FAILED'
                  AND (
                    lower(object_name) = :table_fqn
                    OR lower(object_name) = :table_fqn_clean
                    OR lower(object_name) = :table_name
                    OR lower(object_name) = :table_name_clean
                  )
                ORDER BY loading_finish_dttm DESC NULLS LAST
                LIMIT :limit
            """
            gp_rows = conn.execute(
                text(gp_query),
                {
                    "schema": schema_norm,
                    "table": table_norm,
                    "table_fqn": f"{schema_norm}.{table_norm}",
                    "table_fqn_clean": f"{schema_norm}.{table_clean}" if table_clean else None,
                    "table_name": table_norm,
                    "table_name_clean": table_clean,
                    "limit": limit,
                },
            ).mappings().all()

            click_query = (
                _clickhouse_run_agg_cte(
                    run_filter_sql="",
                    stage_filter_sql="""
                      AND (lower(s.table_name) = lower(:table) OR lower(s.table_name) = lower(:table_clean))
                      AND lower(r.schema_name) = lower(:schema)
                    """,
                )
                + """
                SELECT *
                FROM run_agg
                WHERE upper(status) = 'FAILED'
                ORDER BY end_dttm DESC NULLS LAST
                LIMIT :limit
                """
            )
            click_rows = conn.execute(
                text(click_query),
                {"schema": schema_norm, "table": table_norm, "table_clean": table_clean, "limit": limit},
            ).mappings().all()
        else:
            gp_query = f"""
                SELECT
                    COALESCE(split_part(lower(object_name), '.', 1), '') AS table_schema,
                    COALESCE(NULLIF(split_part(lower(object_name), '.', 2), ''), lower(object_name)) AS table_name,
                    loading_finish_dttm AS error_time,
                    message
                FROM {TABLE_LOADING_HISTORY}
                WHERE object_type = 'table'
                  AND loading_state = 'FAILED'
                ORDER BY loading_finish_dttm DESC NULLS LAST
                LIMIT :limit
            """
            gp_rows = conn.execute(text(gp_query), {"limit": limit}).mappings().all()

    tables = [
        {
            "schema": row.get("table_schema") or schema_norm,
            "table": row.get("table_name") or table_norm,
            "source": "current",
            "fqn": f"{row.get('table_schema') or schema_norm}.{row.get('table_name') or table_norm}",
            "description": f"GP · {serialize_datetime(row.get('error_time')) or '—'} · {str(row.get('message') or '')[:180]}",
        }
        for row in gp_rows
        if (row.get("table_schema") or schema_norm) and (row.get("table_name") or table_norm)
    ]
    tables.extend(
        {
            "schema": row.get("schema_name") or schema_norm,
            "table": row.get("table_name") or table_norm,
            "source": "current",
            "fqn": f"{row.get('schema_name') or schema_norm}.{row.get('table_name') or table_norm}",
            "description": f"ClickHouse · {serialize_datetime(row.get('end_dttm')) or '—'} · {str(row.get('error_text') or '')[:180]}",
        }
        for row in click_rows
        if (row.get("schema_name") or schema_norm) and (row.get("table_name") or table_norm)
    )
    title_target = f" для {schema_norm}.{table_norm}" if schema_norm and table_norm else ""
    return {
        "mode": "errors",
        "title": f"Последние ошибки{title_target}",
        "answer": (
            f"Нашел {len(tables)} последних ошибок{title_target}."
            if tables else f"Последние ошибки{title_target} не найдены."
        ),
        "tables": _assistant_format_table_refs(tables[:limit]),
        "stats": [{"label": "Ошибок", "value": len(tables)}, {"label": "Контекст", "value": "Таблица" if schema_norm and table_norm else "Глобально"}],
        "suggestions": ["Самая долгая загрузка", "От чего зависит таблица?", "На что влияет таблица?"],
    }


def _assistant_find_entity_name(question: str) -> Optional[str]:
    normalized_question = _assistant_normalize_text(question)
    if not normalized_question:
        return None
    snapshot = get_graph_snapshot()
    entities = sorted({entity for values in snapshot.get("table_entity_map", {}).values() for entity in values if entity})
    normalized_entities = [(_assistant_normalize_text(entity), entity) for entity in entities]
    for normalized_entity, entity in normalized_entities:
        if normalized_entity and normalized_entity in normalized_question:
            return entity

    terms = _assistant_extract_terms(question)
    best: tuple[int, str] | None = None
    for normalized_entity, entity in normalized_entities:
        if not normalized_entity:
            continue
        score = sum(1 for term in terms if term in normalized_entity)
        if score and (best is None or score > best[0] or (score == best[0] and len(entity) < len(best[1]))):
            best = (score, entity)
    return best[1] if best else None


def _assistant_answer_entity_usage(question: str) -> dict[str, Any]:
    entity_name = _assistant_find_entity_name(question)
    if not entity_name:
        return {
            "mode": "entity_usage",
            "title": "Влияние сущности",
            "answer": "Не понял название сущности. Напиши его полностью или открой граф сущности и повтори вопрос.",
            "tables": [],
            "stats": [],
            "suggestions": ["Влияет ли сущность TRANSPORTATION на другие сущности?", "Есть зависимости у сущности BI_FI?"],
        }

    snapshot = get_graph_snapshot()
    table_edges = snapshot["table_graph"]["edges"]
    table_nodes = snapshot["table_graph"]["nodes"]
    table_entity_map = {key: set(values or []) for key, values in snapshot["table_entity_map"].items()}
    entity_tables = {table_id for table_id, entities in table_entity_map.items() if entity_name in entities}
    downstream: dict[str, set[str]] = {}
    upstream: dict[str, set[str]] = {}

    for edge in table_edges:
        source = edge.get("source")
        target = edge.get("target")
        if not source or not target:
            continue
        source_entities = table_entity_map.get(source) or set()
        target_entities = table_entity_map.get(target) or set()
        if source in entity_tables:
            for target_entity in target_entities:
                if target_entity and target_entity != entity_name:
                    downstream.setdefault(target_entity, set()).add(target)
        if target in entity_tables:
            for source_entity in source_entities:
                if source_entity and source_entity != entity_name:
                    upstream.setdefault(source_entity, set()).add(source)

    downstream_rows = sorted(downstream.items(), key=lambda item: (-len(item[1]), item[0]))[:8]
    upstream_rows = sorted(upstream.items(), key=lambda item: (-len(item[1]), item[0]))[:6]
    tables = []
    for target_entity, table_ids in downstream_rows:
        sample_id = sorted(table_ids)[0]
        node = table_nodes.get(sample_id) or {}
        tables.append(
            {
                "schema": node.get("schema"),
                "table": node.get("table"),
                "source": "current",
                "fqn": sample_id,
                "entity_name": target_entity,
                "description": f"Используется в сущности: {target_entity} · таблиц: {len(table_ids)}",
            }
        )

    if downstream_rows:
        downstream_text = ", ".join(f"{name} ({len(table_ids)})" for name, table_ids in downstream_rows)
        answer = f"Таблицы сущности {entity_name} используются в других сущностях: {downstream_text}."
    else:
        answer = f"Не нашел, что таблицы сущности {entity_name} используются в других сущностях."
    if upstream_rows:
        upstream_text = ", ".join(f"{name} ({len(table_ids)})" for name, table_ids in upstream_rows)
        answer += f" Также эта сущность зависит от: {upstream_text}."

    return {
        "mode": "entity_usage",
        "title": f"Связи сущности {entity_name}",
        "answer": answer,
        "tables": _assistant_format_table_refs(tables),
        "stats": [
            {"label": "Таблиц сущности", "value": len(entity_tables)},
            {"label": "Куда влияет", "value": len(downstream)},
            {"label": "От кого зависит", "value": len(upstream)},
        ],
        "suggestions": ["Покажи последние ошибки", "Самая долгая загрузка", "Найди таблицы по описанию"],
    }


def _assistant_answer_search(question: str) -> dict[str, Any]:
    tables = _assistant_search_tables(question, limit=10)
    if not tables:
        return {
            "mode": "search",
            "title": "Поиск по описанию",
            "answer": "По запросу ничего не нашел в описаниях таблиц и dbt-моделей.",
            "tables": [],
            "stats": [],
            "suggestions": ["Самая долгая загрузка", "От чего зависит таблица?", "На что влияет таблица?"],
        }
    return {
        "mode": "search",
        "title": "Подходящие таблицы",
        "answer": f"Нашел {len(tables)} таблиц, которые похожи по описанию или названию.",
        "tables": tables,
        "stats": [{"label": "Совпадений", "value": len(tables)}],
        "suggestions": ["От чего зависит таблица?", "На что влияет таблица?", "Самая долгая загрузка"],
    }


def _assistant_answer(question: str, context: AssistantContextPayload | None = None) -> dict[str, Any]:
    question = (question or "").strip()
    source = (getattr(context, "source", None) or "current").strip() or "current"
    ctx_schema = norm(getattr(context, "schema", None))
    ctx_table = norm(getattr(context, "table", None))
    if (not ctx_schema or not ctx_table) and question:
        parsed_schema, parsed_table = _assistant_extract_fqn(question)
        ctx_schema = ctx_schema or parsed_schema
        ctx_table = ctx_table or parsed_table

    normalized = _assistant_normalize_text(question)
    if not normalized and ctx_schema and ctx_table:
        return _assistant_answer_summary(ctx_schema, ctx_table, source)
    if ("последние ошибки" in normalized or "покажи ошибки" in normalized or "ошибк" in normalized):
        return _assistant_answer_recent_errors(ctx_schema, ctx_table, source)
    if ("сущност" in normalized or "entity" in normalized) and ("влия" in normalized or "завис" in normalized or "использ" in normalized):
        return _assistant_answer_entity_usage(question)
    if ("первич" in normalized and "ключ" in normalized) or "key attributes" in normalized or "ключевые атрибуты" in normalized:
        if ctx_schema and ctx_table:
            return _assistant_answer_primary_key(ctx_schema, ctx_table, source)
        return {
            "mode": "primary_key",
            "title": "Первичный ключ",
            "answer": "Укажи таблицу или открой карточку таблицы, и я покажу key attributes.",
            "tables": [],
            "stats": [],
            "suggestions": ["Что это за таблица?", "От чего зависит таблица?", "Покажи последние ошибки"],
        }
    if "клик" in normalized or "clickhouse" in normalized:
        if "долг" in normalized and "загруз" in normalized:
            return _assistant_answer_click_slowest(question, context=context)
    if ("от чего зависит" in normalized or "upstream" in normalized or "зависимости" in normalized) and ctx_schema and ctx_table:
        return _assistant_answer_upstream(ctx_schema, ctx_table, source)
    if ("на что влияет" in normalized or "impact" in normalized or "влияет" in normalized) and ctx_schema and ctx_table:
        return _assistant_answer_impact(ctx_schema, ctx_table, source)
    if "долг" in normalized and "загруз" in normalized:
        return _assistant_answer_slowest(question, context=context)
    if ("что это" in normalized or "опиши" in normalized or "описание" in normalized) and ctx_schema and ctx_table:
        return _assistant_answer_summary(ctx_schema, ctx_table, source)
    return _assistant_answer_search(question)


def iter_meta_dirs(targets: Optional[List[str]] = None):
    """Yield existing metadata directories, searching both root and project/* trees."""
    seen = set()
    for parent in META_PARENT_DIRS:
        if not parent.exists():
            continue
        if targets:
            names = targets
        else:
            names = [p.name for p in parent.iterdir() if p.is_dir()]
        for name in names:
            candidate = parent / name
            if not candidate.exists():
                continue
            real = candidate.resolve()
            if real in seen:
                continue
            seen.add(real)
            yield candidate


def normalize_excel_table_name(value: Optional[str]) -> Optional[str]:
    if not value:
        return None

    cleaned = (
        str(value)
        .strip()
        .replace('"', "")
        .replace("'", "")
        .replace("`", "")
        .replace("[", "")
        .replace("]", "")
        .replace(" ", "")
        .replace("\n", "")
    )

    if not cleaned or cleaned == "-":
        return None

    cleaned = cleaned.lower()
    cleaned = cleaned.replace("..", ".")

    if "." in cleaned:
        schema, table = cleaned.split(".", 1)
        schema = schema.strip()
        table = table.strip()
        if not schema or not table:
            return None
        return f"{schema}.{table}"

    return cleaned or None


def format_excel_datetime(value) -> Optional[str]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")
    try:
        parsed = datetime.fromisoformat(str(value))
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(value)


def parse_ytrek_excel(file_path: Path) -> List[dict]:
    if not file_path.exists():
        raise FileNotFoundError(file_path)

    header_map = {
        "№": "index",
        "id задачи": "issue_id",
        "название задачи": "title",
        "дата начала сбоя": "start_at",
        "дата обнаружения": "detected_at",
        "дата окончания работ": "resolved_at",
        "ссылка на задачу": "link",
        "название таблицы": "table_raw",
        "название сущности": "entity_name",
    }

    def norm_header(value):
        return (value or "").strip().lower()

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    normalized_headers = [norm_header(h) for h in header_row]

    incidents = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for idx, cell_value in enumerate(raw_row):
            header_key = normalized_headers[idx] if idx < len(normalized_headers) else None
            mapped_key = header_map.get(header_key)
            if not mapped_key:
                continue
            record[mapped_key] = cell_value

        issue_id = str(record.get("issue_id") or "").strip()
        title = str(record.get("title") or "").strip()

        if not issue_id and not title:
            continue

        table_raw = str(record.get("table_raw") or "").strip()
        incidents.append(
            {
                "issue_id": issue_id,
                "title": title or "Без названия",
                "start_at": format_excel_datetime(record.get("start_at")),
                "detected_at": format_excel_datetime(record.get("detected_at")),
                "resolved_at": format_excel_datetime(record.get("resolved_at")),
                "link": record.get("link"),
                "table_raw": table_raw,
                "table_normalized": normalize_excel_table_name(table_raw),
                "entity_name": str(record.get("entity_name") or "").strip(),
            }
        )

    incidents.sort(key=lambda x: x.get("start_at") or "", reverse=True)
    return incidents


def build_tables_meta_index() -> tuple[dict, dict]:
    """Return lookup dictionaries for tables_meta: by fqn and by table name."""
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                f"""
                SELECT t.table_id, t.table_schema, t.table_name, t.entity_id, e.entity_name
                FROM {TABLE_TABLES_META} t
                LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                WHERE t.table_schema IS NOT NULL AND t.table_name IS NOT NULL
                """
            )
        ).mappings().all()

    by_fqn = {}
    by_name = {}

    for row in rows:
        schema = (row["table_schema"] or "").lower()
        table = (row["table_name"] or "").lower()
        if not schema or not table:
            continue
        key = f"{schema}.{table}"
        by_fqn[key] = row
        by_name.setdefault(table, []).append(row)

    return by_fqn, by_name


def ensure_ytrek_table_exists(conn) -> None:
    conn.execute(
        text(
            f"""
            CREATE TABLE IF NOT EXISTS {TABLE_YTREK_INCIDENTS} (
                issue_id TEXT PRIMARY KEY,
                title TEXT,
                start_at TIMESTAMP NULL,
                detected_at TIMESTAMP NULL,
                resolved_at TIMESTAMP NULL,
                link TEXT,
                table_raw TEXT,
                table_normalized TEXT,
                table_schema TEXT,
                table_name TEXT,
                table_id BIGINT,
                entity_name TEXT,
                inserted_at TIMESTAMP DEFAULT NOW()
            )
            """
        )
    )


def parse_timestamp_value(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def pick_table_match(table_normalized: Optional[str], entity_hint: Optional[str], by_fqn: dict, by_name: dict):
    if not table_normalized:
        return None

    normalized = table_normalized.lower()
    entity_hint = (entity_hint or "").lower()

    if "." in normalized and normalized in by_fqn:
        return by_fqn[normalized]

    # Fall back to table-name-only match
    table_name = normalized.split(".")[-1]
    candidates = by_name.get(table_name) or []
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    if entity_hint:
        filtered = [c for c in candidates if (c.get("entity_name") or "").lower() == entity_hint]
        if len(filtered) == 1:
            return filtered[0]
    return candidates[0]


def import_ytrek_from_excel(file_path: Union[str, Path]) -> int:
    """Load incidents from an Excel export into the database table."""
    path = Path(file_path)
    incidents = parse_ytrek_excel(path)
    by_fqn, by_name = build_tables_meta_index()

    rows = []
    for record in incidents:
        table_norm = record.get("table_normalized")
        entity_hint = record.get("entity_name")
        match = pick_table_match(table_norm, entity_hint, by_fqn, by_name)

        matched_schema = match.get("table_schema") if match else None
        matched_table = match.get("table_name") if match else None
        consolidated_entity = match.get("entity_name") if match else None
        if not consolidated_entity:
            consolidated_entity = entity_hint or None

        rows.append(
            {
                "issue_id": record.get("issue_id"),
                "title": record.get("title"),
                "start_at": parse_timestamp_value(record.get("start_at")),
                "detected_at": parse_timestamp_value(record.get("detected_at")),
                "resolved_at": parse_timestamp_value(record.get("resolved_at")),
                "link": record.get("link"),
                "table_raw": record.get("table_raw"),
                "table_normalized": table_norm,
                "table_schema": matched_schema,
                "table_name": matched_table,
                "table_id": match.get("table_id") if match else None,
                "entity_name": consolidated_entity,
            }
        )

    with engine.begin() as conn:
        ensure_ytrek_table_exists(conn)
        conn.execute(text(f"TRUNCATE TABLE {TABLE_YTREK_INCIDENTS}"))
        if rows:
            conn.execute(
                text(
                    f"""
                    INSERT INTO {TABLE_YTREK_INCIDENTS} (
                        issue_id, title, start_at, detected_at, resolved_at,
                        link, table_raw, table_normalized, table_schema, table_name,
                        table_id, entity_name
                    ) VALUES (
                        :issue_id, :title, :start_at, :detected_at, :resolved_at,
                        :link, :table_raw, :table_normalized, :table_schema, :table_name,
                        :table_id, :entity_name
                    )
                    """
                ),
                rows,
            )

    return len(rows)


def extract_incident_day(row: dict) -> Optional[date]:
    for key in ("start_at", "detected_at", "resolved_at"):
        value = row.get(key)
        if not value:
            continue
        if isinstance(value, datetime):
            return value.date()
        try:
            return datetime.fromisoformat(str(value)).date()
        except Exception:
            continue
    return None


def fetch_failure_lookup(conn, incident_rows):
    combos = []
    for row in incident_rows:
        table_id = row.get("table_id")
        if not table_id:
            continue
        day = extract_incident_day(row)
        if not day:
            continue
        combos.append((table_id, day))

    if not combos:
        return {}

    unique_ids = sorted({tid for tid, _ in combos})
    min_day = min(day for _, day in combos)
    max_day = max(day for _, day in combos)

    start_bound = datetime.combine(min_day, datetime.min.time())
    end_bound = datetime.combine(max_day + timedelta(days=1), datetime.min.time())

    placeholders = ", ".join(f":id{i}" for i in range(len(unique_ids)))
    params = {f"id{i}": tid for i, tid in enumerate(unique_ids)}
    params.update({"start_bound": start_bound, "end_bound": end_bound})

    query = text(
        f"""
        SELECT object_id, DATE(loading_finish_dttm) AS incident_day, COUNT(*) AS fail_count
        FROM {TABLE_LOADING_HISTORY}
        WHERE object_id IN ({placeholders})
          AND loading_state = 'FAILED'
          AND loading_finish_dttm >= :start_bound
          AND loading_finish_dttm < :end_bound
        GROUP BY object_id, DATE(loading_finish_dttm)
        """
    )

    rows = conn.execute(query, params).mappings().all()
    lookup = {}
    for row in rows:
        lookup[(row["object_id"], row["incident_day"])] = row["fail_count"]

    return lookup


def build_ytrek_dashboard(top_limit: int):
    with engine.begin() as conn:
        ensure_ytrek_table_exists(conn)

    with engine.connect() as conn:
        incident_rows = conn.execute(
            text(
                f"""
                SELECT issue_id, title, start_at, detected_at, resolved_at, link,
                       table_raw, table_normalized, table_schema, table_name,
                       table_id, entity_name
                FROM {TABLE_YTREK_INCIDENTS}
                ORDER BY COALESCE(start_at, detected_at, resolved_at) DESC NULLS LAST,
                         issue_id DESC
                """
            )
        ).mappings().all()

        timeline_rows = conn.execute(
            text(
                f"""
                SELECT DATE(COALESCE(start_at, detected_at, resolved_at)) AS incident_day,
                       COUNT(*) AS incidents_count
                FROM {TABLE_YTREK_INCIDENTS}
                WHERE COALESCE(start_at, detected_at, resolved_at) IS NOT NULL
                GROUP BY DATE(COALESCE(start_at, detected_at, resolved_at))
                ORDER BY incident_day DESC
                LIMIT 30
                """
            )
        ).mappings().all()

        top_tables_rows = conn.execute(
            text(
                f"""
                SELECT
                    table_schema,
                    table_name,
                    table_raw,
                    COUNT(*) AS incidents_count,
                    MAX(COALESCE(start_at, detected_at, resolved_at)) AS last_incident
                FROM {TABLE_YTREK_INCIDENTS}
                GROUP BY table_schema, table_name, table_raw
                ORDER BY incidents_count DESC, last_incident DESC
                LIMIT :limit
                """
            ),
            {"limit": top_limit},
        ).mappings().all()

        top_entities_rows = conn.execute(
            text(
                f"""
                SELECT
                    COALESCE(NULLIF(entity_name, ''), 'Не указано') AS entity_key,
                    COUNT(*) AS incidents_count,
                    MAX(COALESCE(start_at, detected_at, resolved_at)) AS last_incident
                FROM {TABLE_YTREK_INCIDENTS}
                GROUP BY entity_key
                ORDER BY incidents_count DESC, entity_key
                LIMIT :limit
                """
            ),
            {"limit": top_limit},
        ).mappings().all()

        failure_lookup = fetch_failure_lookup(conn, incident_rows)

    incidents = []
    tables_set = set()
    entities_set = set()
    mapped_count = 0
    db_matches = 0

    for row in incident_rows:
        table_schema = row.get("table_schema")
        table_name = row.get("table_name")
        table_fqn = None
        if table_schema and table_name:
            table_fqn = f"{table_schema}.{table_name}"
            tables_set.add(table_fqn)
            mapped_count += 1

        entity_value = row.get("entity_name")
        if isinstance(entity_value, str):
            entity_value = entity_value.strip()
        if entity_value:
            entities_set.add(entity_value)

        incident_day = extract_incident_day(row)
        day_key = incident_day.strftime("%Y-%m-%d") if incident_day else None
        failures = 0
        if row.get("table_id") and incident_day:
            failures = failure_lookup.get((row["table_id"], incident_day), 0)
        if failures:
            db_matches += 1

        incidents.append(
            {
                "issue_id": row["issue_id"],
                "title": row["title"],
                "link": row.get("link"),
                "start_at": serialize_datetime(row.get("start_at")),
                "detected_at": serialize_datetime(row.get("detected_at")),
                "resolved_at": serialize_datetime(row.get("resolved_at")),
                "table_raw": row.get("table_raw"),
                "table_normalized": row.get("table_normalized"),
                "table_schema": table_schema,
                "table_name": table_name,
                "table_fqn": table_fqn,
                "table_id": row.get("table_id"),
                "entity_name": row.get("entity_name"),
                "has_table": bool(table_fqn),
                "incident_day": day_key,
                "has_db_failures": failures > 0,
                "db_failures_count": failures,
            }
        )

    stats = {
        "total": len(incidents),
        "with_table": mapped_count,
        "unique_tables": len(tables_set),
        "unique_entities": len(entities_set),
        "with_db_failures": db_matches,
    }

    timeline = [
        {
            "day": row["incident_day"].strftime("%Y-%m-%d"),
            "count": row["incidents_count"],
        }
        for row in reversed(timeline_rows)
        if row.get("incident_day")
    ]

    top_tables = []
    for row in top_tables_rows:
        schema = row.get("table_schema")
        table = row.get("table_name")
        table_fqn = f"{schema}.{table}" if schema and table else None
        label = table_fqn or row.get("table_raw") or "—"
        top_tables.append(
            {
                "label": label,
                "table_fqn": table_fqn,
                "count": row["incidents_count"],
                "last_incident": serialize_datetime(row.get("last_incident")),
                "has_table": bool(table_fqn),
            }
        )

    top_entities = [
        {
            "label": row["entity_key"],
            "count": row["incidents_count"],
            "last_incident": serialize_datetime(row.get("last_incident")),
        }
        for row in top_entities_rows
    ]

    return {
        "stats": stats,
        "timeline": timeline,
        "top_tables": top_tables,
        "top_entities": top_entities,
        "incidents": incidents,
    }


def serialize_datetime(value):
    if not value:
        return None
    if isinstance(value, str):
        return value
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _duration_minutes(value):
    if value is None:
        return None
    try:
        if isinstance(value, timedelta):
            return round(value.total_seconds() / 60.0, 2)
        if isinstance(value, (int, float)):
            return round(float(value) / 60.0, 2)
        return None
    except Exception:
        return None


def _round_minutes(value):
    if value is None:
        return None
    try:
        return round(float(value), 2)
    except Exception:
        return None


def _normalize_rich_text(value: Any) -> str:
    text_value = str(value or "")
    if not text_value:
        return ""
    text_value = unescape(text_value)
    text_value = re.sub(r"<br\s*/?>", "\n", text_value, flags=re.I)
    text_value = re.sub(r"</p\s*>", "\n", text_value, flags=re.I)
    text_value = re.sub(r"<[^>]+>", " ", text_value)
    text_value = re.sub(r"[ \t]+\n", "\n", text_value)
    text_value = re.sub(r"\n{3,}", "\n\n", text_value)
    text_value = re.sub(r"[ \t]{2,}", " ", text_value)
    return text_value.strip()


def _split_rich_multivalue(value: Any) -> list[str]:
    normalized = _normalize_rich_text(value)
    if not normalized:
        return []
    parts = []
    for raw in re.split(r"\n|;", normalized):
        item = raw.strip().strip(",")
        if item:
            parts.append(item)
    seen = set()
    dedup = []
    for item in parts:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        dedup.append(item)
    return dedup


def _parse_effort_minutes(value: Any) -> int:
    raw = str(value or "").strip().lower()
    if not raw:
        return 0
    normalized = raw.replace(",", ".")
    total = 0.0
    patterns = [
        (r"(\d+(?:\.\d+)?)\s*(?:d|day|days|д|дн|день|дня|дней)", 24 * 60),
        (r"(\d+(?:\.\d+)?)\s*(?:h|hr|hrs|hour|hours|ч|час|часа|часов)", 60),
        (r"(\d+(?:\.\d+)?)\s*(?:m|min|mins|minute|minutes|м|мин|минута|минуты|минут)", 1),
    ]
    matched = False
    for pattern, factor in patterns:
        for number in re.findall(pattern, normalized):
            total += float(number) * factor
            matched = True
    if matched:
        return int(round(total))
    plain = re.search(r"(\d+(?:\.\d+)?)", normalized)
    if plain:
        return int(round(float(plain.group(1)) * 60))
    return 0


def _incident_status_bucket(state_name: Any) -> str:
    state = str(state_name or "").strip().lower()
    if not state:
        return "unknown"
    if "заверш" in state or "resolved" in state or "done" in state or "closed" in state:
        return "resolved"
    if "нов" in state or "зарегистр" in state or "open" in state:
        return "open"
    if "работ" in state or "progress" in state or "исправ" in state:
        return "in_progress"
    return "other"


def _incident_week_start(value: Any) -> Optional[str]:
    if not value:
        return None
    dt_value = value if isinstance(value, datetime) else None
    if dt_value is None:
        try:
            dt_value = datetime.fromisoformat(str(value))
        except Exception:
            return None
    week_start = dt_value - timedelta(days=dt_value.weekday())
    return week_start.date().isoformat()


def _clickhouse_run_metrics(row):
    actual_seconds = float(row.get("actual_duration_seconds") or 0.0)
    elapsed_seconds = float(row.get("elapsed_duration_seconds") or 0.0)
    lag_seconds = max(elapsed_seconds - actual_seconds, 0.0)
    return {
        "actual_duration_min": round(actual_seconds / 60.0, 2),
        "elapsed_duration_min": round(elapsed_seconds / 60.0, 2),
        "lag_duration_min": round(lag_seconds / 60.0, 2),
        # Backward-compatible alias for old UI callers.
        "duration_min": round(actual_seconds / 60.0, 2),
    }


def _clickhouse_run_agg_cte(
    run_filter_sql: str = "",
    stage_filter_sql: str = "",
):
    return f"""
        WITH run_agg AS (
            SELECT
                r.run_uuid,
                r.schema_name,
                s.table_name,
                r.dag_name,
                r.dag_run,
                r.status,
                r.error_text,
                MAX(s.table_id) AS table_id,
                MIN(s.start_dttm) AS start_dttm,
                MAX(s.end_dttm) AS end_dttm,
                SUM(EXTRACT(EPOCH FROM s.duration)) AS actual_duration_seconds,
                EXTRACT(EPOCH FROM (MAX(s.end_dttm) - MIN(s.start_dttm))) AS elapsed_duration_seconds
            FROM {TABLE_CLICK_LOAD_RUN} r
            JOIN {TABLE_CLICK_LOAD_STAGE} s
              ON s.run_uuid = r.run_uuid
             AND s.table_id = r.table_id
            WHERE s.stage_name IN ('UPLOAD_TO_S3', 'CLICKHOUSE_LOAD')
              {run_filter_sql}
              {stage_filter_sql}
            GROUP BY
                r.run_uuid,
                r.schema_name,
                s.table_name,
                r.dag_name,
                r.dag_run,
                r.status,
                r.error_text
        )
    """


@router.get("/api/ytrek/incidents")
def get_ytrek_incidents(top_limit: int = Query(5, ge=1, le=50)):
    return build_ytrek_dashboard(top_limit)

def resolve_dependencies(schema: str, table: str) -> List[DependencyItem]:
    schema = (schema or "").strip().lower()
    table = (table or "").strip().lower()
    all_meta, reverse_index = get_cached_meta_and_index()

    start = (schema, table)
    visited = {start}
    parent = {}
    depth = {start: 0}
    meta_by_node = {}
    order = []

    queue = deque([start])
    while queue:
        node = queue.popleft()
        for dep in reverse_index.get(node, []):
            child = (dep["schema"], dep["table_name"])
            if child in visited:
                continue
            visited.add(child)
            parent[child] = node
            depth[child] = depth[node] + 1
            meta_by_node[child] = dep
            order.append(child)
            queue.append(child)

    out = []
    with engine.connect() as conn:
        for i, node in enumerate(order, 1):
            r = meta_by_node[node]
            avg = None
            if r.get("table_id"):
                avg = conn.execute(
                    text(f"""
                        SELECT round((avg(extract(epoch from (loading_finish_dttm-loading_start_dttm))/60))::numeric,2)
                        FROM {TABLE_LOADING_HISTORY}
                        WHERE object_id=:id AND loading_state IN ('SUCCESS', 'LOADED')
                    """),
                    {"id": r["table_id"]}
                ).scalar()

            path = []
            current = node
            while current != start:
                path.append(f"{current[0]}.{current[1]}")
                current = parent[current]
            path.append(f"{start[0]}.{start[1]}")
            path.reverse()

            out.append(DependencyItem(
                step=i,
                schema=r["schema"],
                table_name=r["table_name"],
                entity_id=r["entity_id"],
                entity_name=r.get("entity_name"),
                avg_duration_minutes=avg,
                depth=depth.get(node, 0),
                path=path,
            ))
    return out

@router.get("/api/routes")
def list_routes():
    return [route.path for route in app.routes]


@router.get("/ping")
def ping():
    return {"pong": True}





def find_all_meta_files(top_dirs: list[str]) -> list[dict]:
    all_meta = []

    for entity_root in iter_meta_dirs(top_dirs):
        for root, _, files in os.walk(entity_root):
            if "meta_data_file.yaml" not in files:
                continue

            path = Path(root) / "meta_data_file.yaml"
            try:
                meta = yaml.safe_load(path.read_text(encoding="utf-8")) or {}

                all_meta.append({
                    "table_schema": norm(meta.get("table_schema")),
                    "table_name": norm(meta.get("table_name")),
                    "entity_id": meta.get("entity_id"),
                    "entity_name": meta.get("entity_name"),
                    "table_id": meta.get("table_id"),
                    "depends_on": {
                        norm(k): [norm(t) for t in v]
                        for k, v in (meta.get("depends_on") or {}).items()
                    },
                })
            except Exception as e:
                print(f"[META ERROR] {path}: {e}")

    return all_meta



def build_reverse_index(all_meta: list[dict]) -> dict[tuple[str, str], list[dict]]:
    reverse = {}

    for m in all_meta:
        consumer = (m["table_schema"], m["table_name"])

        for src_schema, tables in (m.get("depends_on") or {}).items():
            for src_table in tables:
                key = (src_schema, src_table)

                reverse.setdefault(key, []).append({
                    "schema": consumer[0],
                    "table_name": consumer[1],
                    "entity_id": m.get("entity_id"),
                    "entity_name": m.get("entity_name"),
                    "table_id": m.get("table_id"),
                })

    return reverse




def recursive_reverse_search(
    schema: str,
    table: str,
    reverse_index: dict,
    visited: Optional[set] = None
):
    if visited is None:
        visited = set()

    key = (schema, table)
    if key in visited:
        return []

    visited.add(key)

    result = []
    for dep in reverse_index.get(key, []):
        result.append(dep)
        result.extend(
            recursive_reverse_search(
                dep["schema"],
                dep["table_name"],
                reverse_index,
                visited
            )
        )

    return result


@router.get("/api/dependencies", response_model=List[DependencyItem])
def get_dependencies(table: str = Query(...)):
    try:
        schema, table = table.split(".")
    except ValueError:
        return []
    return resolve_dependencies(schema, table)


@router.get("/api/failures")
def get_failed_tables():
    query = f"""
     SELECT
        table_schema as object_schema,
        object_name AS table_name,
        l1.object_type,
        message AS error_message,
        loading_finish_dttm AS error_time,
        (
            SELECT MAX(loading_finish_dttm)
            FROM {TABLE_LOADING_HISTORY} AS l2
            WHERE l2.object_name = l1.object_name
              AND l2.object_type = l1.object_type
              AND l2.loading_state IN ('SUCCESS', 'LOADED')
        ) AS last_success_time
    FROM {TABLE_LOADING_HISTORY} l1
    inner join {TABLE_TABLES_META} tm on l1.object_id = tm.table_id
    WHERE loading_state = 'FAILED' and l1.object_type='table'
    ORDER BY loading_finish_dttm DESC
    LIMIT 10
    """
    try:
        with engine.connect() as conn:
            rows = conn.execute(text(query)).mappings().all()

            cleaned = []
            for r in rows:
                row = dict(r)
                row["schema"] = row["object_schema"]
                row["error_time"] = row["error_time"].strftime("%Y-%m-%d %H:%M:%S") if row["error_time"] else None
                row["last_success_time"] = (
                    row["last_success_time"].strftime("%Y-%m-%d %H:%M:%S") if row["last_success_time"] else None
                )
                cleaned.append(row)

            return JSONResponse(content=cleaned, media_type="application/json; charset=utf-8")

    except Exception as e:
        print("❌ Ошибка при получении данных об ошибках:", str(e))
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/entities")
def get_entities():
    try:
        cleaned = fetch_entities(
            engine,
            table_loading_history=TABLE_LOADING_HISTORY,
            table_tables_meta=TABLE_TABLES_META,
            table_entities_meta=TABLE_ENTITIES_META,
        )
        return JSONResponse(content=cleaned, media_type="application/json; charset=utf-8")

    except Exception as e:
        print("❌ Ошибка при получении данных об ошибках:", str(e))
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/entities/timeline")
def get_entities_timeline(days: int = Query(7, ge=3, le=30)):
    try:
        query = f"""
            WITH latest_table_day_runs AS (
                SELECT
                    l.object_id,
                    DATE(COALESCE(l.loading_finish_dttm, l.loading_start_dttm)) AS load_day,
                    l.loading_start_dttm,
                    COALESCE(l.loading_finish_dttm, l.loading_start_dttm) AS loading_finish_dttm,
                    ROW_NUMBER() OVER (
                        PARTITION BY l.object_id, DATE(COALESCE(l.loading_finish_dttm, l.loading_start_dttm))
                        ORDER BY COALESCE(l.loading_finish_dttm, l.loading_start_dttm) DESC NULLS LAST
                    ) AS rn
                FROM {TABLE_LOADING_HISTORY} l
                WHERE l.object_type = 'table'
                  AND l.loading_state IN ('SUCCESS', 'LOADED')
                  AND l.loading_start_dttm IS NOT NULL
            ),
            base AS (
                SELECT
                    t.entity_id,
                    e.entity_name,
                    l.load_day,
                    MIN(l.loading_start_dttm) AS start_dttm,
                    MAX(l.loading_finish_dttm) AS end_dttm
                FROM latest_table_day_runs l
                JOIN {TABLE_TABLES_META} t
                  ON t.table_id = l.object_id
                JOIN {TABLE_ENTITIES_META} e
                  ON e.entity_id = t.entity_id
                WHERE l.rn = 1
                  AND (e.flag_active OR COALESCE(e.on_new_framework, FALSE))
                GROUP BY t.entity_id, e.entity_name, l.load_day
            ),
            ranked AS (
                SELECT
                    base.*,
                    ROW_NUMBER() OVER (
                        PARTITION BY base.entity_id
                        ORDER BY base.load_day DESC
                    ) AS rn
                FROM base
            )
            SELECT
                entity_id,
                entity_name,
                load_day,
                start_dttm,
                end_dttm,
                EXTRACT(EPOCH FROM (end_dttm - start_dttm)) / 60.0 AS duration_minutes
            FROM ranked
            WHERE rn <= :days
            ORDER BY entity_name, load_day
        """

        with engine.connect() as conn:
            rows = conn.execute(text(query), {"days": days}).mappings().all()

        payload = {}
        for row in rows:
            entity_id = row.get("entity_id")
            if not entity_id:
                continue
            payload.setdefault(str(entity_id), []).append(
                {
                    "day": str(row.get("load_day")),
                    "start_dttm": serialize_datetime(row.get("start_dttm")),
                    "end_dttm": serialize_datetime(row.get("end_dttm")),
                    "duration_minutes": round(float(row.get("duration_minutes") or 0.0), 2),
                }
            )

        return JSONResponse(content={"days": days, "items": payload}, media_type="application/json; charset=utf-8")
    except Exception as e:
        print("❌ Ошибка при получении таймлайна сущностей:", str(e))
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/timeline")
def get_table_timeline(table_name: str):
    query = f"""
    SELECT
        loading_start_dttm,
        loading_finish_dttm,
        loading_state,
        message,
        EXTRACT(EPOCH FROM (loading_finish_dttm - loading_start_dttm)) AS duration_seconds
    FROM {TABLE_LOADING_HISTORY}
    WHERE object_name = :table_name and object_type='table'
    ORDER BY loading_finish_dttm DESC
    LIMIT 5
    """
    with engine.connect() as conn:
        result = conn.execute(text(query), {"table_name": table_name}).fetchall()
        columns = result[0].keys() if result else []
        return [dict(zip(columns, row)) for row in result]


@router.get("/api/metrics")
def get_metrics():
    try:
        with engine.connect() as conn:
            total_tables = conn.execute(text(f"SELECT COUNT(*) FROM {TABLE_TABLES_META} WHERE flag_active = true")).scalar()

            error_count = conn.execute(
                text(
                    f"""
                SELECT COUNT(*)
                FROM {TABLE_LOADING_HISTORY}
                WHERE loading_state = 'FAILED' and object_type='table'
                  AND loading_start_dttm >= now() - interval '24 hours'
            """
                )
            ).scalar()

            avg_duration = conn.execute(
                text(
                    f"""
                SELECT ROUND(cast(AVG(EXTRACT(EPOCH FROM (loading_finish_dttm - loading_start_dttm)) / 60) as numeric), 1)
                FROM {TABLE_LOADING_HISTORY}
                WHERE loading_state IN ('SUCCESS', 'LOADED') and object_type='table'
                  AND loading_start_dttm >= now() - interval '24 hours'
            """
                )
            ).scalar()

            active_entities = conn.execute(text(f"SELECT COUNT(*) FROM {TABLE_ENTITIES_META} WHERE flag_active = true")).scalar()

            return JSONResponse(
                content={
                    "total_tables": total_tables,
                    "error_count": error_count,
                    "avg_duration_minutes": float(avg_duration) if avg_duration is not None else None,
                    "active_entities": active_entities,
                },
                media_type="application/json; charset=utf-8",
            )
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/table-history/{schema}/{table:path}")
def get_table_history(
    schema: str,
    table: str,
    table_id: Optional[int] = None,
    limit: int = Query(10, ge=1, le=50),
):
    schema_norm, table_norm = _normalize_table_param(schema, table)
    table_clean = _clean_table_name(table_norm)

    try:
        with engine.connect() as conn:
            params = {
                "limit": limit,
                "schema": schema_norm,
                "table": table_norm,
                "table_clean": table_clean,
                "table_fqn": f"{schema_norm}.{table_norm}",
                "table_fqn_clean": f"{schema_norm}.{table_clean}" if table_clean else None,
                "table_name": table_norm,
                "table_name_clean": table_clean,
            }
            if table_id:
                where_clause = f"""
                    object_id = :table_id
                    OR object_id IN (
                        SELECT tm.table_id
                        FROM {TABLE_TABLES_META} tm
                        WHERE lower(tm.table_schema) = :schema
                          AND (lower(tm.table_name) = :table OR lower(tm.table_name) = :table_clean)
                    )
                    OR lower(object_name) = :table_fqn
                    OR lower(object_name) = :table_fqn_clean
                    OR lower(object_name) = :table_name
                    OR lower(object_name) = :table_name_clean
                """
                params["table_id"] = table_id
            else:
                where_clause = f"""
                    object_id IN (
                        SELECT tm.table_id
                        FROM {TABLE_TABLES_META} tm
                        WHERE lower(tm.table_schema) = :schema
                          AND (lower(tm.table_name) = :table OR lower(tm.table_name) = :table_clean)
                    )
                    OR lower(object_name) = :table_fqn
                    OR lower(object_name) = :table_fqn_clean
                    OR lower(object_name) = :table_name
                    OR lower(object_name) = :table_name_clean
                """

            rows = conn.execute(
                text(
                    f"""
                    SELECT
                        loading_start_dttm,
                        loading_finish_dttm,
                        loading_state,
                        message,
                        EXTRACT(EPOCH FROM (loading_finish_dttm - loading_start_dttm)) / 60.0 AS duration_minutes
                    FROM {TABLE_LOADING_HISTORY}
                    WHERE object_type = 'table'
                      AND {where_clause}
                    ORDER BY loading_finish_dttm DESC NULLS LAST
                    LIMIT :limit
                    """
                ),
                params,
            ).mappings().all()

        payload = [
            {
                "start": serialize_datetime(row.get("loading_start_dttm")),
                "finish": serialize_datetime(row.get("loading_finish_dttm")),
                "state": row.get("loading_state"),
                "message": row.get("message"),
                "duration_minutes": round(float(row["duration_minutes"]), 2)
                if row.get("duration_minutes") is not None
                else None,
            }
            for row in rows
        ]
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/table-variants/{schema}/{table:path}")
def get_table_variants(schema: str, table: str):
    schema_norm, table_norm = _normalize_table_param(schema, table)
    table_clean = _clean_table_name(table_norm)
    query = f"""
        SELECT
            t.table_id,
            t.table_schema,
            t.table_name,
            t.table_last_load,
            t.entity_id,
            e.entity_name
        FROM {TABLE_TABLES_META} t
        LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
        WHERE lower(t.table_schema) = :schema
          AND (lower(t.table_name) = :table OR lower(t.table_name) = :table_clean)
        ORDER BY t.table_id
    """
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(query),
                {"schema": schema_norm, "table": table_norm, "table_clean": table_clean},
            ).mappings().all()

        payload = []
        for row in rows:
            payload.append(
                {
                    "table_id": row.get("table_id"),
                    "table_schema": row.get("table_schema"),
                    "table_name": row.get("table_name"),
                    "entity_id": row.get("entity_id"),
                    "entity_name": row.get("entity_name"),
                    "table_last_load": serialize_datetime(row.get("table_last_load")),
                }
            )
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/dq/table/{schema}/{table:path}")
def get_table_quality(schema: str, table: str):
    schema_norm, table_norm = _normalize_table_param(schema, table)
    table_clean = _clean_table_name(table_norm)
    try:
        with engine.connect() as conn:
            dup_row = conn.execute(
                text(
                    f"""
                    SELECT metric_result, dt_of_verification
                    FROM {TABLE_DATA_QUALITY}
                    WHERE verification_type = 'duplicate_check'
                      AND lower(table_schema) = :schema
                      AND (lower(table_name) = :table OR lower(table_name) = :table_clean)
                    ORDER BY dt_of_verification DESC NULLS LAST
                    LIMIT 1
                    """
                ),
                {"schema": schema_norm, "table": table_norm, "table_clean": table_clean},
            ).mappings().first()

            rc_rows = conn.execute(
                text(
                    f"""
                    SELECT metric_result, dt_of_verification
                    FROM {TABLE_DATA_QUALITY}
                    WHERE verification_type = 'row_count'
                      AND lower(table_schema) = :schema
                      AND (lower(table_name) = :table OR lower(table_name) = :table_clean)
                    ORDER BY dt_of_verification DESC NULLS LAST
                    LIMIT 8
                    """
                ),
                {"schema": schema_norm, "table": table_norm, "table_clean": table_clean},
            ).mappings().all()

        duplicate_count = _parse_numeric(dup_row.get("metric_result")) if dup_row else None
        duplicate_last = serialize_datetime(dup_row.get("dt_of_verification")) if dup_row else None

        row_counts = []
        for row in rc_rows:
            value = _parse_numeric(row.get("metric_result"))
            if value is None:
                continue
            row_counts.append({
                "value": value,
                "dt": row.get("dt_of_verification"),
            })

        row_counts_sorted = sorted(row_counts, key=lambda r: r["dt"] or datetime.min, reverse=True)
        latest_row = row_counts_sorted[0] if row_counts_sorted else None
        baseline_values = [r["value"] for r in row_counts_sorted[1:8] if r.get("value") is not None]

        baseline = None
        if baseline_values:
            baseline = float(sorted(baseline_values)[len(baseline_values) // 2])

        delta_pct = None
        if baseline and latest_row:
            delta_pct = round(((latest_row["value"] - baseline) / baseline) * 100, 2)

        payload = {
            "duplicate": {
                "count": int(duplicate_count) if duplicate_count is not None else None,
                "last_check": duplicate_last,
            },
            "row_count": {
                "count": int(latest_row["value"]) if latest_row else None,
                "last_check": serialize_datetime(latest_row["dt"]) if latest_row else None,
                "baseline_median": baseline,
                "delta_pct": delta_pct,
                "samples": len(baseline_values),
            },
        }
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/dq/history/{schema}/{table:path}")
def get_table_quality_history(schema: str, table: str, limit: int = Query(20, ge=1, le=200)):
    schema_norm, table_norm = _normalize_table_param(schema, table)
    table_clean = _clean_table_name(table_norm)
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT verification_type, metric_result, dt_of_verification
                    FROM {TABLE_DATA_QUALITY}
                    WHERE lower(table_schema) = :schema
                      AND (lower(table_name) = :table OR lower(table_name) = :table_clean)
                    ORDER BY dt_of_verification DESC NULLS LAST
                    LIMIT :limit
                    """
                ),
                {"schema": schema_norm, "table": table_norm, "table_clean": table_clean, "limit": limit},
            ).mappings().all()

        payload = [
            {
                "verification_type": row.get("verification_type"),
                "metric_result": row.get("metric_result"),
                "value": _parse_numeric(row.get("metric_result")),
                "dt": serialize_datetime(row.get("dt_of_verification")),
            }
            for row in rows
        ]
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


def _collect_dq_alerts(days: int, delta: float) -> list[dict]:
    now = datetime.utcnow()
    entity_map = _entity_map_from_meta()
    with engine.connect() as conn:
        dup_rows = conn.execute(
            text(
                f"""
                WITH ranked AS (
                  SELECT
                    lower(table_schema) AS schema,
                    lower(table_name) AS table_name,
                    metric_result,
                    dt_of_verification,
                    ROW_NUMBER() OVER (
                      PARTITION BY lower(table_schema), lower(table_name)
                      ORDER BY dt_of_verification DESC
                    ) AS rn
                  FROM {TABLE_DATA_QUALITY}
                  WHERE verification_type = 'duplicate_check'
                    AND dt_of_verification >= now() - interval '{days} days'
                )
                SELECT schema, table_name, metric_result, dt_of_verification
                FROM ranked
                WHERE rn = 1
                """
            )
        ).mappings().all()

        rc_rows = conn.execute(
            text(
                f"""
                WITH ranked AS (
                  SELECT
                    lower(table_schema) AS schema,
                    lower(table_name) AS table_name,
                    metric_result,
                    dt_of_verification,
                    ROW_NUMBER() OVER (
                      PARTITION BY lower(table_schema), lower(table_name)
                      ORDER BY dt_of_verification DESC
                    ) AS rn
                  FROM {TABLE_DATA_QUALITY}
                  WHERE verification_type = 'row_count'
                )
                SELECT schema, table_name, metric_result, dt_of_verification, rn
                FROM ranked
                WHERE rn <= 8
                """
            )
        ).mappings().all()

    alerts = []
    for row in dup_rows:
        count = _parse_numeric(row.get("metric_result")) or 0
        if count <= 0:
            continue
        fqn = f"{row.get('schema')}.{row.get('table_name')}"
        alerts.append({
            "type": "duplicate_check",
            "table_schema": row.get("schema"),
            "table_name": row.get("table_name"),
            "entity_name": entity_map.get(fqn),
            "metric_value": int(count),
            "delta_pct": None,
            "dt": serialize_datetime(row.get("dt_of_verification")),
        })

    rc_grouped = {}
    for row in rc_rows:
        key = f"{row.get('schema')}.{row.get('table_name')}"
        rc_grouped.setdefault(key, []).append(row)

    for key, rows in rc_grouped.items():
        rows_sorted = sorted(rows, key=lambda r: r.get("dt_of_verification") or datetime.min, reverse=True)
        latest = rows_sorted[0]
        if latest.get("dt_of_verification") and (now - latest["dt_of_verification"]).days > days:
            continue
        latest_val = _parse_numeric(latest.get("metric_result"))
        baseline_vals = [
            _parse_numeric(r.get("metric_result"))
            for r in rows_sorted[1:8]
            if _parse_numeric(r.get("metric_result")) is not None
        ]
        if latest_val is None or not baseline_vals:
            continue
        baseline = float(sorted(baseline_vals)[len(baseline_vals) // 2])
        if baseline == 0:
            continue
        delta_pct = ((latest_val - baseline) / baseline) * 100
        if abs(delta_pct) < delta:
            continue
            schema, table_name = key.split(".", 1)
            alerts.append({
                "type": "row_count",
                "table_schema": schema,
                "table_name": table_name,
                "entity_name": entity_map.get(key),
                "metric_value": int(latest_val),
                "delta_pct": round(delta_pct, 2),
                "dt": serialize_datetime(latest.get("dt_of_verification")),
            })

    alerts.sort(
        key=lambda a: (
            0 if a["type"] == "duplicate_check" else 1,
            -(a.get("metric_value") or 0),
            -abs(a.get("delta_pct") or 0),
        )
    )
    return alerts


@router.get("/api/dq/summary")
def get_quality_summary(days: int = Query(7, ge=1, le=90), delta: float = Query(10.0, ge=0)):
    try:
        now = datetime.utcnow()
        with engine.connect() as conn:
            dup_rows = conn.execute(
                text(
                    f"""
                    WITH ranked AS (
                      SELECT
                        lower(table_schema) AS schema,
                        lower(table_name) AS table_name,
                        metric_result,
                        dt_of_verification,
                        ROW_NUMBER() OVER (
                          PARTITION BY lower(table_schema), lower(table_name)
                          ORDER BY dt_of_verification DESC
                        ) AS rn
                      FROM {TABLE_DATA_QUALITY}
                      WHERE verification_type = 'duplicate_check'
                        AND dt_of_verification >= now() - interval '{days} days'
                    )
                    SELECT schema, table_name, metric_result, dt_of_verification
                    FROM ranked
                    WHERE rn = 1
                    """
                )
            ).mappings().all()

            rc_rows = conn.execute(
                text(
                    f"""
                    WITH ranked AS (
                      SELECT
                        lower(table_schema) AS schema,
                        lower(table_name) AS table_name,
                        metric_result,
                        dt_of_verification,
                        ROW_NUMBER() OVER (
                          PARTITION BY lower(table_schema), lower(table_name)
                          ORDER BY dt_of_verification DESC
                        ) AS rn
                      FROM {TABLE_DATA_QUALITY}
                      WHERE verification_type = 'row_count'
                    )
                    SELECT schema, table_name, metric_result, dt_of_verification, rn
                    FROM ranked
                    WHERE rn <= 8
                    """
                )
            ).mappings().all()

        dup_issues = 0
        for row in dup_rows:
            count = _parse_numeric(row.get("metric_result")) or 0
            if count > 0:
                dup_issues += 1

        rc_grouped = {}
        for row in rc_rows:
            key = f"{row.get('schema')}.{row.get('table_name')}"
            rc_grouped.setdefault(key, []).append(row)

        rc_issues = 0
        checked_rc = 0
        for key, rows in rc_grouped.items():
            rows_sorted = sorted(rows, key=lambda r: r.get("dt_of_verification") or datetime.min, reverse=True)
            latest = rows_sorted[0]
            if latest.get("dt_of_verification") and (now - latest["dt_of_verification"]).days > days:
                continue
            latest_val = _parse_numeric(latest.get("metric_result"))
            baseline_vals = [
                _parse_numeric(r.get("metric_result"))
                for r in rows_sorted[1:8]
                if _parse_numeric(r.get("metric_result")) is not None
            ]
            if latest_val is None or not baseline_vals:
                continue
            checked_rc += 1
            baseline = float(sorted(baseline_vals)[len(baseline_vals) // 2])
            if baseline == 0:
                continue
            delta_pct = ((latest_val - baseline) / baseline) * 100
            if abs(delta_pct) >= delta:
                rc_issues += 1

        payload = {
            "days": days,
            "duplicate_tables": dup_issues,
            "row_count_tables": rc_issues,
            "row_count_checked": checked_rc,
        }
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/dq/alerts")
def get_quality_alerts(
    days: int = Query(7, ge=1, le=90),
    delta: float = Query(10.0, ge=0),
    limit: int = Query(20, ge=1, le=200),
):
    try:
        alerts = _collect_dq_alerts(days, delta)
        return JSONResponse(content=alerts[:limit], media_type="application/json; charset=utf-8")
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/dq/entity")
def get_quality_by_entity(
    days: int = Query(7, ge=1, le=90),
    delta: float = Query(10.0, ge=0),
    limit: int = Query(12, ge=1, le=50),
):
    alerts = _collect_dq_alerts(days, delta)
    grouped = {}
    for alert in alerts:
        entity = alert.get("entity_name") or "UNKNOWN"
        entry = grouped.setdefault(entity, {"entity": entity, "duplicates": 0, "row_count": 0})
        if alert.get("type") == "duplicate_check":
            entry["duplicates"] += 1
        if alert.get("type") == "row_count":
            entry["row_count"] += 1

    rows = sorted(grouped.values(), key=lambda r: (-(r["duplicates"] + r["row_count"]), r["entity"]))
    return JSONResponse(content=rows[:limit], media_type="application/json; charset=utf-8")


def find_path_case_insensitive(parent_path: Path, name: str) -> Optional[Path]:
    for item in parent_path.iterdir():
        if item.name.lower() == name.lower():
            return item
    return None


@router.get("/api/card/{schema}/{table:path}")
def get_table_card_info_by_path(
    schema: str,
    table: str,
    source: str = Query("current"),
    table_id: Optional[int] = Query(None),
):
    source_name = (source or "current").strip().lower()
    if source_name != "current":
        dbt_card = build_dbt_fallback_card(
            base_dir=BASE_DIR,
            manifest_dir=DBT_MANIFEST_DIR,
            schema_name=schema,
            table_name=table,
            source=source_name,
        )
        if dbt_card:
            return JSONResponse(content=dbt_card, media_type="application/json; charset=utf-8")
        return JSONResponse(status_code=404, content={"error": "Table not found in dbt manifest"})

    table_clean = _clean_table_name(norm(table))
    for entity_folder in iter_meta_dirs():
        schema_folder = find_path_case_insensitive(entity_folder, schema)
        if not schema_folder:
            continue

        table_folder = find_path_case_insensitive(schema_folder, table)
        if not table_folder and table_clean:
            table_folder = find_path_case_insensitive(schema_folder, table_clean)
        if not table_folder:
            continue

        yaml_file = table_folder / "meta_data_file.yaml"
        if not yaml_file.exists():
            return JSONResponse(status_code=404, content={"error": "meta_data_file.yaml not found"})

        try:
            with open(yaml_file, encoding="utf-8") as f:
                meta = yaml.safe_load(f)
        except Exception as e:
            return JSONResponse(status_code=500, content={"error": str(e)})

        if table_id is not None and meta.get("table_id") != table_id:
            continue

        meta["sql_query_insert_init_sql"] = _read_sql_from_meta(
            yaml_file,
            meta,
            "sql_query_insert_init",
            "sql_query_insert_init.sql",
        )
        meta["sql_query_recreate_init_sql"] = _read_sql_from_meta(
            yaml_file,
            meta,
            "sql_query_recreate_init",
            "sql_query_recreate_init.sql",
        )
        meta["sql_query_truncate_sql"] = _read_sql_from_meta(
            yaml_file,
            meta,
            "sql_query_truncate",
            "sql_query_truncate.sql",
        )

        # метрики
        # метрики
        table_id = meta.get("table_id")
        avg_duration = None
        last_success_time = None
        table_size_mb = None

        if table_id:
            try:
                with engine.connect() as conn:
                    duration_result = conn.execute(
                        text(
                            f"""
                        SELECT round(cast(AVG(EXTRACT(EPOCH FROM (loading_finish_dttm - loading_start_dttm)) / 60) as numeric), 1)
                        FROM {TABLE_LOADING_HISTORY}
                        WHERE loading_state IN ('SUCCESS', 'LOADED')
                          AND object_type='table'
                          AND object_id = :object_id
                    """
                        ),
                        {"object_id": table_id},
                    )
                    avg_duration = float(duration_result.scalar() or 0)

                    time_result = conn.execute(
                        text(
                            f"""
                        SELECT table_last_load
                        FROM {TABLE_TABLES_META}
                        WHERE table_id = :object_id
                    """
                        ),
                        {"object_id": table_id},
                    )
                    dt_val = time_result.scalar()
                    if isinstance(dt_val, datetime):
                        last_success_time = dt_val.strftime("%Y-%m-%d %H:%M:%S")

                    # ✅ ВОТ ТУТ — НОВЫЙ БЕЗОПАСНЫЙ КОД
                    size_sql = text("""
                        SELECT
                          pg_total_relation_size(
                            to_regclass(:full_table_name)
                          )::bigint / 1024 / 1024
                    """)

                    schema_name = meta.get("table_schema") or schema
                    table_name = meta.get("table_name") or table
                    schema_name = str(schema_name or "")
                    table_name = str(table_name or "")

                    def quote_ident(value: str) -> str:
                        escaped = value.replace('"', '""')
                        return f"\"{escaped}\""

                    def build_regclass(schema_val: str, table_val: str) -> str:
                        if not schema_val or not table_val:
                            return ""
                        needs_quote = (
                            schema_val.lower() in {"stg", "dict_stg"}
                            or schema_val != schema_val.lower()
                            or table_val != table_val.lower()
                        )
                        if needs_quote:
                            return f"{quote_ident(schema_val)}.{quote_ident(table_val)}"
                        return f"{schema_val.lower()}.{table_val.lower()}"

                    regclass_name = build_regclass(schema_name, table_name)

                    size_result = conn.execute(
                        size_sql,
                        {"full_table_name": regclass_name},
                    ).scalar()

                    table_size_mb = int(size_result) if size_result is not None else None

            except Exception as e:
                print(f"Ошибка при получении метрик: {e}")

        meta["avg_duration_minutes"] = avg_duration
        meta["last_success_time"] = last_success_time
        meta["table_size_mb"] = table_size_mb

        return JSONResponse(content=meta, media_type="application/json; charset=utf-8")

    dbt_card = build_dbt_fallback_card(
        base_dir=BASE_DIR,
        manifest_dir=DBT_MANIFEST_DIR,
        schema_name=schema,
        table_name=table,
        source="ohd",
    )
    if dbt_card:
        return JSONResponse(content=dbt_card, media_type="application/json; charset=utf-8")

    print(f"[WARN] Table {schema}.{table} not found in any of TOP_DIRS")
    return JSONResponse(status_code=404, content={"error": "Table not found in any folder"})


@router.get("/api/dbt/model/{schema}/{table:path}")
def get_dbt_model_info(schema: str, table: str, source: str = "ohd"):
    try:
        model = get_dbt_manifest_model(
            base_dir=BASE_DIR,
            manifest_dir=DBT_MANIFEST_DIR,
            schema_name=schema,
            table_name=table,
            source=source,
        )
        if not model:
            return JSONResponse(status_code=404, content={"error": "dbt model not found"})
        return JSONResponse(content=model, media_type="application/json; charset=utf-8")
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/dbt/history/{schema}/{table:path}")
def get_dbt_model_history(
    schema: str,
    table: str,
    source: str = "ohd",
    limit: int = Query(12, ge=1, le=50),
):
    if not dbt_logs_engine:
        return JSONResponse(
            content={"configured": False, "model": None, "catalog": None, "runs": []},
            media_type="application/json; charset=utf-8",
        )
    try:
        payload = get_dbt_model_run_history(
            engine=dbt_logs_engine,
            base_dir=BASE_DIR,
            manifest_dir=DBT_MANIFEST_DIR,
            schema_name=schema,
            table_name=table,
            source=source,
            limit=limit,
            table_model_catalog=TABLE_DBT_MODEL_CATALOG,
            table_model_log=TABLE_DBT_MODEL_LOG,
            table_run_log=TABLE_DBT_RUN_LOG,
        )
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except OperationalError as e:
        print("❌ /api/dbt/history db unavailable:", e)
        return JSONResponse(
            content={
                "configured": True,
                "available": False,
                "message": "DBT logs database is unavailable",
                "model": None,
                "catalog": None,
                "runs": [],
            },
            media_type="application/json; charset=utf-8",
        )
    except Exception as e:
        print("❌ /api/dbt/history error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/tables")
def list_all_tables(detailed: bool = Query(False)):
    all_meta, _ = get_cached_meta_and_index()
    all_tables = {}
    for meta in all_meta:
        schema = meta.get("table_schema")
        table = meta.get("table_name")
        if schema and table:
            key = f"{schema}.{table}"
            all_tables[f"current:{key.lower()}"] = {
                "fqn": key,
                "schema": schema,
                "table": table,
                "label": key,
                "source": "current",
                "entity_name": meta.get("entity_name"),
            }
    for item in get_dbt_table_catalog(BASE_DIR, DBT_MANIFEST_DIR, source="ohd"):
        all_tables[f"{item['source']}:{item['fqn'].lower()}"] = item
    rows = sorted(
        all_tables.values(),
        key=lambda v: (v["fqn"].lower(), 0 if v.get("source") == "current" else 1, v.get("source") or ""),
    )
    if detailed:
        return JSONResponse(content=rows, media_type="application/json; charset=utf-8")
    return JSONResponse(content=[row["fqn"] for row in rows], media_type="application/json; charset=utf-8")


@router.get("/api/graph/overview")
def get_graph_overview():
    snapshot = get_graph_snapshot()
    nodes = snapshot["entity_graph"]["nodes"]
    edges = snapshot["entity_graph"]["edges"]
    layout = snapshot["layouts"]["entity"]

    nodes, edges, layout, truncated = _cap_graph(
        nodes, edges, layout, max_nodes=150, max_edges=300, sort_key="tables_count"
    )
    return {
        "nodes": list(nodes.values()),
        "edges": edges,
        "layout": layout,
        "truncated": truncated,
        "entity_cycles": snapshot.get("entity_cycles", []),
        "entity_mutual": snapshot.get("entity_mutual_any", []),
        "table_cycles": snapshot.get("table_cycles", []),
    }


@router.get("/api/graph/diagnostics")
def get_graph_diagnostics(include_any: bool = Query(True)):
    snapshot = get_graph_snapshot()
    return {
        "entity_cycles": snapshot.get("entity_cycles", []),
        "entity_mutual": snapshot.get("entity_mutual_any", []) if include_any else snapshot.get("entity_mutual", []),
        "table_cycles": snapshot.get("table_cycles", []),
    }


@router.get("/api/graph/diagnostics/mutual")
def get_graph_mutual_details(
    entity_a: str = Query(...),
    entity_b: str = Query(...),
    strict: bool = Query(True),
):
    snapshot = get_graph_snapshot()
    table_edges = snapshot["table_graph"]["edges"]
    table_entities = {k: set(v) for k, v in snapshot["table_entity_map"].items()}

    def norm_entity(value: str) -> str:
        return (value or "").strip().lower()

    a = norm_entity(entity_a)
    b = norm_entity(entity_b)
    if not a or not b:
        return JSONResponse(status_code=400, content={"error": "invalid entities"})
    if _normalize_entity_group(a) == _normalize_entity_group(b):
        return {"entity_a": entity_a, "entity_b": entity_b, "edges_ab": [], "edges_ba": []}

    edges_ab = []
    edges_ba = []
    for edge in table_edges:
        src = edge.get("source")
        tgt = edge.get("target")
        if not src or not tgt:
            continue
        src_ents = table_entities.get(src) or set()
        tgt_ents = table_entities.get(tgt) or set()
        for src_ent in src_ents:
            for tgt_ent in tgt_ents:
                if not src_ent or not tgt_ent or src_ent == tgt_ent:
                    continue
                src_norm = {norm_entity(e) for e in src_ents if isinstance(e, str)}
                tgt_norm = {norm_entity(e) for e in tgt_ents if isinstance(e, str)}
                allowed = {a, b}
                if strict and (not src_norm.issubset(allowed) or not tgt_norm.issubset(allowed)):
                    continue
                if (
                    norm_entity(src_ent) == a
                    and norm_entity(tgt_ent) == b
                    and norm_entity(tgt_ent) not in {norm_entity(e) for e in src_ents}
                    and norm_entity(src_ent) not in {norm_entity(e) for e in tgt_ents}
                ):
                    edges_ab.append({
                        "source": src,
                        "target": tgt,
                        "source_entities": sorted(src_ents),
                        "target_entities": sorted(tgt_ents),
                    })
                elif (
                    norm_entity(src_ent) == b
                    and norm_entity(tgt_ent) == a
                    and norm_entity(tgt_ent) not in {norm_entity(e) for e in src_ents}
                    and norm_entity(src_ent) not in {norm_entity(e) for e in tgt_ents}
                ):
                    edges_ba.append({
                        "source": src,
                        "target": tgt,
                        "source_entities": sorted(src_ents),
                        "target_entities": sorted(tgt_ents),
                    })

    return {
        "entity_a": entity_a,
        "entity_b": entity_b,
        "edges_ab": edges_ab,
        "edges_ba": edges_ba,
    }


@router.get("/api/graph/entity/{entity_name}")
def get_graph_entity(entity_name: str):
    snapshot = get_graph_snapshot()
    table_nodes = snapshot["table_graph"]["nodes"]
    table_edges = snapshot["table_graph"]["edges"]
    table_entity_map = snapshot["table_entity_map"]

    target = (entity_name or "").strip().lower()
    all_entities = {e for ents in table_entity_map.values() for e in ents}
    entity_key = next((e for e in all_entities if e.lower() == target), None)

    if not entity_key:
        return JSONResponse(status_code=404, content={"error": "entity not found"})

    entity_tables = {t for t, ents in table_entity_map.items() if entity_key in ents}
    if not entity_tables:
        return JSONResponse(status_code=404, content={"error": "entity has no tables"})

    nodes_set = set(entity_tables)
    edges_filtered = []
    for e in table_edges:
        if e["source"] in entity_tables or e["target"] in entity_tables:
            nodes_set.add(e["source"])
            nodes_set.add(e["target"])
            edges_filtered.append(e)

    truncated = False
    if len(nodes_set) > 300:
        keep = set(entity_tables)
        extra = [n for n in nodes_set if n not in keep]
        extra.sort()
        for n in extra:
            if len(keep) >= 300:
                truncated = True
                break
            keep.add(n)
        nodes_set = keep

    edges_filtered = [e for e in edges_filtered if e["source"] in nodes_set and e["target"] in nodes_set]
    if len(edges_filtered) > 500:
        edges_filtered = edges_filtered[:500]
        truncated = True

    nodes_payload = _normalize_layer_widths([table_nodes[n] for n in nodes_set if n in table_nodes])
    layout_payload = _grid_layout_subset(table_nodes, edges_filtered, nodes_set)

    return {
        "entity": {
            "id": f"ENTITY::{entity_key}",
            "label": entity_key,
            "tables_count": len(entity_tables),
        },
        "nodes": nodes_payload,
        "edges": edges_filtered,
        "layout": layout_payload,
        "truncated": truncated,
    }


@router.get("/api/graph/table/{schema}/{table:path}")
def get_graph_table(
    schema: str,
    table: str,
    depth: Optional[int] = Query(None, ge=1, le=20),
    source: str = Query("current"),
    table_id: Optional[int] = Query(None),
):
    snapshot = _get_table_graph_context(source)
    table_nodes = snapshot["table_graph"]["nodes"]
    table_edges = snapshot["table_graph"]["edges"]
    table_fqn_map = snapshot.get("table_fqn_map") or {}

    key = _resolve_table_key(table_nodes, schema, table, table_id=table_id, fqn_map=table_fqn_map)
    if key not in table_nodes:
        return JSONResponse(status_code=404, content={"error": "table not found"})

    rev = {}
    fwd = {}
    for e in table_edges:
        rev.setdefault(e["target"], []).append(e["source"])
        fwd.setdefault(e["source"], []).append(e["target"])

    def traverse(adjacency: dict[str, list[str]]) -> set[str]:
        visited_local = {key}
        queue = deque([(key, 0)])
        while queue:
            node, d = queue.popleft()
            if depth is not None and d >= depth:
                continue
            for nxt in adjacency.get(node, []):
                if nxt in visited_local:
                    continue
                visited_local.add(nxt)
                queue.append((nxt, d + 1))
        return visited_local

    visited = traverse(rev)
    truncated = False

    edges_filtered = [e for e in table_edges if e["source"] in visited and e["target"] in visited]

    nodes_payload = _normalize_layer_widths([table_nodes[n] for n in visited if n in table_nodes])
    layout_payload = _grid_layout_subset(table_nodes, edges_filtered, visited)

    return {
        "table": table_nodes[key],
        "nodes": nodes_payload,
        "edges": edges_filtered,
        "layout": layout_payload,
        "depth": depth,
        "truncated": truncated,
    }


def _layer_label_from_schema(schema: str) -> str:
    if not schema:
        return "OTHER"
    schema = schema.lower()
    if schema.startswith("dict_"):
        return "DICT"
    if schema == "stg":
        return "STG"
    if schema == "ods":
        return "ODS"
    if schema == "dds":
        return "DDS"
    if schema == "dm_calc":
        return "DM_CALC"
    if schema.startswith("dm"):
        return "DM"
    return schema.upper()


def _traverse_forward(
    start: str,
    edges: list[dict],
    depth: int,
    max_nodes: int,
) -> tuple[set[str], dict[str, int], bool]:
    forward = {}
    for edge in edges:
        forward.setdefault(edge["source"], []).append(edge["target"])

    visited = {start}
    depth_map = {start: 0}
    queue = deque([(start, 0)])
    truncated = False

    while queue:
        node, d = queue.popleft()
        if d >= depth:
            continue
        for nxt in forward.get(node, []):
            if nxt in visited:
                continue
            visited.add(nxt)
            depth_map[nxt] = d + 1
            if len(visited) >= max_nodes:
                truncated = True
                queue.clear()
                break
            queue.append((nxt, d + 1))
        if truncated:
            break

    return visited, depth_map, truncated


@router.get("/api/graph/impact/{schema}/{table:path}")
def get_graph_impact(schema: str, table: str, depth: int = Query(3, ge=1, le=4), source: str = Query("current")):
    snapshot = _get_table_graph_context(source)
    table_nodes = snapshot["table_graph"]["nodes"]
    table_edges = snapshot["table_graph"]["edges"]
    table_fqn_map = snapshot.get("table_fqn_map") or {}

    key = _resolve_table_key(table_nodes, schema, table, fqn_map=table_fqn_map)
    if key not in table_nodes:
        return JSONResponse(status_code=404, content={"error": "table not found"})

    visited, _, truncated = _traverse_forward(key, table_edges, depth, max_nodes=300)

    edges_filtered = [e for e in table_edges if e["source"] in visited and e["target"] in visited]
    if len(edges_filtered) > 500:
        edges_filtered = edges_filtered[:500]
        truncated = True

    nodes_payload = _normalize_layer_widths([table_nodes[n] for n in visited if n in table_nodes])
    layout_payload = _grid_layout_subset(table_nodes, edges_filtered, visited)

    return {
        "table": table_nodes[key],
        "nodes": nodes_payload,
        "edges": edges_filtered,
        "layout": layout_payload,
        "depth": depth,
        "truncated": truncated,
    }


@router.get("/api/impact/summary/{schema}/{table:path}")
def get_impact_summary(
    schema: str,
    table: str,
    depth: int = Query(3, ge=1, le=5),
    max_nodes: int = Query(800, ge=50, le=5000),
    limit: int = Query(120, ge=0, le=2000),
    source: str = Query("current"),
):
    snapshot = _get_table_graph_context(source)
    table_nodes = snapshot["table_graph"]["nodes"]
    table_edges = snapshot["table_graph"]["edges"]
    table_fqn_map = snapshot.get("table_fqn_map") or {}

    key = _resolve_table_key(table_nodes, schema, table, fqn_map=table_fqn_map)
    if key not in table_nodes:
        return JSONResponse(status_code=404, content={"error": "table not found"})

    visited, depth_map, truncated = _traverse_forward(key, table_edges, depth, max_nodes=max_nodes)
    visited.discard(key)

    entity_counts = {}
    layer_counts = {}
    table_rows = []
    for node_id in visited:
        node = table_nodes.get(node_id)
        if not node:
            continue
        schema_val = node.get("schema") or ""
        layer = _layer_label_from_schema(schema_val)
        layer_counts[layer] = layer_counts.get(layer, 0) + 1

        entities = node.get("entities") or []
        if not entities:
            entities = [node.get("entity") or "UNKNOWN"]
        for ent in entities:
            if not ent:
                continue
            entity_counts[ent] = entity_counts.get(ent, 0) + 1

        table_rows.append(
            {
                "id": node_id,
                "schema": node.get("schema"),
                "table": node.get("table"),
                "entity": node.get("entity"),
                "entities": entities,
                "layer": layer,
                "depth": depth_map.get(node_id, 0),
            }
        )

    entities_list = [
        {"entity": ent, "count": count}
        for ent, count in sorted(entity_counts.items(), key=lambda x: (-x[1], x[0]))
    ]
    layers_list = [
        {"layer": layer, "count": count}
        for layer, count in sorted(layer_counts.items(), key=lambda x: (-x[1], x[0]))
    ]

    table_rows.sort(key=lambda r: (r.get("depth") or 0, r.get("id") or ""))
    if limit:
        table_rows = table_rows[:limit]

    return {
        "table": table_nodes[key],
        "total_tables": len(visited),
        "total_entities": len(entity_counts),
        "entities": entities_list,
        "layers": layers_list,
        "tables": table_rows,
        "depth": depth,
        "truncated": truncated,
    }


@router.get("/api/impact/list/{schema}/{table:path}")
def get_impact_list(
    schema: str,
    table: str,
    depth: int = Query(4, ge=1, le=6),
    max_nodes: int = Query(2500, ge=100, le=10000),
    source: str = Query("current"),
):
    snapshot = _get_table_graph_context(source)
    table_nodes = snapshot["table_graph"]["nodes"]
    table_edges = snapshot["table_graph"]["edges"]
    table_fqn_map = snapshot.get("table_fqn_map") or {}

    key = _resolve_table_key(table_nodes, schema, table, fqn_map=table_fqn_map)
    if key not in table_nodes:
        return JSONResponse(status_code=404, content={"error": "table not found"})

    visited, depth_map, truncated = _traverse_forward(key, table_edges, depth, max_nodes=max_nodes)
    visited.discard(key)

    table_rows = []
    for node_id in visited:
        node = table_nodes.get(node_id)
        if not node:
            continue
        schema_val = node.get("schema") or ""
        layer = _layer_label_from_schema(schema_val)
        entities = node.get("entities") or []
        if not entities:
            entities = [node.get("entity") or "UNKNOWN"]
        table_rows.append(
            {
                "id": node_id,
                "schema": node.get("schema"),
                "table": node.get("table"),
                "entities": entities,
                "layer": layer,
                "depth": depth_map.get(node_id, 0),
            }
        )

    table_rows.sort(key=lambda r: (r.get("depth") or 0, r.get("id") or ""))
    return {
        "table": table_nodes[key],
        "tables": table_rows,
        "depth": depth,
        "truncated": truncated,
    }


@router.get("/api/inconsistencies")
def get_dependency_violations():
    all_meta, _ = get_cached_meta_and_index()
    dependency_pairs = []

    for meta in all_meta:
        dependent_schema = meta.get("table_schema")
        dependent_table = meta.get("table_name")
        depends_on = meta.get("depends_on", {})
        for source_schema, source_tables in depends_on.items():
            for source_table in source_tables:
                dependency_pairs.append(((source_schema, source_table), (dependent_schema, dependent_table)))

    all_tables = set()
    for src, dep in dependency_pairs:
        all_tables.add(src)
        all_tables.add(dep)

    # Load last N successful runs per table so we can compare the latest source and target refreshes.
    last_loads = {}
    tables_by_schema = {}
    for schema, table in all_tables:
        tables_by_schema.setdefault(schema, set()).add(table)

    batch_size = 250
    query = text(
        f"""
        WITH base AS (
            SELECT
                COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                l.loading_finish_dttm
            FROM {TABLE_LOADING_HISTORY} l
            LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
            WHERE l.object_type = 'table'
              AND l.loading_state IN ('SUCCESS', 'LOADED')
              AND l.loading_finish_dttm IS NOT NULL
              AND t.table_schema = :schema
              AND t.table_name = ANY(:tables)
        )
        SELECT table_schema, table_name, loading_finish_dttm
        FROM (
            SELECT
                table_schema,
                table_name,
                loading_finish_dttm,
                ROW_NUMBER() OVER (
                    PARTITION BY table_schema, table_name
                    ORDER BY loading_finish_dttm DESC
                ) AS rn
            FROM base
        ) x
        WHERE rn <= :limit
        """
    )

    for schema, tables in tables_by_schema.items():
        if not schema or not tables:
            continue

        ordered_tables = sorted(tables)
        for start in range(0, len(ordered_tables), batch_size):
            batch_tables = ordered_tables[start:start + batch_size]
            try:
                with engine.connect() as conn:
                    result = conn.execute(
                        query,
                        {"schema": schema, "tables": batch_tables, "limit": 6},
                    ).mappings().all()
            except OperationalError:
                # Refresh the pooled connection and retry once with a new DB session.
                with engine.connect() as conn:
                    result = conn.execute(
                        query,
                        {"schema": schema, "tables": batch_tables, "limit": 6},
                    ).mappings().all()

            for row in result:
                key = (row.get("table_schema"), row.get("table_name"))
                if not key[0] or not key[1]:
                    continue
                last_loads.setdefault(key, []).append(row.get("loading_finish_dttm"))

    # Sort times DESC for each table
    for key in list(last_loads.keys()):
        times = [t for t in last_loads.get(key, []) if t]
        times.sort(reverse=True)
        last_loads[key] = times

    problems = []
    for (src_schema, src_table), (dep_schema, dep_table) in dependency_pairs:
        dep_times = last_loads.get((dep_schema, dep_table)) or []
        src_times = last_loads.get((src_schema, src_table)) or []
        if not dep_times or not src_times:
            continue
        dep_time = dep_times[0]
        src_time = src_times[0]
        # Violation when the latest upstream refresh finished after the latest dependent refresh.
        if not src_time or src_time <= dep_time:
            continue
        problems.append(
            {
                "source_schema": src_schema,
                "source_table": src_table,
                "source_last_load": src_time.strftime("%Y-%m-%d %H:%M:%S"),
                "dependent_schema": dep_schema,
                "dependent_table": dep_table,
                "dependent_last_load": dep_time.strftime("%Y-%m-%d %H:%M:%S"),
            }
        )

    return JSONResponse(content=problems, media_type="application/json; charset=utf-8")


@router.get("/api/sla")
def get_sla_monitoring():
    query = f"""
        WITH exploded AS (
            SELECT 
                s.report,
                s.source_table,
                s.owner_report,
                s.load_update_table,
                s.load_update_report,
                s.load_interval,
                s.table_name,
                regexp_split_to_table(s.table_name, E'\\n') AS split_table
            FROM {TABLE_YT_SLA} s
        ),
        cleaned AS (
            SELECT *,
                   trim(split_table) AS clean_table,
                   split_part(trim(split_table), '.', 1) AS schema_name,
                   split_part(trim(split_table), '.', 2) AS table_name_only
            FROM exploded
        ),
        click as (
        select 
        	(REGEXP_MATCHES(ddl_clickhouse_view, 'DROP VIEW IF EXISTS\\s+"([^"]+)"\\."([^"]+)"'))[1] as schema_name_view,
        	(REGEXP_MATCHES(ddl_clickhouse_view, 'DROP VIEW IF EXISTS\\s+"([^"]+)"\\."([^"]+)"'))[2] as table_name_view,
        	(REGEXP_MATCHES(ddl_clickhouse_target, 'DROP TABLE IF EXISTS\\s+"([^"]+)"\\."([^"]+)"'))[1] as schema_name_table,
        	(REGEXP_MATCHES(ddl_clickhouse_target, 'DROP TABLE IF EXISTS\\s+"([^"]+)"\\."([^"]+)"'))[2] as table_name_table,
        	table_last_upload
        from {TABLE_TABLES_META_CLICK} ),
        joined AS (
            SELECT 
                c.report,
                c.source_table,
                c.owner_report,
                c.load_update_table,
                c.load_update_report,
                c.load_interval,
                c.table_name AS original_table_name,
                c.clean_table,
                coalesce(tm.table_last_load, cl.table_last_upload, clt.table_last_upload)  as table_last_load
            FROM cleaned c
            LEFT JOIN {TABLE_TABLES_META} tm
              ON tm.table_schema = c.schema_name
             AND tm.table_name = c.table_name_only and source_table='GP'
            left join click cl 
             ON cl.schema_name_view = c.schema_name
             AND cl.table_name_view = c.table_name_only and source_table='Click'
             left join click clt 
             ON clt.schema_name_table = c.schema_name
             AND clt.table_name_table = c.table_name_only and source_table='Click'
        ),
        with_flags AS (
            SELECT *,
                   (table_last_load IS NOT NULL AND
                    (
                        (position('сут' in lower(load_interval)) > 0 AND now() - table_last_load <= interval '24 hours')
                     OR (position('час' in lower(load_interval)) > 0 AND now() - table_last_load <= interval '1 hour')
                    )
                   ) AS sla_ok
            FROM joined
        )
        SELECT 
            report,
            source_table,
            owner_report,
            load_update_table,
            load_update_report,
            load_interval,
            original_table_name,
            json_agg(json_build_object(
                'table_name', clean_table,
                'table_last_load', CASE 
                    WHEN table_last_load IS NOT NULL 
                    THEN to_char(table_last_load, 'YYYY-MM-DD HH24:MI:SS') 
                    ELSE NULL 
                END,
                'sla_ok', sla_ok
            )) AS tables_info
        FROM with_flags
        GROUP BY report, source_table, owner_report, load_update_table, load_update_report, load_interval, original_table_name
        ORDER BY report;
    """
    try:
        with engine.connect() as conn:
            result = conn.execute(text(query))
            rows = [dict(row._mapping) for row in result]

        for row in rows:
            for table in row["tables_info"]:
                try:
                    if not table.get("table_last_load"):
                        table["table_last_load"] = "Нет данных"
                    table["sla_ok"] = bool(table["sla_ok"])
                except Exception as inner_error:
                    print("Ошибка при обработке таблицы:", table)
                    print("Ошибка:", inner_error)
                    table["sla_ok"] = False
                    table["table_last_load"] = "Нет данных"
            row["sla_ok"] = all(t["sla_ok"] for t in row["tables_info"])

        return JSONResponse(content=rows)

    except Exception as e:
        print("🔥 Общая ошибка SLA-эндпоинта 🔥")
        print(e)
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/slowest-tables")
def get_slowest_tables(
    days: int = Query(30, ge=1, le=120),
    limit: int = Query(20, ge=1, le=200),
):
    try:
        query = f"""
            WITH base AS (
                SELECT
                    COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                    COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                    e.entity_name,
                    EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration
                FROM {TABLE_LOADING_HISTORY} l
                LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
                LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                WHERE l.object_type = 'table'
                  AND l.loading_state IN ('SUCCESS', 'LOADED')
                  AND l.loading_start_dttm IS NOT NULL
                  AND l.loading_finish_dttm IS NOT NULL
                  AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
            )
            SELECT
                table_schema,
                table_name,
                entity_name,
                COUNT(*) AS runs_count,
                AVG(duration) AS avg_duration,
                percentile_cont(0.95) WITHIN GROUP (ORDER BY duration) AS p95_duration,
                MAX(duration) AS max_duration,
                STDDEV_SAMP(duration) AS stddev_duration
            FROM base
            GROUP BY table_schema, table_name, entity_name
        """
        with engine.connect() as conn:
            rows = conn.execute(text(query), {"days": days}).mappings().all()

        threshold_minutes = 10.0
        data = []
        for row in rows:
            r = dict(row)
            runs = int(r.get("runs_count") or 0)
            avg = float(r["avg_duration"]) if r.get("avg_duration") is not None else None
            p95 = float(r["p95_duration"]) if r.get("p95_duration") is not None else None
            max_d = float(r["max_duration"]) if r.get("max_duration") is not None else None
            std = float(r["stddev_duration"]) if r.get("stddev_duration") is not None else None

            cv = (std / avg) if (std is not None and avg) else None
            p95_ratio = (p95 / avg) if (p95 is not None and avg) else None

            slow = bool(p95 is not None and p95 > threshold_minutes)
            unstable = bool(cv is not None and cv >= 0.3)
            critical_unstable = bool(cv is not None and cv >= 0.6) or bool(
                p95_ratio is not None and p95_ratio > 2
            )
            low_sample = runs < 5

            if low_sample:
                status = "low_sample"
            elif slow and unstable:
                status = "slow_unstable"
            elif slow and not unstable:
                status = "slow"
            elif unstable:
                status = "unstable"
            else:
                status = "ok"

            if not (slow or unstable):
                continue

            data.append(
                {
                    "table_schema": r.get("table_schema") or "",
                    "table_name": r.get("table_name") or "",
                    "entity_name": r.get("entity_name"),
                    "runs_count": runs,
                    "avg_duration": round(avg, 2) if avg is not None else None,
                    "p95_duration": round(p95, 2) if p95 is not None else None,
                    "max_duration": round(max_d, 2) if max_d is not None else None,
                    "stddev_duration": round(std, 2) if std is not None else None,
                    "cv": round(cv, 4) if cv is not None else None,
                    "p95_avg_ratio": round(p95_ratio, 3) if p95_ratio is not None else None,
                    "slow": slow,
                    "unstable": unstable,
                    "critical_unstable": critical_unstable,
                    "low_sample": low_sample,
                    "status": status,
                }
            )

        status_order = {
            "slow_unstable": 0,
            "slow": 1,
            "unstable": 2,
            "low_sample": 3,
            "ok": 4,
        }
        data.sort(
            key=lambda x: (
                status_order.get(x["status"], 9),
                -(x.get("p95_duration") or 0),
            )
        )
        limited = data[:limit]

        now_dt = datetime.utcnow().date()
        period_from = now_dt - timedelta(days=days)
        payload = {
            "meta": {
                "window_days": days,
                "period_from": period_from.strftime("%Y-%m-%d"),
                "period_to": now_dt.strftime("%Y-%m-%d"),
                "total_tables": len(rows),
                "candidates": len(data),
                "limit": limit,
            },
            "rows": limited,
        }
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        print("❌ /api/slowest-tables error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/table-sizes")
def get_table_sizes(
    limit: int = Query(30, ge=1, le=200),
    schema: str = Query(""),
    owner: str = Query(""),
):
    try:
        schema_value = str(schema or "").strip()
        owner_value = str(owner or "").strip()
        cached = get_cached_table_sizes()
        all_rows = cached.get("rows") or []
        filtered_rows = all_rows
        if schema_value:
            filtered_rows = [row for row in filtered_rows if row.get("table_schema") == schema_value]
        if owner_value:
            filtered_rows = [row for row in filtered_rows if row.get("owner_name") == owner_value]
        rows = filtered_rows[:limit]
        total_size_bytes = sum(int(row.get("size_bytes") or 0) for row in rows)
        payload = {
            "meta": {
                "limit": limit,
                "schema": schema_value or None,
                "owner": owner_value or None,
                "generated_at": cached.get("generated_at"),
                "returned_rows": len(rows),
                "available_rows": len(filtered_rows),
                "total_size_bytes": total_size_bytes,
            },
            "schemas": cached.get("schemas") or [],
            "owners": sorted({row.get("owner_name") for row in all_rows if row.get("owner_name")}),
            "rows": rows,
        }
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        print("❌ /api/table-sizes error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/load-profile")
def get_load_profile(days: int = Query(30, ge=1, le=120)):
    try:
        query = f"""
            SELECT
                EXTRACT(HOUR FROM l.loading_start_dttm) AS hour,
                COUNT(*) AS runs_count,
                SUM(EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0) AS total_duration_minutes
            FROM {TABLE_LOADING_HISTORY} l
            WHERE l.object_type = 'table'
              AND l.loading_state IN ('SUCCESS', 'LOADED')
              AND l.loading_start_dttm IS NOT NULL
              AND l.loading_finish_dttm IS NOT NULL
              AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
            GROUP BY EXTRACT(HOUR FROM l.loading_start_dttm)
            ORDER BY hour
        """
        with engine.connect() as conn:
            rows = conn.execute(text(query), {"days": days}).mappings().all()

        profile_map = {int(r["hour"]): r for r in rows}
        profile = []
        for h in range(24):
            row = profile_map.get(h, {})
            runs = int(row.get("runs_count") or 0)
            total = float(row.get("total_duration_minutes") or 0.0)
            profile.append(
                {
                    "hour": h,
                    "runs_count": runs,
                    "total_duration_minutes": round(total, 2),
                }
            )

        return JSONResponse(
            content={"days": days, "profile": profile},
            media_type="application/json; charset=utf-8",
        )
    except Exception as e:
        print("❌ /api/load-profile error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/night-summary")
def get_night_summary(
    days: int = Query(30, ge=1, le=120),
    limit: int = Query(50, ge=1, le=200),
    start_hour: int = Query(21, ge=0, le=23),
    end_hour: int = Query(8, ge=0, le=23),
    shift_days: int = Query(0, ge=0, le=14),
):
    try:
        with engine.connect() as conn:
            window_row = conn.execute(
                text(
                    """
                    SELECT
                        (date_trunc('day', now()) - interval '1 day' + (:start_hour || ' hours')::interval - (:shift_days || ' days')::interval) AS start_ts,
                        (date_trunc('day', now()) + (:end_hour || ' hours')::interval - (:shift_days || ' days')::interval) AS end_ts
                    """
                ),
                {"start_hour": start_hour, "end_hour": end_hour, "shift_days": shift_days},
            ).mappings().first()

        start_ts = window_row["start_ts"]
        end_ts = window_row["end_ts"]

        base_cte = f"""
            WITH night_runs AS (
                SELECT
                    l.object_id AS table_id,
                    COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                    COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                    e.entity_name,
                    l.loading_start_dttm,
                    l.loading_finish_dttm,
                    EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration,
                    EXTRACT(HOUR FROM l.loading_start_dttm) AS hour
                FROM {TABLE_LOADING_HISTORY} l
                LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
                LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                WHERE l.object_type = 'table'
                  AND l.loading_state IN ('SUCCESS', 'LOADED')
                  AND l.loading_start_dttm IS NOT NULL
                  AND l.loading_finish_dttm IS NOT NULL
                  AND l.loading_start_dttm >= :start_ts
                  AND l.loading_start_dttm < :end_ts
            )
        """
        failed_cte = f"""
            WITH failed_runs AS (
                SELECT
                    l.object_id AS table_id,
                    COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                    COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                    e.entity_name,
                    l.loading_start_dttm,
                    l.loading_finish_dttm,
                    l.message
                FROM {TABLE_LOADING_HISTORY} l
                LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
                LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                WHERE l.object_type = 'table'
                  AND l.loading_state = 'FAILED'
                  AND l.loading_start_dttm IS NOT NULL
                  AND l.loading_start_dttm >= :start_ts
                  AND l.loading_start_dttm < :end_ts
            )
        """

        with engine.connect() as conn:
            summary = conn.execute(
                text(
                    base_cte
                    + """
                    SELECT
                        COUNT(*) AS runs_count,
                        COUNT(DISTINCT table_id) AS tables_count,
                        COUNT(DISTINCT entity_name) AS entities_count,
                        SUM(duration) AS total_duration_minutes,
                        MAX(duration) AS max_duration_minutes
                    FROM night_runs
                    """
                ),
                {"start_ts": start_ts, "end_ts": end_ts},
            ).mappings().first()

            hourly = conn.execute(
                text(
                    base_cte
                    + """
                    SELECT
                        hour,
                        COUNT(*) AS runs_count,
                        SUM(duration) AS total_duration_minutes
                    FROM night_runs
                    GROUP BY hour
                    ORDER BY hour
                    """
                ),
                {"start_ts": start_ts, "end_ts": end_ts},
            ).mappings().all()

            hourly_top = conn.execute(
                text(
                    base_cte
                    + """
                    SELECT hour, table_id, table_schema, table_name, entity_name, duration
                    FROM (
                        SELECT
                            hour,
                            table_id,
                            table_schema,
                            table_name,
                            entity_name,
                            duration,
                            ROW_NUMBER() OVER (PARTITION BY hour ORDER BY duration DESC) AS rn
                        FROM night_runs
                    ) ranked
                    WHERE rn <= :limit
                    ORDER BY hour, duration DESC
                    """
                ),
                {"start_ts": start_ts, "end_ts": end_ts, "limit": limit},
            ).mappings().all()

            top_runs = conn.execute(
                text(
                    base_cte
                    + """
                    SELECT
                        table_id,
                        table_schema,
                        table_name,
                        entity_name,
                        duration,
                        loading_start_dttm,
                        loading_finish_dttm
                    FROM night_runs
                    ORDER BY duration DESC
                    LIMIT :limit
                    """
                ),
                {"start_ts": start_ts, "end_ts": end_ts, "limit": limit},
            ).mappings().all()

            anomalies = conn.execute(
                text(
                    f"""
                    WITH history AS (
                        SELECT
                            l.object_id,
                            percentile_cont(0.95) WITHIN GROUP (ORDER BY EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0) AS p95_duration
                        FROM {TABLE_LOADING_HISTORY} l
                        WHERE l.object_type = 'table'
                          AND l.loading_state IN ('SUCCESS', 'LOADED')
                          AND l.loading_start_dttm IS NOT NULL
                          AND l.loading_finish_dttm IS NOT NULL
                          AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
                        GROUP BY l.object_id
                    ),
                    night_runs AS (
                        SELECT
                            l.object_id AS table_id,
                            COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                            COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                            e.entity_name,
                            l.loading_start_dttm,
                            l.loading_finish_dttm,
                            EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration
                        FROM {TABLE_LOADING_HISTORY} l
                        LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
                        LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                        WHERE l.object_type = 'table'
                          AND l.loading_state IN ('SUCCESS', 'LOADED')
                          AND l.loading_start_dttm IS NOT NULL
                          AND l.loading_finish_dttm IS NOT NULL
                          AND l.loading_start_dttm >= :start_ts
                          AND l.loading_start_dttm < :end_ts
                    )
                    SELECT
                        n.table_id,
                        n.table_schema,
                        n.table_name,
                        n.entity_name,
                        n.duration,
                        n.loading_start_dttm,
                        n.loading_finish_dttm,
                        h.p95_duration,
                        CASE
                            WHEN h.p95_duration > 0 THEN n.duration / h.p95_duration
                            ELSE NULL
                        END AS ratio
                    FROM night_runs n
                    JOIN history h ON h.object_id = n.table_id
                    WHERE n.duration > h.p95_duration * 1.5
                    ORDER BY n.duration DESC
                    LIMIT :limit
                    """
                ),
                {
                    "days": days,
                    "start_ts": start_ts,
                    "end_ts": end_ts,
                    "limit": limit,
                },
            ).mappings().all()

            failed_summary = conn.execute(
                text(
                    failed_cte
                    + """
                    SELECT
                        COUNT(*) AS runs_count,
                        COUNT(DISTINCT table_id) AS tables_count,
                        COUNT(DISTINCT entity_name) AS entities_count
                    FROM failed_runs
                    """
                ),
                {"start_ts": start_ts, "end_ts": end_ts},
            ).mappings().first()

            failed_runs = conn.execute(
                text(
                    failed_cte
                    + """
                    SELECT
                        table_id,
                        table_schema,
                        table_name,
                        entity_name,
                        loading_start_dttm,
                        loading_finish_dttm,
                        message
                    FROM failed_runs
                    ORDER BY loading_finish_dttm DESC NULLS LAST
                    LIMIT :limit
                    """
                ),
                {"start_ts": start_ts, "end_ts": end_ts, "limit": limit},
            ).mappings().all()

        hourly_top_map = {}
        for row in hourly_top:
            hour = int(row["hour"])
            hourly_top_map.setdefault(hour, []).append(
                {
                    "table_id": row.get("table_id"),
                    "table_fqn": f"{row['table_schema']}.{row['table_name']}".strip("."),
                    "entity_name": row.get("entity_name"),
                    "duration_minutes": round(float(row["duration"]), 2) if row["duration"] is not None else None,
                }
            )

        hourly_payload = []
        for row in hourly:
            hour = int(row["hour"])
            total = float(row.get("total_duration_minutes") or 0.0)
            hourly_payload.append(
                {
                    "hour": hour,
                    "runs_count": int(row.get("runs_count") or 0),
                    "total_duration_minutes": round(total, 2),
                    "top_tables": hourly_top_map.get(hour, []),
                }
            )

        payload = {
            "window": {
                "start": serialize_datetime(start_ts),
                "end": serialize_datetime(end_ts),
                "start_hour": start_hour,
                "end_hour": end_hour,
            },
            "summary": {
                "runs_count": int(summary.get("runs_count") or 0),
                "tables_count": int(summary.get("tables_count") or 0),
                "entities_count": int(summary.get("entities_count") or 0),
                "total_duration_minutes": round(float(summary.get("total_duration_minutes") or 0.0), 2),
                "max_duration_minutes": round(float(summary.get("max_duration_minutes") or 0.0), 2)
                if summary.get("max_duration_minutes") is not None
                else None,
            },
            "failed_summary": {
                "runs_count": int(failed_summary.get("runs_count") or 0),
                "tables_count": int(failed_summary.get("tables_count") or 0),
                "entities_count": int(failed_summary.get("entities_count") or 0),
            },
            "hourly": hourly_payload,
            "top_runs": [
                {
                    "table_id": row.get("table_id"),
                    "table_fqn": f"{row['table_schema']}.{row['table_name']}".strip("."),
                    "entity_name": row.get("entity_name"),
                    "duration_minutes": round(float(row["duration"]), 2) if row["duration"] is not None else None,
                    "start": serialize_datetime(row.get("loading_start_dttm")),
                    "end": serialize_datetime(row.get("loading_finish_dttm")),
                }
                for row in top_runs
            ],
            "anomalies": [
                {
                    "table_id": row.get("table_id"),
                    "table_fqn": f"{row['table_schema']}.{row['table_name']}".strip("."),
                    "entity_name": row.get("entity_name"),
                    "duration_minutes": round(float(row["duration"]), 2) if row["duration"] is not None else None,
                    "p95_minutes": round(float(row["p95_duration"]), 2) if row["p95_duration"] is not None else None,
                    "ratio": round(float(row["ratio"]), 2) if row["ratio"] is not None else None,
                    "start": serialize_datetime(row.get("loading_start_dttm")),
                    "end": serialize_datetime(row.get("loading_finish_dttm")),
                }
                for row in anomalies
            ],
            "failed_runs": [
                {
                    "table_id": row.get("table_id"),
                    "table_fqn": f"{row['table_schema']}.{row['table_name']}".strip("."),
                    "entity_name": row.get("entity_name"),
                    "start": serialize_datetime(row.get("loading_start_dttm")),
                    "end": serialize_datetime(row.get("loading_finish_dttm")),
                    "message": row.get("message"),
                }
                for row in failed_runs
            ],
        }
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        print("❌ /api/night-summary error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


def _parse_hhmm_to_minutes(value: str, field_name: str) -> int:
    if not isinstance(value, str):
        raise HTTPException(status_code=400, detail=f"{field_name} must be a string in HH:MM format")
    match = re.fullmatch(r"([01]?\d|2[0-3]):([0-5]\d)", value.strip())
    if not match:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}: {value}. Expected HH:MM")
    hours = int(match.group(1))
    minutes = int(match.group(2))
    return hours * 60 + minutes


@router.get("/api/night/heavy-tables")
def get_night_heavy_tables(
    days: int = Query(30, ge=1, le=120),
    limit: int = Query(25, ge=1, le=200),
    window_start: str = Query("04:30"),
    window_end: str = Query("05:20"),
):
    try:
        start_minutes = _parse_hhmm_to_minutes(window_start, "window_start")
        end_minutes = _parse_hhmm_to_minutes(window_end, "window_end")
        crosses_midnight = start_minutes > end_minutes

        time_filter = """
            (
                (:crosses_midnight = false AND minute_of_day >= :start_minutes AND minute_of_day < :end_minutes)
                OR
                (:crosses_midnight = true AND (minute_of_day >= :start_minutes OR minute_of_day < :end_minutes))
            )
        """

        query = f"""
            WITH base AS (
                SELECT
                    l.object_id AS table_id,
                    COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                    COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                    COALESCE(e.entity_name, 'UNKNOWN') AS entity_name,
                    l.loading_start_dttm,
                    l.loading_finish_dttm,
                    EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration_minutes,
                    (EXTRACT(HOUR FROM l.loading_start_dttm)::int * 60 + EXTRACT(MINUTE FROM l.loading_start_dttm)::int) AS minute_of_day
                FROM {TABLE_LOADING_HISTORY} l
                LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
                LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                WHERE l.object_type = 'table'
                  AND l.loading_state IN ('SUCCESS', 'LOADED')
                  AND l.loading_start_dttm IS NOT NULL
                  AND l.loading_finish_dttm IS NOT NULL
                  AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
            ),
            windowed AS (
                SELECT *
                FROM base
                WHERE {time_filter}
            )
            SELECT
                table_id,
                table_schema,
                table_name,
                entity_name,
                COUNT(*) AS runs_count,
                SUM(duration_minutes) AS total_duration_minutes,
                AVG(duration_minutes) AS avg_duration_minutes,
                MAX(duration_minutes) AS max_duration_minutes,
                MAX(loading_start_dttm) AS last_start
            FROM windowed
            GROUP BY table_id, table_schema, table_name, entity_name
            ORDER BY total_duration_minutes DESC NULLS LAST
            LIMIT :limit
        """

        summary_query = f"""
            WITH base AS (
                SELECT
                    l.object_id AS table_id,
                    (EXTRACT(HOUR FROM l.loading_start_dttm)::int * 60 + EXTRACT(MINUTE FROM l.loading_start_dttm)::int) AS minute_of_day,
                    EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration_minutes
                FROM {TABLE_LOADING_HISTORY} l
                WHERE l.object_type = 'table'
                  AND l.loading_state IN ('SUCCESS', 'LOADED')
                  AND l.loading_start_dttm IS NOT NULL
                  AND l.loading_finish_dttm IS NOT NULL
                  AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
            ),
            windowed AS (
                SELECT *
                FROM base
                WHERE {time_filter}
            )
            SELECT
                COUNT(*) AS runs_count,
                COUNT(DISTINCT table_id) AS tables_count,
                SUM(duration_minutes) AS total_duration_minutes,
                AVG(duration_minutes) AS avg_duration_minutes,
                MAX(duration_minutes) AS max_duration_minutes
            FROM windowed
        """

        params = {
            "days": days,
            "limit": limit,
            "start_minutes": start_minutes,
            "end_minutes": end_minutes,
            "crosses_midnight": crosses_midnight,
        }

        with engine.connect() as conn:
            rows = conn.execute(text(query), params).mappings().all()
            summary = conn.execute(text(summary_query), params).mappings().first() or {}

        payload = {
            "window": {
                "start": window_start,
                "end": window_end,
                "start_minutes": start_minutes,
                "end_minutes": end_minutes,
                "crosses_midnight": crosses_midnight,
                "days": days,
            },
            "summary": {
                "runs_count": int(summary.get("runs_count") or 0),
                "tables_count": int(summary.get("tables_count") or 0),
                "total_duration_minutes": round(float(summary.get("total_duration_minutes") or 0.0), 2),
                "avg_duration_minutes": round(float(summary.get("avg_duration_minutes") or 0.0), 2),
                "max_duration_minutes": round(float(summary.get("max_duration_minutes") or 0.0), 2),
            },
            "rows": [
                {
                    "table_id": row.get("table_id"),
                    "table_fqn": f"{row.get('table_schema')}.{row.get('table_name')}".strip("."),
                    "table_schema": row.get("table_schema"),
                    "table_name": row.get("table_name"),
                    "entity_name": row.get("entity_name"),
                    "runs_count": int(row.get("runs_count") or 0),
                    "total_duration_minutes": round(float(row.get("total_duration_minutes") or 0.0), 2),
                    "avg_duration_minutes": round(float(row.get("avg_duration_minutes") or 0.0), 2),
                    "max_duration_minutes": round(float(row.get("max_duration_minutes") or 0.0), 2),
                    "last_start": serialize_datetime(row.get("last_start")),
                }
                for row in rows
            ],
        }
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except HTTPException:
        raise
    except Exception as e:
        print("❌ /api/night/heavy-tables error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/logic-audit")
def get_logic_audit(
    issue_type: str = Query("all"),
    mode: str = Query("standard"),
    min_score: float = Query(0.72, ge=0.0, le=1.0),
    limit: int = Query(200, ge=1, le=1000),
    search: Optional[str] = Query(None),
):
    payload = _build_logic_audit_cache()
    pairs = payload.get("pairs") or []

    if issue_type != "all":
        pairs = [row for row in pairs if row.get("issue_type") == issue_type]

    if min_score > 0:
        pairs = [row for row in pairs if (row.get("score") or 0) >= min_score]

    if mode == "strict":
        pairs = [
            row for row in pairs
            if (row.get("expression_overlap_count") or 0) >= 1
            and (row.get("score") or 0) >= max(min_score, 0.72)
        ]

    if search:
        term = search.strip().lower()
        if term:
            pairs = [
                row for row in pairs
                if term in (row.get("left_fqn") or "").lower()
                or term in (row.get("right_fqn") or "").lower()
                or term in (row.get("left_entity") or "").lower()
                or term in (row.get("right_entity") or "").lower()
            ]

    pairs = pairs[:limit]
    stats = {
        "duplicate_exact": sum(1 for row in pairs if row.get("issue_type") == "duplicate_exact"),
        "duplicate_candidate": sum(1 for row in pairs if row.get("issue_type") == "duplicate_candidate"),
        "similar_candidate": sum(1 for row in pairs if row.get("issue_type") == "similar_candidate"),
    }
    return {
        "generated_at": payload.get("generated_at"),
        "objects_count": payload.get("objects_count"),
        "pairs_count": payload.get("pairs_count"),
        "mode": mode,
        "returned_count": len(pairs),
        "stats": stats,
        "pairs": pairs,
    }


@router.get("/api/logic-audit/pair/{pair_id}")
def get_logic_audit_pair(pair_id: str):
    payload = _build_logic_audit_cache()
    detail = (payload.get("pair_index") or {}).get(pair_id)
    if not detail:
        return JSONResponse(status_code=404, content={"error": "pair not found"})
    return detail


def _normalize_architecture_fqn(value: Optional[str]) -> Optional[str]:
    raw = str(value or "").strip().lower()
    if not raw or "." not in raw:
        return None
    schema, table = raw.split(".", 1)
    schema = schema.strip()
    table = table.strip()
    if not schema or not table:
        return None
    return f"{schema}.{table}"


def _architecture_fqn_aliases(fqn: str) -> set[str]:
    normalized = _normalize_architecture_fqn(fqn)
    if not normalized or "." not in normalized:
        return set()
    schema, table = normalized.split(".", 1)
    aliases = {normalized}
    cleaned = _clean_table_name(table)
    if cleaned:
        aliases.add(f"{schema}.{cleaned}")
        if not cleaned.startswith("/"):
            aliases.add(f"{schema}./{cleaned}")
    if not table.startswith("/"):
        aliases.add(f"{schema}./{table}")
    return {alias for alias in aliases if alias}


def _build_architecture_alias_map(fqns: list[str]) -> dict[str, str]:
    alias_map: dict[str, str] = {}
    for fqn in fqns:
        for alias in _architecture_fqn_aliases(fqn):
            alias_map[alias] = fqn
    return alias_map


def _empty_architecture_context(fqn: str) -> dict[str, Any]:
    return {
        "fqn": fqn,
        "direct_upstream_count": 0,
        "direct_downstream_count": 0,
        "transitive_downstream_count": 0,
        "downstream_entities_count": 0,
        "downstream_entities": [],
        "releases_count": 0,
        "release_objects_count": 0,
        "release_tasks_count": 0,
        "incidents_count": 0,
        "latest_release": None,
        "latest_incident": None,
        "last_change": None,
    }


def _build_architecture_workbench_enrichment(
    fqns: list[str],
    *,
    release_days: int = 180,
    incident_days: int = 180,
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    normalized_fqns = sorted({_normalize_architecture_fqn(item) for item in fqns if _normalize_architecture_fqn(item)})
    if not normalized_fqns:
        return {}, {
            "objects_with_releases": 0,
            "objects_with_incidents": 0,
            "objects_with_downstream": 0,
            "objects_with_last_change": 0,
        }

    alias_map = _build_architecture_alias_map(normalized_fqns)
    context_map = {fqn: _empty_architecture_context(fqn) for fqn in normalized_fqns}

    all_meta, reverse_index = get_cached_meta_and_index()
    meta_map = {
        f"{m.get('table_schema')}.{m.get('table_name')}": m
        for m in all_meta
        if m.get("table_schema") and m.get("table_name")
    }

    for fqn in normalized_fqns:
        meta = meta_map.get(fqn) or {}
        ctx = context_map[fqn]
        depends_on = meta.get("depends_on") or {}
        direct_upstream = {
            f"{src_schema}.{src_table}"
            for src_schema, tables in depends_on.items()
            for src_table in (tables or [])
            if src_schema and src_table
        }
        if "." in fqn:
            schema_name, table_name = fqn.split(".", 1)
        else:
            schema_name, table_name = "", ""
        direct_downstream = reverse_index.get((schema_name, table_name), []) or []

        visited: set[str] = set()
        queue = [fqn]
        downstream_entities: set[str] = set()
        while queue:
            current = queue.pop(0)
            if current in visited:
                continue
            visited.add(current)
            if "." not in current:
                continue
            current_schema, current_table = current.split(".", 1)
            for consumer in reverse_index.get((current_schema, current_table), []) or []:
                target_schema = str(consumer.get("schema") or "").strip().lower()
                target_table = str(consumer.get("table_name") or "").strip().lower()
                if not target_schema or not target_table:
                    continue
                target_fqn = f"{target_schema}.{target_table}"
                if target_fqn == fqn:
                    continue
                if target_fqn not in visited:
                    queue.append(target_fqn)
                entity_name = str(consumer.get("entity_name") or "").strip()
                if entity_name:
                    downstream_entities.add(entity_name)

        ctx["direct_upstream_count"] = len(direct_upstream)
        ctx["direct_downstream_count"] = len(
            {
                f"{str(item.get('schema') or '').strip().lower()}.{str(item.get('table_name') or '').strip().lower()}"
                for item in direct_downstream
                if item.get("schema") and item.get("table_name")
            }
        )
        transitive_targets = visited - {fqn}
        ctx["transitive_downstream_count"] = len(transitive_targets)
        ctx["downstream_entities_count"] = len(downstream_entities)
        ctx["downstream_entities"] = sorted(downstream_entities)[:8]

    release_rows = []
    snapshot_rows = []
    exec_rows = []
    try:
        with engine.connect() as conn:
            release_query = text(
                f"""
                SELECT
                    lower(ro.schema_name) AS schema_name,
                    lower(ro.table_name) AS table_name,
                    ro.release_id,
                    ro.task_id,
                    ro.change_type,
                    rl.started_at,
                    rl.release_type,
                    rl.status,
                    rl.initiated_by
                FROM {TABLE_RELEASE_OBJECTS} ro
                JOIN {TABLE_RELEASE_LOG} rl ON rl.release_id = ro.release_id
                WHERE rl.started_at >= (now() - (:days || ' days')::interval)
                  AND (lower(ro.schema_name) || '.' || lower(ro.table_name)) IN :fqns
                ORDER BY rl.started_at DESC NULLS LAST, ro.release_id DESC
                """
            ).bindparams(bindparam("fqns", expanding=True))
            for offset in range(0, len(normalized_fqns), 25):
                fqns_chunk = normalized_fqns[offset: offset + 25]
                if not fqns_chunk:
                    continue
                release_rows.extend(
                    conn.execute(
                        release_query,
                        {"days": release_days, "fqns": fqns_chunk},
                    ).mappings().all()
                )

            release_task_ids = sorted(
                {
                    str(row.get("task_id") or "").strip()
                    for row in release_rows
                    if str(row.get("task_id") or "").strip()
                }
            )
            if release_task_ids:
                snapshot_query = text(
                    f"""
                    SELECT
                        issue_id,
                        summary,
                        created_by,
                        assignee,
                        updated_at,
                        created_at,
                        current_state
                    FROM {TABLE_YT_ISSUE_SNAPSHOT}
                    WHERE issue_id IN :issue_ids
                    """
                ).bindparams(bindparam("issue_ids", expanding=True))
                exec_query = text(
                    f"""
                    SELECT DISTINCT ON (issue_id)
                           issue_id,
                           author AS executor,
                           ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE issue_id IN :issue_ids
                      AND event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                    """
                ).bindparams(bindparam("issue_ids", expanding=True))
                for offset in range(0, len(release_task_ids), 100):
                    task_chunk = release_task_ids[offset: offset + 100]
                    if not task_chunk:
                        continue
                    snapshot_rows.extend(conn.execute(snapshot_query, {"issue_ids": task_chunk}).mappings().all())
                    exec_rows.extend(conn.execute(exec_query, {"issue_ids": task_chunk}).mappings().all())
    except OperationalError:
        release_rows = []
        snapshot_rows = []
        exec_rows = []

    snapshot_map = {str(row.get("issue_id") or ""): dict(row) for row in snapshot_rows}
    exec_map = {str(row.get("issue_id") or ""): dict(row) for row in exec_rows}
    release_counts: dict[str, set[str]] = defaultdict(set)
    release_task_counts: dict[str, set[str]] = defaultdict(set)
    for row in release_rows:
        raw_fqn = f"{row.get('schema_name')}.{row.get('table_name')}"
        fqn = alias_map.get(raw_fqn)
        if not fqn:
            continue
        ctx = context_map[fqn]
        release_counts[fqn].add(str(row.get("release_id") or ""))
        if str(row.get("task_id") or "").strip():
            release_task_counts[fqn].add(str(row.get("task_id") or "").strip())
        ctx["release_objects_count"] += 1
        current_latest = ctx.get("latest_release")
        row_started_at = row.get("started_at")
        if not current_latest or (row_started_at and row_started_at > (current_latest.get("_started_at") or datetime.min)):
            task_id = str(row.get("task_id") or "").strip()
            snap = snapshot_map.get(task_id, {})
            exec_meta = exec_map.get(task_id, {})
            actor = (
                exec_meta.get("executor")
                or snap.get("assignee")
                or snap.get("created_by")
                or row.get("initiated_by")
                or "Не указан"
            )
            latest_release = {
                "release_id": row.get("release_id"),
                "started_at": serialize_datetime(row_started_at),
                "release_type": row.get("release_type"),
                "status": row.get("status"),
                "initiated_by": row.get("initiated_by") or "Не указан",
                "task_id": task_id or None,
                "task_link": _build_ytrack_link(task_id) if task_id else None,
                "task_summary": snap.get("summary"),
                "actor": actor,
                "_started_at": row_started_at,
            }
            ctx["latest_release"] = latest_release
            ctx["last_change"] = {
                "changed_at": serialize_datetime(row_started_at),
                "actor": actor,
                "source": "release",
                "release_id": row.get("release_id"),
                "task_id": task_id or None,
                "task_link": _build_ytrack_link(task_id) if task_id else None,
                "task_summary": snap.get("summary"),
            }

    for fqn in normalized_fqns:
        context_map[fqn]["releases_count"] = len(release_counts.get(fqn) or set())
        context_map[fqn]["release_tasks_count"] = len(release_task_counts.get(fqn) or set())
        latest_release = context_map[fqn].get("latest_release")
        if latest_release:
            latest_release.pop("_started_at", None)

    try:
        with engine.connect() as conn:
            incident_rows = conn.execute(
                text(
                    """
                    SELECT
                        issue_id,
                        summary,
                        state_name,
                        author_name,
                        assignee_name,
                        incident_reason_name,
                        alert_source,
                        trigger_dttm,
                        incident_start_dttm,
                        detected_dttm,
                        work_finished_dttm,
                        updated_at_yt,
                        resolved_at_yt,
                        table_schema,
                        table_name,
                        table_name_raw
                    FROM tech_etl.yt_incidents
                    WHERE COALESCE(incident_start_dttm, detected_dttm, trigger_dttm, created_at_yt, dttm_loaded)
                          >= (now() - (:days || ' days')::interval)
                    ORDER BY COALESCE(incident_start_dttm, detected_dttm, trigger_dttm, created_at_yt) DESC NULLS LAST,
                             issue_id DESC
                    """
                ),
                {"days": incident_days},
            ).mappings().all()
    except OperationalError:
        incident_rows = []

    incident_counts: dict[str, set[str]] = defaultdict(set)
    for row in incident_rows:
        table_schema = str(row.get("table_schema") or "").strip().lower()
        table_names = _split_rich_multivalue(row.get("table_name")) or _split_rich_multivalue(row.get("table_name_raw"))
        matched_fqns: set[str] = set()
        if table_schema and table_names:
            for table_name in table_names:
                table_name_norm = str(table_name or "").strip().lower()
                if not table_name_norm:
                    continue
                for alias in _architecture_fqn_aliases(f"{table_schema}.{table_name_norm}"):
                    canonical = alias_map.get(alias)
                    if canonical:
                        matched_fqns.add(canonical)
        if not matched_fqns:
            continue

        incident_at = (
            row.get("incident_start_dttm")
            or row.get("detected_dttm")
            or row.get("trigger_dttm")
            or row.get("updated_at_yt")
            or row.get("resolved_at_yt")
        )
        for fqn in matched_fqns:
            incident_counts[fqn].add(str(row.get("issue_id") or ""))
            ctx = context_map[fqn]
            current_latest = ctx.get("latest_incident")
            if not current_latest or (incident_at and incident_at > (current_latest.get("_incident_at") or datetime.min)):
                issue_id = str(row.get("issue_id") or "").strip()
                ctx["latest_incident"] = {
                    "issue_id": issue_id,
                    "link": _build_ytrack_link(issue_id),
                    "summary": row.get("summary"),
                    "state_name": row.get("state_name"),
                    "incident_reason_name": row.get("incident_reason_name") or "Не указана",
                    "alert_source": row.get("alert_source") or "Не указан",
                    "incident_start_dttm": serialize_datetime(incident_at),
                    "author_name": row.get("author_name") or "Не указан",
                    "assignee_name": row.get("assignee_name") or row.get("author_name") or "Не указан",
                    "_incident_at": incident_at,
                }

    for fqn in normalized_fqns:
        context_map[fqn]["incidents_count"] = len(incident_counts.get(fqn) or set())
        latest_incident = context_map[fqn].get("latest_incident")
        if latest_incident:
            latest_incident.pop("_incident_at", None)

    summary = {
        "objects_with_releases": sum(1 for row in context_map.values() if row.get("releases_count")),
        "objects_with_incidents": sum(1 for row in context_map.values() if row.get("incidents_count")),
        "objects_with_downstream": sum(1 for row in context_map.values() if row.get("transitive_downstream_count")),
        "objects_with_last_change": sum(1 for row in context_map.values() if row.get("last_change")),
    }
    return context_map, summary


@router.get("/api/admin/architecture/workbench")
def get_admin_architecture_workbench(
    request: Request,
    issue_type: str = Query("all"),
    mode: str = Query("standard"),
    min_score: float = Query(0.72, ge=0.0, le=1.0),
    limit: int = Query(500, ge=1, le=1000),
    search: Optional[str] = Query(None),
    release_days: int = Query(180, ge=30, le=3650),
    incident_days: int = Query(180, ge=30, le=3650),
):
    user = get_current_user_from_request(request)
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    payload = _build_logic_audit_cache()
    pairs = payload.get("pairs") or []

    if issue_type != "all":
        pairs = [row for row in pairs if row.get("issue_type") == issue_type]

    if min_score > 0:
        pairs = [row for row in pairs if (row.get("score") or 0) >= min_score]

    if mode == "strict":
        pairs = [
            row for row in pairs
            if (row.get("expression_overlap_count") or 0) >= 1
            and (row.get("score") or 0) >= max(min_score, 0.72)
        ]

    if search:
        term = search.strip().lower()
        if term:
            pairs = [
                row for row in pairs
                if term in (row.get("left_fqn") or "").lower()
                or term in (row.get("right_fqn") or "").lower()
                or term in (row.get("left_entity") or "").lower()
                or term in (row.get("right_entity") or "").lower()
            ]

    pairs = pairs[:limit]
    fqns = sorted(
        {
            item
            for row in pairs
            for item in (_normalize_architecture_fqn(row.get("left_fqn")), _normalize_architecture_fqn(row.get("right_fqn")))
            if item
        }
    )
    enrichment, enrichment_summary = _build_architecture_workbench_enrichment(
        fqns,
        release_days=release_days,
        incident_days=incident_days,
    )
    stats = {
        "duplicate_exact": sum(1 for row in pairs if row.get("issue_type") == "duplicate_exact"),
        "duplicate_candidate": sum(1 for row in pairs if row.get("issue_type") == "duplicate_candidate"),
        "similar_candidate": sum(1 for row in pairs if row.get("issue_type") == "similar_candidate"),
    }
    return {
        "generated_at": payload.get("generated_at"),
        "objects_count": payload.get("objects_count"),
        "pairs_count": payload.get("pairs_count"),
        "mode": mode,
        "returned_count": len(pairs),
        "stats": stats,
        "pairs": pairs,
        "enrichment": enrichment,
        "enrichment_summary": enrichment_summary,
        "block_exact_clusters": payload.get("block_exact_clusters") or [],
        "block_similar_pairs": payload.get("block_similar_pairs") or [],
        "block_summary": payload.get("block_summary") or {},
        "windows": {
            "release_days": release_days,
            "incident_days": incident_days,
        },
    }


@router.get("/api/admin/architecture/block-pair/{pair_id}")
def get_admin_architecture_block_pair(
    pair_id: str,
    request: Request,
):
    user = get_current_user_from_request(request)
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")

    payload = _build_logic_audit_cache()
    detail = (payload.get("block_pair_index") or {}).get(pair_id)
    if not detail:
        return JSONResponse(status_code=404, content={"error": "block pair not found"})
    return detail


@router.get("/api/entity-loads")
def get_entity_loads(
    entity_id: int = Query(..., ge=1),
    days: int = Query(30, ge=1, le=120),
    limit: int = Query(30, ge=1, le=200),
    schema: Optional[str] = Query(None),
):
    try:
        schema = schema.strip() if isinstance(schema, str) else None
        query = f"""
            WITH base AS (
                SELECT
                    COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                    COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                    e.entity_name,
                    EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration,
                    l.loading_finish_dttm
                FROM {TABLE_LOADING_HISTORY} l
                LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
                LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                WHERE l.object_type = 'table'
                  AND l.loading_state IN ('SUCCESS', 'LOADED')
                  AND l.loading_start_dttm IS NOT NULL
                  AND l.loading_finish_dttm IS NOT NULL
                  AND e.entity_id = :entity_id
                  AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
                  AND (
                    :schema IS NULL OR
                    LOWER(COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), ''))) = LOWER(:schema)
                  )
            )
            SELECT
                table_schema,
                table_name,
                entity_name,
                COUNT(*) AS runs_count,
                AVG(duration) AS avg_duration,
                percentile_cont(0.95) WITHIN GROUP (ORDER BY duration) AS p95_duration,
                MAX(duration) AS max_duration,
                MAX(loading_finish_dttm) AS last_finish
            FROM base
            GROUP BY table_schema, table_name, entity_name
            ORDER BY MAX(duration) DESC NULLS LAST
            LIMIT :limit
        """
        with engine.connect() as conn:
            rows = conn.execute(
                text(query),
                {"entity_id": entity_id, "days": days, "limit": limit, "schema": schema},
            ).mappings().all()

        payload = [
            {
                "table_fqn": f"{row['table_schema']}.{row['table_name']}".strip("."),
                "entity_name": row.get("entity_name"),
                "runs_count": int(row.get("runs_count") or 0),
                "avg_duration": round(float(row["avg_duration"]), 2) if row.get("avg_duration") is not None else None,
                "p95_duration": round(float(row["p95_duration"]), 2) if row.get("p95_duration") is not None else None,
                "max_duration": round(float(row["max_duration"]), 2) if row.get("max_duration") is not None else None,
                "last_finish": serialize_datetime(row.get("last_finish")),
            }
            for row in rows
        ]
        return JSONResponse(content=payload, media_type="application/json; charset=utf-8")
    except Exception as e:
        print("❌ /api/entity-loads error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


def load_all_meta():
    all_meta = {}
    for top_path in iter_meta_dirs():
        for schema_dir in top_path.iterdir():
            if not schema_dir.is_dir():
                continue
            for table_dir in schema_dir.iterdir():
                yaml_path = table_dir / "meta_data_file.yaml"
                if yaml_path.exists():
                    try:
                        with open(yaml_path, encoding="utf-8") as f:
                            meta = yaml.safe_load(f)
                            key = f"{meta['table_schema']}.{meta['table_name']}"
                            all_meta[key] = meta
                    except Exception:
                        continue
    return all_meta


def get_downstream_dependencies(start_table: str, all_meta: dict):
    result = set()
    stack = [start_table]

    while stack:
        current = stack.pop()
        if current in all_meta:
            deps = all_meta[current].get("depends_on", {})
            for schema, tables in deps.items():
                for table in tables:
                    full_name = f"{schema}.{table}"
                    if full_name not in result:
                        result.add(full_name)
                        stack.append(full_name)

    return sorted(result)


def get_dependency_edges(start_table: str, all_meta: dict) -> List[Dict[str, str]]:
    edges = []
    visited = set()
    stack = [start_table]

    while stack:
        current = stack.pop()
        if current in visited:
            continue
        visited.add(current)

        meta = all_meta.get(current)
        if not meta:
            continue

        depends_on = meta.get("depends_on", {})
        for source_schema, tables in depends_on.items():
            for source_table in tables:
                source = f"{source_schema}.{source_table}"
                edges.append({"source": source, "target": current})
                stack.append(source)

    return edges


@router.get("/api/dependencies-down/{schema}/{table:path}")
def get_dependencies_down(schema: str, table: str):
    try:
        all_meta = load_all_meta()
        key = _resolve_table_key(all_meta, schema, table)
        if key not in all_meta:
            return JSONResponse(status_code=404, content={"error": "table not found"})

        edges = get_dependency_edges(key, all_meta)
        return {"central_node": key, "edges": edges}

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/dependencies-graph/{schema}/{table:path}")
def get_dependency_graph(
    schema: str,
    table: str,
    max_depth: Optional[int] = Query(None, ge=1),
    max_edges: Optional[int] = Query(None, ge=1),
):
    try:
        now = time.time()

        all_meta_list, _ = get_cached_meta_and_index()
        if _graph_cache_meta_ts != _cache_timestamp:
            _graph_cache.clear()
            globals()["_graph_cache_meta_ts"] = _cache_timestamp
            globals()["_graph_cache_ts"] = now

        all_meta = {
            f"{m.get('table_schema')}.{m.get('table_name')}": m
            for m in all_meta_list
            if m.get("table_schema") and m.get("table_name")
        }
        key = _resolve_table_key(all_meta, schema, table)
        cache_key = (key, max_depth, max_edges)
        if _graph_cache and now - _graph_cache_ts < _GRAPH_CACHE_TTL:
            cached = _graph_cache.get(cache_key)
            if cached is not None:
                return cached
        reverse_index = {}
        for m in all_meta_list:
            consumer = (m.get("table_schema"), m.get("table_name"))
            for src_schema, tables in (m.get("depends_on") or {}).items():
                for src_table in tables or []:
                    reverse_index.setdefault((src_schema, src_table), []).append({
                        "schema": consumer[0],
                        "table_name": consumer[1],
                    })
        visited = set()
        edges_set = set()
        truncated = False

        stack = [(key, 0)]
        while stack:
            current, depth = stack.pop()
            if current in visited:
                continue
            visited.add(current)

            meta = all_meta.get(current)
            if not meta:
                continue
            if max_depth is not None and depth >= max_depth:
                schema_val, table_val = current.split(".", 1)
                if meta.get("depends_on") or reverse_index.get((schema_val, table_val)):
                    truncated = True
                continue

            depends_on = meta.get("depends_on") or {}
            for source_schema, source_tables in depends_on.items():
                if not source_schema:
                    continue
                for source_table in source_tables or []:
                    if not source_table:
                        continue
                    source = f"{source_schema}.{source_table}"
                    edge_key = (source, current)
                    if edge_key not in edges_set:
                        edges_set.add(edge_key)
                        if max_edges is not None and len(edges_set) >= max_edges:
                            truncated = True
                            stack = []
                            break
                    stack.append((source, depth + 1))
                if truncated:
                    break
            if truncated:
                break

            schema_val, table_val = current.split(".", 1)
            for consumer in reverse_index.get((schema_val, table_val), []):
                if not consumer.get("schema") or not consumer.get("table_name"):
                    continue
                target = f"{consumer['schema']}.{consumer['table_name']}"
                edge_key = (current, target)
                if edge_key not in edges_set:
                    edges_set.add(edge_key)
                    if max_edges is not None and len(edges_set) >= max_edges:
                        truncated = True
                        stack = []
                        break
                stack.append((target, depth + 1))
            if truncated:
                break

        edges = [{"source": s, "target": t} for s, t in edges_set]
        payload = {"centralNode": key, "edges": edges, "truncated": truncated}

        if now - _graph_cache_ts >= _GRAPH_CACHE_TTL:
            _graph_cache.clear()
            globals()["_graph_cache_ts"] = now
        _graph_cache[cache_key] = payload
        return payload

    except Exception as e:
        print("Ошибка при построении графа зависимостей:", e)
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/dependencies-nodes/{schema}/{table:path}")
def get_dependency_nodes(
    schema: str,
    table: str,
    max_depth: Optional[int] = Query(None, ge=1),
    max_nodes: Optional[int] = Query(None, ge=1),
):
    try:
        all_meta_list, _ = get_cached_meta_and_index()
        all_meta = {
            f"{m.get('table_schema')}.{m.get('table_name')}": m
            for m in all_meta_list
            if m.get("table_schema") and m.get("table_name")
        }
        key = _resolve_table_key(all_meta, schema, table)
        reverse_index = {}
        for m in all_meta_list:
            consumer = (m.get("table_schema"), m.get("table_name"))
            for src_schema, tables in (m.get("depends_on") or {}).items():
                for src_table in tables or []:
                    reverse_index.setdefault((src_schema, src_table), []).append({
                        "schema": consumer[0],
                        "table_name": consumer[1],
                    })

        visited = set()
        truncated = False
        stack = [(key, 0)]

        while stack:
            current, depth = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            if max_nodes is not None and len(visited) >= max_nodes:
                truncated = True
                break

            meta = all_meta.get(current)
            if not meta:
                continue
            if max_depth is not None and depth >= max_depth:
                schema_val, table_val = current.split(".", 1)
                if meta.get("depends_on") or reverse_index.get((schema_val, table_val)):
                    truncated = True
                continue

            depends_on = meta.get("depends_on") or {}
            for source_schema, source_tables in depends_on.items():
                if not source_schema:
                    continue
                for source_table in source_tables or []:
                    if not source_table:
                        continue
                    source = f"{source_schema}.{source_table}"
                    stack.append((source, depth + 1))

            schema_val, table_val = current.split(".", 1)
            for consumer in reverse_index.get((schema_val, table_val), []):
                if not consumer.get("schema") or not consumer.get("table_name"):
                    continue
                target = f"{consumer['schema']}.{consumer['table_name']}"
                stack.append((target, depth + 1))

        nodes = sorted(visited, key=lambda v: v.lower())
        return {"centralNode": key, "nodes": nodes, "truncated": truncated}
    except Exception as e:
        print("Ошибка при построении списка зависимостей:", e)
        return JSONResponse(status_code=500, content={"error": str(e)})





@router.get("/api/gantt/{schema}/{table:path}")
def get_gantt_data(schema: str, table: str, depth: int = Query(3, ge=1, le=4)):
    try:
        snapshot = get_graph_snapshot()
        table_nodes = snapshot["table_graph"]["nodes"]
        table_edges = snapshot["table_graph"]["edges"]

        start_table = _resolve_table_key(table_nodes, schema, table)
        if start_table not in table_nodes:
            return JSONResponse(status_code=404, content={"error": f"'{start_table}' not found in meta"})

        reverse = {}
        for edge in table_edges:
            reverse.setdefault(edge["target"], []).append(edge["source"])

        visited = {start_table}
        queue = deque([(start_table, 0)])
        while queue:
            node, d = queue.popleft()
            if d >= depth:
                continue
            for src in reverse.get(node, []):
                if src in visited:
                    continue
                visited.add(src)
                queue.append((src, d + 1))

        edges = [e for e in table_edges if e["source"] in visited and e["target"] in visited]
        table_to_id = {
            t: table_nodes[t]["table_id"]
            for t in visited
            if t in table_nodes and table_nodes[t].get("table_id")
        }

        if not table_to_id:
            return JSONResponse(content=[], media_type="application/json")

        id_list = list(table_to_id.values())

        query = (
            text(
                f"""
            WITH cte AS (
                SELECT object_id, loading_start_dttm, loading_finish_dttm,
                       row_number() OVER (PARTITION BY object_id ORDER BY loading_start_dttm DESC) rn
                FROM {TABLE_LOADING_HISTORY}
                WHERE loading_state IN ('SUCCESS', 'LOADED') AND object_type = 'table'
                  AND object_id IN :id_list
            )
            SELECT object_id, loading_start_dttm, loading_finish_dttm
            FROM cte
            WHERE rn = 1
            ORDER BY loading_start_dttm
        """
            )
            .bindparams(bindparam("id_list", expanding=True))
        )

        with engine.connect() as conn:
            rows = conn.execute(query, {"id_list": id_list}).mappings().all()

        loading_times = {
            row["object_id"]: {"start": row["loading_start_dttm"], "end": row["loading_finish_dttm"]} for row in rows
        }

        id_to_table = {v: k for k, v in table_to_id.items()}
        bad_tables = set()

        for edge in edges:
            src = edge["source"]
            tgt = edge["target"]
            src_id = table_to_id.get(src)
            tgt_id = table_to_id.get(tgt)
            if src_id and tgt_id:
                src_end = loading_times.get(src_id, {}).get("end")
                tgt_start = loading_times.get(tgt_id, {}).get("start")
                if src_end and tgt_start and src_end > tgt_start:
                    bad_tables.add(tgt)

        result = []
        for row in rows:
            table_name = id_to_table.get(row["object_id"], str(row["object_id"]))
            result.append(
                {
                    "table_id": row["object_id"],
                    "table_name": table_name,
                    "start": row["loading_start_dttm"].strftime("%Y-%m-%d %H:%M:%S") if row["loading_start_dttm"] else None,
                    "end": row["loading_finish_dttm"].strftime("%Y-%m-%d %H:%M:%S") if row["loading_finish_dttm"] else None,
                    "is_bad": table_name in bad_tables,
                }
            )

        return JSONResponse(content=result, media_type="application/json")

    except Exception as e:
        print("❌ Ошибка при построении Gantt:", str(e))
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/entities/{entity_id}/table-info")
def get_entity_table_info(entity_id: int):
    """
    Возвращает информацию по таблице из tech_etl.tables_meta для конкретной сущности.
    Поля: schema_name, tables_name, last_load, entity_name
    """
    sql = f"""
        WITH entity_tables AS (
            SELECT
                tm.table_id,
                tm.table_schema,
                tm.table_name,
                tm.table_last_load,
                tm.entity_name
            FROM {TABLE_TABLES_META} tm
            WHERE tm.entity_id = :entity_id
        ),
        loading_ranked AS (
            SELECT
                l.object_id,
                l.loading_state,
                CASE
                    WHEN l.loading_start_dttm IS NOT NULL AND l.loading_finish_dttm IS NOT NULL
                        THEN ROUND(CAST(EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS numeric), 2)
                    ELSE NULL
                END AS duration_minutes,
                ROW_NUMBER() OVER (
                    PARTITION BY l.object_id
                    ORDER BY l.loading_finish_dttm DESC NULLS LAST, l.loading_start_dttm DESC NULLS LAST
                ) AS rn_all,
                ROW_NUMBER() OVER (
                    PARTITION BY l.object_id
                    ORDER BY
                        CASE
                            WHEN l.loading_start_dttm IS NOT NULL AND l.loading_finish_dttm IS NOT NULL
                                THEN l.loading_finish_dttm
                        END DESC NULLS LAST,
                        CASE
                            WHEN l.loading_start_dttm IS NOT NULL AND l.loading_finish_dttm IS NOT NULL
                                THEN l.loading_start_dttm
                        END DESC NULLS LAST
                ) AS rn_duration
            FROM {TABLE_LOADING_HISTORY} l
            INNER JOIN entity_tables tm
                ON tm.table_id = l.object_id
            WHERE l.object_type = 'table'
        )
        SELECT
            tm.table_id,
            tm.table_schema,
            tm.table_name,
            tm.table_last_load,
            tm.entity_name,
            MAX(CASE WHEN lr.rn_all = 1 THEN lr.loading_state END) AS last_loading_state,
            MAX(CASE WHEN lr.rn_duration = 1 THEN lr.duration_minutes END) AS current_duration_minutes,
            MAX(CASE WHEN lr.rn_duration = 2 THEN lr.duration_minutes END) AS previous_duration_minutes
        FROM entity_tables tm
        LEFT JOIN loading_ranked lr
            ON lr.object_id = tm.table_id
        GROUP BY
            tm.table_id,
            tm.table_schema,
            tm.table_name,
            tm.table_last_load,
            tm.entity_name
        ORDER BY tm.table_last_load
    """
    try:
        with engine.connect() as conn:
            rows = conn.execute(text(sql), {"entity_id": entity_id}).mappings().all()
            cleaned = []
            for r in rows:
                row = dict(r)
                row["table_schema"] = row["table_schema"]
                row["table_last_load"] = (
                    row["table_last_load"].strftime("%Y-%m-%d %H:%M:%S") if row["table_last_load"] else None
                )
                row["table_name"] = row["table_name"]
                row["entity_name"] = row["entity_name"]
                row["last_loading_state"] = row.get("last_loading_state")
                row["current_duration_minutes"] = (
                    float(row["current_duration_minutes"]) if row.get("current_duration_minutes") is not None else None
                )
                row["previous_duration_minutes"] = (
                    float(row["previous_duration_minutes"]) if row.get("previous_duration_minutes") is not None else None
                )
                cleaned.append(row)

            return JSONResponse(content=cleaned, media_type="application/json; charset=utf-8")

    except Exception as e:
        print("❌ Ошибка при получении данных об ошибках:", str(e))
        return JSONResponse(status_code=500, content={"error": str(e)})





def get_table_id_by_fqn(conn, schema: str, table: str):
    q = text(
        f"""
        SELECT table_id
        FROM {TABLE_TABLES_META}
        WHERE table_schema = :schema
          AND table_name   = :table
        LIMIT 1
    """
    )
    return conn.execute(q, {"schema": schema, "table": table}).scalar()


def build_impact(table_fqn: str, deps: list, sla_resp):
    """
    Строит impact для инцидента:
    - затронутые сущности
    - количество downstream-таблиц
    - отчёты под риском
    - SLA-нарушения
    """

    affected_entities = set()
    blocked_tables = set()
    reports_at_risk = []
    sla_violations = 0

    # --- downstream tables ---
    for d in deps or []:
        schema = d.get("schema")
        table = d.get("table_name") or d.get("table")
        entity = d.get("entity_name")

        if schema and table:
            blocked_tables.add(f"{schema}.{table}")

        if entity:
            affected_entities.add(entity)

    # --- SLA ---
    try:
        if hasattr(sla_resp, "body"):
            sla_rows = json.loads(sla_resp.body.decode("utf-8"))
        else:
            sla_rows = sla_resp or []
    except Exception:
        sla_rows = []

    for row in sla_rows:
        tables = row.get("tables_info", [])
        for t in tables:
            if not t.get("sla_ok", True):
                sla_violations += 1
                reports_at_risk.append(row.get("report"))
                break

    return {
        "affected_entities": sorted(affected_entities),
        "blocked_tables_count": len(blocked_tables),
        "reports_at_risk": sorted(set(reports_at_risk)),
        "sla_violations": sla_violations,
    }

@router.get("/api/incident")
def get_incident(table_fqn: str = Query(...)):
    try:
        schema, table = table_fqn.split(".")
    except ValueError:
        return JSONResponse(status_code=400, content={"error": "schema.table expected"})

    with engine.connect() as conn:
        table_id = conn.execute(
            text(f"""
                SELECT table_id FROM {TABLE_TABLES_META}
                WHERE table_schema=:s AND table_name=:t
            """),
            {"s": schema, "t": table}
        ).scalar()

        if not table_id:
            return JSONResponse(status_code=404, content={"error": "table not found"})

        deps = [d.dict() for d in resolve_dependencies(schema, table)]

        rows = conn.execute(
            text(f"""
                SELECT loading_start_dttm, loading_finish_dttm, loading_state, message,
                       extract(epoch from (loading_finish_dttm-loading_start_dttm)) dur
                FROM {TABLE_LOADING_HISTORY}
                WHERE object_id=:id
                ORDER BY loading_finish_dttm DESC
                LIMIT 15
            """),
            {"id": table_id}
        ).mappings().all()

    timeline = [{
        "start": r["loading_start_dttm"].strftime("%Y-%m-%d %H:%M:%S"),
        "finish": r["loading_finish_dttm"].strftime("%Y-%m-%d %H:%M:%S"),
        "state": r["loading_state"],
        "duration_sec": float(r["dur"]) if r["dur"] else None,
        "message": (r["message"] or "") or None
    } for r in rows]

    return {
        "summary": {
            "table_fqn": table_fqn,
            "state": "FAILING" if timeline and timeline[0]["state"] == "FAILED" else "RECOVERED",
        },
        "timeline": timeline,
        "dependencies": deps,
        "impact": {
            "blocked_tables_count": len(deps),
            "affected_entities": sorted({d["entity_name"] for d in deps if d.get("entity_name")}),
        }
    }


from collections import defaultdict
from datetime import timedelta

INCIDENT_WINDOW_MIN = 60  # минут


def group_failures(failures: list):
    incidents = []
    failures_sorted = sorted(
        [f for f in failures if f.get("error_time")],
        key=lambda x: x["error_time"],
        reverse=True
    )

    used = set()

    for i, f in enumerate(failures_sorted):
        if i in used:
            continue

        entity = f.get("entity_name")

        if not entity:
            all_meta, _ = get_cached_meta_and_index()
            meta = next(
                (m for m in all_meta if m["table_schema"] == f["schema"] and m["table_name"] == f["table_name"]),
                None,
            )
            entity = meta.get("entity_name") if meta else None

        entity = entity or f"{f['schema']}"
        f["entity_name"] = entity  #

        try:
            t0 = datetime.strptime(f["error_time"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue

        group = [f]
        used.add(i)

        for j, other in enumerate(failures_sorted[i + 1 :], start=i + 1):
            if j in used:
                continue

            other_entity = other.get("entity_name")

            if not other_entity:
                all_meta, _ = get_cached_meta_and_index()
                meta = next(
                    (
                        m
                        for m in all_meta
                        if m["table_schema"] == other["schema"] and m["table_name"] == other["table_name"]
                    ),
                    None,
                )
                other_entity = meta.get("entity_name") if meta else None
                other_entity = other_entity or f"{other['schema']}"
                other["entity_name"] = other_entity

            if other_entity != entity:
                continue

            t1 = datetime.strptime(other["error_time"], "%Y-%m-%d %H:%M:%S")

            if abs((t0 - t1).total_seconds()) <= INCIDENT_WINDOW_MIN * 60:
                group.append(other)
                used.add(j)

        incidents.append(group)

    return incidents


from collections import defaultdict


@router.get("/api/incidents/active")
def get_active_incidents():
    query = f"""
        WITH latest AS (
            SELECT
                l.object_id,
                MAX(l.loading_finish_dttm) AS last_event_time
            FROM {TABLE_LOADING_HISTORY} l
            WHERE l.object_type = 'table'
            GROUP BY l.object_id
        ),
        last_state AS (
            SELECT
                l.object_id,
                l.loading_state,
                l.loading_finish_dttm
            FROM {TABLE_LOADING_HISTORY} l
            JOIN latest t
              ON t.object_id = l.object_id
             AND t.last_event_time = l.loading_finish_dttm
            WHERE l.object_type = 'table'
        )
        SELECT
            t.table_schema AS schema,
            t.table_name,
            e.entity_name,
            l.loading_finish_dttm AS error_time
        FROM last_state l
        JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
        LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
        WHERE l.loading_state = 'FAILED'
        ORDER BY l.loading_finish_dttm DESC
    """

    try:
        with engine.connect() as conn:
            rows = conn.execute(text(query)).mappings().all()

        failures = []
        for r in rows:
            row = dict(r)
            row["error_time"] = serialize_datetime(row.get("error_time"))
            failures.append(row)

        if not failures:
            return []

        grouped = group_failures(failures)
        by_entity = defaultdict(list)

        for group in grouped:
            if not group:
                continue
            entity = group[0].get("entity_name") or "UNKNOWN"
            by_entity[entity].extend(group)

        incidents = []
        for entity, group_rows in by_entity.items():
            failed_tables = set()
            last_failure = None
            for r in group_rows:
                table_fqn = f"{r['schema']}.{r['table_name']}"
                failed_tables.add(table_fqn)
                if not last_failure or r["error_time"] > last_failure:
                    last_failure = r["error_time"]

            incidents.append(
                {
                    "entity": entity,
                    "severity": "CRITICAL",
                    "failed_tables": len(failed_tables),
                    "root_tables": sorted(failed_tables)[:3],
                    "last_failure_time": last_failure,
                }
            )

        incidents.sort(key=lambda x: x["last_failure_time"], reverse=True)
        return incidents
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/incidents/history")
def get_incident_history(
    days: int = Query(300, ge=1, le=3650),
    limit: int = Query(10, ge=1, le=100),
):
    query = f"""
        SELECT
            COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) || '.' ||
            COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_fqn,
            COUNT(*) AS incidents_count,
            MAX(l.loading_finish_dttm) AS last_incident
     FROM {TABLE_LOADING_HISTORY} l
      left join {TABLE_TABLES_META} t on t.table_id=l.object_id
        WHERE l.loading_state = 'FAILED'
          AND l.object_type = 'table'
          AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
        GROUP BY
            COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')),
            COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name)
        ORDER BY incidents_count DESC
        LIMIT :limit
    """
    with engine.connect() as conn:
        rows = conn.execute(text(query), {"days": days, "limit": limit}).mappings().all()

    return [
        {
            "table": r["table_fqn"],
            "count": r["incidents_count"],
            "last_incident": r["last_incident"].strftime("%Y-%m-%d %H:%M:%S")
            if r["last_incident"] else None
        }
        for r in rows
    ]


@router.get("/api/incidents/timeline")
def get_incident_timeline(days: int = Query(7, ge=1, le=365)):
    query = f"""
        SELECT
            date_trunc('day', l.loading_finish_dttm) AS day,
            COUNT(*) AS incidents_count
        FROM {TABLE_LOADING_HISTORY} l
        WHERE l.loading_state = 'FAILED'
          AND l.object_type = 'table'
          AND l.loading_finish_dttm >= now() - (:days || ' days')::interval
        GROUP BY date_trunc('day', l.loading_finish_dttm)
        ORDER BY day
    """
    with engine.connect() as conn:
        rows = conn.execute(text(query), {"days": days}).mappings().all()

    return [
        {
            "day": row["day"].strftime("%Y-%m-%d"),
            "count": int(row["incidents_count"] or 0),
        }
        for row in rows
        if row.get("day")
    ]
print("Reg")
@router.get("/api/orderbreaches")
def get_order_breaches():
    return get_cached_order_breaches()


@router.get("/api/click/summary")
def get_clickhouse_summary(
    days: int = Query(7, ge=1, le=365),
    limit: int = Query(8, ge=1, le=50),
):
    try:
        with engine.connect() as conn:
            summary_row = conn.execute(
                text(
                    _clickhouse_run_agg_cte(
                        run_filter_sql="AND r.start_dttm >= now() - (:days || ' days')::interval"
                    )
                    + """
                    SELECT
                        COUNT(*) AS total_runs,
                        COUNT(DISTINCT schema_name || '.' || table_name) AS total_tables,
                        COUNT(DISTINCT CASE WHEN status = 'SUCCESS' THEN schema_name || '.' || table_name END) AS ok_tables,
                        SUM(CASE WHEN status = 'SUCCESS' THEN 1 ELSE 0 END) AS ok_runs,
                        SUM(CASE WHEN status = 'FAILED' THEN 1 ELSE 0 END) AS failed_runs,
                        SUM(CASE WHEN status = 'RUNNING' THEN 1 ELSE 0 END) AS running_runs,
                        SUM(CASE WHEN status = 'UP_FOR_RETRY' THEN 1 ELSE 0 END) AS retry_runs,
                        AVG(actual_duration_seconds) AS avg_actual_seconds,
                        MAX(actual_duration_seconds) AS max_actual_seconds,
                        AVG(GREATEST(elapsed_duration_seconds - actual_duration_seconds, 0)) AS avg_lag_seconds,
                        MAX(GREATEST(elapsed_duration_seconds - actual_duration_seconds, 0)) AS max_lag_seconds,
                        MAX(end_dttm) AS last_finish
                    FROM run_agg
                    """
                ),
                {"days": days},
            ).mappings().first() or {}

            failures = conn.execute(
                text(
                    _clickhouse_run_agg_cte(
                        run_filter_sql="AND r.start_dttm >= now() - (:days || ' days')::interval"
                    )
                    + f""",
                    last_stage AS (
                        SELECT DISTINCT ON (s.run_uuid, s.table_name)
                            s.run_uuid,
                            s.table_name,
                            s.stage_name,
                            s.status AS stage_status,
                            s.error_text AS stage_error
                        FROM {TABLE_CLICK_LOAD_STAGE} s
                        WHERE s.stage_name IN ('UPLOAD_TO_S3', 'CLICKHOUSE_LOAD')
                        ORDER BY s.run_uuid, s.table_name, s.start_dttm DESC NULLS LAST
                    )
                    SELECT
                        r.run_uuid,
                        r.schema_name,
                        r.table_name,
                        r.dag_name,
                        r.dag_run,
                        r.start_dttm,
                        r.end_dttm,
                        r.actual_duration_seconds,
                        r.elapsed_duration_seconds,
                        r.status,
                        r.error_text,
                        st.stage_name,
                        st.stage_status,
                        st.stage_error
                    FROM run_agg r
                    LEFT JOIN last_stage st
                      ON st.run_uuid = r.run_uuid
                     AND st.table_name = r.table_name
                    WHERE r.status IN ('FAILED', 'UP_FOR_RETRY')
                    ORDER BY r.start_dttm DESC
                    LIMIT :limit
                    """
                ),
                {"days": days, "limit": limit},
            ).mappings().all()

        summary = {
            "total_runs": int(summary_row.get("total_runs") or 0),
            "total_tables": int(summary_row.get("total_tables") or 0),
            "ok_tables": int(summary_row.get("ok_tables") or 0),
            "ok_runs": int(summary_row.get("ok_runs") or 0),
            "failed_runs": int(summary_row.get("failed_runs") or 0),
            "running_runs": int(summary_row.get("running_runs") or 0),
            "retry_runs": int(summary_row.get("retry_runs") or 0),
            "avg_duration_min": _round_minutes((summary_row.get("avg_actual_seconds") or 0) / 60.0)
            if summary_row.get("avg_actual_seconds") is not None
            else None,
            "max_duration_min": _round_minutes((summary_row.get("max_actual_seconds") or 0) / 60.0)
            if summary_row.get("max_actual_seconds") is not None
            else None,
            "avg_lag_min": _round_minutes((summary_row.get("avg_lag_seconds") or 0) / 60.0)
            if summary_row.get("avg_lag_seconds") is not None
            else None,
            "max_lag_min": _round_minutes((summary_row.get("max_lag_seconds") or 0) / 60.0)
            if summary_row.get("max_lag_seconds") is not None
            else None,
            "last_finish": serialize_datetime(summary_row.get("last_finish")),
        }

        failure_rows = []
        for row in failures:
            stage_name = row.get("stage_name")
            stage_label = None
            if stage_name == "UPLOAD_TO_S3":
                stage_label = "S3"
            elif stage_name == "CLICKHOUSE_LOAD":
                stage_label = "ClickHouse"

            failure_rows.append(
                {
                    "run_uuid": row.get("run_uuid"),
                    "schema_name": row.get("schema_name"),
                    "table_name": row.get("table_name"),
                    "dag_name": row.get("dag_name"),
                    "dag_run": row.get("dag_run"),
                    "start_dttm": serialize_datetime(row.get("start_dttm")),
                    "end_dttm": serialize_datetime(row.get("end_dttm")),
                    "status": row.get("status"),
                    "error_text": row.get("error_text"),
                    "stage_name": stage_name,
                    "stage_status": row.get("stage_status"),
                    "stage_error": row.get("stage_error"),
                    "problem_area": stage_label,
                    **_clickhouse_run_metrics(row),
                }
            )

        return {"summary": summary, "failures": failure_rows}
    except Exception as e:
        print("❌ /api/click/summary error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/click/table/{schema}/{table:path}")
def get_clickhouse_table_runs(
    schema: str,
    table: str,
    table_id: Optional[int] = None,
    limit: int = Query(6, ge=1, le=50),
):
    try:
        schema_norm = (schema or "").strip()
        table_norm = (table or "").strip()
        table_clean = _clean_table_name(norm(table))
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    _clickhouse_run_agg_cte(
                        run_filter_sql="AND r.schema_name = :schema",
                        stage_filter_sql="""
                          AND (lower(s.table_name) = lower(:table) OR lower(s.table_name) = lower(:table_clean))
                          AND (:table_id IS NULL OR s.table_id = :table_id)
                        """,
                    )
                    + """
                    SELECT *
                    FROM run_agg
                    ORDER BY start_dttm DESC
                    LIMIT :limit
                    """
                ),
                {"schema": schema_norm, "table": table_norm, "table_clean": table_clean, "table_id": table_id, "limit": limit},
            ).mappings().all()

            runs = [
                {
                    "run_uuid": row.get("run_uuid"),
                    "dag_name": row.get("dag_name"),
                    "dag_run": row.get("dag_run"),
                    "start_dttm": serialize_datetime(row.get("start_dttm")),
                    "end_dttm": serialize_datetime(row.get("end_dttm")),
                    "status": row.get("status"),
                    "error_text": row.get("error_text"),
                    **_clickhouse_run_metrics(row),
                }
                for row in rows
            ]

            stages = []
            if runs:
                run_uuid = runs[0].get("run_uuid")
                stages_rows = conn.execute(
                    text(
                        f"""
                        SELECT
                            table_id,
                            table_name,
                            stage_name,
                            start_dttm,
                            end_dttm,
                            duration,
                            status,
                            error_text
                        FROM {TABLE_CLICK_LOAD_STAGE}
                        WHERE run_uuid = :run_uuid
                          AND (lower(table_name) = lower(:table_name) OR lower(table_name) = lower(:table_name_clean))
                          AND stage_name IN ('UPLOAD_TO_S3', 'CLICKHOUSE_LOAD')
                        ORDER BY start_dttm
                        """
                    ),
                    {"run_uuid": run_uuid, "table_name": table_norm, "table_name_clean": table_clean},
                ).mappings().all()
                stages = [
                    {
                        "table_id": row.get("table_id"),
                        "table_name": row.get("table_name"),
                        "stage_name": row.get("stage_name"),
                        "start_dttm": serialize_datetime(row.get("start_dttm")),
                        "end_dttm": serialize_datetime(row.get("end_dttm")),
                        "duration_min": _duration_minutes(row.get("duration")),
                        "status": row.get("status"),
                        "error_text": row.get("error_text"),
                    }
                    for row in stages_rows
                ]

        return {"runs": runs, "stages": stages}
    except Exception as e:
        print("❌ /api/click/table error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/click/meta/{schema}/{table:path}")
def get_clickhouse_meta(schema: str, table: str):
    try:
        schema_norm = (schema or "").strip().lower()
        table_norm = (table or "").strip().lower()
        table_clean = _clean_table_name(table_norm)
        idx = get_click_meta_index()
        meta = idx["meta"].get((schema_norm, table_norm)) or idx["meta"].get((schema_norm, table_clean))
        view_sql = idx["view_sql"].get((schema_norm, table_norm)) or idx["view_sql"].get((schema_norm, table_clean))
        if not meta and not view_sql:
            return {
                "meta": None,
                "view_sql": None,
                "meta_root": idx.get("root"),
            }
        return {
            "meta": meta,
            "view_sql": view_sql,
            "meta_root": idx.get("root"),
        }
    except Exception as e:
        print("❌ /api/click/meta error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/click/view/search")
def search_clickhouse_view(schema: str, table: str, limit: int = Query(10, ge=1, le=50)):
    try:
        schema_norm = (schema or "").strip().lower()
        table_norm = _clean_table_name((table or "").strip().lower())
        idx = get_click_meta_index()
        matches = idx.get("view_refs", {}).get((schema_norm, table_norm), [])

        # also match by view name
        name_key = ("dm_view", table_norm)
        if name_key in idx.get("view_sql", {}):
            matches = matches + [{"view_schema": "dm_view", "view_name": table_norm, "reason": "name_match"}]

        # de-dup
        seen = set()
        uniq = []
        for item in matches:
            key = (item.get("view_schema"), item.get("view_name"))
            if key in seen:
                continue
            seen.add(key)
            uniq.append(item)
            if len(uniq) >= limit:
                break
        return {"matches": uniq}
    except Exception as e:
        print("❌ /api/click/view/search error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/click/history/{schema}/{table:path}")
def get_clickhouse_history(
    schema: str,
    table: str,
    table_id: Optional[int] = None,
    limit: int = Query(20, ge=1, le=200),
):
    try:
        schema_norm = (schema or "").strip()
        table_norm = (table or "").strip()
        table_clean = _clean_table_name(norm(table))
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    _clickhouse_run_agg_cte(
                        run_filter_sql="AND r.schema_name = :schema",
                        stage_filter_sql="""
                          AND (lower(s.table_name) = lower(:table) OR lower(s.table_name) = lower(:table_clean))
                          AND (:table_id IS NULL OR s.table_id = :table_id)
                        """,
                    )
                    + """
                    SELECT *
                    FROM run_agg
                    ORDER BY start_dttm DESC NULLS LAST
                    LIMIT :limit
                    """
                ),
                {"schema": schema_norm, "table": table_norm, "table_clean": table_clean, "table_id": table_id, "limit": limit},
            ).mappings().all()

        history = [
            {
                "run_uuid": row.get("run_uuid"),
                "stage_name": "TOTAL",
                "start_dttm": serialize_datetime(row.get("start_dttm")),
                "end_dttm": serialize_datetime(row.get("end_dttm")),
                "status": row.get("status"),
                "dag_name": row.get("dag_name"),
                "dag_run": row.get("dag_run"),
                **_clickhouse_run_metrics(row),
            }
            for row in rows
        ]
        return history
    except Exception as e:
        print("❌ /api/click/history error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/click/slow-stages")
def get_clickhouse_slow_stages(days: int = Query(7, ge=1, le=365), limit: int = Query(10, ge=1, le=50)):
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    _clickhouse_run_agg_cte(
                        run_filter_sql="AND r.start_dttm >= now() - (:days || ' days')::interval"
                    )
                    + """
                    SELECT
                        run_uuid,
                        schema_name,
                        table_name,
                        dag_name,
                        dag_run,
                        start_dttm,
                        end_dttm,
                        status,
                        actual_duration_seconds,
                        elapsed_duration_seconds
                    FROM run_agg
                    ORDER BY actual_duration_seconds DESC NULLS LAST
                    LIMIT :limit
                    """
                ),
                {"days": days, "limit": limit},
            ).mappings().all()

        result = [
            {
                "run_uuid": row.get("run_uuid"),
                "schema_name": row.get("schema_name"),
                "table_name": row.get("table_name"),
                "start_dttm": serialize_datetime(row.get("start_dttm")),
                "end_dttm": serialize_datetime(row.get("end_dttm")),
                "status": row.get("status"),
                "dag_name": row.get("dag_name"),
                "dag_run": row.get("dag_run"),
                **_clickhouse_run_metrics(row),
            }
            for row in rows
        ]
        return result
    except Exception as e:
        print("❌ /api/click/slow-stages error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/window-runs")
def get_window_runs(
    date: str,
    time_from: str = Query(..., alias="from"),
    time_to: str = Query(..., alias="to"),
    source: str = "both",
):
    """
    Returns runs for GP and/or ClickHouse inside a local-time window.
    date: YYYY-MM-DD
    from/to: HH:MM (local)
    source: gp | click | both
    """
    try:
        window_start = f"{date} {time_from}:00"
        window_end = f"{date} {time_to}:00"
        payload = {"window_start": window_start, "window_end": window_end, "timezone": "local"}

        with engine.connect() as conn:
            if source in ("gp", "both"):
                gp_rows = conn.execute(
                    text(
                        f"""
                        SELECT
                            COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS schema_name,
                            COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                            e.entity_name,
                            l.loading_start_dttm AS start_dttm,
                            l.loading_finish_dttm AS end_dttm,
                            l.loading_state AS status,
                            EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration_min
                        FROM {TABLE_LOADING_HISTORY} l
                        LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
                        LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                        WHERE l.object_type = 'table'
                          AND l.loading_start_dttm >= :window_start
                          AND l.loading_start_dttm <= :window_end
                        ORDER BY l.loading_start_dttm ASC
                        """
                    ),
                    {"window_start": window_start, "window_end": window_end},
                ).mappings().all()
                payload["gp"] = [
                    {
                        "schema_name": row.get("schema_name"),
                        "table_name": row.get("table_name"),
                        "entity_name": row.get("entity_name"),
                        "start_dttm": serialize_datetime(row.get("start_dttm")),
                        "end_dttm": serialize_datetime(row.get("end_dttm")),
                        "duration_min": round(float(row.get("duration_min") or 0), 2) if row.get("duration_min") is not None else None,
                        "status": row.get("status"),
                    }
                    for row in gp_rows
                ]

            if source in ("click", "both"):
                click_rows = conn.execute(
                    text(
                        _clickhouse_run_agg_cte()
                        + f"""
                        SELECT
                            ra.run_uuid,
                            ra.schema_name,
                            ra.table_name,
                            tm.entity_name,
                            ra.start_dttm,
                            ra.end_dttm,
                            ra.status,
                            ra.actual_duration_seconds,
                            ra.elapsed_duration_seconds
                        FROM run_agg ra
                        LEFT JOIN {TABLE_TABLES_META} tm
                          ON tm.table_id = ra.table_id
                        WHERE ra.start_dttm >= :window_start
                          AND ra.start_dttm <= :window_end
                        ORDER BY ra.start_dttm ASC
                        """
                    ),
                    {"window_start": window_start, "window_end": window_end},
                ).mappings().all()
                payload["click"] = [
                    {
                        "run_uuid": row.get("run_uuid"),
                        "schema_name": row.get("schema_name"),
                        "table_name": row.get("table_name"),
                        "entity_name": row.get("entity_name"),
                        "start_dttm": serialize_datetime(row.get("start_dttm")),
                        "end_dttm": serialize_datetime(row.get("end_dttm")),
                        "status": row.get("status"),
                        **_clickhouse_run_metrics(row),
                    }
                    for row in click_rows
                ]

        return payload
    except Exception as e:
        print("❌ /api/window-runs error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/api/load-compare")
def get_load_compare(
    date_a: str = Query(..., alias="date_a"),
    date_b: str = Query(..., alias="date_b"),
    entity_id: Optional[int] = Query(None, ge=1),
):
    try:
        start_a = f"{date_a} 00:00:00"
        end_a = f"{date_a} 23:59:59"
        start_b = f"{date_b} 00:00:00"
        end_b = f"{date_b} 23:59:59"

        query = f"""
            WITH base AS (
                SELECT
                    l.object_id AS table_id,
                    COALESCE(t.table_schema, NULLIF(split_part(l.object_name, '.', 1), '')) AS table_schema,
                    COALESCE(t.table_name, NULLIF(split_part(l.object_name, '.', 2), ''), l.object_name) AS table_name,
                    e.entity_id,
                    COALESCE(e.entity_name, 'UNKNOWN') AS entity_name,
                    l.loading_start_dttm,
                    l.loading_finish_dttm,
                    EXTRACT(EPOCH FROM (l.loading_finish_dttm - l.loading_start_dttm)) / 60.0 AS duration_minutes
                FROM {TABLE_LOADING_HISTORY} l
                LEFT JOIN {TABLE_TABLES_META} t ON t.table_id = l.object_id
                LEFT JOIN {TABLE_ENTITIES_META} e ON e.entity_id = t.entity_id
                WHERE l.object_type = 'table'
                  AND l.loading_state IN ('SUCCESS', 'LOADED')
                  AND l.loading_start_dttm IS NOT NULL
                  AND l.loading_finish_dttm IS NOT NULL
                  AND (
                    (l.loading_start_dttm >= :start_a AND l.loading_start_dttm <= :end_a)
                    OR
                    (l.loading_start_dttm >= :start_b AND l.loading_start_dttm <= :end_b)
                  )
                  AND (:entity_id IS NULL OR e.entity_id = :entity_id)
            ),
            ranked_a AS (
                SELECT
                    *,
                    ROW_NUMBER() OVER (
                        PARTITION BY table_schema, table_name
                        ORDER BY loading_start_dttm DESC NULLS LAST
                    ) AS rn
                FROM base
                WHERE loading_start_dttm >= :start_a AND loading_start_dttm <= :end_a
            ),
            ranked_b AS (
                SELECT
                    *,
                    ROW_NUMBER() OVER (
                        PARTITION BY table_schema, table_name
                        ORDER BY loading_start_dttm DESC NULLS LAST
                    ) AS rn
                FROM base
                WHERE loading_start_dttm >= :start_b AND loading_start_dttm <= :end_b
            ),
            latest_a AS (
                SELECT * FROM ranked_a WHERE rn = 1
            ),
            latest_b AS (
                SELECT * FROM ranked_b WHERE rn = 1
            )
            SELECT
                COALESCE(a.table_id, b.table_id) AS table_id,
                COALESCE(a.table_schema, b.table_schema) AS table_schema,
                COALESCE(a.table_name, b.table_name) AS table_name,
                COALESCE(a.entity_id, b.entity_id) AS entity_id,
                COALESCE(a.entity_name, b.entity_name) AS entity_name,
                a.duration_minutes AS duration_a,
                a.loading_start_dttm AS start_a,
                a.loading_finish_dttm AS end_a,
                b.duration_minutes AS duration_b,
                b.loading_start_dttm AS start_b,
                b.loading_finish_dttm AS end_b
            FROM latest_a a
            FULL OUTER JOIN latest_b b
              ON a.table_schema = b.table_schema
             AND a.table_name = b.table_name
            ORDER BY ABS(COALESCE(b.duration_minutes, 0) - COALESCE(a.duration_minutes, 0)) DESC NULLS LAST,
                     COALESCE(a.table_schema, b.table_schema),
                     COALESCE(a.table_name, b.table_name)
        """

        params = {
            "start_a": start_a,
            "end_a": end_a,
            "start_b": start_b,
            "end_b": end_b,
            "entity_id": entity_id,
        }

        with engine.connect() as conn:
            rows = conn.execute(text(query), params).mappings().all()

        result_rows = []
        for row in rows:
            duration_a = float(row.get("duration_a")) if row.get("duration_a") is not None else None
            duration_b = float(row.get("duration_b")) if row.get("duration_b") is not None else None
            delta_minutes = None
            delta_pct = None
            if duration_a is not None and duration_b is not None:
                delta_minutes = round(duration_b - duration_a, 2)
                if duration_a:
                    delta_pct = round(((duration_b - duration_a) / duration_a) * 100, 2)

            result_rows.append(
                {
                    "table_id": row.get("table_id"),
                    "table_fqn": f"{row.get('table_schema')}.{row.get('table_name')}".strip("."),
                    "table_schema": row.get("table_schema"),
                    "table_name": row.get("table_name"),
                    "entity_id": row.get("entity_id"),
                    "entity_name": row.get("entity_name"),
                    "duration_a": round(duration_a, 2) if duration_a is not None else None,
                    "duration_b": round(duration_b, 2) if duration_b is not None else None,
                    "delta_minutes": delta_minutes,
                    "delta_pct": delta_pct,
                    "start_a": serialize_datetime(row.get("start_a")),
                    "end_a": serialize_datetime(row.get("end_a")),
                    "start_b": serialize_datetime(row.get("start_b")),
                    "end_b": serialize_datetime(row.get("end_b")),
                }
            )

        return {
            "date_a": date_a,
            "date_b": date_b,
            "entity_id": entity_id,
            "rows": result_rows,
        }
    except Exception as e:
        print("❌ /api/load-compare error:", e)
        print(traceback.format_exc())
        return JSONResponse(status_code=500, content={"error": str(e)})


def _build_ytrack_link(task_id: Optional[str]) -> Optional[str]:
    if not task_id:
        return None
    base = (YTRACK_ISSUE_URL or "").strip()
    if not base:
        return None
    return base.replace("{task_id}", task_id).replace("{id}", task_id)


def _load_pdf_font(size: int, bold: bool = False):
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    ]
    for path in candidates:
        if path and os.path.exists(path):
            try:
                return ImageFont.truetype(path, size=size)
            except Exception:
                continue
    return ImageFont.load_default()


def _pdf_wrap_text(draw, text_value: Any, font, max_width: int) -> list[str]:
    text = str(text_value or "").strip()
    if not text:
        return ["—"]
    lines: list[str] = []
    for paragraph in text.splitlines() or [""]:
        words = paragraph.split()
        if not words:
            lines.append("")
            continue
        current = words[0]
        for word in words[1:]:
            candidate = f"{current} {word}"
            bbox = draw.textbbox((0, 0), candidate, font=font)
            if bbox[2] - bbox[0] <= max_width:
                current = candidate
            else:
                lines.append(current)
                current = word
        lines.append(current)
    return lines or ["—"]


def _draw_pdf_text_block(draw, x: int, y: int, text_value: Any, font, fill: str, max_width: int, line_gap: int = 5):
    lines = _pdf_wrap_text(draw, text_value, font, max_width)
    line_height = font.size + line_gap if hasattr(font, "size") else 20
    for index, line in enumerate(lines):
        draw.text((x, y + index * line_height), line, font=font, fill=fill)
    return y + max(len(lines), 1) * line_height


def _render_report_pdf_raster(report: dict[str, Any]) -> bytes:
    page_width, page_height = 2480, 3508
    margin_x, margin_y = 132, 132
    content_width = page_width - margin_x * 2
    background = "#08111f"
    surface = "#0f172a"
    surface_alt = "#111c31"
    line_color = "#243247"
    text = "#f8fbff"
    muted = "#93a4bd"
    accent = "#38bdf8"

    title_font = _load_pdf_font(58, bold=True)
    subtitle_font = _load_pdf_font(30)
    section_font = _load_pdf_font(38, bold=True)
    body_font = _load_pdf_font(26)
    small_font = _load_pdf_font(22)
    kpi_value_font = _load_pdf_font(42, bold=True)
    card_value_font = _load_pdf_font(34, bold=True)

    pages: list[Image.Image] = []

    def new_page():
        image = Image.new("RGB", (page_width, page_height), background)
        draw = ImageDraw.Draw(image)
        return image, draw, margin_y

    image, draw, current_y = new_page()

    def ensure_space(required_height: int):
        nonlocal image, draw, current_y
        if current_y + required_height <= page_height - margin_y:
            return
        pages.append(image)
        image, draw, current_y = new_page()

    def draw_title_block():
        nonlocal current_y
        current_y = _draw_pdf_text_block(draw, margin_x, current_y, report.get("title") or "Отчёт", title_font, text, content_width)
        current_y += 6
        current_y = _draw_pdf_text_block(draw, margin_x, current_y, report.get("subtitle") or "", subtitle_font, muted, content_width)
        current_y += 20

    def draw_kpis(items: list[dict[str, Any]], section_title: Optional[str]):
        nonlocal current_y
        ensure_space(420)
        if section_title:
            current_y = _draw_pdf_text_block(draw, margin_x, current_y, section_title, section_font, text, content_width)
            current_y += 12
        columns = 3 if len(items) >= 6 else 2
        gap = 28
        card_width = (content_width - gap) // columns
        card_height = 208
        for index, item in enumerate(items):
            if index and index % columns == 0:
                current_y += card_height + gap
                ensure_space(card_height + 40)
            col = index % columns
            x = margin_x + col * (card_width + gap)
            y = current_y
            draw.rounded_rectangle((x, y, x + card_width, y + card_height), radius=24, fill=surface, outline=line_color, width=2)
            draw.text((x + 30, y + 26), str(item.get("label") or "Показатель"), font=small_font, fill=muted)
            draw.text((x + 30, y + 76), str(item.get("value") or "—"), font=kpi_value_font, fill=text)
            hint = str(item.get("hint") or "").strip()
            if hint:
                _draw_pdf_text_block(draw, x + 30, y + 136, hint, small_font, accent, card_width - 60, line_gap=6)
        current_y += card_height + 28

    def draw_cards(items: list[dict[str, Any]], section_title: Optional[str]):
        nonlocal current_y
        if not items:
            return
        ensure_space(360)
        if section_title:
            current_y = _draw_pdf_text_block(draw, margin_x, current_y, section_title, section_font, text, content_width)
            current_y += 12
        gap = 24
        columns = min(3, max(1, len(items)))
        card_width = (content_width - gap * (columns - 1)) // columns
        card_height = 196
        start_y = current_y
        for index, item in enumerate(items):
            col = index % columns
            row = index // columns
            x = margin_x + col * (card_width + gap)
            y = start_y + row * (card_height + gap)
            draw.rounded_rectangle((x, y, x + card_width, y + card_height), radius=24, fill=surface_alt, outline=line_color, width=2)
            draw.text((x + 28, y + 24), str(item.get("title") or "Фокус"), font=small_font, fill=muted)
            value_bottom = _draw_pdf_text_block(draw, x + 28, y + 74, item.get("value") or "—", card_value_font, text, card_width - 56, line_gap=5)
            meta_lines = item.get("meta") or []
            meta_y = value_bottom + 10
            for meta in meta_lines[:3]:
                meta_y = _draw_pdf_text_block(draw, x + 28, meta_y, meta, small_font, accent, card_width - 56, line_gap=5)
        rows = ((len(items) - 1) // columns) + 1
        current_y = start_y + rows * card_height + max(0, rows - 1) * gap + 28

    def draw_table(section: dict[str, Any]):
        nonlocal current_y
        rows = section.get("rows") or []
        columns = section.get("columns") or []
        if not columns:
            return
        title = section.get("title")
        subtitle = section.get("subtitle")
        table_body_font = _load_pdf_font(23)
        table_header_font = _load_pdf_font(21, bold=True)
        estimate = 220 + max(1, len(rows)) * 84
        ensure_space(min(max(estimate, 220), 1000))
        if title:
            current_y = _draw_pdf_text_block(draw, margin_x, current_y, title, section_font, text, content_width)
            current_y += 6
        if subtitle:
            current_y = _draw_pdf_text_block(draw, margin_x, current_y, subtitle, small_font, muted, content_width)
            current_y += 10
        table_top = current_y
        col_count = len(columns)
        col_widths = []
        if col_count == 2:
            col_widths = [int(content_width * 0.74), int(content_width * 0.26)]
        elif col_count == 4:
            col_widths = [int(content_width * 0.44), int(content_width * 0.16), int(content_width * 0.16), int(content_width * 0.24)]
        elif col_count == 5:
            col_widths = [int(content_width * 0.16), int(content_width * 0.36), int(content_width * 0.16), int(content_width * 0.16), int(content_width * 0.16)]
        else:
            flexible = content_width - 210 * (col_count - 1)
            first_col_width = max(560, flexible)
            for idx in range(col_count):
                col_widths.append(first_col_width if idx == 0 else 210)
        if title == "Источники сигналов":
            col_widths = [
                int(content_width * 0.48),
                int(content_width * 0.12),
                int(content_width * 0.16),
                int(content_width * 0.24),
            ]
        elif title == "Связи с delivery":
            col_widths = [
                int(content_width * 0.78),
                int(content_width * 0.22),
            ]
        total_width = sum(col_widths)
        if total_width > content_width:
            scale = content_width / total_width
            col_widths = [int(width * scale) for width in col_widths]
        header_height = 62
        draw.rounded_rectangle((margin_x, table_top, margin_x + content_width, table_top + header_height), radius=14, fill=surface, outline=line_color, width=2)
        x = margin_x
        for idx, col in enumerate(columns):
            draw.text((x + 18, table_top + 18), str(col), font=table_header_font, fill=accent)
            x += col_widths[idx]
        current_y = table_top + header_height
        for row_index, row in enumerate(rows[:14]):
            values = list(row)[:col_count]
            wrapped_cells = []
            row_height = 42
            for idx, value in enumerate(values):
                lines = _pdf_wrap_text(draw, value, table_body_font, col_widths[idx] - 36)
                wrapped_cells.append(lines)
                row_height = max(row_height, len(lines) * 30 + 18)
            ensure_space(row_height + 8)
            fill = surface_alt if row_index % 2 == 0 else background
            draw.rounded_rectangle((margin_x, current_y, margin_x + content_width, current_y + row_height), radius=12, fill=fill, outline=line_color, width=1)
            x = margin_x
            for idx, lines in enumerate(wrapped_cells):
                cell_y = current_y + 12
                for line in lines:
                    draw.text((x + 18, cell_y), line, font=table_body_font, fill=text)
                    cell_y += 30
                x += col_widths[idx]
            current_y += row_height + 10
        current_y += 24

    draw_title_block()
    for section in report.get("sections") or []:
        section_type = section.get("type")
        if section_type == "kpis":
            draw_kpis(section.get("items") or [], section.get("title"))
        elif section_type == "cards":
            draw_cards(section.get("items") or [], section.get("title"))
        elif section_type == "table":
            draw_table(section)

    pages.append(image)
    pdf_buffer = BytesIO()
    first_page, rest = pages[0], pages[1:]
    first_page.save(pdf_buffer, format="PDF", save_all=True, append_images=rest, resolution=300.0)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()


def _render_report_pdf(report: dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        return _render_report_pdf_raster(report)

    font_regular_path = None
    font_bold_path = None
    font_candidates = [
        (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ),
        (
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
        ),
    ]
    for regular_path, bold_path in font_candidates:
        if os.path.exists(regular_path) and os.path.exists(bold_path):
            font_regular_path = regular_path
            font_bold_path = bold_path
            break
    if not font_regular_path or not font_bold_path:
        return _render_report_pdf_raster(report)

    regular_font_name = "ReportsPdfRegular"
    bold_font_name = "ReportsPdfBold"
    registered_fonts = set(pdfmetrics.getRegisteredFontNames())
    if regular_font_name not in registered_fonts:
        pdfmetrics.registerFont(TTFont(regular_font_name, font_regular_path))
    if bold_font_name not in registered_fonts:
        pdfmetrics.registerFont(TTFont(bold_font_name, font_bold_path))

    def pdf_text(value: Any) -> str:
        text_value = str(value or "").strip()
        if not text_value:
            return "—"
        return (
            text_value.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br/>")
        )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportsTitle",
        parent=styles["Heading1"],
        fontName=bold_font_name,
        fontSize=24,
        leading=28,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "ReportsSubtitle",
        parent=styles["Normal"],
        fontName=regular_font_name,
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#475569"),
        spaceAfter=10,
    )
    section_title_style = ParagraphStyle(
        "ReportsSectionTitle",
        parent=styles["Heading2"],
        fontName=bold_font_name,
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#0f172a"),
        spaceBefore=8,
        spaceAfter=6,
    )
    section_subtitle_style = ParagraphStyle(
        "ReportsSectionSubtitle",
        parent=styles["Normal"],
        fontName=regular_font_name,
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=5,
    )
    card_label_style = ParagraphStyle(
        "ReportsCardLabel",
        parent=styles["Normal"],
        fontName=bold_font_name,
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
    )
    card_value_style = ParagraphStyle(
        "ReportsCardValue",
        parent=styles["Normal"],
        fontName=bold_font_name,
        fontSize=18,
        leading=21,
        textColor=colors.HexColor("#0f172a"),
    )
    card_hint_style = ParagraphStyle(
        "ReportsCardHint",
        parent=styles["Normal"],
        fontName=regular_font_name,
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#475569"),
    )
    table_header_style = ParagraphStyle(
        "ReportsTableHeader",
        parent=styles["Normal"],
        fontName=bold_font_name,
        fontSize=10,
        leading=12,
        textColor=colors.white,
    )
    table_cell_style = ParagraphStyle(
        "ReportsTableCell",
        parent=styles["Normal"],
        fontName=regular_font_name,
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
    )

    page_width, page_height = landscape(A4)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(page_width, page_height),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=10 * mm,
    )

    story = [
        Paragraph(pdf_text(report.get("title") or "Отчёт"), title_style),
        Paragraph(pdf_text(report.get("subtitle") or ""), subtitle_style),
        Spacer(1, 4),
    ]

    usable_width = page_width - doc.leftMargin - doc.rightMargin

    def chunk_items(items: list[dict[str, Any]], size: int):
        for start in range(0, len(items), size):
            yield items[start:start + size]

    def build_card_table(items: list[dict[str, Any]], *, columns: int = 2):
        rows = []
        for chunk in chunk_items(items, columns):
            row = []
            for item in chunk:
                parts = [
                    Paragraph(pdf_text(item.get("label") or item.get("title") or "Показатель"), card_label_style),
                    Spacer(1, 2),
                    Paragraph(pdf_text(item.get("value") or "—"), card_value_style),
                ]
                hint = item.get("hint")
                if not hint and item.get("meta"):
                    hint = " • ".join(str(part) for part in item.get("meta") if part)
                if hint:
                    parts.extend([Spacer(1, 4), Paragraph(pdf_text(hint), card_hint_style)])
                row.append(parts)
            while len(row) < columns:
                row.append("")
            rows.append(row)
        widths = [usable_width / columns] * columns
        table = Table(rows, colWidths=widths, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eff6ff")),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#bfdbfe")),
            ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#dbeafe")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        return table

    def build_section_table(title: str, columns: list[str], rows: list[list[Any]]):
        if not columns:
            return None
        if title == "Источники сигналов":
            col_widths = [usable_width * 0.48, usable_width * 0.12, usable_width * 0.16, usable_width * 0.24]
        elif title == "Связи с delivery":
            col_widths = [usable_width * 0.78, usable_width * 0.22]
        elif len(columns) == 5:
            col_widths = [usable_width * 0.14, usable_width * 0.38, usable_width * 0.16, usable_width * 0.16, usable_width * 0.16]
        elif len(columns) == 4:
            col_widths = [usable_width * 0.48, usable_width * 0.14, usable_width * 0.16, usable_width * 0.22]
        elif len(columns) == 2:
            col_widths = [usable_width * 0.76, usable_width * 0.24]
        else:
            col_widths = [usable_width / len(columns)] * len(columns)
        table_rows = [[Paragraph(pdf_text(col), table_header_style) for col in columns]]
        for row in rows:
            table_rows.append([Paragraph(pdf_text(cell), table_cell_style) for cell in row[:len(columns)]])
        table = Table(table_rows, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        return table

    for section in report.get("sections") or []:
        section_type = section.get("type")
        section_title = section.get("title")
        section_subtitle = section.get("subtitle")
        force_new_page = section_type == "table" and section_title in {
            "Причины",
            "Направления",
            "Источники сигналов",
            "Связи с delivery",
            "Поток инцидентов",
        }
        if force_new_page:
            story.append(PageBreak())
        if section_title:
            story.append(Paragraph(pdf_text(section_title), section_title_style))
        if section_subtitle:
            story.append(Paragraph(pdf_text(section_subtitle), section_subtitle_style))
        if section_type == "kpis":
            items = section.get("items") or []
            if items:
                story.append(build_card_table(items, columns=2))
                story.append(Spacer(1, 8))
        elif section_type == "cards":
            items = section.get("items") or []
            if items:
                story.append(build_card_table(items, columns=min(2, max(1, len(items)))))
                story.append(Spacer(1, 8))
        elif section_type == "table":
            table = build_section_table(section_title or "", section.get("columns") or [], (section.get("rows") or [])[:8])
            if table:
                story.append(table)
                story.append(Spacer(1, 8))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def _resolve_date_window(date_from: Optional[str], date_to: Optional[str], days: int):
    # date_from/date_to are expected as YYYY-MM-DD; fallback to last N days
    params = {"days": days}
    clause = "r.started_at >= (now() - (:days || ' days')::interval)"
    if date_from:
        clause = "r.started_at >= :date_from"
        params = {"date_from": date_from, "days": days}
    if date_to:
        clause = clause + " AND r.started_at <= :date_to"
        params["date_to"] = date_to
    return clause, params


@router.get("/api/releases")
def get_releases(days: int = Query(30, ge=1, le=365), limit: int = Query(50, ge=1, le=200)):
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    SELECT
                        l.release_id,
                        l.release_type,
                        l.initiated_by,
                        l.started_at,
                        l.finished_at,
                        l.status,
                        l.total_objects,
                        l.ready_to_release,
                        EXTRACT(EPOCH FROM (COALESCE(l.finished_at, now()) - l.started_at)) / 60.0 AS duration_minutes,
                        COALESCE(o.objects_count, 0) AS objects_count,
                        COALESCE(o.failed_count, 0) AS failed_count,
                        COALESCE(o.failed_any, false) AS failed_any,
                        COALESCE(o.task_ids, ARRAY[]::text[]) AS task_ids,
                        COALESCE(w.hours_total, 0) AS hours_total
                    FROM {TABLE_RELEASE_LOG} l
                    LEFT JOIN (
                        SELECT
                            release_id,
                            COUNT(*) AS objects_count,
                            COUNT(*) FILTER (WHERE lower(final_status) NOT IN ('success', 'succeeded', 'ok')) AS failed_count,
                            BOOL_OR(failed_objects) AS failed_any,
                            ARRAY_AGG(DISTINCT task_id) FILTER (WHERE task_id IS NOT NULL AND task_id <> '') AS task_ids
                        FROM {TABLE_RELEASE_OBJECTS}
                        GROUP BY release_id
                    ) o ON o.release_id = l.release_id
                    LEFT JOIN (
                        SELECT ro.release_id,
                               COALESCE(SUM(w.minutes), 0) / 60.0 AS hours_total
                        FROM (
                            SELECT DISTINCT release_id, task_id
                            FROM {TABLE_RELEASE_OBJECTS}
                            WHERE task_id IS NOT NULL AND task_id <> ''
                        ) ro
                        LEFT JOIN (
                            SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                            FROM {TABLE_YT_ISSUE_WORKLOG}
                            GROUP BY issue_id
                        ) w ON w.issue_id = ro.task_id
                        GROUP BY ro.release_id
                    ) w ON w.release_id = l.release_id
                    WHERE l.started_at >= (now() - (:days || ' days')::interval)
                    ORDER BY l.started_at DESC
                    LIMIT :limit
                    """
                ),
                {"days": days, "limit": limit},
            )
            payload = []
            for row in rows:
                record = dict(row._mapping)
                task_ids = record.get("task_ids") or []
                record["task_ids"] = task_ids
                record["task_links"] = [link for link in (_build_ytrack_link(t) for t in task_ids) if link]
                payload.append(record)
            return {"items": payload}
    except Exception as e:
        print("❌ /api/releases error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить релизы")


@router.get("/api/releases/{release_id}")
def get_release_details(release_id: str, limit: int = Query(500, ge=1, le=2000)):
    try:
        with engine.connect() as conn:
            log_row = conn.execute(
                text(
                    f"""
                    SELECT release_id, release_type, initiated_by, started_at, finished_at,
                           status, total_objects, ready_to_release,
                           EXTRACT(EPOCH FROM (COALESCE(finished_at, now()) - started_at)) / 60.0 AS duration_minutes
                    FROM {TABLE_RELEASE_LOG}
                    WHERE release_id = :release_id
                    LIMIT 1
                    """
                ),
                {"release_id": release_id},
            ).fetchone()
            if not log_row:
                raise HTTPException(status_code=404, detail="Релиз не найден")
            objects = conn.execute(
                text(
                    f"""
                    SELECT
                        release_id,
                        task_id,
                        target_system,
                        schema_name,
                        table_name,
                        entity_id,
                        entity_name,
                        change_type,
                        final_status,
                        attempts_count,
                        created_at,
                        failed_objects,
                        attempt_no,
                        error_message,
                        error_stacktrace
                    FROM {TABLE_RELEASE_OBJECTS}
                    WHERE release_id = :release_id
                    ORDER BY created_at DESC
                    LIMIT :limit
                    """
                ),
                {"release_id": release_id, "limit": limit},
            ).fetchall()
            obj_payload = []
            for row in objects:
                record = dict(row._mapping)
                record["task_link"] = _build_ytrack_link(record.get("task_id"))
                obj_payload.append(record)
            return {"release": dict(log_row._mapping), "objects": obj_payload}
    except HTTPException:
        raise
    except Exception as e:
        print("❌ /api/releases/{id} error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить детали релиза")


@router.get("/api/releases/table/{schema}/{table:path}")
def get_table_releases(
    schema: str,
    table: str,
    target_system: Optional[str] = None,
    limit: int = Query(30, ge=1, le=200),
):
    try:
        table_clean = _clean_table_name(norm(table))
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    WITH issue_snapshot AS (
                        SELECT
                            issue_id,
                            assignee,
                            created_by
                        FROM {TABLE_YT_ISSUE_SNAPSHOT}
                    ),
                    issue_executor AS (
                        SELECT DISTINCT ON (issue_id)
                               issue_id,
                               author AS executor
                        FROM {TABLE_YT_ISSUE_TIMELINE}
                        WHERE event_type = 'State change'
                          AND value_to IN ('Ожидание релиза', 'В работе')
                        ORDER BY issue_id, ts DESC NULLS LAST
                    )
                    SELECT
                        o.release_id,
                        o.task_id,
                        o.target_system,
                        o.schema_name,
                        o.table_name,
                        o.entity_id,
                        o.entity_name,
                        o.change_type,
                        o.final_status,
                        o.attempts_count,
                        o.created_at,
                        o.failed_objects,
                        o.attempt_no,
                        o.error_message,
                        o.error_stacktrace,
                        l.release_type,
                        l.initiated_by,
                        l.started_at,
                        l.finished_at,
                        l.status AS release_status,
                        COALESCE(exec.executor, snap.assignee, snap.created_by, '—') AS task_executor
                    FROM {TABLE_RELEASE_OBJECTS} o
                    LEFT JOIN {TABLE_RELEASE_LOG} l ON l.release_id = o.release_id
                    LEFT JOIN issue_snapshot snap ON snap.issue_id = o.task_id
                    LEFT JOIN issue_executor exec ON exec.issue_id = o.task_id
                    WHERE o.schema_name = :schema
                      AND (lower(o.table_name) = lower(:table) OR lower(o.table_name) = lower(:table_clean))
                      AND (:target_system IS NULL OR lower(o.target_system) = lower(:target_system))
                    ORDER BY o.created_at DESC
                    LIMIT :limit
                    """
                ),
                {"schema": schema, "table": table, "table_clean": table_clean, "target_system": target_system, "limit": limit},
            ).fetchall()
            payload = []
            for row in rows:
                record = dict(row._mapping)
                record["task_link"] = _build_ytrack_link(record.get("task_id"))
                payload.append(record)
            return {"items": payload}
    except Exception as e:
        print("❌ /api/releases/table error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить релизы по объекту")


@router.get("/api/ytrek/table/{schema}/{table:path}")
def get_ytrek_table_info(schema: str, table: str):
    try:
        table_clean = _clean_table_name(norm(table))
        with engine.connect() as conn:
            task_rows = conn.execute(
                text(
                    f"""
                    SELECT DISTINCT task_id
                    FROM {TABLE_RELEASE_OBJECTS}
                    WHERE schema_name = :schema
                      AND (lower(table_name) = lower(:table) OR lower(table_name) = lower(:table_clean))
                      AND task_id IS NOT NULL
                    """
                ),
                {"schema": schema, "table": table, "table_clean": table_clean},
            ).fetchall()
            task_ids = [r[0] for r in task_rows if r[0]]
            if not task_ids:
                return {"tasks": [], "timeline": [], "worklog": [], "stats": {}}

            snapshots = conn.execute(
                text(
                    f"""
                    SELECT issue_id, summary, project_name, project_key, created_by, assignee,
                           created_at, updated_at, resolved_at, current_state
                    FROM {TABLE_YT_ISSUE_SNAPSHOT}
                    WHERE issue_id = ANY(:ids)
                    ORDER BY updated_at DESC NULLS LAST
                    """
                ),
                {"ids": task_ids},
            ).mappings().all()

            worklog = conn.execute(
                text(
                    f"""
                    SELECT issue_id,
                           COALESCE(SUM(minutes), 0) AS minutes,
                           COUNT(*) AS entries
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    WHERE issue_id = ANY(:ids)
                    GROUP BY issue_id
                    """
                ),
                {"ids": task_ids},
            ).mappings().all()
            worklog_map = {r["issue_id"]: r for r in worklog}

            timeline = conn.execute(
                text(
                    f"""
                    SELECT issue_id, ts, author, event_type, field_name, value_from, value_to
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE issue_id = ANY(:ids)
                    ORDER BY ts DESC NULLS LAST
                    LIMIT 200
                    """
                ),
                {"ids": task_ids},
            ).mappings().all()

            last_assignee = conn.execute(
                text(
                    f"""
                    SELECT DISTINCT ON (issue_id)
                           issue_id, ts, author, value_from, value_to
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE issue_id = ANY(:ids)
                      AND event_type = 'Assignee change'
                    ORDER BY issue_id, ts DESC NULLS LAST
                    """
                ),
                {"ids": task_ids},
            ).mappings().all()
            last_assignee_map = {r["issue_id"]: r for r in last_assignee}

            last_state = conn.execute(
                text(
                    f"""
                    SELECT DISTINCT ON (issue_id)
                           issue_id, ts, author, value_from, value_to
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE issue_id = ANY(:ids)
                      AND event_type = 'State change'
                    ORDER BY issue_id, ts DESC NULLS LAST
                    """
                ),
                {"ids": task_ids},
            ).mappings().all()
            last_state_map = {r["issue_id"]: r for r in last_state}

            last_wait = conn.execute(
                text(
                    f"""
                    SELECT DISTINCT ON (issue_id)
                           issue_id, ts, author, value_from, value_to
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE issue_id = ANY(:ids)
                      AND event_type = 'State change'
                      AND value_to = 'Ожидание релиза'
                    ORDER BY issue_id, ts DESC NULLS LAST
                    """
                ),
                {"ids": task_ids},
            ).mappings().all()
            last_wait_map = {r["issue_id"]: r for r in last_wait}

            last_work = conn.execute(
                text(
                    f"""
                    SELECT DISTINCT ON (issue_id)
                           issue_id, ts, author, value_from, value_to
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE issue_id = ANY(:ids)
                      AND event_type = 'State change'
                      AND value_to = 'В работе'
                    ORDER BY issue_id, ts DESC NULLS LAST
                    """
                ),
                {"ids": task_ids},
            ).mappings().all()
            last_work_map = {r["issue_id"]: r for r in last_work}

            custom = conn.execute(
                text(
                    f"""
                    SELECT issue_id, field_name, field_value
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE issue_id = ANY(:ids)
                      AND field_name IN ('Subsystem', 'Дашборд КХД/Направление', 'Дашборд SD', 'Тип карточки', 'Дата релиза')
                    """
                ),
                {"ids": task_ids},
            ).mappings().all()
            custom_map = {}
            for row in custom:
                custom_map.setdefault(row["issue_id"], {})[row["field_name"]] = row["field_value"]

            tasks_payload = []
            for s in snapshots:
                issue_id = s["issue_id"]
                wl = worklog_map.get(issue_id, {})
                effective_assignee = None
                effective_reason = None
                if issue_id in last_wait_map:
                    effective_assignee = last_wait_map[issue_id].get("author")
                    effective_reason = "Ожидание релиза"
                elif issue_id in last_work_map:
                    effective_assignee = last_work_map[issue_id].get("author")
                    effective_reason = "В работе"
                else:
                    effective_assignee = s.get("assignee")
                    effective_reason = "Текущий"
                tasks_payload.append(
                    {
                        **s,
                        "work_minutes": wl.get("minutes", 0),
                        "work_entries": wl.get("entries", 0),
                        "last_assignee_change": last_assignee_map.get(issue_id),
                        "last_state_change": last_state_map.get(issue_id),
                        "custom": custom_map.get(issue_id, {}),
                        "effective_assignee": effective_assignee,
                        "effective_assignee_reason": effective_reason,
                    }
                )

            stats = {
                "tasks_count": len(task_ids),
                "work_minutes_total": sum(r.get("minutes", 0) for r in worklog_map.values()),
            }

            return {
                "tasks": tasks_payload,
                "timeline": [
                    {
                        "issue_id": r["issue_id"],
                        "ts": serialize_datetime(r["ts"]),
                        "author": r["author"],
                        "event_type": r["event_type"],
                        "field_name": r["field_name"],
                        "value_from": r["value_from"],
                        "value_to": r["value_to"],
                    }
                    for r in timeline
                ],
                "stats": stats,
            }
    except Exception as e:
        print("❌ /api/ytrek/table error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить данные YouTrack")


@router.get("/api/ytrek/analytics")
def get_ytrek_analytics(days: int = Query(365, ge=1, le=3650)):
    try:
        with engine.connect() as conn:
            base = f"""
                WITH tasks AS (
                    SELECT DISTINCT ro.task_id
                    FROM {TABLE_RELEASE_OBJECTS} ro
                    WHERE ro.task_id IS NOT NULL
                ),
                snap AS (
                    SELECT s.*
                    FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                    JOIN tasks t ON t.task_id = s.issue_id
                    WHERE s.created_at >= (now() - (:days || ' days')::interval)
                ),
                exec AS (
                    SELECT DISTINCT ON (issue_id)
                           issue_id, author AS executor, ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                team AS (
                    SELECT issue_id, field_value AS team
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Subsystem'
                ),
                dashboard AS (
                    SELECT issue_id, field_value AS dashboard_direction
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Дашборд КХД/Направление'
                )
            """

            by_team = conn.execute(
                text(
                    base
                    + """
                    SELECT COALESCE(t.team, 'Не указана') AS team,
                           COUNT(*) AS tasks_count,
                           COALESCE(SUM(w.minutes), 0) AS minutes
                    FROM snap s
                    LEFT JOIN work w ON w.issue_id = s.issue_id
                    LEFT JOIN team t ON t.issue_id = s.issue_id
                    GROUP BY t.team
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                {"days": days},
            ).mappings().all()

            by_dashboard = conn.execute(
                text(
                    base
                    + """
                    SELECT COALESCE(d.dashboard_direction, 'Не указан') AS dashboard_direction,
                           COUNT(*) AS tasks_count,
                           COALESCE(SUM(w.minutes), 0) AS minutes
                    FROM snap s
                    LEFT JOIN work w ON w.issue_id = s.issue_id
                    LEFT JOIN dashboard d ON d.issue_id = s.issue_id
                    GROUP BY d.dashboard_direction
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                {"days": days},
            ).mappings().all()

            by_creator = conn.execute(
                text(
                    base
                    + """
                    SELECT COALESCE(s.created_by, 'Не указан') AS creator,
                           COUNT(*) AS tasks_count,
                           COALESCE(SUM(w.minutes), 0) AS minutes
                    FROM snap s
                    LEFT JOIN work w ON w.issue_id = s.issue_id
                    GROUP BY s.created_by
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                {"days": days},
            ).mappings().all()

            by_assignee = conn.execute(
                text(
                    base
                    + """
                    SELECT COALESCE(exec.executor, s.assignee, s.created_by, 'Не указан') AS assignee,
                           COUNT(*) AS tasks_count,
                           COALESCE(SUM(w.minutes), 0) AS minutes
                    FROM snap s
                    LEFT JOIN exec ON exec.issue_id = s.issue_id
                    LEFT JOIN work w ON w.issue_id = s.issue_id
                    GROUP BY 1
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                {"days": days},
            ).mappings().all()

            return {
                "by_team": by_team,
                "by_dashboard": by_dashboard,
                "by_creator": by_creator,
                "by_assignee": by_assignee,
            }
    except Exception as e:
        print("❌ /api/ytrek/analytics error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить аналитику YouTrack")


@router.get("/api/ytrek/tasks")
def get_ytrek_tasks(days: int = Query(365, ge=1, le=3650)):
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    WITH tasks AS (
                        SELECT DISTINCT ro.task_id
                        FROM {TABLE_RELEASE_OBJECTS} ro
                        WHERE ro.task_id IS NOT NULL
                    ),
                    snap AS (
                        SELECT s.*
                        FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                        JOIN tasks t ON t.task_id = s.issue_id
                        WHERE s.created_at >= (now() - (:days || ' days')::interval)
                    ),
                    exec AS (
                        SELECT DISTINCT ON (issue_id)
                               issue_id, author AS executor, ts
                        FROM {TABLE_YT_ISSUE_TIMELINE}
                        WHERE event_type = 'State change'
                          AND value_to IN ('Ожидание релиза', 'В работе')
                        ORDER BY issue_id, ts DESC NULLS LAST
                    ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                team AS (
                    SELECT issue_id, field_value AS team
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Subsystem'
                ),
                dashboard AS (
                    SELECT issue_id, field_value AS dashboard_direction
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Дашборд КХД/Направление'
                )
                SELECT s.issue_id,
                       COALESCE(s.created_by, 'Не указан') AS created_by,
                       COALESCE(exec.executor, s.assignee, s.created_by, 'Не указан') AS assignee,
                       s.current_state,
                       COALESCE(t.team, 'Не указана') AS team,
                       COALESCE(d.dashboard_direction, 'Не указан') AS dashboard_direction,
                       COALESCE(w.minutes, 0) AS minutes
                FROM snap s
                LEFT JOIN exec ON exec.issue_id = s.issue_id
                LEFT JOIN work w ON w.issue_id = s.issue_id
                LEFT JOIN team t ON t.issue_id = s.issue_id
                LEFT JOIN dashboard d ON d.issue_id = s.issue_id
                ORDER BY w.minutes DESC NULLS LAST
                """
            ),
                {"days": days},
            ).mappings().all()
            return rows
    except Exception as e:
        print("❌ /api/ytrek/tasks error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить задачи YouTrack")


@router.get("/api/ytrek/workload")
def get_ytrek_workload(days: int = Query(90, ge=14, le=3650)):
    try:
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    f"""
                    WITH tasks AS (
                        SELECT DISTINCT ro.task_id
                        FROM {TABLE_RELEASE_OBJECTS} ro
                        WHERE ro.task_id IS NOT NULL
                    ),
                    day_series AS (
                        SELECT generate_series(
                            CURRENT_DATE - (:days - 1) * INTERVAL '1 day',
                            CURRENT_DATE,
                            INTERVAL '1 day'
                        )::date AS day
                    ),
                    snap AS (
                        SELECT s.issue_id, s.created_at
                        FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                        JOIN tasks t ON t.task_id = s.issue_id
                    ),
                    first_assignee AS (
                        SELECT issue_id, MIN(ts) AS ts
                        FROM {TABLE_YT_ISSUE_TIMELINE}
                        WHERE event_type = 'Assignee change'
                        GROUP BY issue_id
                    ),
                    first_work AS (
                        SELECT issue_id, MIN(ts) AS ts
                        FROM {TABLE_YT_ISSUE_TIMELINE}
                        WHERE event_type = 'State change'
                          AND value_to = 'В работе'
                        GROUP BY issue_id
                    ),
                    first_wait_release AS (
                        SELECT issue_id, MIN(ts) AS ts
                        FROM {TABLE_YT_ISSUE_TIMELINE}
                        WHERE event_type = 'State change'
                          AND value_to = 'Ожидание релиза'
                        GROUP BY issue_id
                    ),
                    created_daily AS (
                        SELECT DATE(created_at) AS day, COUNT(*) AS cnt
                        FROM snap
                        WHERE created_at >= CURRENT_DATE - (:days - 1) * INTERVAL '1 day'
                        GROUP BY 1
                    ),
                    assignee_daily AS (
                        SELECT DATE(a.ts) AS day, COUNT(*) AS cnt
                        FROM first_assignee a
                        JOIN tasks t ON t.task_id = a.issue_id
                        WHERE a.ts >= CURRENT_DATE - (:days - 1) * INTERVAL '1 day'
                        GROUP BY 1
                    ),
                    work_daily AS (
                        SELECT DATE(w.ts) AS day, COUNT(*) AS cnt
                        FROM first_work w
                        JOIN tasks t ON t.task_id = w.issue_id
                        WHERE w.ts >= CURRENT_DATE - (:days - 1) * INTERVAL '1 day'
                        GROUP BY 1
                    ),
                    wait_daily AS (
                        SELECT DATE(w.ts) AS day, COUNT(*) AS cnt
                        FROM first_wait_release w
                        JOIN tasks t ON t.task_id = w.issue_id
                        WHERE w.ts >= CURRENT_DATE - (:days - 1) * INTERVAL '1 day'
                        GROUP BY 1
                    ),
                    release_daily AS (
                        SELECT DATE(started_at) AS day, COUNT(*) AS cnt
                        FROM {TABLE_RELEASE_LOG}
                        WHERE started_at >= CURRENT_DATE - (:days - 1) * INTERVAL '1 day'
                        GROUP BY 1
                    )
                    SELECT
                        ds.day::text AS day,
                        COALESCE(cd.cnt, 0) AS created_count,
                        COALESCE(ad.cnt, 0) AS assigned_count,
                        COALESCE(wd.cnt, 0) AS in_work_count,
                        COALESCE(rd.cnt, 0) AS release_ready_count,
                        COALESCE(ld.cnt, 0) AS release_count,
                        COALESCE(cd.cnt, 0)
                          + COALESCE(ad.cnt, 0)
                          + COALESCE(wd.cnt, 0)
                          + COALESCE(rd.cnt, 0) AS total_activity
                    FROM day_series ds
                    LEFT JOIN created_daily cd ON cd.day = ds.day
                    LEFT JOIN assignee_daily ad ON ad.day = ds.day
                    LEFT JOIN work_daily wd ON wd.day = ds.day
                    LEFT JOIN wait_daily rd ON rd.day = ds.day
                    LEFT JOIN release_daily ld ON ld.day = ds.day
                    ORDER BY ds.day
                    """
                ),
                {"days": days},
            ).mappings().all()
            return list(rows)
    except Exception as e:
        print("❌ /api/ytrek/workload error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить нагрузку по задачам")


@router.get("/api/analytics/dashboard")
def get_analytics_dashboard(
    days: int = Query(30, ge=1, le=3650),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    try:
        date_clause, params = _resolve_date_window(date_from, date_to, days)
        with engine.connect() as conn:
            base = f"""
                WITH rel AS (
                    SELECT r.release_id, r.started_at, r.finished_at, r.initiated_by, r.status
                    FROM {TABLE_RELEASE_LOG} r
                    WHERE {date_clause}
                ),
                ro AS (
                    SELECT ro.*
                    FROM {TABLE_RELEASE_OBJECTS} ro
                    JOIN rel r ON r.release_id = ro.release_id
                ),
                tasks AS (
                    SELECT DISTINCT task_id
                    FROM ro
                    WHERE task_id IS NOT NULL
                ),
                snap AS (
                    SELECT s.*
                    FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                    JOIN tasks t ON t.task_id = s.issue_id
                ),
                exec AS (
                    SELECT DISTINCT ON (issue_id)
                           issue_id, author AS executor, ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                direction AS (
                    SELECT issue_id, field_value AS direction
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Направление'
                )
            """

            summary = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        (SELECT COUNT(*) FROM rel) AS releases,
                        (SELECT COUNT(*) FROM tasks) AS tasks,
                        (SELECT COUNT(*) FROM ro) AS objects,
                        (SELECT COALESCE(SUM(minutes), 0) FROM work) AS minutes,
                        (SELECT COUNT(DISTINCT COALESCE(exec.executor, snap.assignee, snap.created_by))
                         FROM snap
                         LEFT JOIN exec ON exec.issue_id = snap.issue_id) AS executors
                    """
                ),
                params,
            ).mappings().first()

            by_executor = conn.execute(
                text(
                    base
                    + """
                    SELECT COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан') AS executor,
                           COUNT(DISTINCT snap.issue_id) AS tasks_count,
                           COUNT(ro.table_name) AS tables_count,
                           COALESCE(SUM(work.minutes), 0) AS minutes
                    FROM snap
                    LEFT JOIN exec ON exec.issue_id = snap.issue_id
                    LEFT JOIN work ON work.issue_id = snap.issue_id
                    LEFT JOIN ro ON ro.task_id = snap.issue_id
                    GROUP BY COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан')
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

            by_creator = conn.execute(
                text(
                    base
                    + """
                    SELECT COALESCE(snap.created_by, 'Не указан') AS creator,
                           COUNT(DISTINCT snap.issue_id) AS tasks_count,
                           COUNT(ro.table_name) AS tables_count,
                           COALESCE(SUM(work.minutes), 0) AS minutes
                    FROM snap
                    LEFT JOIN work ON work.issue_id = snap.issue_id
                    LEFT JOIN ro ON ro.task_id = snap.issue_id
                    GROUP BY COALESCE(snap.created_by, 'Не указан')
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

            by_direction = conn.execute(
                text(
                    base
                    + """
                    SELECT COALESCE(direction.direction, 'Не указан') AS direction,
                           COUNT(DISTINCT snap.issue_id) AS tasks_count,
                           COUNT(ro.table_name) AS tables_count,
                           COALESCE(SUM(work.minutes), 0) AS minutes
                    FROM snap
                    LEFT JOIN direction ON direction.issue_id = snap.issue_id
                    LEFT JOIN work ON work.issue_id = snap.issue_id
                    LEFT JOIN ro ON ro.task_id = snap.issue_id
                    GROUP BY COALESCE(direction.direction, 'Не указан')
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

            top_tables = conn.execute(
                text(
                    base
                    + """
                    SELECT ro.schema_name, ro.table_name,
                           COUNT(*) AS changes,
                           COALESCE(SUM(work.minutes), 0) AS minutes
                    FROM ro
                    LEFT JOIN work ON work.issue_id = ro.task_id
                    GROUP BY ro.schema_name, ro.table_name
                    ORDER BY changes DESC, minutes DESC
                    LIMIT 20
                    """
                ),
                params,
            ).mappings().all()

            return {
                "summary": summary,
                "by_executor": by_executor,
                "by_creator": by_creator,
                "by_direction": by_direction,
                "top_tables": top_tables,
            }
    except Exception as e:
        print("❌ /api/analytics/dashboard error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить дашборд")


@router.get("/api/analytics/release/{release_id}")
def get_analytics_release(release_id: str):
    try:
        with engine.connect() as conn:
            rel = conn.execute(
                text(
                    f"""
                    SELECT release_id, started_at, finished_at, initiated_by, status, total_objects
                    FROM {TABLE_RELEASE_LOG}
                    WHERE release_id = :release_id
                    """
                ),
                {"release_id": release_id},
            ).mappings().first()
            if not rel:
                raise HTTPException(status_code=404, detail="Релиз не найден")

            base = f"""
                WITH ro AS (
                    SELECT * FROM {TABLE_RELEASE_OBJECTS}
                    WHERE release_id = :release_id
                ),
                tasks AS (
                    SELECT DISTINCT task_id FROM ro WHERE task_id IS NOT NULL
                ),
                snap AS (
                    SELECT s.*
                    FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                    JOIN tasks t ON t.task_id = s.issue_id
                ),
                exec AS (
                    SELECT DISTINCT ON (issue_id)
                           issue_id, author AS executor, ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                direction AS (
                    SELECT issue_id, field_value AS direction
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Направление'
                )
            """

            tasks_list = conn.execute(
                text(
                    base
                    + """
                    SELECT snap.issue_id AS task_id,
                           snap.created_by AS creator,
                           COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан') AS executor,
                           COALESCE(direction.direction, 'Не указан') AS direction,
                           COUNT(ro.table_name) AS tables_count,
                           COALESCE(SUM(work.minutes), 0) AS minutes
                    FROM snap
                    LEFT JOIN exec ON exec.issue_id = snap.issue_id
                    LEFT JOIN direction ON direction.issue_id = snap.issue_id
                    LEFT JOIN work ON work.issue_id = snap.issue_id
                    LEFT JOIN ro ON ro.task_id = snap.issue_id
                    GROUP BY snap.issue_id, snap.created_by, exec.executor, snap.assignee, direction.direction
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                {"release_id": release_id},
            ).mappings().all()

            tables_list = conn.execute(
                text(
                    base
                    + """
                    SELECT ro.schema_name, ro.table_name, ro.change_type, ro.task_id,
                           COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан') AS executor,
                           ro.final_status
                    FROM ro
                    LEFT JOIN snap ON snap.issue_id = ro.task_id
                    LEFT JOIN exec ON exec.issue_id = ro.task_id
                    ORDER BY ro.schema_name, ro.table_name
                    """
                ),
                {"release_id": release_id},
            ).mappings().all()

            summary = {
                "release_id": rel["release_id"],
                "status": rel["status"],
                "started_at": rel["started_at"],
                "finished_at": rel["finished_at"],
                "initiated_by": rel["initiated_by"],
                "tasks": len({t["task_id"] for t in tasks_list}),
                "tables": len(tables_list),
                "hours": round(sum(t["minutes"] for t in tasks_list) / 60, 2),
            }

            return {"summary": summary, "tasks": tasks_list, "tables": tables_list}
    except HTTPException:
        raise
    except Exception as e:
        print("❌ /api/analytics/release error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить релиз")


@router.get("/api/analytics/table/{schema}/{table:path}")
def get_analytics_table(schema: str, table: str, days: int = Query(365, ge=1, le=3650)):
    try:
        date_clause, params = _resolve_date_window(None, None, days)
        params.update({"schema": schema, "table": table, "table_clean": _clean_table_name(norm(table))})
        with engine.connect() as conn:
            base = f"""
                WITH rel AS (
                    SELECT r.release_id, r.started_at
                    FROM {TABLE_RELEASE_LOG} r
                    WHERE {date_clause}
                ),
                ro AS (
                    SELECT ro.*
                    FROM {TABLE_RELEASE_OBJECTS} ro
                    JOIN rel r ON r.release_id = ro.release_id
                    WHERE ro.schema_name = :schema
                      AND (lower(ro.table_name) = lower(:table) OR lower(ro.table_name) = lower(:table_clean))
                ),
                tasks AS (
                    SELECT DISTINCT task_id FROM ro WHERE task_id IS NOT NULL
                ),
                snap AS (
                    SELECT s.*
                    FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                    JOIN tasks t ON t.task_id = s.issue_id
                ),
                exec AS (
                    SELECT DISTINCT ON (issue_id)
                           issue_id, author AS executor, ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                direction AS (
                    SELECT issue_id, field_value AS direction
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Направление'
                )
            """

            history = conn.execute(
                text(
                    base
                    + """
                    SELECT ro.schema_name, ro.table_name, ro.task_id, ro.release_id,
                           COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан') AS executor,
                           snap.created_by AS creator,
                           COALESCE(direction.direction, 'Не указан') AS direction,
                           COALESCE(work.minutes, 0) AS minutes,
                           rel.started_at AS changed_at
                    FROM ro
                    LEFT JOIN snap ON snap.issue_id = ro.task_id
                    LEFT JOIN exec ON exec.issue_id = ro.task_id
                    LEFT JOIN work ON work.issue_id = ro.task_id
                    LEFT JOIN direction ON direction.issue_id = ro.task_id
                    LEFT JOIN rel ON rel.release_id = ro.release_id
                    ORDER BY rel.started_at DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()

            summary = {
                "schema": schema,
                "table": table,
                "changes": len(history),
                "hours": round(sum(r["minutes"] for r in history) / 60, 2) if history else 0,
                "last_change": history[0]["changed_at"] if history else None,
            }

            return {"summary": summary, "history": history}
    except Exception as e:
        print("❌ /api/analytics/table error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить аналитику таблицы")


@router.get("/api/analytics/workload")
def get_analytics_workload(
    days: int = Query(30, ge=1, le=3650),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    group_by: str = "executor",
):
    try:
        date_clause, params = _resolve_date_window(date_from, date_to, days)
        with engine.connect() as conn:
            base = f"""
                WITH rel AS (
                    SELECT r.release_id, r.started_at
                    FROM {TABLE_RELEASE_LOG} r
                    WHERE {date_clause}
                ),
                ro AS (
                    SELECT ro.*
                    FROM {TABLE_RELEASE_OBJECTS} ro
                    JOIN rel r ON r.release_id = ro.release_id
                ),
                tasks AS (
                    SELECT DISTINCT task_id
                    FROM ro
                    WHERE task_id IS NOT NULL
                ),
                snap AS (
                    SELECT s.*
                    FROM {TABLE_YT_ISSUE_SNAPSHOT} s
                    JOIN tasks t ON t.task_id = s.issue_id
                ),
                exec AS (
                    SELECT DISTINCT ON (issue_id)
                           issue_id, author AS executor, ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                direction AS (
                    SELECT issue_id, field_value AS direction
                    FROM {TABLE_YT_ISSUE_CUSTOM}
                    WHERE field_name = 'Направление'
                )
            """

            summary = conn.execute(
                text(
                    base
                    + """
                    SELECT
                        COUNT(DISTINCT snap.issue_id) AS tasks_count,
                        COUNT(DISTINCT ro.schema_name || '.' || ro.table_name) AS tables_count,
                        COUNT(DISTINCT COALESCE(exec.executor, snap.assignee, snap.created_by)) AS executors_count,
                        COALESCE(SUM(work.minutes), 0) AS minutes
                    FROM snap
                    LEFT JOIN exec ON exec.issue_id = snap.issue_id
                    LEFT JOIN work ON work.issue_id = snap.issue_id
                    LEFT JOIN direction ON direction.issue_id = snap.issue_id
                    LEFT JOIN ro ON ro.task_id = snap.issue_id
                    """
                ),
                params,
            ).mappings().first()

            group_expr = "COALESCE(exec.executor, snap.assignee, snap.created_by, 'Не указан')"
            label = "executor"
            if group_by == "creator":
                group_expr = "COALESCE(snap.created_by, 'Не указан')"
                label = "creator"
            elif group_by == "direction":
                group_expr = "COALESCE(direction.direction, 'Не указан')"
                label = "direction"

            rows = conn.execute(
                text(
                    base
                    + f"""
                    SELECT {group_expr} AS {label},
                           COUNT(DISTINCT snap.issue_id) AS tasks_count,
                           COUNT(ro.table_name) AS tables_count,
                           COALESCE(SUM(work.minutes), 0) AS minutes,
                           MAX(COALESCE(exec.ts, snap.updated_at, snap.created_at)) AS last_activity
                    FROM snap
                    LEFT JOIN exec ON exec.issue_id = snap.issue_id
                    LEFT JOIN work ON work.issue_id = snap.issue_id
                    LEFT JOIN direction ON direction.issue_id = snap.issue_id
                    LEFT JOIN ro ON ro.task_id = snap.issue_id
                    GROUP BY 1
                    ORDER BY minutes DESC NULLS LAST
                    """
                ),
                params,
            ).mappings().all()
            return {"group_by": group_by, "summary": summary, "items": rows}
    except Exception as e:
        print("❌ /api/analytics/workload error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить нагрузку")


@router.get("/api/analytics/hot-tables")
def get_analytics_hot_tables(
    days: int = Query(90, ge=1, le=3650),
    min_changes: int = Query(3, ge=1, le=1000),
):
    try:
        date_clause, params = _resolve_date_window(None, None, days)
        with engine.connect() as conn:
            base = f"""
                WITH rel AS (
                    SELECT r.release_id, r.started_at
                    FROM {TABLE_RELEASE_LOG} r
                    WHERE {date_clause}
                ),
                ro AS (
                    SELECT ro.*
                    FROM {TABLE_RELEASE_OBJECTS} ro
                    JOIN rel r ON r.release_id = ro.release_id
                ),
                exec AS (
                    SELECT DISTINCT ON (issue_id)
                           issue_id, author AS executor, ts
                    FROM {TABLE_YT_ISSUE_TIMELINE}
                    WHERE event_type = 'State change'
                      AND value_to IN ('Ожидание релиза', 'В работе')
                    ORDER BY issue_id, ts DESC NULLS LAST
                ),
                work AS (
                    SELECT issue_id, COALESCE(SUM(minutes), 0) AS minutes
                    FROM {TABLE_YT_ISSUE_WORKLOG}
                    GROUP BY issue_id
                ),
                last_change AS (
                    SELECT DISTINCT ON (ro.schema_name, ro.table_name)
                           ro.schema_name, ro.table_name, ro.task_id, rel.started_at AS changed_at
                    FROM ro
                    JOIN rel ON rel.release_id = ro.release_id
                    ORDER BY ro.schema_name, ro.table_name, rel.started_at DESC NULLS LAST
                )
            """

            rows = conn.execute(
                text(
                    base
                    + """
                    SELECT ro.schema_name, ro.table_name,
                           COUNT(*) AS changes,
                           COALESCE(SUM(work.minutes), 0) AS minutes,
                           MAX(rel.started_at) AS last_change_at,
                           COALESCE(exec.executor, 'Не указан') AS last_executor
                    FROM ro
                    JOIN rel ON rel.release_id = ro.release_id
                    LEFT JOIN work ON work.issue_id = ro.task_id
                    LEFT JOIN last_change lc ON lc.schema_name = ro.schema_name AND lc.table_name = ro.table_name
                    LEFT JOIN exec ON exec.issue_id = lc.task_id
                    GROUP BY ro.schema_name, ro.table_name, exec.executor
                    HAVING COUNT(*) >= :min_changes
                    ORDER BY changes DESC, minutes DESC
                    LIMIT 50
                    """
                ),
                {"days": days, "min_changes": min_changes},
            ).mappings().all()
            return {"items": rows}
    except Exception as e:
        print("❌ /api/analytics/hot-tables error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось получить hot tables")


@router.get("/api/search")
def search_entities(q: str):
    try:
        q = (q or "").strip()
        if not q:
            return {"releases": [], "tasks": [], "tables": []}
        like = f"%{q}%"
        with engine.connect() as conn:
            releases = conn.execute(
                text(
                    f"""
                    SELECT release_id, started_at, status
                    FROM {TABLE_RELEASE_LOG}
                    WHERE release_id ILIKE :q
                    ORDER BY started_at DESC NULLS LAST
                    LIMIT 20
                    """
                ),
                {"q": like},
            ).mappings().all()

            tasks = conn.execute(
                text(
                    f"""
                    SELECT issue_id, summary, created_by, updated_at, current_state
                    FROM {TABLE_YT_ISSUE_SNAPSHOT}
                    WHERE issue_id ILIKE :q OR summary ILIKE :q
                    ORDER BY updated_at DESC NULLS LAST
                    LIMIT 20
                    """
                ),
                {"q": like},
            ).mappings().all()

            tables = conn.execute(
                text(
                    f"""
                    SELECT DISTINCT schema_name, table_name
                    FROM {TABLE_RELEASE_OBJECTS}
                    WHERE (schema_name || '.' || table_name) ILIKE :q
                       OR table_name ILIKE :q
                    ORDER BY schema_name, table_name
                    LIMIT 40
                    """
                ),
                {"q": like},
            ).mappings().all()

            table_rows = [
                {
                    "schema_name": row.get("schema_name"),
                    "table_name": row.get("table_name"),
                    "source": "current",
                    "label": f"{row.get('schema_name')}.{row.get('table_name')}",
                }
                for row in tables
            ]
            seen = {(row["schema_name"], row["table_name"], row["source"]) for row in table_rows}
            term = q.lower()
            for item in get_dbt_table_catalog(BASE_DIR, DBT_MANIFEST_DIR, source="ohd"):
                if term in item["fqn"].lower() or term in (item.get("description") or "").lower():
                    key = (item["schema"], item["table"], item["source"])
                    if key in seen:
                        continue
                    seen.add(key)
                    table_rows.append(
                        {
                            "schema_name": item["schema"],
                            "table_name": item["table"],
                            "source": item["source"],
                            "label": item["label"],
                            "description": item.get("description") or "",
                        }
                    )

            return {"releases": releases, "tasks": tasks, "tables": table_rows[:40]}
    except Exception as e:
        print("❌ /api/search error:", e)
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Не удалось выполнить поиск")


app.include_router(router)
