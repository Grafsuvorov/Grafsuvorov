#!/usr/bin/env python3
import argparse
import json
from collections import defaultdict, deque
from statistics import median
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from sqlalchemy import create_engine, text

from api.config import DATABASE_URL, TABLE_LOADING_HISTORY
from api.main import get_cached_meta_and_index, norm


def build_meta_entries():
    all_meta, _ = get_cached_meta_and_index()
    entries = []
    meta_tables = set()
    for m in all_meta:
        schema = norm(m.get("table_schema"))
        table = norm(m.get("table_name"))
        entity = m.get("entity_name") or "UNKNOWN"
        if not schema or not table:
            continue
        if schema in ("raw_ext", "dict_raw_ext"):
            continue
        if isinstance(entity, str) and entity.lower() == "raw_ext":
            continue

        meta_tables.add(f"{schema}.{table}")
        depends = {}
        for src_schema, tables in (m.get("depends_on") or {}).items():
            src_schema_norm = norm(src_schema)
            if not src_schema_norm or src_schema_norm in ("raw_ext", "dict_raw_ext"):
                continue
            cleaned = [norm(t) for t in (tables or []) if t]
            depends[src_schema_norm] = [t for t in cleaned if t]

        entries.append(
            {
                "table_schema": schema,
                "table_name": table,
                "entity_name": entity,
                "table_id": m.get("table_id"),
                "depends_on": depends,
            }
        )
    return entries, meta_tables


def build_table_graph(entries, meta_tables):
    table_entities = defaultdict(set)
    table_ids = {}

    for m in entries:
        fqn = f"{m['table_schema']}.{m['table_name']}"
        table_entities[fqn].add(m["entity_name"])
        if m.get("table_id"):
            table_ids.setdefault(fqn, set()).add(m["table_id"])

    edges = []
    for m in entries:
        target = f"{m['table_schema']}.{m['table_name']}"
        for src_schema, tables in (m.get("depends_on") or {}).items():
            for src_table in tables:
                source = f"{src_schema}.{src_table}"
                if source not in meta_tables or target not in meta_tables:
                    continue
                edges.append((source, target))

    return table_entities, table_ids, edges


def build_adjacency(edges):
    forward = defaultdict(list)
    reverse = defaultdict(list)
    for src, tgt in edges:
        forward[src].append(tgt)
        reverse[tgt].append(src)
    return forward, reverse


def load_durations(days, engine):
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                f"""
                SELECT
                    object_id,
                    AVG(EXTRACT(EPOCH FROM (loading_finish_dttm - loading_start_dttm)) / 60.0) AS avg_minutes,
                    COUNT(*) AS samples
                FROM {TABLE_LOADING_HISTORY}
                WHERE object_type = 'table'
                  AND loading_state = 'SUCCESS'
                  AND loading_start_dttm >= now() - interval '{days} days'
                GROUP BY object_id
                """
            )
        ).mappings().all()

    return {int(r["object_id"]): {"avg": float(r["avg_minutes"]), "samples": int(r["samples"])} for r in rows}


def table_duration_map(table_ids, duration_by_id, min_samples):
    durations = {}
    all_values = []
    for fqn, ids in table_ids.items():
        values = []
        for tid in ids:
            data = duration_by_id.get(int(tid))
            if not data:
                continue
            if data["samples"] < min_samples:
                continue
            values.append(data["avg"])
        if values:
            avg_val = sum(values) / len(values)
            durations[fqn] = avg_val
            all_values.extend(values)
    default = median(all_values) if all_values else 5.0
    return durations, default


def get_layer(schema):
    return schema or "unknown"


def dm_tables_by_entity(meta_tables, table_entities):
    result = defaultdict(list)
    for fqn in meta_tables:
        schema, table = fqn.split(".", 1)
        if schema != "dm":
            continue
        for ent in table_entities.get(fqn, []):
            result[ent].append(fqn)
    return result


def upstream_closure(start_nodes, reverse):
    seen = set()
    stack = list(start_nodes)
    while stack:
        node = stack.pop()
        if node in seen:
            continue
        seen.add(node)
        for parent in reverse.get(node, []):
            if parent not in seen:
                stack.append(parent)
    return seen


def build_entity_dependencies(dm_by_entity, reverse, table_entities):
    deps = defaultdict(lambda: defaultdict(set))
    dm_closure = {}
    for entity, dm_tables in dm_by_entity.items():
        closure = upstream_closure(dm_tables, reverse)
        dm_closure[entity] = closure
        for table in closure:
            owners = table_entities.get(table, set())
            if entity in owners:
                continue
            for owner in owners:
                deps[entity][owner].add(table)
    return deps, dm_closure


def entity_schedule_levels(entities, deps):
    graph = defaultdict(set)
    indeg = {e: 0 for e in entities}
    for entity, external in deps.items():
        for owner in external.keys():
            if owner == entity:
                continue
            graph[owner].add(entity)
            indeg[entity] += 1

    queue = deque(sorted([e for e, d in indeg.items() if d == 0]))
    levels = []
    visited = set()
    while queue:
        level = list(queue)
        levels.append(level)
        queue = deque()
        for e in level:
            visited.add(e)
            for nxt in graph.get(e, []):
                indeg[nxt] -= 1
                if indeg[nxt] == 0:
                    queue.append(nxt)
    remaining = [e for e in entities if e not in visited]
    if remaining:
        levels.append(sorted(remaining))
    return levels


def estimate_entity_duration(entity, closure, durations, default_duration, concurrency):
    layers = defaultdict(list)
    for fqn in closure:
        schema, _ = fqn.split(".", 1)
        layers[get_layer(schema)].append(fqn)

    dm_calc_total = sum(durations.get(t, default_duration) for t in layers.get("dm_calc", []))
    dm_total = sum(durations.get(t, default_duration) for t in layers.get("dm", []))
    seq_total = dm_calc_total + dm_total

    parallel_total = 0.0
    for schema, tables in layers.items():
        if schema in ("dm_calc", "dm"):
            continue
        layer_sum = sum(durations.get(t, default_duration) for t in tables)
        parallel_total += layer_sum / max(concurrency, 1)

    return round(parallel_total + seq_total, 2), {
        "parallel_minutes": round(parallel_total, 2),
        "sequential_minutes": round(seq_total, 2),
        "tables": sum(len(v) for v in layers.values()),
    }


def recommend_entities(meta_tables, forward, table_entities, dm_entities):
    recommendations = []
    dm_nodes = set()
    dm_owner_map = {}
    for entity, tables in dm_entities.items():
        for t in tables:
            dm_nodes.add(t)
            dm_owner_map.setdefault(t, set()).add(entity)

    excluded_entities = {"DICT_LOADER"}

    for fqn in sorted(meta_tables):
        owners = table_entities.get(fqn, set())
        if not owners:
            continue
        if owners.intersection(excluded_entities):
            continue
        queue = deque([fqn])
        seen = set([fqn])
        hits = defaultdict(int)
        while queue:
            node = queue.popleft()
            if node in dm_nodes:
                for ent in dm_owner_map.get(node, []):
                    hits[ent] += 1
            for nxt in forward.get(node, []):
                if nxt not in seen:
                    seen.add(nxt)
                    queue.append(nxt)
        if not hits:
            continue
        best_ent = max(hits.items(), key=lambda item: item[1])[0]
        if best_ent in excluded_entities:
            continue
        current_owner = sorted(owners)[0]
        if best_ent != current_owner:
            recommendations.append(
                {
                    "table": fqn,
                    "current_owner": current_owner,
                    "suggested_owner": best_ent,
                    "dm_hits": hits[best_ent],
                    "owner_set": sorted(owners),
                }
            )
    return recommendations


def write_excel_report(path, report):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"

    headers = [
        "table",
        "schema",
        "current_owner",
        "suggested_owner",
        "dm_hits",
        "current_owner_est_min",
        "suggested_owner_est_min",
        "current_owner_ext_deps",
        "suggested_owner_ext_deps",
        "current_owner_dm_tables",
        "suggested_owner_dm_tables",
        "owner_set",
    ]
    ws.append(headers)

    entities = report.get("entities", {})
    for row in report.get("recommendations", []):
        table = row.get("table") or ""
        schema = table.split(".", 1)[0] if "." in table else ""
        current_owner = row.get("current_owner") or ""
        suggested_owner = row.get("suggested_owner") or ""
        current_stats = entities.get(current_owner, {})
        suggested_stats = entities.get(suggested_owner, {})
        current_ext = current_stats.get("external_dependencies", {}) if isinstance(current_stats, dict) else {}
        suggested_ext = suggested_stats.get("external_dependencies", {}) if isinstance(suggested_stats, dict) else {}
        ws.append(
            [
                table,
                schema,
                current_owner,
                suggested_owner,
                row.get("dm_hits"),
                current_stats.get("estimated_minutes"),
                suggested_stats.get("estimated_minutes"),
                sum(current_ext.values()) if isinstance(current_ext, dict) else None,
                sum(suggested_ext.values()) if isinstance(suggested_ext, dict) else None,
                current_stats.get("dm_tables"),
                suggested_stats.get("dm_tables"),
                ", ".join(row.get("owner_set") or []),
            ]
        )

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Schedule advisor for entity parallelization.")
    parser.add_argument("--days", type=int, default=14, help="History window in days.")
    parser.add_argument("--min-samples", type=int, default=3, help="Min samples to trust duration.")
    parser.add_argument("--concurrency", type=int, default=6, help="Max tables per entity in parallel.")
    parser.add_argument("--max-blocking", type=int, default=8, help="Max blocking tables per entity pair.")
    parser.add_argument("--output", type=str, default="", help="Write JSON report to file.")
    parser.add_argument("--excel", type=str, default="schedule_report.xlsx", help="Write Excel report to file.")
    args = parser.parse_args()

    entries, meta_tables = build_meta_entries()
    table_entities, table_ids, edges = build_table_graph(entries, meta_tables)
    forward, reverse = build_adjacency(edges)
    dm_by_entity = dm_tables_by_entity(meta_tables, table_entities)
    deps, closures = build_entity_dependencies(dm_by_entity, reverse, table_entities)

    engine = create_engine(DATABASE_URL)
    duration_by_id = load_durations(args.days, engine)
    table_durations, default_duration = table_duration_map(table_ids, duration_by_id, args.min_samples)

    entity_stats = {}
    for entity, closure in closures.items():
        total, detail = estimate_entity_duration(
            entity, closure, table_durations, default_duration, args.concurrency
        )
        entity_stats[entity] = {
            "estimated_minutes": total,
            "details": detail,
            "dm_tables": len(dm_by_entity.get(entity, [])),
            "external_dependencies": {k: len(v) for k, v in deps.get(entity, {}).items()},
        }

    levels = entity_schedule_levels(sorted(closures.keys()), deps)

    blocking = []
    for entity, external in deps.items():
        for owner, tables in external.items():
            ranked = sorted(
                tables,
                key=lambda t: table_durations.get(t, default_duration),
                reverse=True,
            )
            blocking.append(
                {
                    "needs": entity,
                    "from": owner,
                    "count": len(tables),
                    "top_tables": ranked[: args.max_blocking],
                }
            )

    recommendations = recommend_entities(meta_tables, forward, table_entities, dm_by_entity)
    recommendations = sorted(recommendations, key=lambda r: r["dm_hits"], reverse=True)[:200]

    report = {
        "days": args.days,
        "default_duration_minutes": round(default_duration, 2),
        "entities": entity_stats,
        "parallel_levels": levels,
        "blocking_dependencies": blocking,
        "recommendations": recommendations,
    }

    excel_path = Path(args.excel).resolve()
    write_excel_report(excel_path, report)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=True, indent=2)
    else:
        print(json.dumps(report, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
